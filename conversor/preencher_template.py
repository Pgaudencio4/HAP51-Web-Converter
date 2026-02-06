"""
Preencher Template RSECE a partir do ficheiro HAP 5.2

Lê os dados de 'Folha HAP_5_2_2026.xlsx' (ou similar) e preenche
o template 'HAP_Template_RSECE.xlsx' com os valores.

Usage:
    python preencher_template.py <input_hap52.xlsx> <output.xlsx>

Exemplo:
    python preencher_template.py "Folha HAP_5_2_2026.xlsx" "MeuProjecto.xlsx"
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os
import warnings
warnings.filterwarnings('ignore')

# Caminho do template
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'templates', 'HAP_Template_RSECE.xlsx')


def preencher_template(input_path, output_path):
    """Preenche o template com dados do ficheiro HAP 5.2"""

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print()

    # Abrir ficheiro de entrada
    print("A ler ficheiro de entrada...")
    wb_in = openpyxl.load_workbook(input_path, data_only=True)

    # Detectar aba de espaços
    if 'INPUT SPACES HAP' in wb_in.sheetnames:
        ws_in = wb_in['INPUT SPACES HAP']
        print("  Aba encontrada: INPUT SPACES HAP")
    elif 'Espacos' in wb_in.sheetnames:
        ws_in = wb_in['Espacos']
        print("  Aba encontrada: Espacos")
    else:
        print("ERRO: Não encontrei aba 'INPUT SPACES HAP' nem 'Espacos'")
        return False

    # Abrir template
    print("A abrir template...")
    wb_out = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_out = wb_out['Espacos']

    # Copiar dados dos espaços (linha 4 em diante)
    print("A copiar espaços...")
    row_in = 4
    row_out = 4
    count = 0

    while True:
        # Verificar se há dados (nome do espaço na coluna 1)
        name = ws_in.cell(row=row_in, column=1).value
        if not name or str(name).strip() == '':
            break

        # Copiar todas as colunas (1-147)
        for col in range(1, 148):
            value = ws_in.cell(row=row_in, column=col).value
            ws_out.cell(row=row_out, column=col, value=value)

        count += 1
        print(f"  {count}. {name}")

        row_in += 1
        row_out += 1

    print(f"\nTotal: {count} espaços copiados")

    # Copiar Walls se existir aba INPUT WALLS HAP
    if 'INPUT WALLS HAP' in wb_in.sheetnames:
        print("\nA copiar Walls...")
        ws_walls_in = wb_in['INPUT WALLS HAP']

        if 'Walls' in wb_out.sheetnames:
            ws_walls_out = wb_out['Walls']
        else:
            ws_walls_out = wb_out.create_sheet('Walls')
            # Copiar headers
            ws_walls_out.cell(1, 1, 'WALLS')
            ws_walls_out.cell(3, 1, 'Nome')
            ws_walls_out.cell(3, 2, 'U-Value (W/m²K)')
            ws_walls_out.cell(3, 3, 'Peso (kg/m²)')
            ws_walls_out.cell(3, 4, 'Espessura (m)')

        row_in = 4
        row_out = 4
        wall_count = 0
        while True:
            name = ws_walls_in.cell(row=row_in, column=1).value
            if not name or str(name).strip() == '':
                break
            for col in range(1, 5):
                value = ws_walls_in.cell(row=row_in, column=col).value
                ws_walls_out.cell(row=row_out, column=col, value=value)
            wall_count += 1
            print(f"  {name}")
            row_in += 1
            row_out += 1
        print(f"  Total: {wall_count} walls")

    # Copiar Roofs se existir aba INPUT ROOFS HAP
    if 'INPUT ROOFS HAP' in wb_in.sheetnames:
        print("\nA copiar Roofs...")
        ws_roofs_in = wb_in['INPUT ROOFS HAP']

        if 'Roofs' in wb_out.sheetnames:
            ws_roofs_out = wb_out['Roofs']
        else:
            ws_roofs_out = wb_out.create_sheet('Roofs')
            ws_roofs_out.cell(1, 1, 'ROOFS')
            ws_roofs_out.cell(3, 1, 'Nome')
            ws_roofs_out.cell(3, 2, 'U-Value (W/m²K)')
            ws_roofs_out.cell(3, 3, 'Peso (kg/m²)')
            ws_roofs_out.cell(3, 4, 'Espessura (m)')

        row_in = 4
        row_out = 4
        roof_count = 0
        while True:
            name = ws_roofs_in.cell(row=row_in, column=1).value
            if not name or str(name).strip() == '':
                break
            for col in range(1, 5):
                value = ws_roofs_in.cell(row=row_in, column=col).value
                ws_roofs_out.cell(row=row_out, column=col, value=value)
            roof_count += 1
            print(f"  {name}")
            row_in += 1
            row_out += 1
        print(f"  Total: {roof_count} roofs")

    # Copiar Windows se existir aba INPUT VIDROS HAP
    if 'INPUT VIDROS HAP' in wb_in.sheetnames:
        print("\nA copiar Windows...")
        ws_win_in = wb_in['INPUT VIDROS HAP']

        if 'Windows' in wb_out.sheetnames:
            ws_win_out = wb_out['Windows']
        else:
            ws_win_out = wb_out.create_sheet('Windows')
            ws_win_out.cell(1, 1, 'WINDOWS')
            ws_win_out.cell(3, 1, 'Nome')
            ws_win_out.cell(3, 2, 'U-Value (W/m²K)')
            ws_win_out.cell(3, 3, 'SHGC')
            ws_win_out.cell(3, 4, 'Altura (m)')
            ws_win_out.cell(3, 5, 'Largura (m)')

        # No HAP 5.2, os dados começam na linha 6 (linha 5 tem headers REF_ID, U, g, alt, Lar)
        row_in = 6
        row_out = 4
        win_count = 0
        while True:
            name = ws_win_in.cell(row=row_in, column=1).value
            if not name or str(name).strip() == '':
                break
            for col in range(1, 6):
                value = ws_win_in.cell(row=row_in, column=col).value
                ws_win_out.cell(row=row_out, column=col, value=value)
            win_count += 1
            row_in += 1
            row_out += 1
        print(f"  Total: {win_count} windows")

    # Guardar
    print(f"\nA guardar: {output_path}")
    wb_out.save(output_path)

    wb_in.close()
    wb_out.close()

    print("\nConcluído!")
    return True


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        print("\nUsage:")
        print("  python preencher_template.py <input.xlsx> <output.xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.exists(input_file):
        print(f"ERRO: Ficheiro não encontrado: {input_file}")
        sys.exit(1)

    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERRO: Template não encontrado: {TEMPLATE_PATH}")
        sys.exit(1)

    preencher_template(input_file, output_file)
