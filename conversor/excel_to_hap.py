"""
Converter Excel para ficheiro HAP 5.1 (.E3A)
Lê o Excel com os espaços e gera o ficheiro binário.

Usage:
    python excel_to_hap.py <input.xlsx> <modelo_base.E3A> <output.E3A>

Exemplo:
    python excel_to_hap.py MeusDados.xlsx Modelo_RSECE.E3A Output.E3A
"""
import zipfile
import tempfile
import shutil
import struct
import os
import sys
import math
import openpyxl

# =============================================================================
# CONFIGURAÇÃO (valores default, podem ser substituídos por argumentos)
# =============================================================================

# Ficheiro base (template) - deve ter a estrutura MDB correcta
BASE_FILE = None

# Excel de entrada
EXCEL_FILE = None

# Ficheiro de saída
OUTPUT_FILE = None

# =============================================================================
# CONSTANTES
# =============================================================================

RECORD_SIZE = 682
WALL_BLOCK_SIZE = 34
WALL_BLOCK_START = 72
ROOF_BLOCK_SIZE = 24
ROOF_BLOCK_START = 344

DIRECTION_CODES = {
    'N': 1, 'NNE': 2, 'NE': 3, 'ENE': 4,
    'E': 5, 'ESE': 6, 'SE': 7, 'SSE': 8,
    'S': 9, 'SSW': 10, 'SW': 11, 'WSW': 12,
    'W': 13, 'WNW': 14, 'NW': 15, 'NNW': 16,
    'H': 17, 'HORIZ': 17, 'HORIZONTAL': 17  # Horizontal para coberturas
}

ACTIVITY_CODES = {
    'User-defined': 0,
    'Seated at Rest': 1,
    'Office Work': 2,
    'Sedentary Work': 3,
    'Medium Work': 4,
    'Heavy Work': 5,
    'Dancing': 6,
    'Athletics': 7,
}

FIXTURE_CODES = {
    'Recessed Unvented': 0,
    'Vented to Return Air': 1,
    'Vented to Supply & Return': 2,
    'Surface Mount/Pendant': 3,
}

FLOOR_TYPE_CODES = {
    'Floor Above Cond Space': 1,
    'Floor Above Uncond Space': 2,
    'Slab Floor On Grade': 3,
    'Slab Floor Below Grade': 4,
}

OA_UNIT_CODES = {
    'L/s': 1,
    'L/s/m2': 2,
    'L/s/person': 3,
    '%': 4,
}

# OA encoding - exact closed-form formula (2026-02-05)
# HAP 5.1 uses a piecewise-linear "fast_exp2" approximation:
#   y = Y0 * fast_exp2(k * (x - 4))
#   k=4 for x<4 (base 16), k=2 for x>=4 (base 4)
#   Y0 = 512 CFM in L/s = 241.637...
_OA_Y0 = 512.0 * (28.316846592 / 60.0)

def _fast_exp2(t):
    n = math.floor(t)
    f = t - n
    return (2.0 ** n) * (1.0 + f)

def _fast_log2(v):
    n = math.floor(math.log2(v))
    f = v / (2.0 ** n) - 1.0
    if f < 0:
        n -= 1
        f = v / (2.0 ** n) - 1.0
    if f >= 1.0:
        n += 1
        f = v / (2.0 ** n) - 1.0
    return n + f

# =============================================================================
# CONVERSÕES
# =============================================================================

def m2_to_ft2(m2):
    if m2 is None or m2 == '':
        return 0.0
    return float(m2) * 10.7639

def m_to_ft(m):
    if m is None or m == '':
        return 0.0
    return float(m) * 3.28084

def kg_m2_to_lb_ft2(kg):
    if kg is None or kg == '':
        return 0.0
    return float(kg) / 4.8824

def c_to_f(c):
    if c is None or c == '':
        return 0.0
    return float(c) * 1.8 + 32

def w_to_btu(w):
    if w is None or w == '':
        return 0.0
    return float(w) * 3.412

def w_m2_to_w_ft2(w):
    if w is None or w == '':
        return 0.0
    return float(w) / 10.764

def u_si_to_ip(u):
    if u is None or u == '':
        return 0.0
    return float(u) / 5.678

def r_si_to_ip(r):
    if r is None or r == '':
        return 0.0
    return float(r) * 5.678

def encode_oa(value, unit_code):
    if value is None or value == '' or float(value) <= 0:
        return 0.0
    v = float(value) / _OA_Y0
    t = _fast_log2(v)
    if t < 0:
        return t / 4.0 + 4.0
    else:
        return t / 2.0 + 4.0

def safe_int(val, default=0):
    if val is None or val == '':
        return default
    try:
        return int(val)
    except:
        return default

def safe_float(val, default=0.0):
    if val is None or val == '':
        return default
    try:
        return float(val)
    except:
        return default

# =============================================================================
# CRIAR WALL/ROOF ASSEMBLIES COM LAYERS
# =============================================================================

# Constantes para assemblies
ASSEMBLY_SIZE = 3187
LAYER_SIZE = 281
LAYER_START = 377  # Onde começa o primeiro layer (Inside surface)

# R-Values fixos das superfícies (em IP: ft²·°F·hr/BTU)
R_INSIDE_IP = 0.68    # 0.12 SI
R_OUTSIDE_IP = 0.33   # 0.06 SI

def fill_assembly_layers(data, offset, u_value_si, weight_si, absorptivity=0.9):
    """
    Preenche as layers de um Wall/Roof assembly para obter o U-Value e Weight desejados.

    Args:
        data: bytearray do ficheiro DAT
        offset: offset do início do assembly
        u_value_si: U-Value desejado em W/m²K
        weight_si: Weight desejado em kg/m²
        absorptivity: Absorptivity (0-1), default 0.9
    """
    # Converter para IP
    u_value_ip = u_value_si / 5.678  # W/m²K -> BTU/hr·ft²·°F
    weight_ip = weight_si * 0.2048   # kg/m² -> lb/ft²

    # Calcular R-Value do material
    r_total_ip = 1.0 / u_value_ip if u_value_ip > 0 else 10.0
    r_material_ip = r_total_ip - R_INSIDE_IP - R_OUTSIDE_IP
    if r_material_ip < 0.01:
        r_material_ip = 0.01  # Mínimo

    # Usar thickness fixo de 0.1 ft (~30mm) e calcular density para dar o weight
    thickness_ft = 0.1
    density_lb_ft3 = weight_ip / thickness_ft if thickness_ft > 0 else 100.0

    # Absorptivity (offset 255)
    struct.pack_into('<f', data, offset + 255, absorptivity)

    # Surface Color = 2 (Dark) - offset 259
    data[offset + 259] = 2

    # Layer 0: Inside surface resistance (offset 377)
    layer0 = offset + LAYER_START
    inside_name = b'Inside surface resistance' + b' ' * (255 - 25)
    data[layer0:layer0+255] = inside_name
    struct.pack_into('<f', data, layer0 + 257, 0.0)        # Thickness
    struct.pack_into('<f', data, layer0 + 261, 0.0)        # Conductivity
    struct.pack_into('<f', data, layer0 + 265, 0.0)        # Density
    struct.pack_into('<f', data, layer0 + 269, 0.0)        # Specific Heat
    struct.pack_into('<f', data, layer0 + 273, R_INSIDE_IP) # R-Value
    struct.pack_into('<f', data, layer0 + 277, 0.0)        # Weight

    # Layer 1: Material principal (offset 377 + 281 = 658)
    layer1 = offset + LAYER_START + LAYER_SIZE
    material_name = b'Insulation' + b' ' * (255 - 10)
    data[layer1:layer1+255] = material_name
    struct.pack_into('<f', data, layer1 + 257, thickness_ft)      # Thickness
    struct.pack_into('<f', data, layer1 + 261, 0.02)              # Conductivity
    struct.pack_into('<f', data, layer1 + 265, density_lb_ft3)    # Density
    struct.pack_into('<f', data, layer1 + 269, 0.2)               # Specific Heat
    struct.pack_into('<f', data, layer1 + 273, r_material_ip)     # R-Value
    struct.pack_into('<f', data, layer1 + 277, weight_ip)         # Weight

    # Layer 2: Outside surface resistance (offset 377 + 281*2 = 939)
    layer2 = offset + LAYER_START + LAYER_SIZE * 2
    outside_name = b'Outside surface resistance' + b' ' * (255 - 26)
    data[layer2:layer2+255] = outside_name
    struct.pack_into('<f', data, layer2 + 257, 0.0)         # Thickness
    struct.pack_into('<f', data, layer2 + 261, 0.0)         # Conductivity
    struct.pack_into('<f', data, layer2 + 265, 0.0)         # Density
    struct.pack_into('<f', data, layer2 + 269, 0.0)         # Specific Heat
    struct.pack_into('<f', data, layer2 + 273, R_OUTSIDE_IP) # R-Value
    struct.pack_into('<f', data, layer2 + 277, 0.0)         # Weight

    # Limpar layers 3+ (preencher com nulls para o HAP não os detectar)
    for layer_idx in range(3, 9):
        layer_off = offset + LAYER_START + LAYER_SIZE * layer_idx
        if layer_off + LAYER_SIZE <= offset + ASSEMBLY_SIZE:
            data[layer_off:layer_off+255] = b'\x00' * 255
            for i in range(255, LAYER_SIZE):
                data[layer_off + i] = 0

# =============================================================================
# LER EXCEL
# =============================================================================

def read_excel_spaces(excel_path):
    """Lê os espaços do Excel e retorna lista de dicionários.

    Formato do template (HAP_Exemplo_5Espacos.xlsx / HAP_Template_RSECE.xlsx):
    - Linhas 1-2: Categorias (GENERAL, INTERNALS, etc.)
    - Linha 3: Headers das colunas
    - Linha 4+: Dados

    Colunas principais:
    1: Space Name, 2: Floor Area, 3: Ceiling Ht, 4: Building Wt, 5: OA valor, 6: OA Unit
    7: Occupancy, 8: Activity, 9: Sensible, 10: Latent, 11: People Schedule
    12: Task Light, 13: General Light, 14: Fixture, 15: Ballast, 16: Light Schedule
    17: Equipment, 18: Equip Schedule
    19: Misc Sens, 20: Misc Lat, 21: Misc Sens Sch, 22: Misc Lat Sch
    23-25: ACH values (Design Clg, Design Htg, Energy)
    26-38: Floor data
    39-44: Ceiling partition
    45-50: Wall partition
    51+: Walls (8 x 9 cols), Roofs (4 x 6 cols)
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Espacos']

    spaces = []

    # Dados começam na linha 4 (linhas 1-3 são headers)
    for row in range(4, ws.max_row + 1):
        # Verificar se a linha tem dados (nome do espaço na coluna 1)
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == '':
            continue

        space = {}

        # GENERAL (cols 1-6)
        space['name'] = str(name)[:24]
        space['area'] = ws.cell(row=row, column=2).value
        space['height'] = ws.cell(row=row, column=3).value
        space['weight'] = ws.cell(row=row, column=4).value
        space['oa'] = ws.cell(row=row, column=5).value
        space['oa_unit'] = ws.cell(row=row, column=6).value

        # INTERNALS - People (cols 7-11)
        space['occupancy'] = ws.cell(row=row, column=7).value
        space['activity'] = ws.cell(row=row, column=8).value
        space['sensible'] = ws.cell(row=row, column=9).value
        space['latent'] = ws.cell(row=row, column=10).value
        space['people_sch'] = ws.cell(row=row, column=11).value

        # INTERNALS - Lighting (cols 12-16)
        space['task_light'] = ws.cell(row=row, column=12).value
        space['general_light'] = ws.cell(row=row, column=13).value
        space['fixture'] = ws.cell(row=row, column=14).value
        space['ballast'] = ws.cell(row=row, column=15).value
        space['light_sch'] = ws.cell(row=row, column=16).value

        # INTERNALS - Equipment (cols 17-18)
        space['equipment'] = ws.cell(row=row, column=17).value
        space['equip_sch'] = ws.cell(row=row, column=18).value

        # INTERNALS - Misc (cols 19-22)
        space['misc_sens'] = ws.cell(row=row, column=19).value
        space['misc_lat'] = ws.cell(row=row, column=20).value
        space['misc_sens_sch'] = ws.cell(row=row, column=21).value
        space['misc_lat_sch'] = ws.cell(row=row, column=22).value

        # INFILTRATION (cols 23-26)
        space['infil_method'] = ws.cell(row=row, column=23).value
        space['ach_clg'] = ws.cell(row=row, column=24).value
        space['ach_htg'] = ws.cell(row=row, column=25).value
        space['ach_energy'] = ws.cell(row=row, column=26).value

        # FLOORS (cols 27-39)
        space['floor_type'] = ws.cell(row=row, column=27).value
        space['floor_area'] = ws.cell(row=row, column=28).value
        space['floor_u'] = ws.cell(row=row, column=29).value
        space['floor_perim'] = ws.cell(row=row, column=30).value
        space['floor_edge_r'] = ws.cell(row=row, column=31).value
        space['floor_depth'] = ws.cell(row=row, column=32).value
        space['bsmt_u'] = ws.cell(row=row, column=33).value
        space['wall_ins_r'] = ws.cell(row=row, column=34).value
        space['ins_depth'] = ws.cell(row=row, column=35).value
        space['floor_unc_max'] = ws.cell(row=row, column=36).value
        space['floor_out_max'] = ws.cell(row=row, column=37).value
        space['floor_unc_min'] = ws.cell(row=row, column=38).value
        space['floor_out_min'] = ws.cell(row=row, column=39).value

        # PARTITIONS - Ceiling (cols 40-45)
        space['ceil_area'] = ws.cell(row=row, column=40).value
        space['ceil_u'] = ws.cell(row=row, column=41).value
        space['ceil_unc_max'] = ws.cell(row=row, column=42).value
        space['ceil_out_max'] = ws.cell(row=row, column=43).value
        space['ceil_unc_min'] = ws.cell(row=row, column=44).value
        space['ceil_out_min'] = ws.cell(row=row, column=45).value

        # PARTITIONS - Wall (cols 46-51)
        space['wall_part_area'] = ws.cell(row=row, column=46).value
        space['wall_part_u'] = ws.cell(row=row, column=47).value
        space['wall_unc_max'] = ws.cell(row=row, column=48).value
        space['wall_out_max'] = ws.cell(row=row, column=49).value
        space['wall_unc_min'] = ws.cell(row=row, column=50).value
        space['wall_out_min'] = ws.cell(row=row, column=51).value

        # WALLS (8 walls x 9 cols = cols 52-123)
        space['walls'] = []
        wall_start = 52
        for w in range(8):
            col = wall_start + w * 9
            wall = {
                'exposure': ws.cell(row=row, column=col).value,
                'area': ws.cell(row=row, column=col+1).value,
                'type': ws.cell(row=row, column=col+2).value,
                'win1': ws.cell(row=row, column=col+3).value,
                'win1_qty': ws.cell(row=row, column=col+4).value,
                'win2': ws.cell(row=row, column=col+5).value,
                'win2_qty': ws.cell(row=row, column=col+6).value,
                'door': ws.cell(row=row, column=col+7).value,
                'door_qty': ws.cell(row=row, column=col+8).value,
            }
            space['walls'].append(wall)

        # ROOFS (4 roofs x 6 cols = cols 124-147)
        space['roofs'] = []
        roof_start = 124
        for r in range(4):
            col = roof_start + r * 6
            roof = {
                'exposure': ws.cell(row=row, column=col).value,
                'area': ws.cell(row=row, column=col+1).value,
                'slope': ws.cell(row=row, column=col+2).value,
                'type': ws.cell(row=row, column=col+3).value,
                'sky': ws.cell(row=row, column=col+4).value,
                'sky_qty': ws.cell(row=row, column=col+5).value,
            }
            space['roofs'].append(roof)

        spaces.append(space)
        print(f"  Lido: {space['name']}")

    # Ler sheet Tipos para mapear nomes -> IDs (opcional)
    types = {
        'walls': {},
        'windows': {},
        'doors': {},
        'roofs': {},
        'schedules': {},
    }

    if 'Tipos' in wb.sheetnames:
        ws_tipos = wb['Tipos']

        # Wall Types (cols 1-2)
        for row in range(3, ws_tipos.max_row + 1):
            id_val = ws_tipos.cell(row=row, column=1).value
            name = ws_tipos.cell(row=row, column=2).value
            if id_val and name:
                types['walls'][str(name).strip()] = int(id_val)

        # Window Types (cols 4-5)
        for row in range(3, ws_tipos.max_row + 1):
            id_val = ws_tipos.cell(row=row, column=4).value
            name = ws_tipos.cell(row=row, column=5).value
            if id_val and name:
                types['windows'][str(name).strip()] = int(id_val)

        # Door Types (cols 7-8)
        for row in range(3, ws_tipos.max_row + 1):
            id_val = ws_tipos.cell(row=row, column=7).value
            name = ws_tipos.cell(row=row, column=8).value
            if id_val and name:
                types['doors'][str(name).strip()] = int(id_val)

        # Roof Types (cols 10-11)
        for row in range(3, ws_tipos.max_row + 1):
            id_val = ws_tipos.cell(row=row, column=10).value
            name = ws_tipos.cell(row=row, column=11).value
            if id_val and name:
                types['roofs'][str(name).strip()] = int(id_val)

        # Schedules (cols 13-14) - IGNORADO!
        # Os IDs na sheet Tipos podem estar errados.
        # Os schedules sao carregados directamente do modelo base (HAP51SCH.DAT)
        # para garantir que os IDs estao correctos.
        pass

    # Ler tipos das sheets Windows, Walls, Roofs (dados para criar no HAP)
    type_definitions = {
        'windows': [],
        'walls': [],
        'roofs': [],
    }

    # Sheet Windows: Nome, U-Value, SHGC, Altura, Largura
    if 'Windows' in wb.sheetnames:
        ws_win = wb['Windows']
        for row in range(4, ws_win.max_row + 1):
            name = ws_win.cell(row=row, column=1).value
            if name and str(name).strip():
                win_def = {
                    'name': str(name).strip(),
                    'u_value': safe_float(ws_win.cell(row=row, column=2).value, 2.8),
                    'shgc': safe_float(ws_win.cell(row=row, column=3).value, 0.7),
                    'height': safe_float(ws_win.cell(row=row, column=4).value, 1.2),
                    'width': safe_float(ws_win.cell(row=row, column=5).value, 1.0),
                }
                type_definitions['windows'].append(win_def)

    # Sheet Walls: Nome, U-Value, Peso, Espessura, Absorptivity
    if 'Walls' in wb.sheetnames:
        ws_wal = wb['Walls']
        for row in range(4, ws_wal.max_row + 1):
            name = ws_wal.cell(row=row, column=1).value
            if name and str(name).strip():
                wall_def = {
                    'name': str(name).strip(),
                    'u_value': safe_float(ws_wal.cell(row=row, column=2).value, 0.5),
                    'weight': safe_float(ws_wal.cell(row=row, column=3).value, 200),
                    'thickness': safe_float(ws_wal.cell(row=row, column=4).value, 0.3),
                    'absorptivity': safe_float(ws_wal.cell(row=row, column=5).value, 0.9),
                }
                type_definitions['walls'].append(wall_def)

    # Sheet Roofs: Nome, U-Value, Peso, Espessura, Absorptivity
    if 'Roofs' in wb.sheetnames:
        ws_rof = wb['Roofs']
        for row in range(4, ws_rof.max_row + 1):
            name = ws_rof.cell(row=row, column=1).value
            if name and str(name).strip():
                roof_def = {
                    'name': str(name).strip(),
                    'u_value': safe_float(ws_rof.cell(row=row, column=2).value, 0.4),
                    'weight': safe_float(ws_rof.cell(row=row, column=3).value, 300),
                    'thickness': safe_float(ws_rof.cell(row=row, column=4).value, 0.3),
                    'absorptivity': safe_float(ws_rof.cell(row=row, column=5).value, 0.9),
                }
                type_definitions['roofs'].append(roof_def)

    return spaces, types, type_definitions

def normalize_name(name):
    """Normaliza nome para matching (remove variações comuns)."""
    if not name:
        return ''
    n = str(name).lower().strip()
    n = n.replace('estrelas', 'estrela')
    n = n.replace('  ', ' ')
    return n

def get_type_id(name, type_dict, default=1):
    """Obtém o ID de um tipo pelo nome."""
    if name is None or name == '':
        return 0
    name_str = str(name).strip()

    # Match exacto
    if name_str in type_dict:
        return type_dict[name_str]

    # Match normalizado
    name_norm = normalize_name(name_str)
    for key, val in type_dict.items():
        if normalize_name(key) == name_norm:
            return val

    # Tentar encontrar por substring
    for key, val in type_dict.items():
        if name_str.lower() in key.lower() or key.lower() in name_str.lower():
            return val
    return default

# =============================================================================
# CRIAR TIPOS BINÁRIOS (Windows, Walls, Roofs)
# =============================================================================

WINDOW_RECORD_SIZE = 555
WALL_RECORD_SIZE = 910  # Aproximado baseado em análise
ROOF_RECORD_SIZE = 910  # Igual ao wall

def create_window_binary(win_def, template_record):
    """Cria registo binário de 555 bytes para uma window."""
    data = bytearray(template_record)

    # Nome (0-255)
    name_bytes = win_def['name'].encode('latin-1')[:255].ljust(255, b' ')
    data[0:255] = name_bytes

    # Altura (257-260) em ft
    struct.pack_into('<f', data, 257, m_to_ft(win_def['height']))

    # Largura (261-264) em ft
    struct.pack_into('<f', data, 261, m_to_ft(win_def['width']))

    # U-Value (269-272) em BTU/hr.ft2.F
    struct.pack_into('<f', data, 269, u_si_to_ip(win_def['u_value']))

    # SHGC (273-276)
    struct.pack_into('<f', data, 273, win_def['shgc'])

    return bytes(data)

# =============================================================================
# CRIAR ESPAÇO BINÁRIO
# =============================================================================

def create_space_binary(space, types, template_record):
    """Cria o registo binário de 682 bytes para um espaço."""
    # Começar com o template
    data = bytearray(template_record)

    # Nome (0-24)
    name_bytes = space['name'].encode('latin-1')[:24].ljust(24, b'\x00')
    data[0:24] = name_bytes

    # Área, Altura, Peso (24-36)
    struct.pack_into('<f', data, 24, m2_to_ft2(space['area']))
    struct.pack_into('<f', data, 28, m_to_ft(space['height']))
    struct.pack_into('<f', data, 32, kg_m2_to_lb_ft2(space['weight']))

    # OA (46-52)
    oa_unit = OA_UNIT_CODES.get(space['oa_unit'], 3)
    oa_internal = encode_oa(space['oa'], oa_unit)
    struct.pack_into('<f', data, 46, oa_internal)
    struct.pack_into('<H', data, 50, oa_unit)

    # WALLS (72-344)
    for i in range(8):
        wall_start = WALL_BLOCK_START + i * WALL_BLOCK_SIZE
        # Limpar bloco
        for j in range(WALL_BLOCK_SIZE):
            data[wall_start + j] = 0

        if i < len(space['walls']):
            wall = space['walls'][i]
            exp = wall.get('exposure')
            wall_type_id = get_type_id(wall.get('type'), types['walls'], 0)
            # Só escrever wall block se tiver exposição E tipo válido
            if exp and exp in DIRECTION_CODES and wall_type_id > 0:
                struct.pack_into('<H', data, wall_start, DIRECTION_CODES[exp])
                struct.pack_into('<f', data, wall_start + 2, m2_to_ft2(wall.get('area')))
                struct.pack_into('<H', data, wall_start + 6, wall_type_id)
                struct.pack_into('<H', data, wall_start + 8, get_type_id(wall.get('win1'), types['windows'], 0))
                struct.pack_into('<H', data, wall_start + 10, get_type_id(wall.get('win2'), types['windows'], 0))
                struct.pack_into('<H', data, wall_start + 12, safe_int(wall.get('win1_qty')))
                struct.pack_into('<H', data, wall_start + 14, safe_int(wall.get('win2_qty')))
                struct.pack_into('<H', data, wall_start + 16, get_type_id(wall.get('door'), types['doors'], 0))
                struct.pack_into('<H', data, wall_start + 18, safe_int(wall.get('door_qty')))

    # ROOFS (344-440)
    for i in range(4):
        roof_start = ROOF_BLOCK_START + i * ROOF_BLOCK_SIZE
        # Limpar bloco
        for j in range(ROOF_BLOCK_SIZE):
            data[roof_start + j] = 0

        if i < len(space['roofs']):
            roof = space['roofs'][i]
            exp = roof.get('exposure')
            roof_type_id = get_type_id(roof.get('type'), types['roofs'], 0)
            # CRÍTICO: só escrever roof block se tiver exposição E tipo válido
            # Exposição sem tipo (type=0) causa crash no HAP (division by zero)
            if exp and exp in DIRECTION_CODES and roof_type_id > 0:
                struct.pack_into('<H', data, roof_start, DIRECTION_CODES[exp])
                struct.pack_into('<H', data, roof_start + 2, safe_int(roof.get('slope')))
                struct.pack_into('<f', data, roof_start + 4, m2_to_ft2(roof.get('area')))
                struct.pack_into('<H', data, roof_start + 8, roof_type_id)
                struct.pack_into('<H', data, roof_start + 10, get_type_id(roof.get('sky'), types['windows'], 0))
                struct.pack_into('<H', data, roof_start + 12, safe_int(roof.get('sky_qty')))

    # PARTITIONS - Ceiling (440-466) - só escreve se houver dados
    if space.get('ceil_area'):
        struct.pack_into('<H', data, 440, 1)  # Type 1 = Ceiling
        struct.pack_into('<f', data, 442, m2_to_ft2(space.get('ceil_area')))
        if space.get('ceil_u'):
            struct.pack_into('<f', data, 446, u_si_to_ip(space.get('ceil_u')))
        if space.get('ceil_unc_max'):
            struct.pack_into('<f', data, 450, c_to_f(space.get('ceil_unc_max')))
        if space.get('ceil_out_max'):
            struct.pack_into('<f', data, 454, c_to_f(space.get('ceil_out_max')))
        if space.get('ceil_unc_min'):
            struct.pack_into('<f', data, 458, c_to_f(space.get('ceil_unc_min')))
        if space.get('ceil_out_min'):
            struct.pack_into('<f', data, 462, c_to_f(space.get('ceil_out_min')))

    # PARTITIONS - Wall (466-492) - só escreve se houver dados
    if space.get('wall_part_area'):
        struct.pack_into('<H', data, 466, 2)  # Type 2 = Wall
        struct.pack_into('<f', data, 468, m2_to_ft2(space.get('wall_part_area')))
        if space.get('wall_part_u'):
            struct.pack_into('<f', data, 472, u_si_to_ip(space.get('wall_part_u')))
        if space.get('wall_unc_max'):
            struct.pack_into('<f', data, 476, c_to_f(space.get('wall_unc_max')))
        if space.get('wall_out_max'):
            struct.pack_into('<f', data, 480, c_to_f(space.get('wall_out_max')))
        if space.get('wall_unc_min'):
            struct.pack_into('<f', data, 484, c_to_f(space.get('wall_unc_min')))
        if space.get('wall_out_min'):
            struct.pack_into('<f', data, 488, c_to_f(space.get('wall_out_min')))

    # FLOOR (492-542) - só escreve se houver dados
    if space.get('floor_type'):
        floor_type = FLOOR_TYPE_CODES.get(space.get('floor_type'), 2)
        struct.pack_into('<H', data, 492, floor_type)
        if space.get('floor_area'):
            struct.pack_into('<f', data, 494, m2_to_ft2(space.get('floor_area')))
        if space.get('floor_u'):
            struct.pack_into('<f', data, 498, u_si_to_ip(space.get('floor_u')))
        if space.get('floor_perim'):
            struct.pack_into('<f', data, 502, m_to_ft(space.get('floor_perim')))
        if space.get('floor_edge_r'):
            struct.pack_into('<f', data, 506, r_si_to_ip(space.get('floor_edge_r')))
        if space.get('floor_depth'):
            struct.pack_into('<f', data, 510, m_to_ft(space.get('floor_depth')))
        if space.get('bsmt_u'):
            struct.pack_into('<f', data, 514, u_si_to_ip(space.get('bsmt_u')))
        if space.get('wall_ins_r'):
            struct.pack_into('<f', data, 518, r_si_to_ip(space.get('wall_ins_r')))
        if space.get('ins_depth'):
            struct.pack_into('<f', data, 522, m_to_ft(space.get('ins_depth')))
        if space.get('floor_unc_max'):
            struct.pack_into('<f', data, 526, c_to_f(space.get('floor_unc_max')))
        if space.get('floor_out_max'):
            struct.pack_into('<f', data, 530, c_to_f(space.get('floor_out_max')))
        if space.get('floor_unc_min'):
            struct.pack_into('<f', data, 534, c_to_f(space.get('floor_unc_min')))
        if space.get('floor_out_min'):
            struct.pack_into('<f', data, 538, c_to_f(space.get('floor_out_min')))

    # INFILTRATION (554-572)
    # Offsets 554, 560, 566 são FLAGS de modo (2 = ACH mode)
    # Seguidos de float com o valor ACH
    ACH_MODE_FLAG = 2
    struct.pack_into('<H', data, 554, ACH_MODE_FLAG)  # Design Cooling flag
    struct.pack_into('<f', data, 556, safe_float(space.get('ach_clg')))
    struct.pack_into('<H', data, 560, ACH_MODE_FLAG)  # Design Heating flag
    struct.pack_into('<f', data, 562, safe_float(space.get('ach_htg')))
    struct.pack_into('<H', data, 566, ACH_MODE_FLAG)  # Energy Analysis flag
    struct.pack_into('<f', data, 568, safe_float(space.get('ach_energy')))

    # PEOPLE (580-596)
    struct.pack_into('<f', data, 580, safe_float(space.get('occupancy')))
    activity = ACTIVITY_CODES.get(space.get('activity'), 3)
    struct.pack_into('<H', data, 584, activity)
    struct.pack_into('<f', data, 586, w_to_btu(space.get('sensible')))
    struct.pack_into('<f', data, 590, w_to_btu(space.get('latent')))
    people_sch = get_type_id(space.get('people_sch'), types['schedules'])
    struct.pack_into('<H', data, 594, people_sch)

    # LIGHTING (600-623)
    struct.pack_into('<f', data, 600, safe_float(space.get('task_light')))
    fixture = FIXTURE_CODES.get(space.get('fixture'), 0)
    struct.pack_into('<H', data, 604, fixture)
    struct.pack_into('<f', data, 606, safe_float(space.get('general_light')))
    struct.pack_into('<f', data, 610, safe_float(space.get('ballast'), 1.0))
    light_sch = get_type_id(space.get('light_sch'), types['schedules'])
    struct.pack_into('<H', data, 616, light_sch)  # Offset 616 (confirmado no README)

    # MISC (632-646)
    struct.pack_into('<f', data, 632, w_to_btu(space.get('misc_sens')))
    struct.pack_into('<f', data, 636, w_to_btu(space.get('misc_lat')))
    misc_sens_sch = get_type_id(space.get('misc_sens_sch'), types['schedules'], 0)
    struct.pack_into('<H', data, 640, misc_sens_sch)
    misc_lat_sch = get_type_id(space.get('misc_lat_sch'), types['schedules'], 0)
    struct.pack_into('<H', data, 644, misc_lat_sch)

    # EQUIPMENT (656-662)
    struct.pack_into('<f', data, 656, w_m2_to_w_ft2(space.get('equipment')))
    equip_sch = get_type_id(space.get('equip_sch'), types['schedules'])
    struct.pack_into('<H', data, 660, equip_sch)

    return bytes(data)

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("EXCEL -> HAP 5.1 CONVERTER")
    print("=" * 60)

    # Verificar ficheiros
    if not os.path.exists(BASE_FILE):
        print(f"ERRO: Ficheiro base não encontrado: {BASE_FILE}")
        return

    if not os.path.exists(EXCEL_FILE):
        print(f"ERRO: Excel não encontrado: {EXCEL_FILE}")
        return

    print(f"\nBase: {BASE_FILE}")
    print(f"Excel: {EXCEL_FILE}")
    print(f"Output: {OUTPUT_FILE}")

    # Ler Excel
    print("\n--- Lendo Excel ---")
    spaces, types, type_definitions = read_excel_spaces(EXCEL_FILE)
    print(f"\nTotal: {len(spaces)} espaços")
    print(f"Types definidos: {len(type_definitions['walls'])} walls, {len(type_definitions['windows'])} windows, {len(type_definitions['roofs'])} roofs")

    # Extrair base
    print("\n--- Extraindo base ---")
    temp_dir = tempfile.mkdtemp()

    try:
        with zipfile.ZipFile(BASE_FILE, 'r') as zf:
            zf.extractall(temp_dir)

        # Ler template record
        spc_path = os.path.join(temp_dir, 'HAP51SPC.DAT')
        with open(spc_path, 'rb') as f:
            spc_data = f.read()

        default_record = spc_data[0:RECORD_SIZE]  # Record 0 (default)
        # Usar o default como template se só houver 1 registo
        if len(spc_data) >= RECORD_SIZE * 2:
            template_record = spc_data[RECORD_SIZE:RECORD_SIZE*2]
        else:
            template_record = default_record

        print(f"Template record size: {len(template_record)} bytes")

        # Ler schedules do modelo base (HAP51SCH.DAT)
        sch_path = os.path.join(temp_dir, 'HAP51SCH.DAT')
        if os.path.exists(sch_path):
            with open(sch_path, 'rb') as f:
                sch_data = f.read()

            SCHEDULE_RECORD_SIZE = 792
            num_schedules = len(sch_data) // SCHEDULE_RECORD_SIZE

            print(f"Schedules no modelo base: {num_schedules}")

            # Adicionar schedules do modelo aos types (se não existirem já)
            for i in range(num_schedules):
                offset = i * SCHEDULE_RECORD_SIZE
                name_bytes = sch_data[offset:offset+24]
                name = name_bytes.rstrip(b'\x00').decode('latin-1', errors='ignore').strip()
                if name and name not in types['schedules']:
                    types['schedules'][name] = i

            print(f"Total schedules disponíveis: {len(types['schedules'])}")

        # Criar novos tipos de Windows se definidos no Excel
        if type_definitions['windows']:
            print("\n--- Criando Windows ---")
            win_path = os.path.join(temp_dir, 'HAP51WIN.DAT')
            with open(win_path, 'rb') as f:
                win_data = bytearray(f.read())

            num_existing_windows = len(win_data) // WINDOW_RECORD_SIZE
            win_template = win_data[0:WINDOW_RECORD_SIZE]  # Primeiro como template

            for win_def in type_definitions['windows']:
                new_win = create_window_binary(win_def, win_template)
                win_data.extend(new_win)
                # Mapear nome -> ID (índice)
                new_id = len(win_data) // WINDOW_RECORD_SIZE - 1
                types['windows'][win_def['name']] = new_id
                print(f"  Window {new_id}: {win_def['name']}")

            with open(win_path, 'wb') as f:
                f.write(bytes(win_data))

            print(f"HAP51WIN.DAT: {len(win_data)} bytes ({len(win_data) // WINDOW_RECORD_SIZE} windows)")

        # Criar novos tipos de Walls se definidos no Excel
        # Walls são assemblies com layers - preenchemos as layers para obter U-Value e Weight correctos
        if type_definitions['walls']:
            print("\n--- Criando Walls ---")
            wal_path = os.path.join(temp_dir, 'HAP51WAL.DAT')
            with open(wal_path, 'rb') as f:
                wal_data = bytearray(f.read())

            # Usar tamanho fixo de 3187 bytes por assembly
            num_existing_walls = len(wal_data) // ASSEMBLY_SIZE

            for i, wall_def in enumerate(type_definitions['walls']):
                # Criar novo assembly baseado no template (primeiro assembly)
                new_wall = bytearray(wal_data[0:ASSEMBLY_SIZE])

                # Modificar nome (0-255)
                name_bytes = wall_def['name'].encode('latin-1')[:255].ljust(255, b' ')
                new_wall[0:255] = name_bytes

                # Preencher layers para obter U-Value e Weight correctos
                u_value = safe_float(wall_def.get('u_value'), 1.0)
                weight = safe_float(wall_def.get('weight'), 100.0)
                absorptivity = safe_float(wall_def.get('absorptivity'), 0.9)

                fill_assembly_layers(new_wall, 0, u_value, weight, absorptivity)

                wal_data.extend(new_wall)
                new_id = num_existing_walls + i
                types['walls'][wall_def['name']] = new_id
                print(f"  Wall {new_id}: {wall_def['name']} (U={u_value:.2f}, W={weight:.0f}, A={absorptivity:.1f})")

            with open(wal_path, 'wb') as f:
                f.write(bytes(wal_data))

            print(f"HAP51WAL.DAT: {len(wal_data)} bytes ({len(wal_data)//ASSEMBLY_SIZE} walls)")

        # Criar novos tipos de Roofs se definidos no Excel
        # Roofs têm a mesma estrutura que Walls - assemblies com layers
        if type_definitions['roofs']:
            print("\n--- Criando Roofs ---")
            rof_path = os.path.join(temp_dir, 'HAP51ROF.DAT')
            with open(rof_path, 'rb') as f:
                rof_data = bytearray(f.read())

            # Usar tamanho fixo de 3187 bytes por assembly
            num_existing_roofs = len(rof_data) // ASSEMBLY_SIZE

            for i, roof_def in enumerate(type_definitions['roofs']):
                # Criar novo roof como cópia exacta do Default Roof Assembly (record 0)
                # Só alterar nome e absorptivity - manter layers e CTF intactos
                new_roof = bytearray(rof_data[0:ASSEMBLY_SIZE])

                # Modificar nome (0-255)
                name_bytes = roof_def['name'].encode('latin-1')[:255].ljust(255, b' ')
                new_roof[0:255] = name_bytes

                # Absorptivity (offset 255)
                absorptivity = safe_float(roof_def.get('absorptivity'), 0.9)
                struct.pack_into('<f', new_roof, 255, absorptivity)

                rof_data.extend(new_roof)
                new_id = num_existing_roofs + i
                types['roofs'][roof_def['name']] = new_id
                print(f"  Roof {new_id}: {roof_def['name']} (cópia Default, A={absorptivity:.1f})")

            with open(rof_path, 'wb') as f:
                f.write(bytes(rof_data))

            print(f"HAP51ROF.DAT: {len(rof_data)} bytes ({len(rof_data)//ASSEMBLY_SIZE} roofs)")

        # Criar novos espaços
        print("\n--- Criando espaços ---")
        new_spc_data = bytearray(default_record)  # Começar com default

        for i, space in enumerate(spaces):
            space_binary = create_space_binary(space, types, template_record)
            new_spc_data.extend(space_binary)
            print(f"  {i+1}. {space['name']} - {len(space_binary)} bytes")

        # Escrever novo HAP51SPC.DAT
        with open(spc_path, 'wb') as f:
            f.write(bytes(new_spc_data))

        print(f"\nHAP51SPC.DAT: {len(new_spc_data)} bytes ({len(spaces)+1} records)")

        # Actualizar MDB (SpaceIndex)
        print("\n--- Actualizando MDB ---")
        try:
            import pyodbc
            mdb_path = os.path.join(temp_dir, 'HAP51INX.MDB')
            conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={mdb_path};'
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Limpar SpaceIndex existente
            cursor.execute("DELETE FROM SpaceIndex")

            # Corrigir ScheduleIndex (MDB pode ter nomes deslocados)
            # Ler nomes correctos do HAP51SCH.DAT e actualizar MDB
            cursor.execute("DELETE FROM ScheduleIndex")
            sch_path_mdb = os.path.join(temp_dir, 'HAP51SCH.DAT')
            with open(sch_path_mdb, 'rb') as f:
                sch_data_mdb = f.read()
            SCHEDULE_RECORD_SIZE_MDB = 792
            num_sch = len(sch_data_mdb) // SCHEDULE_RECORD_SIZE_MDB
            sch_count = 0
            for i in range(num_sch):
                sch_offset = i * SCHEDULE_RECORD_SIZE_MDB
                sch_name = sch_data_mdb[sch_offset:sch_offset+24].rstrip(b'\x00').decode('latin-1', errors='ignore').strip()
                if sch_name and i > 0:  # MDB nao permite nIndex=0
                    # Escapar aspas simples no nome
                    sch_name_safe = sch_name.replace("'", "''")
                    cursor.execute(f"INSERT INTO ScheduleIndex (nIndex, szName) VALUES ({i}, '{sch_name_safe}')")
                    sch_count += 1
            print(f"  ScheduleIndex: {sch_count} schedules actualizados")

            # Inserir novos espaços
            # SpaceIndex tem colunas: nIndex, szName, fFloorArea, fNumPeople, fLightingDensity
            for i, space in enumerate(spaces):
                space_id = i + 1
                name = space['name'][:24]
                area_ft2 = m2_to_ft2(space.get('area', 0))
                people = safe_float(space.get('occupancy', 0))
                lighting = safe_float(space.get('general_light', 0))
                cursor.execute(f"INSERT INTO SpaceIndex (nIndex, szName, fFloorArea, fNumPeople, fLightingDensity) VALUES ({space_id}, '{name}', {area_ft2}, {people}, {lighting})")
                print(f"  SpaceIndex: {space_id} = {name}")

            # Inserir novas Windows no WindowIndex
            # WindowIndex: nIndex, szName, fOverallUValue, fOverallShadeCo, fHeight, fWidth
            if type_definitions['windows']:
                for win_def in type_definitions['windows']:
                    win_id = types['windows'].get(win_def['name'])
                    if win_id:
                        name = win_def['name'][:255]
                        u_val = u_si_to_ip(win_def['u_value'])
                        shgc = win_def['shgc']
                        height = m_to_ft(win_def['height'])
                        width = m_to_ft(win_def['width'])
                        cursor.execute(f"INSERT INTO WindowIndex (nIndex, szName, fOverallUValue, fOverallShadeCo, fHeight, fWidth) VALUES ({win_id}, '{name}', {u_val}, {shgc}, {height}, {width})")
                        print(f"  WindowIndex: {win_id} = {name}")

            # Inserir novas Walls no WallIndex
            # WallIndex: nIndex, szName, fOverallUValue, fOverallWeight, fThickness
            if type_definitions['walls']:
                for wall_def in type_definitions['walls']:
                    wall_id = types['walls'].get(wall_def['name'])
                    if wall_id is not None:
                        name = wall_def['name'][:255]
                        u_val = u_si_to_ip(wall_def['u_value'])
                        weight = kg_m2_to_lb_ft2(wall_def['weight'])
                        thickness = m_to_ft(wall_def['thickness']) * 12  # metros para inches
                        cursor.execute(f"INSERT INTO WallIndex (nIndex, szName, fOverallUValue, fOverallWeight, fThickness) VALUES ({wall_id}, '{name}', {u_val}, {weight}, {thickness})")
                        print(f"  WallIndex: {wall_id} = {name}")

            # Inserir novos Roofs no RoofIndex
            # RoofIndex: nIndex, szName, fOverallUValue, fOverallWeight, fThickness
            if type_definitions['roofs']:
                for roof_def in type_definitions['roofs']:
                    roof_id = types['roofs'].get(roof_def['name'])
                    if roof_id is not None:
                        name = roof_def['name'][:255]
                        u_val = u_si_to_ip(roof_def['u_value'])
                        weight = kg_m2_to_lb_ft2(roof_def['weight'])
                        thickness = m_to_ft(roof_def['thickness']) * 12  # metros para inches
                        cursor.execute(f"INSERT INTO RoofIndex (nIndex, szName, fOverallUValue, fOverallWeight, fThickness) VALUES ({roof_id}, '{name}', {u_val}, {weight}, {thickness})")
                        print(f"  RoofIndex: {roof_id} = {name}")

            # Actualizar links de schedules, walls, windows, etc.
            cursor.execute("DELETE FROM Space_Schedule_Links")
            cursor.execute("DELETE FROM Space_Wall_Links")
            cursor.execute("DELETE FROM Space_Window_Links")
            cursor.execute("DELETE FROM Space_Door_Links")
            cursor.execute("DELETE FROM Space_Roof_Links")

            for i, space in enumerate(spaces):
                space_id = i + 1

                # Schedule links
                schedules_used = set()
                for sch_field in ['people_sch', 'light_sch', 'equip_sch', 'misc_sens_sch', 'misc_lat_sch']:
                    sch = space.get(sch_field)
                    if sch:
                        sch_id = get_type_id(sch, types['schedules'], 0)
                        if sch_id > 0:
                            schedules_used.add(sch_id)

                for sch_id in schedules_used:
                    cursor.execute(f"INSERT INTO Space_Schedule_Links (Space_ID, Schedule_ID) VALUES ({space_id}, {sch_id})")

                # Wall/Window/Door links
                walls_used = set()
                windows_used = set()
                doors_used = set()

                for wall in space['walls']:
                    wall_type = get_type_id(wall.get('type'), types['walls'], 0)
                    if wall_type > 0:
                        walls_used.add(wall_type)

                    for win_field in ['win1', 'win2']:
                        win = wall.get(win_field)
                        if win:
                            win_id = get_type_id(win, types['windows'], 0)
                            if win_id > 0:
                                windows_used.add(win_id)

                    door = wall.get('door')
                    if door:
                        door_id = get_type_id(door, types['doors'], 0)
                        if door_id > 0:
                            doors_used.add(door_id)

                for wall_id in walls_used:
                    cursor.execute(f"INSERT INTO Space_Wall_Links (Space_ID, Wall_ID) VALUES ({space_id}, {wall_id})")

                for win_id in windows_used:
                    cursor.execute(f"INSERT INTO Space_Window_Links (Space_ID, Window_ID) VALUES ({space_id}, {win_id})")

                for door_id in doors_used:
                    cursor.execute(f"INSERT INTO Space_Door_Links (Space_ID, Door_ID) VALUES ({space_id}, {door_id})")

                # Roof links
                roofs_used = set()
                for roof in space['roofs']:
                    roof_type = get_type_id(roof.get('type'), types['roofs'], 0)
                    if roof_type > 0:
                        roofs_used.add(roof_type)

                    sky = roof.get('sky')
                    if sky:
                        sky_id = get_type_id(sky, types['windows'], 0)
                        if sky_id > 0:
                            windows_used.add(sky_id)

                for roof_id in roofs_used:
                    cursor.execute(f"INSERT INTO Space_Roof_Links (Space_ID, Roof_ID) VALUES ({space_id}, {roof_id})")

            conn.commit()
            conn.close()
            print("  MDB actualizado com sucesso")

        except ImportError:
            print("  AVISO: pyodbc não disponível - MDB não actualizado")
        except Exception as e:
            print(f"  ERRO MDB: {e}")

        # Criar ZIP final
        print("\n--- Criando ficheiro E3A ---")
        with zipfile.ZipFile(OUTPUT_FILE, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, temp_dir)
                    zf.write(file_path, arc_name)

        print(f"\n{'='*60}")
        print(f"FICHEIRO CRIADO: {OUTPUT_FILE}")
        print(f"{'='*60}")
        print(f"\nEspaços: {len(spaces)}")
        for i, space in enumerate(spaces):
            print(f"  {i+1}. {space['name']}")

        print("\n>>> Abre o ficheiro no HAP 5.1 para verificar!")

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

if __name__ == '__main__':
    # Parse command line arguments
    if len(sys.argv) >= 4:
        EXCEL_FILE = sys.argv[1]
        BASE_FILE = sys.argv[2]
        OUTPUT_FILE = sys.argv[3]
    elif len(sys.argv) == 2 and sys.argv[1] in ['-h', '--help']:
        print(__doc__)
        sys.exit(0)
    else:
        print("Usage: python excel_to_hap.py <input.xlsx> <modelo_base.E3A> <output.E3A>")
        print()
        print("Exemplo:")
        print("  python excel_to_hap.py MeusDados.xlsx Modelo_RSECE.E3A Output.E3A")
        print()
        print("Use -h para mais informação.")
        sys.exit(1)

    main()
