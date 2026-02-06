"""
Adaptador para converter o formato HAP 5.2 (Folha HAP_5_2_2026.xlsx)
para o formato esperado pelo excel_to_hap.py

Este adaptador lê as abas INPUT do formato HAP 5.2:
- INPUT SPACES HAP
- INPUT WALLS HAP
- INPUT ROOFS HAP
- INPUT VIDROS HAP

E gera um Excel temporário no formato HAP_Template_RSECE.xlsx

Usage:
    python adapter_hap52.py <input_hap52.xlsx> <output_template.xlsx>

Ou usar directamente no excel_to_hap.py:
    python excel_to_hap.py --format hap52 <input.xlsx> <modelo.E3A> <output.E3A>
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os

# =============================================================================
# MAPEAMENTO DE COLUNAS HAP 5.2 -> TEMPLATE
# =============================================================================

# Colunas do INPUT SPACES HAP (1-indexed)
HAP52_COLS = {
    # GENERAL
    'name': 1,           # Space Name
    'area': 2,           # Floor Area (m2)
    'height': 3,         # Avg Ceiling Ht (m)
    'weight': 4,         # Building Wt (kg/m2)
    'oa': 5,             # Outdoor Air (valor)
    'oa_unit': 6,        # OA Unit

    # PEOPLE
    'occupancy': 7,      # Occupancy (people)
    'activity': 8,       # Activity Level
    'sensible': 9,       # Sensible (W/person)
    'latent': 10,        # Latent (W/person)
    'people_sch': 11,    # Schedule

    # LIGHTING
    'task_light': 12,    # Task Lighting (W)
    'general_light': 13, # General Ltg (W)
    'fixture': 14,       # Fixture Type
    'ballast': 15,       # Ballast Mult
    'light_sch': 16,     # Schedule

    # EQUIPMENT
    'equipment': 17,     # Equipment (W/m2)
    'equip_sch': 18,     # Schedule

    # MISC
    'misc_sens': 19,     # Sensible (W)
    'misc_lat': 20,      # Latent (W)
    'misc_sens_sch': 21, # Sens Sch
    'misc_lat_sch': 22,  # Lat Sch

    # INFILTRATION
    'infil_method': 23,  # Infil Method
    'ach_clg': 24,       # Design Clg (ACH)
    'ach_htg': 25,       # Design Htg (ACH)
    'ach_energy': 26,    # Energy (ACH)

    # FLOORS
    'floor_type': 27,    # Floor Type
    'floor_area': 28,    # Floor Area (m2)
    'floor_u': 29,       # U-Value (W/m2K)
    'floor_perim': 30,   # Exp Perim (m)
    'floor_edge_r': 31,  # Edge R (m2K/W)
    'floor_depth': 32,   # Depth (m)
    'bsmt_u': 33,        # Bsmt Wall U (W/m2K)
    'wall_ins_r': 34,    # Wall Ins R (m2K/W)
    'ins_depth': 35,     # Ins Depth (m)
    'floor_unc_max': 36, # Unc Max (C)
    'floor_out_max': 37, # Out Max (C)
    'floor_unc_min': 38, # Unc Min (C)
    'floor_out_min': 39, # Out Min (C)

    # PARTITIONS - Ceiling
    'ceil_area': 40,     # Area (m2)
    'ceil_u': 41,        # U-Value (W/m2K)
    'ceil_unc_max': 42,  # Unc Max (C)
    'ceil_out_max': 43,  # Out Max (C)
    'ceil_unc_min': 44,  # Unc Min (C)
    'ceil_out_min': 45,  # Out Min (C)

    # PARTITIONS - Wall
    'wall_part_area': 46,  # Area (m2)
    'wall_part_u': 47,     # U-Value (W/m2K)
    'wall_unc_max': 48,    # Unc Max (C)
    'wall_out_max': 49,    # Out Max (C)
    'wall_unc_min': 50,    # Unc Min (C)
    'wall_out_min': 51,    # Out Min (C)

    # WALLS (8 walls, cada um com 9 colunas)
    # Wall 1: 53-61, Wall 2: 62-70, etc.
    'wall1_exposure': 53,
    'wall1_area': 54,
    'wall1_type': 55,
    'wall1_win1': 56,
    'wall1_win1_qty': 57,
    'wall1_win2': 58,
    'wall1_win2_qty': 59,
    'wall1_door': 60,
    'wall1_door_qty': 61,

    # ROOFS (4 roofs, cada um com 6 colunas)
    # Roof 1: 125-130, Roof 2: 131-136, etc.
    'roof1_exposure': 125,
    'roof1_area': 126,
    'roof1_slope': 127,
    'roof1_type': 128,
    'roof1_sky': 129,
    'roof1_sky_qty': 130,
}

# Offsets para paredes adicionais (9 colunas cada)
WALL_OFFSET = 9
WALL_COUNT = 8
WALL_START = 53

# Offsets para coberturas adicionais (6 colunas cada)
ROOF_OFFSET = 6
ROOF_COUNT = 4
ROOF_START = 125

def get_wall_col(wall_idx, field):
    """Retorna a coluna para um campo de parede específica (0-indexed wall)"""
    base = WALL_START + wall_idx * WALL_OFFSET
    fields = ['exposure', 'area', 'type', 'win1', 'win1_qty', 'win2', 'win2_qty', 'door', 'door_qty']
    return base + fields.index(field)

def get_roof_col(roof_idx, field):
    """Retorna a coluna para um campo de cobertura específica (0-indexed roof)"""
    base = ROOF_START + roof_idx * ROOF_OFFSET
    fields = ['exposure', 'area', 'slope', 'type', 'sky', 'sky_qty']
    return base + fields.index(field)


def read_hap52_format(excel_path):
    """Lê o Excel no formato HAP 5.2 e retorna dados estruturados"""

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    data = {
        'spaces': [],
        'walls': [],
        'roofs': [],
        'windows': [],
    }

    # =========================================================================
    # LER INPUT WALLS HAP
    # =========================================================================
    if 'INPUT WALLS HAP' in wb.sheetnames:
        ws = wb['INPUT WALLS HAP']
        for row in ws.iter_rows(min_row=4, max_col=4, values_only=True):
            if row[0] and str(row[0]).strip():
                data['walls'].append({
                    'name': str(row[0]).strip(),
                    'u_value': row[1] if row[1] else 0.5,
                    'weight': row[2] if row[2] else 200,
                    'thickness': row[3] if row[3] else 0.3,
                })
        print(f"  Lidas {len(data['walls'])} paredes")

    # =========================================================================
    # LER INPUT ROOFS HAP
    # =========================================================================
    if 'INPUT ROOFS HAP' in wb.sheetnames:
        ws = wb['INPUT ROOFS HAP']
        for row in ws.iter_rows(min_row=4, max_col=4, values_only=True):
            if row[0] and str(row[0]).strip():
                data['roofs'].append({
                    'name': str(row[0]).strip(),
                    'u_value': row[1] if row[1] else 0.5,
                    'weight': row[2] if row[2] else 300,
                    'thickness': row[3] if row[3] else 0.3,
                })
        print(f"  Lidas {len(data['roofs'])} coberturas")

    # =========================================================================
    # LER INPUT VIDROS HAP
    # =========================================================================
    if 'INPUT VIDROS HAP' in wb.sheetnames:
        ws = wb['INPUT VIDROS HAP']
        # Linha 5 tem os headers reais: REF_ID, U, g, alt, Lar
        for row in ws.iter_rows(min_row=6, max_col=5, values_only=True):
            if row[0] and str(row[0]).strip():
                data['windows'].append({
                    'name': str(row[0]).strip(),
                    'u_value': row[1] if row[1] else 2.0,
                    'shgc': row[2] if row[2] else 0.4,
                    'height': row[3] if row[3] else 1.5,
                    'width': row[4] if row[4] else 1.0,
                })
        print(f"  Lidas {len(data['windows'])} janelas")

    # =========================================================================
    # LER INPUT SPACES HAP
    # =========================================================================
    if 'INPUT SPACES HAP' in wb.sheetnames:
        ws = wb['INPUT SPACES HAP']

        # Ler todas as linhas de dados (a partir da linha 4)
        for row_num in range(4, ws.max_row + 1):
            name = ws.cell(row=row_num, column=1).value
            if not name or str(name).strip() == '':
                continue

            space = {'name': str(name).strip()[:24]}

            # GENERAL
            space['area'] = ws.cell(row=row_num, column=2).value
            space['height'] = ws.cell(row=row_num, column=3).value
            space['weight'] = ws.cell(row=row_num, column=4).value
            space['oa'] = ws.cell(row=row_num, column=5).value
            space['oa_unit'] = ws.cell(row=row_num, column=6).value

            # PEOPLE
            space['occupancy'] = ws.cell(row=row_num, column=7).value
            space['activity'] = ws.cell(row=row_num, column=8).value
            space['sensible'] = ws.cell(row=row_num, column=9).value
            space['latent'] = ws.cell(row=row_num, column=10).value
            space['people_sch'] = ws.cell(row=row_num, column=11).value

            # LIGHTING
            space['task_light'] = ws.cell(row=row_num, column=12).value
            space['general_light'] = ws.cell(row=row_num, column=13).value
            space['fixture'] = ws.cell(row=row_num, column=14).value
            space['ballast'] = ws.cell(row=row_num, column=15).value
            space['light_sch'] = ws.cell(row=row_num, column=16).value

            # EQUIPMENT
            space['equipment'] = ws.cell(row=row_num, column=17).value
            space['equip_sch'] = ws.cell(row=row_num, column=18).value

            # MISC
            space['misc_sens'] = ws.cell(row=row_num, column=19).value
            space['misc_lat'] = ws.cell(row=row_num, column=20).value
            space['misc_sens_sch'] = ws.cell(row=row_num, column=21).value
            space['misc_lat_sch'] = ws.cell(row=row_num, column=22).value

            # INFILTRATION
            space['infil_method'] = ws.cell(row=row_num, column=23).value
            space['ach_clg'] = ws.cell(row=row_num, column=24).value
            space['ach_htg'] = ws.cell(row=row_num, column=25).value
            space['ach_energy'] = ws.cell(row=row_num, column=26).value

            # FLOORS
            space['floor_type'] = ws.cell(row=row_num, column=27).value
            space['floor_area'] = ws.cell(row=row_num, column=28).value
            space['floor_u'] = ws.cell(row=row_num, column=29).value
            space['floor_perim'] = ws.cell(row=row_num, column=30).value
            space['floor_edge_r'] = ws.cell(row=row_num, column=31).value
            space['floor_depth'] = ws.cell(row=row_num, column=32).value
            space['bsmt_u'] = ws.cell(row=row_num, column=33).value
            space['wall_ins_r'] = ws.cell(row=row_num, column=34).value
            space['ins_depth'] = ws.cell(row=row_num, column=35).value
            space['floor_unc_max'] = ws.cell(row=row_num, column=36).value
            space['floor_out_max'] = ws.cell(row=row_num, column=37).value
            space['floor_unc_min'] = ws.cell(row=row_num, column=38).value
            space['floor_out_min'] = ws.cell(row=row_num, column=39).value

            # PARTITIONS - Ceiling
            space['ceil_area'] = ws.cell(row=row_num, column=40).value
            space['ceil_u'] = ws.cell(row=row_num, column=41).value
            space['ceil_unc_max'] = ws.cell(row=row_num, column=42).value
            space['ceil_out_max'] = ws.cell(row=row_num, column=43).value
            space['ceil_unc_min'] = ws.cell(row=row_num, column=44).value
            space['ceil_out_min'] = ws.cell(row=row_num, column=45).value

            # PARTITIONS - Wall
            space['wall_part_area'] = ws.cell(row=row_num, column=46).value
            space['wall_part_u'] = ws.cell(row=row_num, column=47).value
            space['wall_unc_max'] = ws.cell(row=row_num, column=48).value
            space['wall_out_max'] = ws.cell(row=row_num, column=49).value
            space['wall_unc_min'] = ws.cell(row=row_num, column=50).value
            space['wall_out_min'] = ws.cell(row=row_num, column=51).value

            # WALLS (8 paredes)
            space['walls'] = []
            for w in range(WALL_COUNT):
                col = WALL_START + w * WALL_OFFSET
                wall = {
                    'exposure': ws.cell(row=row_num, column=col).value,
                    'area': ws.cell(row=row_num, column=col+1).value,
                    'type': ws.cell(row=row_num, column=col+2).value,
                    'win1': ws.cell(row=row_num, column=col+3).value,
                    'win1_qty': ws.cell(row=row_num, column=col+4).value,
                    'win2': ws.cell(row=row_num, column=col+5).value,
                    'win2_qty': ws.cell(row=row_num, column=col+6).value,
                    'door': ws.cell(row=row_num, column=col+7).value,
                    'door_qty': ws.cell(row=row_num, column=col+8).value,
                }
                space['walls'].append(wall)

            # ROOFS (4 coberturas)
            space['roofs'] = []
            for r in range(ROOF_COUNT):
                col = ROOF_START + r * ROOF_OFFSET
                roof = {
                    'exposure': ws.cell(row=row_num, column=col).value,
                    'area': ws.cell(row=row_num, column=col+1).value,
                    'slope': ws.cell(row=row_num, column=col+2).value,
                    'type': ws.cell(row=row_num, column=col+3).value,
                    'sky': ws.cell(row=row_num, column=col+4).value,
                    'sky_qty': ws.cell(row=row_num, column=col+5).value,
                }
                space['roofs'].append(roof)

            data['spaces'].append(space)
            print(f"    Espaço: {space['name']}")

    print(f"  Total: {len(data['spaces'])} espaços lidos")

    wb.close()
    return data


def write_template_format(data, output_path):
    """Escreve os dados no formato HAP_Template_RSECE.xlsx"""

    wb = openpyxl.Workbook()

    # =========================================================================
    # SHEET: Espacos
    # =========================================================================
    ws = wb.active
    ws.title = 'Espacos'

    # Headers (linha 1: categorias, linha 2: subcategorias, linha 3: nomes)
    # Row 1 - Categorias
    ws.cell(row=1, column=1, value='GENERAL')
    ws.cell(row=1, column=7, value='INTERNALS')
    ws.cell(row=1, column=23, value='INFILTRATION')
    ws.cell(row=1, column=27, value='FLOORS')
    ws.cell(row=1, column=40, value='PARTITIONS')
    ws.cell(row=1, column=52, value='WALLS')
    ws.cell(row=1, column=124, value='ROOFS')

    # Row 2 - Subcategorias
    ws.cell(row=2, column=7, value='PEOPLE')
    ws.cell(row=2, column=12, value='LIGHTING')
    ws.cell(row=2, column=17, value='EQUIPMENT')
    ws.cell(row=2, column=19, value='MISC')
    ws.cell(row=2, column=40, value='CEILING')
    ws.cell(row=2, column=46, value='WALL')

    for w in range(8):
        ws.cell(row=2, column=52 + w*9, value=f'WALL {w+1}')
    for r in range(4):
        ws.cell(row=2, column=124 + r*6, value=f'ROOF {r+1}')

    # Row 3 - Headers das colunas
    headers = [
        'Space Name', 'Floor Area (m2)', 'Avg Ceiling Ht (m)', 'Building Wt (kg/m2)',
        'Outdoor Air', 'OA Unit', 'Occupancy', 'Activity Level', 'Sensible', 'Latent',
        'Schedule', 'Task Light', 'General Light', 'Fixture', 'Ballast', 'Schedule',
        'Equipment', 'Schedule', 'Sensible', 'Latent', 'Sens Sch', 'Lat Sch',
        'Infil Method', 'Design Clg', 'Design Htg', 'Energy',
        'Floor Type', 'Floor Area', 'U-Value', 'Exp Perim', 'Edge R', 'Depth',
        'Bsmt U', 'Wall Ins R', 'Ins Depth', 'Unc Max', 'Out Max', 'Unc Min', 'Out Min',
        'Area', 'U-Value', 'Unc Max', 'Out Max', 'Unc Min', 'Out Min',
        'Area', 'U-Value', 'Unc Max', 'Out Max', 'Unc Min', 'Out Min',
    ]

    # Wall headers (9 colunas x 8 paredes)
    wall_headers = ['Exposure', 'Gross Area', 'Wall Type', 'Window 1', 'Win1 Qty',
                    'Window 2', 'Win2 Qty', 'Door', 'Door Qty']
    for w in range(8):
        headers.extend(wall_headers)

    # Roof headers (6 colunas x 4 coberturas)
    roof_headers = ['Exposure', 'Gross Area', 'Slope', 'Roof Type', 'Skylight', 'Sky Qty']
    for r in range(4):
        headers.extend(roof_headers)

    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)

    # Dados dos espaços (linha 4+)
    for row_idx, space in enumerate(data['spaces'], 4):
        ws.cell(row=row_idx, column=1, value=space.get('name', ''))
        ws.cell(row=row_idx, column=2, value=space.get('area'))
        ws.cell(row=row_idx, column=3, value=space.get('height'))
        ws.cell(row=row_idx, column=4, value=space.get('weight'))
        ws.cell(row=row_idx, column=5, value=space.get('oa'))
        ws.cell(row=row_idx, column=6, value=space.get('oa_unit'))

        ws.cell(row=row_idx, column=7, value=space.get('occupancy'))
        ws.cell(row=row_idx, column=8, value=space.get('activity'))
        ws.cell(row=row_idx, column=9, value=space.get('sensible'))
        ws.cell(row=row_idx, column=10, value=space.get('latent'))
        ws.cell(row=row_idx, column=11, value=space.get('people_sch'))

        ws.cell(row=row_idx, column=12, value=space.get('task_light'))
        ws.cell(row=row_idx, column=13, value=space.get('general_light'))
        ws.cell(row=row_idx, column=14, value=space.get('fixture'))
        ws.cell(row=row_idx, column=15, value=space.get('ballast'))
        ws.cell(row=row_idx, column=16, value=space.get('light_sch'))

        ws.cell(row=row_idx, column=17, value=space.get('equipment'))
        ws.cell(row=row_idx, column=18, value=space.get('equip_sch'))

        ws.cell(row=row_idx, column=19, value=space.get('misc_sens'))
        ws.cell(row=row_idx, column=20, value=space.get('misc_lat'))
        ws.cell(row=row_idx, column=21, value=space.get('misc_sens_sch'))
        ws.cell(row=row_idx, column=22, value=space.get('misc_lat_sch'))

        ws.cell(row=row_idx, column=23, value=space.get('infil_method'))
        ws.cell(row=row_idx, column=24, value=space.get('ach_clg'))
        ws.cell(row=row_idx, column=25, value=space.get('ach_htg'))
        ws.cell(row=row_idx, column=26, value=space.get('ach_energy'))

        ws.cell(row=row_idx, column=27, value=space.get('floor_type'))
        ws.cell(row=row_idx, column=28, value=space.get('floor_area'))
        ws.cell(row=row_idx, column=29, value=space.get('floor_u'))
        ws.cell(row=row_idx, column=30, value=space.get('floor_perim'))
        ws.cell(row=row_idx, column=31, value=space.get('floor_edge_r'))
        ws.cell(row=row_idx, column=32, value=space.get('floor_depth'))
        ws.cell(row=row_idx, column=33, value=space.get('bsmt_u'))
        ws.cell(row=row_idx, column=34, value=space.get('wall_ins_r'))
        ws.cell(row=row_idx, column=35, value=space.get('ins_depth'))
        ws.cell(row=row_idx, column=36, value=space.get('floor_unc_max'))
        ws.cell(row=row_idx, column=37, value=space.get('floor_out_max'))
        ws.cell(row=row_idx, column=38, value=space.get('floor_unc_min'))
        ws.cell(row=row_idx, column=39, value=space.get('floor_out_min'))

        ws.cell(row=row_idx, column=40, value=space.get('ceil_area'))
        ws.cell(row=row_idx, column=41, value=space.get('ceil_u'))
        ws.cell(row=row_idx, column=42, value=space.get('ceil_unc_max'))
        ws.cell(row=row_idx, column=43, value=space.get('ceil_out_max'))
        ws.cell(row=row_idx, column=44, value=space.get('ceil_unc_min'))
        ws.cell(row=row_idx, column=45, value=space.get('ceil_out_min'))

        ws.cell(row=row_idx, column=46, value=space.get('wall_part_area'))
        ws.cell(row=row_idx, column=47, value=space.get('wall_part_u'))
        ws.cell(row=row_idx, column=48, value=space.get('wall_unc_max'))
        ws.cell(row=row_idx, column=49, value=space.get('wall_out_max'))
        ws.cell(row=row_idx, column=50, value=space.get('wall_unc_min'))
        ws.cell(row=row_idx, column=51, value=space.get('wall_out_min'))

        # Walls
        for w, wall in enumerate(space.get('walls', [])):
            col = 52 + w * 9
            ws.cell(row=row_idx, column=col, value=wall.get('exposure'))
            ws.cell(row=row_idx, column=col+1, value=wall.get('area'))
            ws.cell(row=row_idx, column=col+2, value=wall.get('type'))
            ws.cell(row=row_idx, column=col+3, value=wall.get('win1'))
            ws.cell(row=row_idx, column=col+4, value=wall.get('win1_qty'))
            ws.cell(row=row_idx, column=col+5, value=wall.get('win2'))
            ws.cell(row=row_idx, column=col+6, value=wall.get('win2_qty'))
            ws.cell(row=row_idx, column=col+7, value=wall.get('door'))
            ws.cell(row=row_idx, column=col+8, value=wall.get('door_qty'))

        # Roofs
        for r, roof in enumerate(space.get('roofs', [])):
            col = 124 + r * 6
            ws.cell(row=row_idx, column=col, value=roof.get('exposure'))
            ws.cell(row=row_idx, column=col+1, value=roof.get('area'))
            ws.cell(row=row_idx, column=col+2, value=roof.get('slope'))
            ws.cell(row=row_idx, column=col+3, value=roof.get('type'))
            ws.cell(row=row_idx, column=col+4, value=roof.get('sky'))
            ws.cell(row=row_idx, column=col+5, value=roof.get('sky_qty'))

    # =========================================================================
    # SHEET: Walls
    # =========================================================================
    ws_walls = wb.create_sheet('Walls')
    ws_walls.cell(row=1, column=1, value='WALLS')
    ws_walls.cell(row=2, column=1, value='IDENTIFICAÇÃO')
    ws_walls.cell(row=2, column=2, value='PROPRIEDADES')
    ws_walls.cell(row=3, column=1, value='Nome')
    ws_walls.cell(row=3, column=2, value='U-Value (W/m²K)')
    ws_walls.cell(row=3, column=3, value='Peso (kg/m²)')
    ws_walls.cell(row=3, column=4, value='Espessura (m)')

    for row_idx, wall in enumerate(data['walls'], 4):
        ws_walls.cell(row=row_idx, column=1, value=wall['name'])
        ws_walls.cell(row=row_idx, column=2, value=wall['u_value'])
        ws_walls.cell(row=row_idx, column=3, value=wall['weight'])
        ws_walls.cell(row=row_idx, column=4, value=wall['thickness'])

    # =========================================================================
    # SHEET: Roofs
    # =========================================================================
    ws_roofs = wb.create_sheet('Roofs')
    ws_roofs.cell(row=1, column=1, value='ROOFS')
    ws_roofs.cell(row=2, column=1, value='IDENTIFICAÇÃO')
    ws_roofs.cell(row=2, column=2, value='PROPRIEDADES')
    ws_roofs.cell(row=3, column=1, value='Nome')
    ws_roofs.cell(row=3, column=2, value='U-Value (W/m²K)')
    ws_roofs.cell(row=3, column=3, value='Peso (kg/m²)')
    ws_roofs.cell(row=3, column=4, value='Espessura (m)')

    for row_idx, roof in enumerate(data['roofs'], 4):
        ws_roofs.cell(row=row_idx, column=1, value=roof['name'])
        ws_roofs.cell(row=row_idx, column=2, value=roof['u_value'])
        ws_roofs.cell(row=row_idx, column=3, value=roof['weight'])
        ws_roofs.cell(row=row_idx, column=4, value=roof['thickness'])

    # =========================================================================
    # SHEET: Windows
    # =========================================================================
    ws_windows = wb.create_sheet('Windows')
    ws_windows.cell(row=1, column=1, value='WINDOWS')
    ws_windows.cell(row=2, column=1, value='IDENTIFICAÇÃO')
    ws_windows.cell(row=2, column=2, value='PROPRIEDADES TÉRMICAS')
    ws_windows.cell(row=2, column=4, value='DIMENSÕES')
    ws_windows.cell(row=3, column=1, value='Nome')
    ws_windows.cell(row=3, column=2, value='U-Value (W/m²K)')
    ws_windows.cell(row=3, column=3, value='SHGC')
    ws_windows.cell(row=3, column=4, value='Altura (m)')
    ws_windows.cell(row=3, column=5, value='Largura (m)')

    for row_idx, window in enumerate(data['windows'], 4):
        ws_windows.cell(row=row_idx, column=1, value=window['name'])
        ws_windows.cell(row=row_idx, column=2, value=window['u_value'])
        ws_windows.cell(row=row_idx, column=3, value=window['shgc'])
        ws_windows.cell(row=row_idx, column=4, value=window['height'])
        ws_windows.cell(row=row_idx, column=5, value=window['width'])

    # Guardar
    wb.save(output_path)
    print(f"\nFicheiro guardado: {output_path}")
    print(f"  - {len(data['spaces'])} espaços")
    print(f"  - {len(data['walls'])} paredes")
    print(f"  - {len(data['roofs'])} coberturas")
    print(f"  - {len(data['windows'])} janelas")


def detect_format(excel_path):
    """Detecta automaticamente o formato do Excel"""
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    # Formato HAP 5.2: tem abas INPUT
    if 'INPUT SPACES HAP' in sheets:
        return 'hap52'

    # Formato Template RSECE: tem aba Espacos
    if 'Espacos' in sheets:
        return 'template'

    return 'unknown'


def convert_hap52_to_template(input_path, output_path=None):
    """Converte ficheiro HAP 5.2 para formato Template"""

    if output_path is None:
        base = os.path.splitext(input_path)[0]
        output_path = f"{base}_converted.xlsx"

    print(f"\n=== Converter HAP 5.2 -> Template ===")
    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print()

    data = read_hap52_format(input_path)
    write_template_format(data, output_path)

    return output_path


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUsage:")
        print("  python adapter_hap52.py <input.xlsx> [output.xlsx]")
        print("\nExemplo:")
        print("  python adapter_hap52.py 'Folha HAP_5_2_2026.xlsx' dados_convertidos.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_file):
        print(f"Erro: Ficheiro não encontrado: {input_file}")
        sys.exit(1)

    fmt = detect_format(input_file)
    print(f"Formato detectado: {fmt}")

    if fmt == 'hap52':
        convert_hap52_to_template(input_file, output_file)
    elif fmt == 'template':
        print("Ficheiro já está no formato Template!")
    else:
        print("Formato desconhecido. Tentando converter como HAP 5.2...")
        convert_hap52_to_template(input_file, output_file)
