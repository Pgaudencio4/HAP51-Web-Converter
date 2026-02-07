"""
Adaptador para converter o formato HAP 5.2 (Folha HAP_5_2_2026.xlsx)
para o formato esperado pelo excel_to_hap.py

Este adaptador lê as abas INPUT do formato HAP 5.2:
- INPUT SPACES HAP  (147 colunas, mesma estrutura que o template)
- INPUT WALLS HAP   (Nome, U-Value, Peso, Espessura)
- INPUT ROOFS HAP   (Nome, U-Value, Peso, Espessura)
- INPUT VIDROS HAP  (Nome, U-Value, SHGC, Altura, Largura)

E gera um Excel no formato do template (Espacos, Walls, Roofs, Windows)
que pode ser usado directamente com excel_to_hap.py

Usage:
    python adapter_hap52.py <input.xlsx> [output.xlsx]

Exemplo:
    python adapter_hap52.py "Folha HAP_5_2_2026.xlsx" dados_convertidos.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import os

# =============================================================================
# MAPEAMENTO DE COLUNAS DO FORMATO HAP 5.2
# =============================================================================
# As colunas em "INPUT SPACES HAP" seguem o mesmo layout de 147 colunas
# do template original. Mapeamento campo -> coluna (1-indexed).

HAP52_COLS = {
    # GENERAL (cols 1-6)
    'name': 1,
    'area': 2,
    'height': 3,
    'weight': 4,
    'oa': 5,
    'oa_unit': 6,
    # PEOPLE (cols 7-11)
    'occupancy': 7,
    'activity': 8,
    'sensible': 9,
    'latent': 10,
    'people_sch': 11,
    # LIGHTING (cols 12-16)
    'task_light': 12,
    'general_light': 13,
    'fixture': 14,
    'ballast': 15,
    'light_sch': 16,
    # EQUIPMENT (cols 17-18)
    'equipment': 17,
    'equip_sch': 18,
    # MISC (cols 19-22)
    'misc_sens': 19,
    'misc_lat': 20,
    'misc_sens_sch': 21,
    'misc_lat_sch': 22,
    # INFILTRATION (cols 23-26)
    'infil_method': 23,
    'ach_clg': 24,
    'ach_htg': 25,
    'ach_energy': 26,
    # FLOORS (cols 27-39)
    'floor_type': 27,
    'floor_area': 28,
    'floor_u': 29,
    'floor_perim': 30,
    'floor_edge_r': 31,
    'floor_depth': 32,
    'bsmt_u': 33,
    'wall_ins_r': 34,
    'ins_depth': 35,
    'floor_unc_max': 36,
    'floor_out_max': 37,
    'floor_unc_min': 38,
    'floor_out_min': 39,
    # PARTITIONS - CEILING (cols 40-45)
    'ceil_area': 40,
    'ceil_u': 41,
    'ceil_unc_max': 42,
    'ceil_out_max': 43,
    'ceil_unc_min': 44,
    'ceil_out_min': 45,
    # PARTITIONS - WALL (cols 46-51)
    'wall_part_area': 46,
    'wall_part_u': 47,
    'wall_unc_max': 48,
    'wall_out_max': 49,
    'wall_unc_min': 50,
    'wall_out_min': 51,
}

# Walls: 8 walls x 9 campos, começando na coluna 52
WALL_OFFSET = 52   # Coluna onde começa WALL 1
WALL_COUNT = 8     # Máximo de walls por espaço
WALL_FIELDS = 9    # Campos por wall: Exposure, Area, Type, Win1, Win1Qty, Win2, Win2Qty, Door, DoorQty

# Roofs: 4 roofs x 6 campos, começando na coluna 124
ROOF_OFFSET = 124  # Coluna onde começa ROOF 1
ROOF_COUNT = 4     # Máximo de roofs por espaço
ROOF_FIELDS = 6    # Campos por roof: Exposure, Area, Slope, Type, Skylight, SkyQty


def get_wall_col(wall_index, field_index):
    """Retorna a coluna para um campo de parede específica (0-indexed wall)"""
    return WALL_OFFSET + wall_index * WALL_FIELDS + field_index


def get_roof_col(roof_index, field_index):
    """Retorna a coluna para um campo de cobertura específica (0-indexed roof)"""
    return ROOF_OFFSET + roof_index * ROOF_FIELDS + field_index


# =============================================================================
# LEITURA DO FORMATO HAP 5.2
# =============================================================================

def read_hap52_format(filepath):
    """Lê o Excel no formato HAP 5.2 e retorna dados estruturados"""

    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {'spaces': [], 'walls': [], 'roofs': [], 'windows': []}

    # --- WALLS ---
    if 'INPUT WALLS HAP' in wb.sheetnames:
        ws = wb['INPUT WALLS HAP']
        for row in range(4, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or str(name).strip() == '':
                continue
            data['walls'].append({
                'name': str(name).strip(),
                'u_value': ws.cell(row, 2).value or 0,
                'weight': ws.cell(row, 3).value or 0,
                'thickness': ws.cell(row, 4).value or 0,
            })
        print(f"  Lidas {len(data['walls'])} paredes")

    # --- ROOFS ---
    if 'INPUT ROOFS HAP' in wb.sheetnames:
        ws = wb['INPUT ROOFS HAP']
        for row in range(4, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or str(name).strip() == '':
                continue
            data['roofs'].append({
                'name': str(name).strip(),
                'u_value': ws.cell(row, 2).value or 0,
                'weight': ws.cell(row, 3).value or 0,
                'thickness': ws.cell(row, 4).value or 0,
            })
        print(f"  Lidas {len(data['roofs'])} coberturas")

    # --- WINDOWS (VIDROS) ---
    if 'INPUT VIDROS HAP' in wb.sheetnames:
        ws = wb['INPUT VIDROS HAP']
        for row in range(4, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or str(name).strip() == '':
                continue
            name_str = str(name).strip()
            # Ignorar sub-headers (ex: "REF_ID", "Nome", etc.)
            if name_str in ('REF_ID', 'Nome', 'Name'):
                continue
            u_val = ws.cell(row, 2).value
            # Ignorar linhas onde U-Value nao e numerico (headers)
            try:
                u_val = float(u_val)
            except (TypeError, ValueError):
                continue
            data['windows'].append({
                'name': name_str,
                'u_value': u_val,
                'shgc': ws.cell(row, 3).value or 0,
                'height': ws.cell(row, 4).value or 0,
                'width': ws.cell(row, 5).value or 0,
            })
        print(f"  Lidas {len(data['windows'])} janelas")

    # --- SPACES ---
    if 'INPUT SPACES HAP' in wb.sheetnames:
        ws = wb['INPUT SPACES HAP']
        for row in range(4, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or str(name).strip() == '':
                continue

            space = {}

            # Ler campos gerais (cols 1-51)
            for field, col in HAP52_COLS.items():
                space[field] = ws.cell(row, col).value

            # Ler walls (8 walls x 9 campos)
            space['walls'] = []
            for w in range(WALL_COUNT):
                wall = {
                    'exposure': ws.cell(row, get_wall_col(w, 0)).value,
                    'area': ws.cell(row, get_wall_col(w, 1)).value,
                    'type': ws.cell(row, get_wall_col(w, 2)).value,
                    'win1': ws.cell(row, get_wall_col(w, 3)).value,
                    'win1_qty': ws.cell(row, get_wall_col(w, 4)).value,
                    'win2': ws.cell(row, get_wall_col(w, 5)).value,
                    'win2_qty': ws.cell(row, get_wall_col(w, 6)).value,
                    'door': ws.cell(row, get_wall_col(w, 7)).value,
                    'door_qty': ws.cell(row, get_wall_col(w, 8)).value,
                }
                space['walls'].append(wall)

            # Ler roofs (4 roofs x 6 campos)
            space['roofs'] = []
            for r in range(ROOF_COUNT):
                roof = {
                    'exposure': ws.cell(row, get_roof_col(r, 0)).value,
                    'area': ws.cell(row, get_roof_col(r, 1)).value,
                    'slope': ws.cell(row, get_roof_col(r, 2)).value,
                    'type': ws.cell(row, get_roof_col(r, 3)).value,
                    'sky': ws.cell(row, get_roof_col(r, 4)).value,
                    'sky_qty': ws.cell(row, get_roof_col(r, 5)).value,
                }
                space['roofs'].append(roof)

            data['spaces'].append(space)
            print(f"    Espaço: {space['name']}")

        print(f"  Total: {len(data['spaces'])} espaços lidos")

    wb.close()
    return data


# =============================================================================
# ESCRITA NO FORMATO TEMPLATE
# =============================================================================

def write_template_format(data, output_path):
    """Escreve os dados no formato HAP_Template_RSECE.xlsx"""

    wb = openpyxl.Workbook()

    # =========================================================================
    # SHEET ESPACOS (147 colunas)
    # =========================================================================
    ws = wb.active
    ws.title = 'Espacos'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    sub_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # Linha 1: Categorias
    categories = [
        (1, 'GENERAL'), (7, 'INTERNALS'), (23, 'INFILTRATION'),
        (27, 'FLOORS'), (40, 'PARTITIONS'), (52, 'WALLS'), (124, 'ROOFS')
    ]
    for col, name in categories:
        cell = ws.cell(1, col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    # Linha 2: Sub-categorias
    subcategories = [
        (7, 'PEOPLE'), (12, 'LIGHTING'), (17, 'EQUIPMENT'), (19, 'MISC'),
        (40, 'CEILING'), (46, 'WALL'),
    ]
    for w in range(WALL_COUNT):
        subcategories.append((WALL_OFFSET + w * WALL_FIELDS, f'WALL {w+1}'))
    for r in range(ROOF_COUNT):
        subcategories.append((ROOF_OFFSET + r * ROOF_FIELDS, f'ROOF {r+1}'))

    for col, name in subcategories:
        cell = ws.cell(2, col, value=name)
        cell.fill = sub_fill
        cell.font = Font(bold=True)

    # Linha 3: Headers dos campos
    headers = [
        'Space Name', 'Floor Area\n(m2)', 'Avg Ceiling Ht\n(m)', 'Building Wt\n(kg/m2)',
        'Outdoor Air\n(valor)', 'OA Unit',
        'Occupancy\n(people)', 'Activity Level', 'Sensible\n(W/person)', 'Latent\n(W/person)', 'Schedule',
        'Task Lighting\n(W)', 'General Ltg\n(W)', 'Fixture Type', 'Ballast Mult', 'Schedule',
        'Equipment\n(W/m2)', 'Schedule',
        'Sensible\n(W)', 'Latent\n(W)', 'Sens Sch', 'Lat Sch',
        'Infil Method', 'Design Clg\n(ACH)', 'Design Htg\n(ACH)', 'Energy\n(ACH)',
        'Floor Type', 'Floor Area\n(m2)', 'U-Value\n(W/m2K)', 'Exp Perim\n(m)',
        'Edge R\n(m2K/W)', 'Depth\n(m)', 'Bsmt Wall U\n(W/m2K)', 'Wall Ins R\n(m2K/W)',
        'Ins Depth\n(m)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
        'Area\n(m2)', 'U-Value\n(W/m2K)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
        'Area\n(m2)', 'U-Value\n(W/m2K)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
    ]

    wall_headers = ['Exposure', 'Gross Area\n(m2)', 'Wall Type', 'Window 1', 'Win1 Qty',
                    'Window 2', 'Win2 Qty', 'Door', 'Door Qty']
    for w in range(WALL_COUNT):
        headers.extend(wall_headers)

    roof_headers = ['Exposure', 'Gross Area\n(m2)', 'Slope\n(deg)', 'Roof Type', 'Skylight', 'Sky Qty']
    for r in range(ROOF_COUNT):
        headers.extend(roof_headers)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, value=h)
        cell.fill = sub_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Dados dos espaços
    for row_idx, space in enumerate(data['spaces'], 4):
        # Campos gerais (cols 1-51)
        for field, col in HAP52_COLS.items():
            val = space.get(field)
            if val is not None:
                ws.cell(row_idx, col, value=val)

        # Walls (cols 52-123)
        for w, wall in enumerate(space.get('walls', [])):
            ws.cell(row_idx, get_wall_col(w, 0), value=wall.get('exposure'))
            ws.cell(row_idx, get_wall_col(w, 1), value=wall.get('area'))
            ws.cell(row_idx, get_wall_col(w, 2), value=wall.get('type'))
            ws.cell(row_idx, get_wall_col(w, 3), value=wall.get('win1'))
            ws.cell(row_idx, get_wall_col(w, 4), value=wall.get('win1_qty'))
            ws.cell(row_idx, get_wall_col(w, 5), value=wall.get('win2'))
            ws.cell(row_idx, get_wall_col(w, 6), value=wall.get('win2_qty'))
            ws.cell(row_idx, get_wall_col(w, 7), value=wall.get('door'))
            ws.cell(row_idx, get_wall_col(w, 8), value=wall.get('door_qty'))

        # Roofs (cols 124-147)
        for r, roof in enumerate(space.get('roofs', [])):
            ws.cell(row_idx, get_roof_col(r, 0), value=roof.get('exposure'))
            ws.cell(row_idx, get_roof_col(r, 1), value=roof.get('area'))
            ws.cell(row_idx, get_roof_col(r, 2), value=roof.get('slope'))
            ws.cell(row_idx, get_roof_col(r, 3), value=roof.get('type'))
            ws.cell(row_idx, get_roof_col(r, 4), value=roof.get('sky'))
            ws.cell(row_idx, get_roof_col(r, 5), value=roof.get('sky_qty'))

    # Ajustar larguras
    ws.column_dimensions['A'].width = 20
    for col in range(2, 148):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = 'B4'

    # =========================================================================
    # SHEET WALLS
    # =========================================================================
    ws_w = wb.create_sheet('Walls')
    ws_w.cell(1, 1, value='WALLS')
    ws_w.cell(2, 1, value='IDENTIFICAÇÃO')
    ws_w.cell(2, 2, value='PROPRIEDADES')
    ws_w.cell(3, 1, value='Nome')
    ws_w.cell(3, 2, value='U-Value (W/m²K)')
    ws_w.cell(3, 3, value='Peso (kg/m²)')
    ws_w.cell(3, 4, value='Espessura (m)')

    for i, wall in enumerate(data.get('walls', []), 4):
        ws_w.cell(i, 1, value=wall['name'])
        ws_w.cell(i, 2, value=wall['u_value'])
        ws_w.cell(i, 3, value=wall['weight'])
        ws_w.cell(i, 4, value=wall.get('thickness', 0))

    ws_w.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws_w.column_dimensions[col].width = 15

    # =========================================================================
    # SHEET ROOFS
    # =========================================================================
    ws_r = wb.create_sheet('Roofs')
    ws_r.cell(1, 1, value='ROOFS')
    ws_r.cell(2, 1, value='IDENTIFICAÇÃO')
    ws_r.cell(2, 2, value='PROPRIEDADES')
    ws_r.cell(3, 1, value='Nome')
    ws_r.cell(3, 2, value='U-Value (W/m²K)')
    ws_r.cell(3, 3, value='Peso (kg/m²)')
    ws_r.cell(3, 4, value='Espessura (m)')

    for i, roof in enumerate(data.get('roofs', []), 4):
        ws_r.cell(i, 1, value=roof['name'])
        ws_r.cell(i, 2, value=roof['u_value'])
        ws_r.cell(i, 3, value=roof['weight'])
        ws_r.cell(i, 4, value=roof.get('thickness', 0))

    ws_r.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws_r.column_dimensions[col].width = 15

    # =========================================================================
    # SHEET WINDOWS
    # =========================================================================
    ws_win = wb.create_sheet('Windows')
    ws_win.cell(1, 1, value='WINDOWS')
    ws_win.cell(2, 1, value='IDENTIFICAÇÃO')
    ws_win.cell(2, 2, value='PROPRIEDADES TÉRMICAS')
    ws_win.cell(2, 4, value='DIMENSÕES')
    ws_win.cell(3, 1, value='Nome')
    ws_win.cell(3, 2, value='U-Value (W/m²K)')
    ws_win.cell(3, 3, value='SHGC')
    ws_win.cell(3, 4, value='Altura (m)')
    ws_win.cell(3, 5, value='Largura (m)')

    for i, win in enumerate(data.get('windows', []), 4):
        ws_win.cell(i, 1, value=win['name'])
        ws_win.cell(i, 2, value=win['u_value'])
        ws_win.cell(i, 3, value=win['shgc'])
        ws_win.cell(i, 4, value=win['height'])
        ws_win.cell(i, 5, value=win['width'])

    ws_win.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E']:
        ws_win.column_dimensions[col].width = 15

    # Guardar
    wb.save(output_path)
    print(f"\nFicheiro guardado: {output_path}")
    print(f"  - {len(data['spaces'])} espaços")
    print(f"  - {len(data['walls'])} paredes")
    print(f"  - {len(data['roofs'])} coberturas")
    print(f"  - {len(data['windows'])} janelas")


# =============================================================================
# DETECÇÃO DE FORMATO
# =============================================================================

def detect_format(filepath):
    """Detecta automaticamente o formato do Excel"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    if 'INPUT SPACES HAP' in sheets:
        return 'hap52'
    elif 'Espacos' in sheets:
        return 'template'
    else:
        return 'unknown'


# =============================================================================
# CONVERSÃO
# =============================================================================

def convert_hap52_to_template(input_path, output_path=None):
    """Converte ficheiro HAP 5.2 para formato Template"""

    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = base + '_converted.xlsx'

    print(f"\n=== Converter HAP 5.2 -> Template ===")
    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    data = read_hap52_format(input_path)
    write_template_format(data, output_path)

    return output_path


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_file):
        print(f"Erro: Ficheiro não encontrado: {input_file}")
        sys.exit(1)

    fmt = detect_format(input_file)
    print(f"Formato detectado: {fmt}")

    if fmt == 'hap52':
        convert_hap52_to_template(input_file, output_file)
    elif fmt == 'template':
        print("Ficheiro já está no formato Template!")
    else:
        print("Formato desconhecido. Tentando converter como HAP 5.2...")
        convert_hap52_to_template(input_file, output_file)
