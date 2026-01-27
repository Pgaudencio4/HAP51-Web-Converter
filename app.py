"""
HAP 5.1 Web Application
=======================

Aplicacao web para converter entre Excel e ficheiros HAP 5.1 (.E3A).

Funcionalidades:
- Upload Excel -> Download E3A
- Upload E3A -> Download Excel

Executar:
    python app.py

Abrir no browser:
    http://localhost:5000
"""

import os
import tempfile
import shutil
import zipfile
import struct
from flask import Flask, render_template_string, request, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename

# Importar bibliotecas do projecto
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

app = Flask(__name__)
app.secret_key = 'hap51_converter_secret_key'

# Configuracao
UPLOAD_FOLDER = tempfile.mkdtemp()
ALLOWED_EXTENSIONS_EXCEL = {'xlsx', 'xls'}
ALLOWED_EXTENSIONS_E3A = {'e3a'}
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODELO_RSECE = os.path.join(BASE_DIR, 'Template_Limpo_RSECE.E3A')

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

# =============================================================================
# CONSTANTES E CONVERSOES (copiadas de excel_to_hap.py)
# =============================================================================

RECORD_SIZE = 682
SCHEDULE_RECORD_SIZE = 792
WALL_BLOCK_SIZE = 34
WALL_BLOCK_START = 72
ROOF_BLOCK_SIZE = 24
ROOF_BLOCK_START = 344

DIRECTION_CODES = {
    'N': 1, 'NNE': 2, 'NE': 3, 'ENE': 4,
    'E': 5, 'ESE': 6, 'SE': 7, 'SSE': 8,
    'S': 9, 'SSW': 10, 'SW': 11, 'WSW': 12,
    'W': 13, 'WNW': 14, 'NW': 15, 'NNW': 16
}

DIRECTION_NAMES = {v: k for k, v in DIRECTION_CODES.items()}

ACTIVITY_CODES = {
    'Seated at Rest': 0, 'Office Work': 3, 'Sedentary Work': 4,
    'Light Bench Work': 4, 'Medium Work': 5, 'Heavy Work': 6,
    'Dancing': 7, 'Athletics': 8,
}

FIXTURE_CODES = {
    'Recessed Unvented': 0, 'Vented to Return Air': 1,
    'Vented to Supply & Return': 2, 'Surface Mount/Pendant': 3,
}

FLOOR_TYPE_CODES = {
    'Floor Above Cond Space': 1, 'Floor Above Uncond Space': 2,
    'Slab Floor On Grade': 3, 'Slab Floor Below Grade': 4,
}

OA_UNIT_CODES = {'L/s': 1, 'L/s/m2': 2, 'L/s/person': 3, '%': 4}
OA_UNIT_NAMES = {v: k for k, v in OA_UNIT_CODES.items()}

OA_A = 0.00470356
OA_B = 2.71147770

import math

def m2_to_ft2(m2):
    return float(m2 or 0) * 10.7639

def ft2_to_m2(ft2):
    return float(ft2 or 0) / 10.7639

def m_to_ft(m):
    return float(m or 0) * 3.28084

def ft_to_m(ft):
    return float(ft or 0) / 3.28084

def kg_m2_to_lb_ft2(kg):
    return float(kg or 0) / 4.8824

def lb_ft2_to_kg_m2(lb):
    return float(lb or 0) * 4.8824

def u_si_to_ip(u):
    return float(u or 0) / 5.678

def u_ip_to_si(u):
    return float(u or 0) * 5.678

def w_to_btu(w):
    return float(w or 0) * 3.412

def btu_to_w(btu):
    return float(btu or 0) / 3.412

def w_m2_to_w_ft2(w):
    return float(w or 0) / 10.764

def w_ft2_to_w_m2(w):
    return float(w or 0) * 10.764

def encode_oa(value, unit_code):
    if not value or float(value) <= 0:
        return 0.0
    return math.log(float(value) / OA_A) / OA_B

def decode_oa(internal):
    if internal <= 0:
        return 0.0
    return OA_A * math.exp(OA_B * internal)

def safe_float(val, default=0.0):
    try:
        return float(val) if val else default
    except:
        return default

def safe_int(val, default=0):
    try:
        return int(val) if val else default
    except:
        return default

# =============================================================================
# FUNCOES DE CONVERSAO
# =============================================================================

def read_excel_spaces(excel_path):
    """Le espacos do Excel."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Espacos']

    spaces = []
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == '':
            continue

        space = {
            'name': str(name)[:24],
            'area': ws.cell(row=row, column=2).value,
            'height': ws.cell(row=row, column=3).value,
            'weight': ws.cell(row=row, column=4).value,
            'oa': ws.cell(row=row, column=5).value,
            'oa_unit': ws.cell(row=row, column=6).value,
            'occupancy': ws.cell(row=row, column=7).value,
            'activity': ws.cell(row=row, column=8).value,
            'sensible': ws.cell(row=row, column=9).value,
            'latent': ws.cell(row=row, column=10).value,
            'people_sch': ws.cell(row=row, column=11).value,
            'task_light': ws.cell(row=row, column=12).value,
            'general_light': ws.cell(row=row, column=13).value,
            'fixture': ws.cell(row=row, column=14).value,
            'ballast': ws.cell(row=row, column=15).value,
            'light_sch': ws.cell(row=row, column=16).value,
            'equipment': ws.cell(row=row, column=17).value,
            'equip_sch': ws.cell(row=row, column=18).value,
            'ach_clg': ws.cell(row=row, column=23).value,
            'ach_htg': ws.cell(row=row, column=24).value,
            'ach_energy': ws.cell(row=row, column=25).value,
        }

        # Walls
        space['walls'] = []
        for w in range(8):
            col = 51 + w * 9
            wall = {
                'exposure': ws.cell(row=row, column=col).value,
                'area': ws.cell(row=row, column=col+1).value,
                'type': ws.cell(row=row, column=col+2).value,
                'win1': ws.cell(row=row, column=col+3).value,
                'win1_qty': ws.cell(row=row, column=col+4).value,
                'win2': ws.cell(row=row, column=col+5).value,
                'win2_qty': ws.cell(row=row, column=col+6).value,
                'door': ws.cell(row=row, column=col+7).value,
                'door_qty': ws.cell(row=row, column=col+8).value,
            }
            space['walls'].append(wall)

        # Roofs (6 campos cada: exposure, area, slope, type, sky, sky_qty)
        space['roofs'] = []
        for r in range(4):
            col = 123 + r * 6
            roof = {
                'exposure': ws.cell(row=row, column=col).value,
                'area': ws.cell(row=row, column=col+1).value,
                'slope': ws.cell(row=row, column=col+2).value,
                'type': ws.cell(row=row, column=col+3).value,
                'sky': ws.cell(row=row, column=col+4).value,       # Skylight type
                'sky_qty': ws.cell(row=row, column=col+5).value,   # Skylight quantity
            }
            space['roofs'].append(roof)

        spaces.append(space)

    # Ler tipos
    types = {'walls': {}, 'windows': {}, 'doors': {}, 'roofs': {}, 'schedules': {}}
    if 'Tipos' in wb.sheetnames:
        ws_tipos = wb['Tipos']
        for row in range(3, ws_tipos.max_row + 1):
            # Schedules (cols 13-14)
            id_val = ws_tipos.cell(row=row, column=13).value
            name = ws_tipos.cell(row=row, column=14).value
            if id_val and name:
                types['schedules'][str(name).strip()] = int(id_val)

    return spaces, types


def read_e3a_spaces(e3a_path):
    """Le espacos de um ficheiro E3A."""
    spaces = []
    schedules = {}

    with zipfile.ZipFile(e3a_path, 'r') as zf:
        # Ler espacos
        spc_data = zf.read('HAP51SPC.DAT')
        num_spaces = len(spc_data) // RECORD_SIZE

        for i in range(1, num_spaces):  # Skip record 0 (default)
            offset = i * RECORD_SIZE
            record = spc_data[offset:offset + RECORD_SIZE]

            name = record[0:24].rstrip(b'\x00').decode('latin-1', errors='ignore').strip()
            if not name:
                continue

            area_ft2 = struct.unpack_from('<f', record, 24)[0]
            height_ft = struct.unpack_from('<f', record, 28)[0]
            weight_lb = struct.unpack_from('<f', record, 32)[0]
            oa_internal = struct.unpack_from('<f', record, 46)[0]
            oa_unit = struct.unpack_from('<H', record, 50)[0]

            occupancy = struct.unpack_from('<f', record, 580)[0]
            activity = struct.unpack_from('<H', record, 584)[0]
            sensible_btu = struct.unpack_from('<f', record, 586)[0]
            latent_btu = struct.unpack_from('<f', record, 590)[0]
            people_sch = struct.unpack_from('<H', record, 594)[0]

            task_light = struct.unpack_from('<f', record, 600)[0]
            fixture = struct.unpack_from('<H', record, 604)[0]
            general_light = struct.unpack_from('<f', record, 606)[0]
            ballast = struct.unpack_from('<f', record, 610)[0]
            light_sch = struct.unpack_from('<H', record, 614)[0]

            equip_w_ft2 = struct.unpack_from('<f', record, 656)[0]
            equip_sch = struct.unpack_from('<H', record, 660)[0]

            ach_clg = struct.unpack_from('<f', record, 556)[0]
            ach_htg = struct.unpack_from('<f', record, 562)[0]
            ach_energy = struct.unpack_from('<f', record, 568)[0]

            # Floor (offset 492-541)
            floor_type = struct.unpack_from('<H', record, 492)[0]
            floor_area = struct.unpack_from('<f', record, 494)[0]
            floor_u = struct.unpack_from('<f', record, 498)[0]

            # Ceiling Partition (offset 440-465)
            ceil_area = struct.unpack_from('<f', record, 442)[0]
            ceil_u = struct.unpack_from('<f', record, 446)[0]

            # Wall Partition (offset 466-491)
            part_wall_area = struct.unpack_from('<f', record, 468)[0]
            part_wall_u = struct.unpack_from('<f', record, 472)[0]

            # Walls
            walls = []
            for w in range(8):
                wall_offset = WALL_BLOCK_START + w * WALL_BLOCK_SIZE
                exp_code = struct.unpack_from('<H', record, wall_offset)[0]
                wall_area = struct.unpack_from('<f', record, wall_offset + 2)[0]
                wall_type = struct.unpack_from('<H', record, wall_offset + 6)[0]
                win1_type = struct.unpack_from('<H', record, wall_offset + 8)[0]
                win2_type = struct.unpack_from('<H', record, wall_offset + 10)[0]
                win1_qty = struct.unpack_from('<H', record, wall_offset + 12)[0]
                win2_qty = struct.unpack_from('<H', record, wall_offset + 14)[0]
                door_type = struct.unpack_from('<H', record, wall_offset + 16)[0]
                door_qty = struct.unpack_from('<H', record, wall_offset + 18)[0]
                if exp_code > 0:
                    walls.append({
                        'exposure': DIRECTION_NAMES.get(exp_code, ''),
                        'area': ft2_to_m2(wall_area),
                        'type_id': wall_type,
                        'win1_id': win1_type,
                        'win1_qty': win1_qty if win1_qty > 0 else None,
                        'win2_id': win2_type,
                        'win2_qty': win2_qty if win2_qty > 0 else None,
                        'door_id': door_type,
                        'door_qty': door_qty if door_qty > 0 else None,
                    })

            # Roofs
            roofs = []
            for r in range(4):
                roof_offset = ROOF_BLOCK_START + r * ROOF_BLOCK_SIZE
                exp_code = struct.unpack_from('<H', record, roof_offset)[0]
                slope = struct.unpack_from('<H', record, roof_offset + 2)[0]
                roof_area = struct.unpack_from('<f', record, roof_offset + 4)[0]
                roof_type = struct.unpack_from('<H', record, roof_offset + 8)[0]
                sky_type = struct.unpack_from('<H', record, roof_offset + 10)[0]
                sky_qty = struct.unpack_from('<H', record, roof_offset + 12)[0]
                if exp_code > 0:
                    roofs.append({
                        'exposure': DIRECTION_NAMES.get(exp_code, ''),
                        'area': ft2_to_m2(roof_area),
                        'slope': slope,
                        'type_id': roof_type,
                        'sky_id': sky_type,
                        'sky_qty': sky_qty if sky_qty > 0 else None,
                    })

            space = {
                'name': name,
                'area': ft2_to_m2(area_ft2),
                'height': ft_to_m(height_ft),
                'weight': lb_ft2_to_kg_m2(weight_lb),
                'oa': decode_oa(oa_internal),
                'oa_unit': OA_UNIT_NAMES.get(oa_unit, 'L/s'),
                'occupancy': occupancy,
                'sensible': btu_to_w(sensible_btu),
                'latent': btu_to_w(latent_btu),
                'people_sch_id': people_sch,
                'task_light': task_light,
                'general_light': general_light,
                'ballast': ballast,
                'light_sch_id': light_sch,
                'equipment': w_ft2_to_w_m2(equip_w_ft2),
                'equip_sch_id': equip_sch,
                'ach_clg': ach_clg,
                'ach_htg': ach_htg,
                'ach_energy': ach_energy,
                'floor_area': ft2_to_m2(floor_area),
                'floor_u': u_ip_to_si(floor_u),
                'ceil_area': ft2_to_m2(ceil_area),
                'ceil_u': u_ip_to_si(ceil_u),
                'part_wall_area': ft2_to_m2(part_wall_area),
                'part_wall_u': u_ip_to_si(part_wall_u),
                'walls': walls,
                'roofs': roofs,
            }
            spaces.append(space)

        # Ler schedules
        if 'HAP51SCH.DAT' in zf.namelist():
            sch_data = zf.read('HAP51SCH.DAT')
            num_schedules = len(sch_data) // SCHEDULE_RECORD_SIZE
            for i in range(num_schedules):
                offset = i * SCHEDULE_RECORD_SIZE
                name = sch_data[offset:offset+24].rstrip(b'\x00').decode('latin-1', errors='ignore').strip()
                if name:
                    schedules[i] = name

        # Ler Windows (HAP51WIN.DAT - 555 bytes cada)
        windows = {}
        if 'HAP51WIN.DAT' in zf.namelist():
            win_data = zf.read('HAP51WIN.DAT')
            WIN_RECORD_SIZE = 555
            num_windows = len(win_data) // WIN_RECORD_SIZE
            for i in range(num_windows):
                offset = i * WIN_RECORD_SIZE
                name = win_data[offset:offset+255].rstrip(b'\x00').rstrip(b' ').decode('latin-1', errors='ignore').strip()
                if name:
                    height_ft = struct.unpack_from('<f', win_data, offset + 257)[0]
                    width_ft = struct.unpack_from('<f', win_data, offset + 261)[0]
                    u_value_ip = struct.unpack_from('<f', win_data, offset + 269)[0]
                    shgc = struct.unpack_from('<f', win_data, offset + 273)[0]
                    windows[i] = {
                        'name': name,
                        'u_value': u_ip_to_si(u_value_ip),
                        'shgc': shgc,
                        'height': ft_to_m(height_ft),
                        'width': ft_to_m(width_ft),
                    }

        # Ler Walls (HAP51WAL.DAT) - usar MDB para nomes
        walls_types = {}

        # Ler Roofs (HAP51ROF.DAT) - usar MDB para nomes
        roofs_types = {}

        # Tentar ler nomes do MDB (mais fiavel)
        try:
            import pyodbc
            # Extrair MDB temporariamente
            mdb_data = zf.read('HAP51INX.MDB')
            mdb_temp = tempfile.mktemp(suffix='.mdb')
            with open(mdb_temp, 'wb') as f:
                f.write(mdb_data)

            conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={mdb_temp};'
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Ler WallIndex
            try:
                cursor.execute("SELECT nIndex, szName, fOverallUValue, fOverallWeight, fThickness FROM WallIndex")
                for row in cursor.fetchall():
                    if row.szName:
                        walls_types[row.nIndex] = {
                            'name': row.szName.strip(),
                            'u_value': u_ip_to_si(row.fOverallUValue) if row.fOverallUValue else None,
                            'weight': lb_ft2_to_kg_m2(row.fOverallWeight) if row.fOverallWeight else None,
                            'thickness': ft_to_m(row.fThickness) if row.fThickness else None,
                        }
            except:
                pass

            # Ler RoofIndex
            try:
                cursor.execute("SELECT nIndex, szName, fOverallUValue, fOverallWeight, fThickness FROM RoofIndex")
                for row in cursor.fetchall():
                    if row.szName:
                        roofs_types[row.nIndex] = {
                            'name': row.szName.strip(),
                            'u_value': u_ip_to_si(row.fOverallUValue) if row.fOverallUValue else None,
                            'weight': lb_ft2_to_kg_m2(row.fOverallWeight) if row.fOverallWeight else None,
                            'thickness': ft_to_m(row.fThickness) if row.fThickness else None,
                        }
            except:
                pass

            conn.close()
            os.remove(mdb_temp)
        except:
            pass

    return spaces, schedules, windows, walls_types, roofs_types


def create_space_binary(space, types, template_record):
    """Cria registo binario de 682 bytes para um espaco."""
    data = bytearray(template_record)

    # Nome
    name_bytes = space['name'].encode('latin-1', errors='ignore')[:24].ljust(24, b'\x00')
    data[0:24] = name_bytes

    # Area, Altura, Peso
    struct.pack_into('<f', data, 24, m2_to_ft2(space.get('area')))
    struct.pack_into('<f', data, 28, m_to_ft(space.get('height')))
    struct.pack_into('<f', data, 32, kg_m2_to_lb_ft2(space.get('weight')))

    # OA
    oa_unit = OA_UNIT_CODES.get(space.get('oa_unit'), 3)
    oa_internal = encode_oa(space.get('oa'), oa_unit)
    struct.pack_into('<f', data, 46, oa_internal)
    struct.pack_into('<H', data, 50, oa_unit)

    # Walls (34 bytes cada: exp(2) + area(4) + ?(2) + wall_type(2) + win1(2) + win2(2) + qty1(2) + qty2(2) + door(2) + door_qty(2) + ...)
    for i in range(8):
        wall_start = WALL_BLOCK_START + i * WALL_BLOCK_SIZE
        for j in range(WALL_BLOCK_SIZE):
            data[wall_start + j] = 0

        if i < len(space.get('walls', [])):
            wall = space['walls'][i]
            exp = wall.get('exposure')
            if exp and exp in DIRECTION_CODES:
                struct.pack_into('<H', data, wall_start, DIRECTION_CODES[exp])
                struct.pack_into('<f', data, wall_start + 2, m2_to_ft2(wall.get('area')))

                # Wall Type ID (offset +6)
                wall_type_name = wall.get('type')
                if wall_type_name and 'walls' in types and wall_type_name in types['walls']:
                    struct.pack_into('<H', data, wall_start + 6, types['walls'][wall_type_name])

                # Window 1 Type ID (offset +8) e Quantity (offset +10)
                win1_name = wall.get('win1')
                if win1_name and 'windows' in types and win1_name in types['windows']:
                    struct.pack_into('<H', data, wall_start + 8, types['windows'][win1_name])
                    struct.pack_into('<H', data, wall_start + 10, safe_int(wall.get('win1_qty', 1)))

                # Window 2 Type ID (offset +12) e Quantity (offset +14)
                win2_name = wall.get('win2')
                if win2_name and 'windows' in types and win2_name in types['windows']:
                    struct.pack_into('<H', data, wall_start + 12, types['windows'][win2_name])
                    struct.pack_into('<H', data, wall_start + 14, safe_int(wall.get('win2_qty', 1)))

                # Door Type ID (offset +16) e Quantity (offset +18)
                door_name = wall.get('door')
                if door_name and 'doors' in types and door_name in types['doors']:
                    struct.pack_into('<H', data, wall_start + 16, types['doors'][door_name])
                    struct.pack_into('<H', data, wall_start + 18, safe_int(wall.get('door_qty', 1)))

    # Roofs (24 bytes cada: exp(2) + slope(2) + area(4) + roof_type(2) + sky_type(2) + sky_qty(2) + ...)
    for i in range(4):
        roof_start = ROOF_BLOCK_START + i * ROOF_BLOCK_SIZE
        for j in range(ROOF_BLOCK_SIZE):
            data[roof_start + j] = 0

        if i < len(space.get('roofs', [])):
            roof = space['roofs'][i]
            exp = roof.get('exposure')
            if exp and exp in DIRECTION_CODES:
                struct.pack_into('<H', data, roof_start, DIRECTION_CODES[exp])
                struct.pack_into('<H', data, roof_start + 2, safe_int(roof.get('slope')))
                struct.pack_into('<f', data, roof_start + 4, m2_to_ft2(roof.get('area')))

                # Roof Type ID (offset +8)
                roof_type_name = roof.get('type')
                if roof_type_name and 'roofs' in types and roof_type_name in types['roofs']:
                    struct.pack_into('<H', data, roof_start + 8, types['roofs'][roof_type_name])

                # Skylight Type ID (offset +10) e Quantity (offset +12)
                # NOTA: Skylights usam os mesmos Window Types
                sky_name = roof.get('sky')
                if sky_name and 'windows' in types and sky_name in types['windows']:
                    struct.pack_into('<H', data, roof_start + 10, types['windows'][sky_name])
                    struct.pack_into('<H', data, roof_start + 12, safe_int(roof.get('sky_qty', 1)))

    # Infiltration
    struct.pack_into('<H', data, 554, 2)
    struct.pack_into('<f', data, 556, safe_float(space.get('ach_clg')))
    struct.pack_into('<H', data, 560, 2)
    struct.pack_into('<f', data, 562, safe_float(space.get('ach_htg')))
    struct.pack_into('<H', data, 566, 2)
    struct.pack_into('<f', data, 568, safe_float(space.get('ach_energy')))

    # People
    struct.pack_into('<f', data, 580, safe_float(space.get('occupancy')))
    struct.pack_into('<H', data, 584, ACTIVITY_CODES.get(space.get('activity'), 3))
    struct.pack_into('<f', data, 586, w_to_btu(space.get('sensible')))
    struct.pack_into('<f', data, 590, w_to_btu(space.get('latent')))

    # People Schedule (offset 594)
    people_sch = space.get('people_sch')
    if people_sch and 'schedules' in types and people_sch in types['schedules']:
        struct.pack_into('<H', data, 594, types['schedules'][people_sch])

    # Lighting
    struct.pack_into('<f', data, 600, safe_float(space.get('task_light')))
    struct.pack_into('<H', data, 604, FIXTURE_CODES.get(space.get('fixture'), 0))
    struct.pack_into('<f', data, 606, safe_float(space.get('general_light')))
    struct.pack_into('<f', data, 610, safe_float(space.get('ballast'), 1.0))

    # Light Schedule (offset 616) - IMPORTANTE: É 616, NÃO 614!
    light_sch = space.get('light_sch')
    if light_sch and 'schedules' in types and light_sch in types['schedules']:
        struct.pack_into('<H', data, 616, types['schedules'][light_sch])

    # Equipment
    struct.pack_into('<f', data, 656, w_m2_to_w_ft2(space.get('equipment')))

    # Equipment Schedule (offset 660)
    equip_sch = space.get('equip_sch')
    if equip_sch and 'schedules' in types and equip_sch in types['schedules']:
        struct.pack_into('<H', data, 660, types['schedules'][equip_sch])

    return bytes(data)


def read_excel_windows(wb):
    """Le tipos de janelas do Excel."""
    windows = {}
    if 'Windows' not in wb.sheetnames:
        return windows
    ws = wb['Windows']
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == '':
            continue
        windows[str(name).strip()] = {
            'u_value': ws.cell(row=row, column=2).value,  # W/m2K
            'shgc': ws.cell(row=row, column=3).value,
            'height': ws.cell(row=row, column=4).value,   # m
            'width': ws.cell(row=row, column=5).value,    # m
        }
    return windows


def read_excel_walls(wb):
    """Le tipos de paredes do Excel."""
    walls = {}
    if 'Walls' not in wb.sheetnames:
        return walls
    ws = wb['Walls']
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == '':
            continue
        walls[str(name).strip()] = {
            'u_value': ws.cell(row=row, column=2).value,    # W/m2K
            'weight': ws.cell(row=row, column=3).value,     # kg/m2
            'thickness': ws.cell(row=row, column=4).value,  # m
        }
    return walls


def read_excel_roofs(wb):
    """Le tipos de coberturas do Excel."""
    roofs = {}
    if 'Roofs' not in wb.sheetnames:
        return roofs
    ws = wb['Roofs']
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).strip() == '':
            continue
        roofs[str(name).strip()] = {
            'u_value': ws.cell(row=row, column=2).value,    # W/m2K
            'weight': ws.cell(row=row, column=3).value,     # kg/m2
            'thickness': ws.cell(row=row, column=4).value,  # m
        }
    return roofs


def create_window_record(name, data, template_record):
    """Cria registo binario de 555 bytes para uma janela."""
    WIN_RECORD_SIZE = 555
    record = bytearray(template_record[:WIN_RECORD_SIZE] if len(template_record) >= WIN_RECORD_SIZE else template_record.ljust(WIN_RECORD_SIZE, b'\x00'))

    # Nome (255 bytes)
    name_bytes = name.encode('latin-1', errors='ignore')[:255].ljust(255, b'\x00')
    record[0:255] = name_bytes

    # Height (offset 257)
    height_ft = m_to_ft(data.get('height') or 1.5)
    struct.pack_into('<f', record, 257, height_ft)

    # Width (offset 261)
    width_ft = m_to_ft(data.get('width') or 1.2)
    struct.pack_into('<f', record, 261, width_ft)

    # U-value (offset 269)
    u_ip = u_si_to_ip(data.get('u_value') or 3.0)
    struct.pack_into('<f', record, 269, u_ip)

    # SHGC (offset 273)
    shgc = float(data.get('shgc') or 0.4)
    struct.pack_into('<f', record, 273, shgc)

    return bytes(record)


def excel_to_e3a(excel_path, output_path):
    """Converte Excel para E3A, criando Walls/Windows/Roofs novos."""
    if not os.path.exists(MODELO_RSECE):
        raise FileNotFoundError(f"Modelo nao encontrado: {MODELO_RSECE}")

    # Ler Excel
    wb = openpyxl.load_workbook(excel_path)
    spaces, types = read_excel_spaces(excel_path)
    if not spaces:
        raise ValueError("Nenhum espaco encontrado no Excel")

    # Ler tipos do Excel
    excel_windows = read_excel_windows(wb)
    excel_walls = read_excel_walls(wb)
    excel_roofs = read_excel_roofs(wb)

    temp_dir = tempfile.mkdtemp()
    try:
        # Extrair modelo
        with zipfile.ZipFile(MODELO_RSECE, 'r') as zf:
            zf.extractall(temp_dir)

        # Ler template e schedules
        spc_path = os.path.join(temp_dir, 'HAP51SPC.DAT')
        with open(spc_path, 'rb') as f:
            spc_data = f.read()

        default_record = spc_data[0:RECORD_SIZE]
        # Usar default_record como template (template limpo só tem 1 registo)
        template_record = default_record if len(spc_data) <= RECORD_SIZE else spc_data[RECORD_SIZE:RECORD_SIZE*2]

        # Ler schedules do modelo
        sch_path = os.path.join(temp_dir, 'HAP51SCH.DAT')
        if os.path.exists(sch_path):
            with open(sch_path, 'rb') as f:
                sch_data = f.read()
            num_schedules = len(sch_data) // SCHEDULE_RECORD_SIZE
            for i in range(num_schedules):
                offset = i * SCHEDULE_RECORD_SIZE
                name = sch_data[offset:offset+24].rstrip(b'\x00').decode('latin-1', errors='ignore').strip()
                if name and name not in types['schedules']:
                    types['schedules'][name] = i

        # =====================================================================
        # CRIAR WINDOWS no HAP51WIN.DAT
        # =====================================================================
        win_path = os.path.join(temp_dir, 'HAP51WIN.DAT')
        WIN_RECORD_SIZE = 555

        # Ler template de window existente
        with open(win_path, 'rb') as f:
            win_template_data = f.read()
        win_template = win_template_data[:WIN_RECORD_SIZE]  # Primeiro registo como template

        # Criar novo ficheiro de windows
        new_win_data = bytearray(win_template)  # Manter o "Sample Window Assembly"
        window_ids = {'Sample Window Assembly': 0}

        for idx, (win_name, win_data) in enumerate(excel_windows.items(), 1):
            win_record = create_window_record(win_name, win_data, win_template)
            new_win_data.extend(win_record)
            window_ids[win_name] = idx
            types['windows'] = types.get('windows', {})
            types['windows'][win_name] = idx

        with open(win_path, 'wb') as f:
            f.write(bytes(new_win_data))

        # =====================================================================
        # CRIAR WALLS e ROOFS no MDB
        # =====================================================================
        mdb_path = os.path.join(temp_dir, 'HAP51INX.MDB')
        wall_ids = {}
        roof_ids = {}
        conn = None  # Inicializar conexão

        try:
            import pyodbc
            conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={mdb_path};'
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Limpar walls existentes (excepto Default)
            cursor.execute("DELETE FROM WallIndex WHERE nIndex > 0")

            # Criar walls novos
            for idx, (wall_name, wall_data) in enumerate(excel_walls.items(), 1):
                u_ip = u_si_to_ip(wall_data.get('u_value') or 0.5)
                weight_lb = kg_m2_to_lb_ft2(wall_data.get('weight') or 200)
                thickness_ft = m_to_ft(wall_data.get('thickness') or 0.3)

                cursor.execute("""
                    INSERT INTO WallIndex (nIndex, szName, fOverallUValue, fOverallWeight, fThickness)
                    VALUES (?, ?, ?, ?, ?)
                """, (idx, wall_name, u_ip, weight_lb, thickness_ft))

                wall_ids[wall_name] = idx
                types['walls'] = types.get('walls', {})
                types['walls'][wall_name] = idx

            # Limpar roofs existentes (excepto Default)
            cursor.execute("DELETE FROM RoofIndex WHERE nIndex > 0")

            # Criar roofs novos
            for idx, (roof_name, roof_data) in enumerate(excel_roofs.items(), 1):
                u_ip = u_si_to_ip(roof_data.get('u_value') or 0.4)
                weight_lb = kg_m2_to_lb_ft2(roof_data.get('weight') or 300)
                thickness_ft = m_to_ft(roof_data.get('thickness') or 0.4)

                cursor.execute("""
                    INSERT INTO RoofIndex (nIndex, szName, fOverallUValue, fOverallWeight, fThickness)
                    VALUES (?, ?, ?, ?, ?)
                """, (idx, roof_name, u_ip, weight_lb, thickness_ft))

                roof_ids[roof_name] = idx
                types['roofs'] = types.get('roofs', {})
                types['roofs'][roof_name] = idx

            # Limpar windows existentes (excepto Sample)
            cursor.execute("DELETE FROM WindowIndex WHERE nIndex > 0")

            # Criar windows novos no MDB
            for idx, (win_name, win_data) in enumerate(excel_windows.items(), 1):
                u_ip = u_si_to_ip(win_data.get('u_value') or 2.8)
                shgc = win_data.get('shgc') or 0.6
                height_ft = m_to_ft(win_data.get('height') or 1.2)
                width_ft = m_to_ft(win_data.get('width') or 1.0)

                cursor.execute("""
                    INSERT INTO WindowIndex (nIndex, szName, fOverallUValue, fOverallShadeCo, fHeight, fWidth)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (idx, win_name, u_ip, shgc, height_ft, width_ft))

            conn.commit()

            # NÃO FECHAR CONEXÃO AQUI - será usada para SpaceIndex também
        except Exception as e:
            print(f"ERRO MDB Walls/Roofs/Windows: {e}")
            conn = None

        # =====================================================================
        # CRIAR ESPACOS
        # =====================================================================
        new_spc_data = bytearray(default_record)
        for space in spaces:
            space_binary = create_space_binary(space, types, template_record)
            new_spc_data.extend(space_binary)

        with open(spc_path, 'wb') as f:
            f.write(bytes(new_spc_data))

        # =====================================================================
        # ACTUALIZAR SpaceIndex no MDB (reutilizar conexão se existir)
        # =====================================================================
        try:
            if conn is None:
                import pyodbc
                conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={mdb_path};'
                conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            # Limpar SpaceIndex existente (TODOS os registos!)
            cursor.execute("DELETE FROM SpaceIndex")

            # Limpar tabelas de links existentes
            cursor.execute("DELETE FROM Space_Schedule_Links")
            cursor.execute("DELETE FROM Space_Wall_Links")
            cursor.execute("DELETE FROM Space_Window_Links")
            cursor.execute("DELETE FROM Space_Roof_Links")

            # Adicionar espacos novos com todos os campos obrigatorios
            for idx, space in enumerate(spaces, 1):
                # Nome com padding de espacos (24 chars)
                name_padded = space['name'][:24].ljust(24)
                # Area em ft2
                area_ft2 = m2_to_ft2(space.get('area') or 0)
                # Numero de pessoas
                num_people = float(space.get('occupancy') or 0)
                # Iluminacao total (task + general) em W, nao densidade
                task_w = float(space.get('task_light') or 0)
                general_w = float(space.get('general_light') or 0)
                lighting_total = task_w + general_w

                cursor.execute("""
                    INSERT INTO SpaceIndex (nIndex, szName, fFloorArea, fNumPeople, fLightingDensity)
                    VALUES (?, ?, ?, ?, ?)
                """, (idx, name_padded, area_ft2, num_people, lighting_total))

                # =========================================================
                # CRIAR LINKS PARA ESTE ESPACO
                # NOTA: Schedule IDs no MDB = DAT index + 1
                # =========================================================

                # Space_Schedule_Links (People, Light, Equip schedules)
                # Os IDs no MDB são offset +1 em relação ao DAT
                people_sch = space.get('people_sch')
                if people_sch and people_sch in types.get('schedules', {}):
                    mdb_sch_id = types['schedules'][people_sch] + 1  # +1 para MDB
                    cursor.execute("INSERT INTO Space_Schedule_Links (Space_ID, Schedule_ID) VALUES (?, ?)",
                                   (idx, mdb_sch_id))

                light_sch = space.get('light_sch')
                if light_sch and light_sch in types.get('schedules', {}):
                    mdb_sch_id = types['schedules'][light_sch] + 1  # +1 para MDB
                    cursor.execute("INSERT INTO Space_Schedule_Links (Space_ID, Schedule_ID) VALUES (?, ?)",
                                   (idx, mdb_sch_id))

                equip_sch = space.get('equip_sch')
                if equip_sch and equip_sch in types.get('schedules', {}):
                    mdb_sch_id = types['schedules'][equip_sch] + 1  # +1 para MDB
                    cursor.execute("INSERT INTO Space_Schedule_Links (Space_ID, Schedule_ID) VALUES (?, ?)",
                                   (idx, mdb_sch_id))

                # Space_Wall_Links (evitar duplicados)
                wall_links_added = set()
                for wall in space.get('walls', []):
                    wall_type_name = wall.get('type')
                    if wall_type_name and wall_type_name in types.get('walls', {}):
                        wall_id = types['walls'][wall_type_name]
                        if wall_id not in wall_links_added:
                            cursor.execute("INSERT INTO Space_Wall_Links (Space_ID, Wall_ID) VALUES (?, ?)",
                                           (idx, wall_id))
                            wall_links_added.add(wall_id)

                # Space_Window_Links (evitar duplicados)
                window_links_added = set()
                for wall in space.get('walls', []):
                    win1_name = wall.get('win1')
                    if win1_name and win1_name in types.get('windows', {}):
                        win_id = types['windows'][win1_name]
                        if win_id not in window_links_added:
                            cursor.execute("INSERT INTO Space_Window_Links (Space_ID, Window_ID) VALUES (?, ?)",
                                           (idx, win_id))
                            window_links_added.add(win_id)
                    win2_name = wall.get('win2')
                    if win2_name and win2_name in types.get('windows', {}):
                        win_id = types['windows'][win2_name]
                        if win_id not in window_links_added:
                            cursor.execute("INSERT INTO Space_Window_Links (Space_ID, Window_ID) VALUES (?, ?)",
                                           (idx, win_id))
                            window_links_added.add(win_id)

                # Space_Roof_Links (evitar duplicados)
                roof_links_added = set()
                for roof in space.get('roofs', []):
                    roof_type_name = roof.get('type')
                    if roof_type_name and roof_type_name in types.get('roofs', {}):
                        roof_id = types['roofs'][roof_type_name]
                        if roof_id not in roof_links_added:
                            cursor.execute("INSERT INTO Space_Roof_Links (Space_ID, Roof_ID) VALUES (?, ?)",
                                           (idx, roof_id))
                            roof_links_added.add(roof_id)
                    # Skylights usam Window types (evitar duplicados)
                    sky_name = roof.get('sky')
                    if sky_name and sky_name in types.get('windows', {}):
                        win_id = types['windows'][sky_name]
                        if win_id not in window_links_added:
                            cursor.execute("INSERT INTO Space_Window_Links (Space_ID, Window_ID) VALUES (?, ?)",
                                           (idx, win_id))
                            window_links_added.add(win_id)

            conn.commit()
        except Exception as e:
            # Log do erro para debug
            print(f"ERRO MDB SpaceIndex: {e}")
        finally:
            # Garantir que conexão é sempre fechada
            if conn:
                try:
                    conn.close()
                except:
                    pass

        # Aguardar um pouco para o ficheiro ser libertado
        import time
        time.sleep(0.5)

        # Remover ficheiro de lock do MDB se existir
        ldb_path = os.path.join(temp_dir, 'HAP51INX.ldb')
        if os.path.exists(ldb_path):
            try:
                os.remove(ldb_path)
            except:
                pass  # Ignorar se não conseguir remover

        # Criar ZIP (excluir ficheiros .ldb)
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith('.ldb'):
                        continue  # Ignorar ficheiros de lock
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, temp_dir)
                    zf.write(file_path, arc_name)

        return len(spaces)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def read_e3a_full(e3a_path):
    """Le todos os dados de um ficheiro E3A de forma completa."""
    spaces = []
    schedules = {}
    windows = {}
    walls_types = {}
    roofs_types = {}
    doors = {}

    with zipfile.ZipFile(e3a_path, 'r') as zf:
        # Ler espacos
        spc_data = zf.read('HAP51SPC.DAT')
        num_spaces = len(spc_data) // RECORD_SIZE

        for i in range(num_spaces):
            offset = i * RECORD_SIZE
            record = spc_data[offset:offset + RECORD_SIZE]

            name = record[0:24].rstrip(b'\x00').decode('latin-1', errors='ignore').strip()

            area_ft2 = struct.unpack_from('<f', record, 24)[0]
            height_ft = struct.unpack_from('<f', record, 28)[0]
            weight_lb = struct.unpack_from('<f', record, 32)[0]
            oa_internal = struct.unpack_from('<f', record, 46)[0]
            oa_unit = struct.unpack_from('<H', record, 50)[0]

            occupancy = struct.unpack_from('<f', record, 580)[0]
            activity = struct.unpack_from('<H', record, 584)[0]
            sensible_btu = struct.unpack_from('<f', record, 586)[0]
            latent_btu = struct.unpack_from('<f', record, 590)[0]
            people_sch = struct.unpack_from('<H', record, 594)[0]

            task_light = struct.unpack_from('<f', record, 600)[0]
            fixture = struct.unpack_from('<H', record, 604)[0]
            general_light = struct.unpack_from('<f', record, 606)[0]
            ballast = struct.unpack_from('<f', record, 610)[0]
            light_sch = struct.unpack_from('<H', record, 614)[0]

            equip_w_ft2 = struct.unpack_from('<f', record, 656)[0]
            equip_sch = struct.unpack_from('<H', record, 660)[0]

            ach_clg = struct.unpack_from('<f', record, 556)[0]
            ach_htg = struct.unpack_from('<f', record, 562)[0]
            ach_energy = struct.unpack_from('<f', record, 568)[0]

            # Floor
            floor_type = struct.unpack_from('<H', record, 492)[0]
            floor_area = struct.unpack_from('<f', record, 494)[0]
            floor_u = struct.unpack_from('<f', record, 498)[0]

            # Ceiling Partition
            ceil_area = struct.unpack_from('<f', record, 442)[0]
            ceil_u = struct.unpack_from('<f', record, 446)[0]

            # Wall Partition
            part_wall_area = struct.unpack_from('<f', record, 468)[0]
            part_wall_u = struct.unpack_from('<f', record, 472)[0]

            # Walls
            walls = []
            for w in range(8):
                wall_offset = WALL_BLOCK_START + w * WALL_BLOCK_SIZE
                exp_code = struct.unpack_from('<H', record, wall_offset)[0]
                wall_area = struct.unpack_from('<f', record, wall_offset + 2)[0]
                wall_type = struct.unpack_from('<H', record, wall_offset + 6)[0]
                win1_type = struct.unpack_from('<H', record, wall_offset + 8)[0]
                win2_type = struct.unpack_from('<H', record, wall_offset + 10)[0]
                win1_qty = struct.unpack_from('<H', record, wall_offset + 12)[0]
                win2_qty = struct.unpack_from('<H', record, wall_offset + 14)[0]
                door_type = struct.unpack_from('<H', record, wall_offset + 16)[0]
                door_qty = struct.unpack_from('<H', record, wall_offset + 18)[0]

                walls.append({
                    'exposure_code': exp_code,
                    'exposure': DIRECTION_NAMES.get(exp_code, ''),
                    'area_ft2': wall_area,
                    'area': ft2_to_m2(wall_area),
                    'type_id': wall_type,
                    'win1_id': win1_type,
                    'win1_qty': win1_qty,
                    'win2_id': win2_type,
                    'win2_qty': win2_qty,
                    'door_id': door_type,
                    'door_qty': door_qty,
                })

            # Roofs
            roofs = []
            for r in range(4):
                roof_offset = ROOF_BLOCK_START + r * ROOF_BLOCK_SIZE
                exp_code = struct.unpack_from('<H', record, roof_offset)[0]
                slope = struct.unpack_from('<H', record, roof_offset + 2)[0]
                roof_area = struct.unpack_from('<f', record, roof_offset + 4)[0]
                roof_type = struct.unpack_from('<H', record, roof_offset + 8)[0]
                sky_type = struct.unpack_from('<H', record, roof_offset + 10)[0]
                sky_qty = struct.unpack_from('<H', record, roof_offset + 12)[0]

                roofs.append({
                    'exposure_code': exp_code,
                    'exposure': DIRECTION_NAMES.get(exp_code, ''),
                    'slope': slope,
                    'area_ft2': roof_area,
                    'area': ft2_to_m2(roof_area),
                    'type_id': roof_type,
                    'sky_id': sky_type,
                    'sky_qty': sky_qty,
                })

            space = {
                'index': i,
                'name': name,
                'area_ft2': area_ft2,
                'area': ft2_to_m2(area_ft2),
                'height_ft': height_ft,
                'height': ft_to_m(height_ft),
                'weight_lb': weight_lb,
                'weight': lb_ft2_to_kg_m2(weight_lb),
                'oa_internal': oa_internal,
                'oa': decode_oa(oa_internal),
                'oa_unit_code': oa_unit,
                'oa_unit': OA_UNIT_NAMES.get(oa_unit, 'L/s'),
                'occupancy': occupancy,
                'activity_code': activity,
                'sensible_btu': sensible_btu,
                'sensible': btu_to_w(sensible_btu),
                'latent_btu': latent_btu,
                'latent': btu_to_w(latent_btu),
                'people_sch_id': people_sch,
                'task_light': task_light,
                'fixture_code': fixture,
                'general_light': general_light,
                'ballast': ballast,
                'light_sch_id': light_sch,
                'equipment_ft2': equip_w_ft2,
                'equipment': w_ft2_to_w_m2(equip_w_ft2),
                'equip_sch_id': equip_sch,
                'ach_clg': ach_clg,
                'ach_htg': ach_htg,
                'ach_energy': ach_energy,
                'floor_type': floor_type,
                'floor_area_ft2': floor_area,
                'floor_area': ft2_to_m2(floor_area),
                'floor_u_ip': floor_u,
                'floor_u': u_ip_to_si(floor_u),
                'ceil_area_ft2': ceil_area,
                'ceil_area': ft2_to_m2(ceil_area),
                'ceil_u_ip': ceil_u,
                'ceil_u': u_ip_to_si(ceil_u),
                'part_wall_area_ft2': part_wall_area,
                'part_wall_area': ft2_to_m2(part_wall_area),
                'part_wall_u_ip': part_wall_u,
                'part_wall_u': u_ip_to_si(part_wall_u),
                'walls': walls,
                'roofs': roofs,
            }
            spaces.append(space)

        # Ler schedules
        if 'HAP51SCH.DAT' in zf.namelist():
            sch_data = zf.read('HAP51SCH.DAT')
            num_schedules = len(sch_data) // SCHEDULE_RECORD_SIZE
            for i in range(num_schedules):
                offset = i * SCHEDULE_RECORD_SIZE
                name = sch_data[offset:offset+24].rstrip(b'\x00').decode('latin-1', errors='ignore').strip()
                if name:
                    schedules[i] = name

        # Ler Windows
        if 'HAP51WIN.DAT' in zf.namelist():
            win_data = zf.read('HAP51WIN.DAT')
            WIN_RECORD_SIZE = 555
            num_windows = len(win_data) // WIN_RECORD_SIZE
            for i in range(num_windows):
                offset = i * WIN_RECORD_SIZE
                name = win_data[offset:offset+255].rstrip(b'\x00').rstrip(b' ').decode('latin-1', errors='ignore').strip()
                if name:
                    height_ft = struct.unpack_from('<f', win_data, offset + 257)[0]
                    width_ft = struct.unpack_from('<f', win_data, offset + 261)[0]
                    u_value_ip = struct.unpack_from('<f', win_data, offset + 269)[0]
                    shgc = struct.unpack_from('<f', win_data, offset + 273)[0]
                    windows[i] = {
                        'name': name,
                        'u_value_ip': u_value_ip,
                        'u_value': u_ip_to_si(u_value_ip),
                        'shgc': shgc,
                        'height_ft': height_ft,
                        'height': ft_to_m(height_ft),
                        'width_ft': width_ft,
                        'width': ft_to_m(width_ft),
                    }

        # Ler Walls e Roofs do MDB
        try:
            import pyodbc
            mdb_data = zf.read('HAP51INX.MDB')
            mdb_temp = tempfile.mktemp(suffix='.mdb')
            with open(mdb_temp, 'wb') as f:
                f.write(mdb_data)

            conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={mdb_temp};'
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            try:
                cursor.execute("SELECT nIndex, szName, fOverallUValue, fOverallWeight, fThickness FROM WallIndex")
                for row in cursor.fetchall():
                    if row.szName:
                        walls_types[row.nIndex] = {
                            'name': row.szName.strip(),
                            'u_value_ip': row.fOverallUValue,
                            'u_value': u_ip_to_si(row.fOverallUValue) if row.fOverallUValue else None,
                            'weight_lb': row.fOverallWeight,
                            'weight': lb_ft2_to_kg_m2(row.fOverallWeight) if row.fOverallWeight else None,
                            'thickness_ft': row.fThickness,
                            'thickness': ft_to_m(row.fThickness) if row.fThickness else None,
                        }
            except:
                pass

            try:
                cursor.execute("SELECT nIndex, szName, fOverallUValue, fOverallWeight, fThickness FROM RoofIndex")
                for row in cursor.fetchall():
                    if row.szName:
                        roofs_types[row.nIndex] = {
                            'name': row.szName.strip(),
                            'u_value_ip': row.fOverallUValue,
                            'u_value': u_ip_to_si(row.fOverallUValue) if row.fOverallUValue else None,
                            'weight_lb': row.fOverallWeight,
                            'weight': lb_ft2_to_kg_m2(row.fOverallWeight) if row.fOverallWeight else None,
                            'thickness_ft': row.fThickness,
                            'thickness': ft_to_m(row.fThickness) if row.fThickness else None,
                        }
            except:
                pass

            conn.close()
            os.remove(mdb_temp)
        except:
            pass

    return {
        'spaces': spaces,
        'schedules': schedules,
        'windows': windows,
        'walls': walls_types,
        'roofs': roofs_types,
    }


def apply_excel_changes_to_e3a(e3a_path, excel_path, output_path):
    """
    Aplica alteracoes do Excel ao E3A original.
    Compara os dados do Excel com o E3A e aplica apenas as diferencas.
    Retorna relatorio de alteracoes.
    """
    # Ler dados do E3A original
    e3a_data = read_e3a_full(e3a_path)
    e3a_spaces = {s['name']: s for s in e3a_data['spaces']}

    # Ler dados do Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Espacos']

    changes = []

    # Extrair E3A para pasta temporaria
    temp_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(e3a_path, 'r') as zf:
            zf.extractall(temp_dir)

        # Ler ficheiro de espacos
        spc_path = os.path.join(temp_dir, 'HAP51SPC.DAT')
        with open(spc_path, 'rb') as f:
            spc_data = bytearray(f.read())

        # Processar cada linha do Excel
        for row in range(4, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if not name or str(name).strip() == '':
                continue

            name = str(name).strip()
            if name not in e3a_spaces:
                continue

            space = e3a_spaces[name]
            idx = space['index']
            offset = idx * RECORD_SIZE

            space_changes = []

            # Comparar e aplicar alteracoes

            # Area (col 2)
            new_area = ws.cell(row=row, column=2).value
            if new_area is not None:
                new_area_ft2 = m2_to_ft2(new_area)
                if abs(new_area_ft2 - space['area_ft2']) > 0.1:
                    struct.pack_into('<f', spc_data, offset + 24, new_area_ft2)
                    space_changes.append(f"Area: {space['area']:.1f} -> {new_area:.1f} m2")

            # Altura (col 3)
            new_height = ws.cell(row=row, column=3).value
            if new_height is not None:
                new_height_ft = m_to_ft(new_height)
                if abs(new_height_ft - space['height_ft']) > 0.01:
                    struct.pack_into('<f', spc_data, offset + 28, new_height_ft)
                    space_changes.append(f"Altura: {space['height']:.2f} -> {new_height:.2f} m")

            # Peso (col 4)
            new_weight = ws.cell(row=row, column=4).value
            if new_weight is not None:
                new_weight_lb = kg_m2_to_lb_ft2(new_weight)
                if abs(new_weight_lb - space['weight_lb']) > 0.1:
                    struct.pack_into('<f', spc_data, offset + 32, new_weight_lb)
                    space_changes.append(f"Peso: {space['weight']:.1f} -> {new_weight:.1f} kg/m2")

            # OA (col 5-6)
            new_oa = ws.cell(row=row, column=5).value
            new_oa_unit = ws.cell(row=row, column=6).value
            if new_oa is not None:
                oa_unit_code = OA_UNIT_CODES.get(new_oa_unit, space['oa_unit_code'])
                new_oa_internal = encode_oa(new_oa, oa_unit_code)
                if abs(new_oa_internal - space['oa_internal']) > 0.001:
                    struct.pack_into('<f', spc_data, offset + 46, new_oa_internal)
                    struct.pack_into('<H', spc_data, offset + 50, oa_unit_code)
                    space_changes.append(f"OA: {space['oa']:.1f} -> {new_oa:.1f}")

            # Ocupacao (col 7)
            new_occ = ws.cell(row=row, column=7).value
            if new_occ is not None and abs(float(new_occ) - space['occupancy']) > 0.01:
                struct.pack_into('<f', spc_data, offset + 580, float(new_occ))
                space_changes.append(f"Ocupacao: {space['occupancy']:.0f} -> {new_occ:.0f}")

            # Sensible (col 9)
            new_sens = ws.cell(row=row, column=9).value
            if new_sens is not None:
                new_sens_btu = w_to_btu(new_sens)
                if abs(new_sens_btu - space['sensible_btu']) > 0.1:
                    struct.pack_into('<f', spc_data, offset + 586, new_sens_btu)
                    space_changes.append(f"Sensible: {space['sensible']:.0f} -> {new_sens:.0f} W")

            # Latent (col 10)
            new_lat = ws.cell(row=row, column=10).value
            if new_lat is not None:
                new_lat_btu = w_to_btu(new_lat)
                if abs(new_lat_btu - space['latent_btu']) > 0.1:
                    struct.pack_into('<f', spc_data, offset + 590, new_lat_btu)
                    space_changes.append(f"Latent: {space['latent']:.0f} -> {new_lat:.0f} W")

            # People Schedule (col 11)
            new_ppl_sch = ws.cell(row=row, column=11).value
            if new_ppl_sch:
                sch_id = None
                for sid, sname in e3a_data['schedules'].items():
                    if sname == new_ppl_sch:
                        sch_id = sid
                        break
                if sch_id is not None and sch_id != space['people_sch_id']:
                    struct.pack_into('<H', spc_data, offset + 594, sch_id)
                    old_sch = e3a_data['schedules'].get(space['people_sch_id'], '?')
                    space_changes.append(f"People Sch: {old_sch} -> {new_ppl_sch}")

            # Task Light (col 12)
            new_task = ws.cell(row=row, column=12).value
            if new_task is not None and abs(float(new_task) - space['task_light']) > 0.01:
                struct.pack_into('<f', spc_data, offset + 600, float(new_task))
                space_changes.append(f"Task Light: {space['task_light']:.1f} -> {new_task:.1f}")

            # General Light (col 13)
            new_gen = ws.cell(row=row, column=13).value
            if new_gen is not None and abs(float(new_gen) - space['general_light']) > 0.01:
                struct.pack_into('<f', spc_data, offset + 606, float(new_gen))
                space_changes.append(f"General Light: {space['general_light']:.1f} -> {new_gen:.1f}")

            # Ballast (col 15)
            new_bal = ws.cell(row=row, column=15).value
            if new_bal is not None and abs(float(new_bal) - space['ballast']) > 0.001:
                struct.pack_into('<f', spc_data, offset + 610, float(new_bal))
                space_changes.append(f"Ballast: {space['ballast']:.2f} -> {new_bal:.2f}")

            # Light Schedule (col 16)
            new_light_sch = ws.cell(row=row, column=16).value
            if new_light_sch:
                sch_id = None
                for sid, sname in e3a_data['schedules'].items():
                    if sname == new_light_sch:
                        sch_id = sid
                        break
                if sch_id is not None and sch_id != space['light_sch_id']:
                    struct.pack_into('<H', spc_data, offset + 616, sch_id)
                    old_sch = e3a_data['schedules'].get(space['light_sch_id'], '?')
                    space_changes.append(f"Light Sch: {old_sch} -> {new_light_sch}")

            # Equipment (col 17)
            new_equip = ws.cell(row=row, column=17).value
            if new_equip is not None:
                new_equip_ft2 = w_m2_to_w_ft2(new_equip)
                if abs(new_equip_ft2 - space['equipment_ft2']) > 0.001:
                    struct.pack_into('<f', spc_data, offset + 656, new_equip_ft2)
                    space_changes.append(f"Equipment: {space['equipment']:.1f} -> {new_equip:.1f} W/m2")

            # Equip Schedule (col 18)
            new_equip_sch = ws.cell(row=row, column=18).value
            if new_equip_sch:
                sch_id = None
                for sid, sname in e3a_data['schedules'].items():
                    if sname == new_equip_sch:
                        sch_id = sid
                        break
                if sch_id is not None and sch_id != space['equip_sch_id']:
                    struct.pack_into('<H', spc_data, offset + 660, sch_id)
                    old_sch = e3a_data['schedules'].get(space['equip_sch_id'], '?')
                    space_changes.append(f"Equip Sch: {old_sch} -> {new_equip_sch}")

            # Infiltration ACH (cols 23-25)
            new_ach_clg = ws.cell(row=row, column=23).value
            if new_ach_clg is not None and abs(float(new_ach_clg) - space['ach_clg']) > 0.001:
                struct.pack_into('<f', spc_data, offset + 556, float(new_ach_clg))
                space_changes.append(f"ACH Clg: {space['ach_clg']:.2f} -> {new_ach_clg:.2f}")

            new_ach_htg = ws.cell(row=row, column=24).value
            if new_ach_htg is not None and abs(float(new_ach_htg) - space['ach_htg']) > 0.001:
                struct.pack_into('<f', spc_data, offset + 562, float(new_ach_htg))
                space_changes.append(f"ACH Htg: {space['ach_htg']:.2f} -> {new_ach_htg:.2f}")

            new_ach_energy = ws.cell(row=row, column=25).value
            if new_ach_energy is not None and abs(float(new_ach_energy) - space['ach_energy']) > 0.001:
                struct.pack_into('<f', spc_data, offset + 568, float(new_ach_energy))
                space_changes.append(f"ACH Energy: {space['ach_energy']:.2f} -> {new_ach_energy:.2f}")

            # Floor U-value (col 28)
            new_floor_u = ws.cell(row=row, column=28).value
            if new_floor_u is not None:
                new_floor_u_ip = u_si_to_ip(new_floor_u)
                if abs(new_floor_u_ip - space['floor_u_ip']) > 0.001:
                    struct.pack_into('<f', spc_data, offset + 498, new_floor_u_ip)
                    space_changes.append(f"Floor U: {space['floor_u']:.3f} -> {new_floor_u:.3f} W/m2K")

            # Ceiling Partition U-value (col 40)
            new_ceil_u = ws.cell(row=row, column=40).value
            if new_ceil_u is not None:
                new_ceil_u_ip = u_si_to_ip(new_ceil_u)
                if abs(new_ceil_u_ip - space['ceil_u_ip']) > 0.001:
                    struct.pack_into('<f', spc_data, offset + 446, new_ceil_u_ip)
                    space_changes.append(f"Ceiling U: {space['ceil_u']:.3f} -> {new_ceil_u:.3f} W/m2K")

            # Wall Partition U-value (col 46)
            new_part_u = ws.cell(row=row, column=46).value
            if new_part_u is not None:
                new_part_u_ip = u_si_to_ip(new_part_u)
                if abs(new_part_u_ip - space['part_wall_u_ip']) > 0.001:
                    struct.pack_into('<f', spc_data, offset + 472, new_part_u_ip)
                    space_changes.append(f"Part Wall U: {space['part_wall_u']:.3f} -> {new_part_u:.3f} W/m2K")

            if space_changes:
                changes.append({
                    'space': name,
                    'changes': space_changes
                })

        # Guardar ficheiro de espacos modificado
        with open(spc_path, 'wb') as f:
            f.write(bytes(spc_data))

        # =====================================================================
        # WINDOWS - Editar U-values e SHGC no HAP51WIN.DAT
        # =====================================================================
        win_path = os.path.join(temp_dir, 'HAP51WIN.DAT')
        WIN_RECORD_SIZE = 555
        if os.path.exists(win_path) and 'Windows' in wb.sheetnames:
            ws_win = wb['Windows']
            with open(win_path, 'rb') as f:
                win_data = bytearray(f.read())

            # Criar dicionario de windows do E3A por nome
            e3a_windows = {}
            num_windows = len(win_data) // WIN_RECORD_SIZE
            for i in range(num_windows):
                offset = i * WIN_RECORD_SIZE
                name = win_data[offset:offset+255].rstrip(b'\x00').rstrip(b' ').decode('latin-1', errors='ignore').strip()
                if name:
                    e3a_windows[name] = {
                        'index': i,
                        'u_value_ip': struct.unpack_from('<f', win_data, offset + 269)[0],
                        'shgc': struct.unpack_from('<f', win_data, offset + 273)[0],
                    }

            # Processar alteracoes do Excel
            for row in range(4, ws_win.max_row + 1):
                win_name = ws_win.cell(row=row, column=1).value
                if not win_name or str(win_name).strip() == '':
                    continue
                win_name = str(win_name).strip()

                if win_name not in e3a_windows:
                    continue

                win_info = e3a_windows[win_name]
                win_offset = win_info['index'] * WIN_RECORD_SIZE
                win_changes = []

                # U-value (col 2)
                new_u = ws_win.cell(row=row, column=2).value
                if new_u is not None:
                    new_u_ip = u_si_to_ip(new_u)
                    if abs(new_u_ip - win_info['u_value_ip']) > 0.0001:
                        struct.pack_into('<f', win_data, win_offset + 269, new_u_ip)
                        old_u_si = u_ip_to_si(win_info['u_value_ip'])
                        win_changes.append(f"U: {old_u_si:.3f} -> {new_u:.3f}")

                # SHGC (col 3)
                new_shgc = ws_win.cell(row=row, column=3).value
                if new_shgc is not None and abs(float(new_shgc) - win_info['shgc']) > 0.001:
                    struct.pack_into('<f', win_data, win_offset + 273, float(new_shgc))
                    win_changes.append(f"SHGC: {win_info['shgc']:.3f} -> {new_shgc:.3f}")

                if win_changes:
                    changes.append({
                        'space': f"Window: {win_name}",
                        'changes': win_changes
                    })

            with open(win_path, 'wb') as f:
                f.write(bytes(win_data))

        # =====================================================================
        # WALLS e ROOFS - Editar U-values no MDB (Access Database)
        # =====================================================================
        mdb_path = os.path.join(temp_dir, 'HAP51INX.MDB')
        if os.path.exists(mdb_path):
            try:
                import pyodbc
                conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={mdb_path};'
                conn = pyodbc.connect(conn_str)
                cursor = conn.cursor()

                # WALLS
                if 'Walls' in wb.sheetnames:
                    ws_wal = wb['Walls']
                    # Ler walls actuais do MDB
                    cursor.execute("SELECT nIndex, szName, fOverallUValue FROM WallIndex")
                    e3a_walls = {}
                    for row in cursor.fetchall():
                        if row.szName:
                            e3a_walls[row.szName.strip()] = {
                                'index': row.nIndex,
                                'u_value_ip': row.fOverallUValue or 0,
                            }

                    for row in range(4, ws_wal.max_row + 1):
                        wall_name = ws_wal.cell(row=row, column=1).value
                        if not wall_name or str(wall_name).strip() == '':
                            continue
                        wall_name = str(wall_name).strip()

                        if wall_name not in e3a_walls:
                            continue

                        wall_info = e3a_walls[wall_name]
                        wall_changes = []

                        # U-value (col 2)
                        new_u = ws_wal.cell(row=row, column=2).value
                        if new_u is not None:
                            new_u_ip = u_si_to_ip(new_u)
                            if abs(new_u_ip - wall_info['u_value_ip']) > 0.0001:
                                cursor.execute(
                                    "UPDATE WallIndex SET fOverallUValue = ? WHERE nIndex = ?",
                                    (new_u_ip, wall_info['index'])
                                )
                                old_u_si = u_ip_to_si(wall_info['u_value_ip'])
                                wall_changes.append(f"U: {old_u_si:.3f} -> {new_u:.3f}")

                        if wall_changes:
                            changes.append({
                                'space': f"Wall: {wall_name}",
                                'changes': wall_changes
                            })

                # ROOFS
                if 'Roofs' in wb.sheetnames:
                    ws_rof = wb['Roofs']
                    # Ler roofs actuais do MDB
                    cursor.execute("SELECT nIndex, szName, fOverallUValue FROM RoofIndex")
                    e3a_roofs = {}
                    for row in cursor.fetchall():
                        if row.szName:
                            e3a_roofs[row.szName.strip()] = {
                                'index': row.nIndex,
                                'u_value_ip': row.fOverallUValue or 0,
                            }

                    for row in range(4, ws_rof.max_row + 1):
                        roof_name = ws_rof.cell(row=row, column=1).value
                        if not roof_name or str(roof_name).strip() == '':
                            continue
                        roof_name = str(roof_name).strip()

                        if roof_name not in e3a_roofs:
                            continue

                        roof_info = e3a_roofs[roof_name]
                        roof_changes = []

                        # U-value (col 2)
                        new_u = ws_rof.cell(row=row, column=2).value
                        if new_u is not None:
                            new_u_ip = u_si_to_ip(new_u)
                            if abs(new_u_ip - roof_info['u_value_ip']) > 0.0001:
                                cursor.execute(
                                    "UPDATE RoofIndex SET fOverallUValue = ? WHERE nIndex = ?",
                                    (new_u_ip, roof_info['index'])
                                )
                                old_u_si = u_ip_to_si(roof_info['u_value_ip'])
                                roof_changes.append(f"U: {old_u_si:.3f} -> {new_u:.3f}")

                        if roof_changes:
                            changes.append({
                                'space': f"Roof: {roof_name}",
                                'changes': roof_changes
                            })

                conn.commit()
                conn.close()
            except Exception as e:
                # Se falhar pyodbc, continua sem editar MDB
                changes.append({
                    'space': 'AVISO',
                    'changes': [f'Nao foi possivel editar Walls/Roofs no MDB: {str(e)}']
                })

        # Criar novo ZIP
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, temp_dir)
                    zf.write(file_path, arc_name)

        return changes

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def e3a_to_excel(e3a_path, output_path):
    """Converte E3A para Excel usando o template completo HAP_Template_RSECE.xlsx."""
    spaces, schedules, windows, walls_types, roofs_types = read_e3a_spaces(e3a_path)
    if not spaces:
        raise ValueError("Nenhum espaco encontrado no E3A")

    # Usar o template como base
    template_path = os.path.join(BASE_DIR, 'HAP_Template_RSECE.xlsx')
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template nao encontrado: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Espacos']

    # Limpar dados existentes (linhas 4 em diante)
    for row in range(4, ws.max_row + 1):
        for col in range(1, 148):
            ws.cell(row=row, column=col).value = None

    # Preencher dados dos espacos (a partir da linha 4)
    for row_idx, space in enumerate(spaces, 4):
        # GENERAL (cols 1-6)
        ws.cell(row=row_idx, column=1, value=space['name'])
        ws.cell(row=row_idx, column=2, value=round(space['area'], 2) if space['area'] else None)
        ws.cell(row=row_idx, column=3, value=round(space['height'], 2) if space['height'] else None)
        ws.cell(row=row_idx, column=4, value=round(space['weight'], 2) if space['weight'] else None)
        ws.cell(row=row_idx, column=5, value=round(space['oa'], 2) if space['oa'] else None)
        ws.cell(row=row_idx, column=6, value=space['oa_unit'])

        # PEOPLE (cols 7-11)
        ws.cell(row=row_idx, column=7, value=space.get('occupancy'))
        # col 8 = Activity Level (nao temos)
        ws.cell(row=row_idx, column=9, value=round(space['sensible'], 1) if space.get('sensible') else None)
        ws.cell(row=row_idx, column=10, value=round(space['latent'], 1) if space.get('latent') else None)
        ws.cell(row=row_idx, column=11, value=schedules.get(space.get('people_sch_id'), ''))

        # LIGHTING (cols 12-16)
        ws.cell(row=row_idx, column=12, value=space.get('task_light'))
        ws.cell(row=row_idx, column=13, value=space.get('general_light'))
        # col 14 = Fixture Type (nao temos)
        ws.cell(row=row_idx, column=15, value=space.get('ballast'))
        ws.cell(row=row_idx, column=16, value=schedules.get(space.get('light_sch_id'), ''))

        # EQUIPMENT (cols 17-18)
        ws.cell(row=row_idx, column=17, value=round(space['equipment'], 2) if space.get('equipment') else None)
        ws.cell(row=row_idx, column=18, value=schedules.get(space.get('equip_sch_id'), ''))

        # MISC (cols 19-22) - nao temos dados

        # INFILTRATION (cols 23-25)
        ws.cell(row=row_idx, column=23, value=space.get('ach_clg'))
        ws.cell(row=row_idx, column=24, value=space.get('ach_htg'))
        ws.cell(row=row_idx, column=25, value=space.get('ach_energy'))

        # FLOORS (cols 26-38)
        floor_area = space.get('floor_area')
        floor_u = space.get('floor_u')
        ws.cell(row=row_idx, column=27, value=round(floor_area, 2) if floor_area is not None and floor_area > 0 else None)
        ws.cell(row=row_idx, column=28, value=round(floor_u, 3) if floor_u is not None and floor_u > 0 else None)

        # PARTITIONS - CEILING (cols 39-44)
        ceil_area = space.get('ceil_area')
        ceil_u = space.get('ceil_u')
        ws.cell(row=row_idx, column=39, value=round(ceil_area, 2) if ceil_area is not None and ceil_area > 0 else None)
        ws.cell(row=row_idx, column=40, value=round(ceil_u, 3) if ceil_u is not None and ceil_u > 0 else None)

        # PARTITIONS - WALL (cols 45-50)
        part_wall_area = space.get('part_wall_area')
        part_wall_u = space.get('part_wall_u')
        ws.cell(row=row_idx, column=45, value=round(part_wall_area, 2) if part_wall_area is not None and part_wall_area > 0 else None)
        ws.cell(row=row_idx, column=46, value=round(part_wall_u, 3) if part_wall_u is not None and part_wall_u > 0 else None)

        # WALLS (cols 51-122: 8 walls x 9 cols)
        for w_idx, wall in enumerate(space.get('walls', [])):
            col_start = 51 + w_idx * 9
            ws.cell(row=row_idx, column=col_start, value=wall.get('exposure'))
            ws.cell(row=row_idx, column=col_start + 1, value=round(wall['area'], 2) if wall.get('area') else None)
            ws.cell(row=row_idx, column=col_start + 2, value=wall.get('type_name', ''))
            ws.cell(row=row_idx, column=col_start + 3, value=wall.get('win1_name', ''))
            ws.cell(row=row_idx, column=col_start + 4, value=wall.get('win1_qty'))
            ws.cell(row=row_idx, column=col_start + 5, value=wall.get('win2_name', ''))
            ws.cell(row=row_idx, column=col_start + 6, value=wall.get('win2_qty'))
            ws.cell(row=row_idx, column=col_start + 7, value=wall.get('door_name', ''))
            ws.cell(row=row_idx, column=col_start + 8, value=wall.get('door_qty'))

        # ROOFS (cols 123-146: 4 roofs x 6 cols)
        for r_idx, roof in enumerate(space.get('roofs', [])):
            col_start = 123 + r_idx * 6
            ws.cell(row=row_idx, column=col_start, value=roof.get('exposure'))
            ws.cell(row=row_idx, column=col_start + 1, value=round(roof['area'], 2) if roof.get('area') else None)
            ws.cell(row=row_idx, column=col_start + 2, value=roof.get('slope'))
            ws.cell(row=row_idx, column=col_start + 3, value=roof.get('type_name', ''))
            ws.cell(row=row_idx, column=col_start + 4, value=roof.get('sky_name', ''))
            ws.cell(row=row_idx, column=col_start + 5, value=roof.get('sky_qty'))

    # Actualizar sheet Tipos com os schedules do E3A
    if 'Tipos' in wb.sheetnames:
        ws_tipos = wb['Tipos']
        # Limpar schedules existentes (cols 13-14)
        for row in range(3, ws_tipos.max_row + 1):
            ws_tipos.cell(row=row, column=13).value = None
            ws_tipos.cell(row=row, column=14).value = None
        # Adicionar schedules do E3A
        for i, (sch_id, sch_name) in enumerate(sorted(schedules.items()), 3):
            ws_tipos.cell(row=i, column=13, value=sch_id)
            ws_tipos.cell(row=i, column=14, value=sch_name)

    # Preencher sheet Windows
    if 'Windows' in wb.sheetnames and windows:
        ws_win = wb['Windows']
        # Limpar dados existentes (linha 4 em diante)
        for row in range(4, ws_win.max_row + 1):
            for col in range(1, 6):
                ws_win.cell(row=row, column=col).value = None
        # Preencher com dados do E3A
        for i, (win_id, win_data) in enumerate(sorted(windows.items()), 4):
            ws_win.cell(row=i, column=1, value=win_data['name'])
            ws_win.cell(row=i, column=2, value=round(win_data['u_value'], 3) if win_data['u_value'] else None)
            ws_win.cell(row=i, column=3, value=round(win_data['shgc'], 3) if win_data['shgc'] else None)
            ws_win.cell(row=i, column=4, value=round(win_data['height'], 3) if win_data['height'] else None)
            ws_win.cell(row=i, column=5, value=round(win_data['width'], 3) if win_data['width'] else None)

    # Preencher sheet Walls
    if 'Walls' in wb.sheetnames and walls_types:
        ws_wal = wb['Walls']
        # Limpar dados existentes (linha 4 em diante)
        for row in range(4, ws_wal.max_row + 1):
            for col in range(1, 5):
                ws_wal.cell(row=row, column=col).value = None
        # Preencher com dados do E3A
        for i, (wal_id, wal_data) in enumerate(sorted(walls_types.items()), 4):
            ws_wal.cell(row=i, column=1, value=wal_data['name'])
            # U-Value, Peso, Espessura - se disponivel
            if 'u_value' in wal_data:
                ws_wal.cell(row=i, column=2, value=round(wal_data['u_value'], 3))
            if 'weight' in wal_data:
                ws_wal.cell(row=i, column=3, value=round(wal_data['weight'], 2))
            if 'thickness' in wal_data:
                ws_wal.cell(row=i, column=4, value=round(wal_data['thickness'], 3))

    # Preencher sheet Roofs
    if 'Roofs' in wb.sheetnames and roofs_types:
        ws_rof = wb['Roofs']
        # Limpar dados existentes (linha 4 em diante)
        for row in range(4, ws_rof.max_row + 1):
            for col in range(1, 5):
                ws_rof.cell(row=row, column=col).value = None
        # Preencher com dados do E3A
        for i, (rof_id, rof_data) in enumerate(sorted(roofs_types.items()), 4):
            ws_rof.cell(row=i, column=1, value=rof_data['name'])
            # U-Value, Peso, Espessura - se disponivel
            if 'u_value' in rof_data:
                ws_rof.cell(row=i, column=2, value=round(rof_data['u_value'], 3))
            if 'weight' in rof_data:
                ws_rof.cell(row=i, column=3, value=round(rof_data['weight'], 2))
            if 'thickness' in rof_data:
                ws_rof.cell(row=i, column=4, value=round(rof_data['thickness'], 3))

    wb.save(output_path)
    return len(spaces)


# =============================================================================
# ROTAS WEB
# =============================================================================

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="pt">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HAP 5.1 Converter</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            min-height: 100vh;
            padding: 20px;
            color: #fff;
        }
        .container { max-width: 1100px; margin: 0 auto; }
        header { text-align: center; margin-bottom: 40px; }
        h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
            background: linear-gradient(90deg, #00d2ff, #3a7bd5);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        .subtitle { color: #888; font-size: 1.1em; }
        .cards {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
            gap: 25px;
            margin-bottom: 30px;
        }
        .card {
            background: rgba(255, 255, 255, 0.05);
            border-radius: 20px;
            padding: 25px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            transition: transform 0.3s, box-shadow 0.3s;
        }
        .card:hover {
            transform: translateY(-5px);
            box-shadow: 0 20px 40px rgba(0, 0, 0, 0.3);
        }
        .card h2 {
            display: flex;
            align-items: center;
            gap: 10px;
            margin-bottom: 15px;
            font-size: 1.3em;
        }
        .card.excel h2 { color: #00d26a; }
        .card.e3a h2 { color: #ff6b6b; }
        .card.edit h2 { color: #ffa500; }
        .icon { font-size: 1.4em; }
        .description { color: #aaa; margin-bottom: 20px; line-height: 1.5; font-size: 0.95em; }
        .upload-area {
            border: 2px dashed rgba(255, 255, 255, 0.2);
            border-radius: 12px;
            padding: 20px;
            text-align: center;
            margin-bottom: 15px;
            transition: border-color 0.3s, background 0.3s;
            cursor: pointer;
        }
        .upload-area:hover {
            border-color: rgba(255, 255, 255, 0.4);
            background: rgba(255, 255, 255, 0.02);
        }
        .upload-area.dragover {
            border-color: #00d2ff;
            background: rgba(0, 210, 255, 0.1);
        }
        .upload-area.small { padding: 15px; }
        .upload-icon { font-size: 2.5em; margin-bottom: 10px; opacity: 0.5; }
        .upload-text { color: #888; font-size: 0.9em; }
        .upload-text strong { color: #00d2ff; }
        input[type="file"] { display: none; }
        .file-name {
            margin-top: 8px;
            padding: 8px;
            background: rgba(0, 210, 255, 0.1);
            border-radius: 6px;
            color: #00d2ff;
            display: none;
            font-size: 0.85em;
        }
        .file-name.show { display: block; }
        button {
            width: 100%;
            padding: 12px 20px;
            font-size: 1em;
            font-weight: 600;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            transition: all 0.3s;
        }
        .card.excel button { background: linear-gradient(90deg, #00d26a, #00b359); color: white; }
        .card.e3a button { background: linear-gradient(90deg, #ff6b6b, #ee5a5a); color: white; }
        .card.edit button { background: linear-gradient(90deg, #ffa500, #ff8c00); color: white; }
        button:hover { transform: scale(1.02); box-shadow: 0 5px 20px rgba(0, 0, 0, 0.3); }
        button:disabled { opacity: 0.5; cursor: not-allowed; transform: none; }
        .messages { margin-bottom: 30px; }
        .alert {
            padding: 15px 20px;
            border-radius: 10px;
            margin-bottom: 15px;
            display: flex;
            align-items: flex-start;
            gap: 10px;
        }
        .alert.success { background: rgba(0, 210, 106, 0.2); border: 1px solid #00d26a; color: #00d26a; }
        .alert.error { background: rgba(255, 107, 107, 0.2); border: 1px solid #ff6b6b; color: #ff6b6b; }
        .alert.info { background: rgba(255, 165, 0, 0.2); border: 1px solid #ffa500; color: #ffa500; }
        .alert-content { flex: 1; }
        .alert-content ul { margin: 10px 0 0 20px; font-size: 0.9em; }
        .alert-content li { margin: 3px 0; }
        footer { text-align: center; color: #666; padding: 20px; }
        footer a { color: #00d2ff; text-decoration: none; margin: 0 10px; }
        .info-box {
            background: rgba(0, 210, 255, 0.1);
            border: 1px solid rgba(0, 210, 255, 0.3);
            border-radius: 10px;
            padding: 20px;
            margin-top: 20px;
        }
        .info-box h3 { color: #00d2ff; margin-bottom: 10px; }
        .info-box ul { margin-left: 20px; color: #aaa; }
        .info-box li { margin: 5px 0; }
        .workflow {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
            margin: 15px 0;
            font-size: 0.85em;
            color: #888;
        }
        .workflow span {
            background: rgba(255,165,0,0.2);
            padding: 5px 12px;
            border-radius: 15px;
            color: #ffa500;
        }
        .workflow .arrow { background: none; color: #666; }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>HAP 5.1 Converter</h1>
            <p class="subtitle">Converta e edite ficheiros HAP facilmente</p>
        </header>

        <div class="messages">
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="alert {{ category }}">
                            {% if category == 'success' %}&#10004;{% elif category == 'info' %}&#9432;{% else %}&#10006;{% endif %}
                            <div class="alert-content">{{ message|safe }}</div>
                        </div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
        </div>

        <div class="cards">
            <div class="card edit" style="grid-column: span 2;">
                <h2><span class="icon">&#9998;</span> Editar Projecto (Prev/Ref)</h2>
                <p class="description">
                    Edita um projecto HAP existente: exporta para Excel, faz alteracoes, e aplica de volta ao E3A.
                    O sistema detecta e aplica <strong>apenas os campos alterados</strong>.
                </p>

                <div style="display: grid; grid-template-columns: 1fr auto 1fr; gap: 20px; align-items: start;">
                    <!-- PASSO 1: Exportar -->
                    <div style="background: rgba(255,165,0,0.1); border-radius: 12px; padding: 20px;">
                        <h3 style="color: #ffa500; margin-bottom: 15px; font-size: 1.1em;">1. Exportar E3A para Excel</h3>
                        <form action="/e3a-to-excel" method="post" enctype="multipart/form-data" id="form-e3a">
                            <div class="upload-area small" onclick="document.getElementById('e3a-file').click()">
                                <p class="upload-text">Carrega o <strong>.E3A</strong> a editar</p>
                                <input type="file" name="file" id="e3a-file" accept=".e3a,.E3A" required>
                                <div class="file-name" id="e3a-name"></div>
                            </div>
                            <button type="submit" id="btn-e3a" disabled>Exportar para Excel</button>
                        </form>
                        <p style="color: #888; font-size: 0.8em; margin-top: 10px; text-align: center;">
                            Recebe um Excel editavel com todos os dados
                        </p>
                    </div>

                    <!-- Seta -->
                    <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; padding-top: 60px;">
                        <div style="font-size: 2em; color: #ffa500;">&#10148;</div>
                        <p style="color: #666; font-size: 0.75em; text-align: center; margin-top: 5px;">Edita o<br>Excel</p>
                    </div>

                    <!-- PASSO 2: Aplicar -->
                    <div style="background: rgba(255,165,0,0.1); border-radius: 12px; padding: 20px;">
                        <h3 style="color: #ffa500; margin-bottom: 15px; font-size: 1.1em;">2. Aplicar Alteracoes</h3>
                        <form action="/apply-changes" method="post" enctype="multipart/form-data" id="form-edit">
                            <div class="upload-area small" onclick="document.getElementById('edit-e3a').click()">
                                <p class="upload-text"><strong>.E3A</strong> Original</p>
                                <input type="file" name="e3a_file" id="edit-e3a" accept=".e3a,.E3A" required>
                                <div class="file-name" id="edit-e3a-name"></div>
                            </div>
                            <div class="upload-area small" onclick="document.getElementById('edit-xlsx').click()">
                                <p class="upload-text"><strong>.xlsx</strong> Editado</p>
                                <input type="file" name="excel_file" id="edit-xlsx" accept=".xlsx,.xls" required>
                                <div class="file-name" id="edit-xlsx-name"></div>
                            </div>
                            <button type="submit" id="btn-edit" disabled>Aplicar Alteracoes</button>
                        </form>
                        <p style="color: #888; font-size: 0.8em; margin-top: 10px; text-align: center;">
                            Recebe E3A modificado (so campos alterados)
                        </p>
                    </div>
                </div>
            </div>

            <div class="card excel">
                <h2><span class="icon">&#128196;</span> Excel para E3A (Novo)</h2>
                <p class="description">
                    Cria um novo ficheiro HAP a partir de um Excel.
                    Usa o modelo RSECE como base (82 schedules).
                </p>
                <form action="/excel-to-e3a" method="post" enctype="multipart/form-data" id="form-excel">
                    <div class="upload-area" onclick="document.getElementById('excel-file').click()">
                        <div class="upload-icon">&#128194;</div>
                        <p class="upload-text">Carrega o ficheiro <strong>.xlsx</strong></p>
                        <input type="file" name="file" id="excel-file" accept=".xlsx,.xls" required>
                        <div class="file-name" id="excel-name"></div>
                    </div>
                    <button type="submit" id="btn-excel" disabled>Criar E3A Novo</button>
                </form>
            </div>
        </div>

        <div class="info-box">
            <h3>Como usar a funcionalidade Prev/Ref</h3>
            <ul>
                <li><strong>Passo 1:</strong> Usa "E3A para Excel" para exportar o projecto</li>
                <li><strong>Passo 2:</strong> Edita o Excel (U-values, iluminacao, schedules, etc.)</li>
                <li><strong>Passo 3:</strong> Usa "Aplicar Alteracoes" com o E3A original + Excel editado</li>
                <li><strong>Resultado:</strong> Obtem um E3A modificado apenas nos campos alterados</li>
            </ul>
        </div>

        <footer>
            <a href="/download-template">Descarregar Template</a> |
            <span>HAP 5.1 Tools 2026</span>
        </footer>
    </div>

    <script>
        function setupFileInput(inputId, nameId, btnId, otherInputId) {
            document.getElementById(inputId).addEventListener('change', function() {
                const fileName = this.files[0]?.name || '';
                const nameDiv = document.getElementById(nameId);
                const btn = document.getElementById(btnId);
                if (fileName) {
                    nameDiv.textContent = fileName;
                    nameDiv.classList.add('show');
                    if (otherInputId) {
                        const other = document.getElementById(otherInputId);
                        btn.disabled = !other.files[0];
                    } else {
                        btn.disabled = false;
                    }
                } else {
                    nameDiv.classList.remove('show');
                    btn.disabled = true;
                }
            });
        }

        setupFileInput('excel-file', 'excel-name', 'btn-excel', null);
        setupFileInput('e3a-file', 'e3a-name', 'btn-e3a', null);
        setupFileInput('edit-e3a', 'edit-e3a-name', 'btn-edit', 'edit-xlsx');
        setupFileInput('edit-xlsx', 'edit-xlsx-name', 'btn-edit', 'edit-e3a');

        document.querySelectorAll('.upload-area').forEach(area => {
            area.addEventListener('dragover', e => { e.preventDefault(); area.classList.add('dragover'); });
            area.addEventListener('dragleave', e => { area.classList.remove('dragover'); });
            area.addEventListener('drop', e => {
                e.preventDefault();
                area.classList.remove('dragover');
                const input = area.querySelector('input[type="file"]');
                input.files = e.dataTransfer.files;
                input.dispatchEvent(new Event('change'));
            });
        });
    </script>
</body>
</html>
'''

def allowed_file(filename, extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/excel-to-e3a', methods=['POST'])
def convert_excel_to_e3a():
    if 'file' not in request.files:
        flash('Nenhum ficheiro seleccionado', 'error')
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        flash('Nenhum ficheiro seleccionado', 'error')
        return redirect(url_for('index'))

    if not allowed_file(file.filename, ALLOWED_EXTENSIONS_EXCEL):
        flash('Tipo de ficheiro invalido. Use .xlsx ou .xls', 'error')
        return redirect(url_for('index'))

    try:
        # Guardar ficheiro temporario
        filename = secure_filename(file.filename)
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(input_path)

        # Converter
        output_filename = filename.rsplit('.', 1)[0] + '.E3A'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)

        num_spaces = excel_to_e3a(input_path, output_path)

        # Limpar input
        os.remove(input_path)

        # Enviar ficheiro
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/octet-stream'
        )

    except Exception as e:
        flash(f'Erro na conversao: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/e3a-to-excel', methods=['POST'])
def convert_e3a_to_excel():
    if 'file' not in request.files:
        flash('Nenhum ficheiro seleccionado', 'error')
        return redirect(url_for('index'))

    file = request.files['file']
    if file.filename == '':
        flash('Nenhum ficheiro seleccionado', 'error')
        return redirect(url_for('index'))

    if not allowed_file(file.filename, ALLOWED_EXTENSIONS_E3A):
        flash('Tipo de ficheiro invalido. Use .E3A', 'error')
        return redirect(url_for('index'))

    try:
        # Guardar ficheiro temporario
        filename = secure_filename(file.filename)
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(input_path)

        # Converter
        output_filename = filename.rsplit('.', 1)[0] + '_export.xlsx'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)

        num_spaces = e3a_to_excel(input_path, output_path)

        # Limpar input
        os.remove(input_path)

        # Enviar ficheiro
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Erro na conversao: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/apply-changes', methods=['POST'])
def apply_changes():
    """Aplica alteracoes do Excel ao E3A original (Prev/Ref)."""
    if 'e3a_file' not in request.files or 'excel_file' not in request.files:
        flash('Faltam ficheiros. Carrega o E3A original e o Excel editado.', 'error')
        return redirect(url_for('index'))

    e3a_file = request.files['e3a_file']
    excel_file = request.files['excel_file']

    if e3a_file.filename == '' or excel_file.filename == '':
        flash('Faltam ficheiros. Carrega o E3A original e o Excel editado.', 'error')
        return redirect(url_for('index'))

    if not allowed_file(e3a_file.filename, ALLOWED_EXTENSIONS_E3A):
        flash('Ficheiro E3A invalido.', 'error')
        return redirect(url_for('index'))

    if not allowed_file(excel_file.filename, ALLOWED_EXTENSIONS_EXCEL):
        flash('Ficheiro Excel invalido.', 'error')
        return redirect(url_for('index'))

    try:
        # Guardar ficheiros temporarios
        e3a_filename = secure_filename(e3a_file.filename)
        excel_filename = secure_filename(excel_file.filename)

        e3a_path = os.path.join(app.config['UPLOAD_FOLDER'], e3a_filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)

        e3a_file.save(e3a_path)
        excel_file.save(excel_path)

        # Aplicar alteracoes
        output_filename = e3a_filename.rsplit('.', 1)[0] + '_modificado.E3A'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)

        changes = apply_excel_changes_to_e3a(e3a_path, excel_path, output_path)

        # Limpar inputs
        os.remove(e3a_path)
        os.remove(excel_path)

        if not changes:
            flash('Nenhuma alteracao detectada. O Excel e identico ao E3A original.', 'info')
            return redirect(url_for('index'))

        # Guardar relatorio de alteracoes na sessao para mostrar na proxima pagina
        changes_html = f'<strong>{len(changes)} espaco(s) modificado(s):</strong><ul>'
        for c in changes[:10]:  # Mostrar max 10
            changes_html += f'<li><strong>{c["space"]}</strong>: {", ".join(c["changes"][:3])}'
            if len(c["changes"]) > 3:
                changes_html += f' (+{len(c["changes"])-3} mais)'
            changes_html += '</li>'
        if len(changes) > 10:
            changes_html += f'<li>... e mais {len(changes)-10} espaco(s)</li>'
        changes_html += '</ul>'

        flash(changes_html, 'success')

        # Enviar ficheiro
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/octet-stream'
        )

    except Exception as e:
        flash(f'Erro ao aplicar alteracoes: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/download-template')
def download_template():
    template_path = os.path.join(BASE_DIR, 'HAP_Template_RSECE.xlsx')
    if os.path.exists(template_path):
        return send_file(
            template_path,
            as_attachment=True,
            download_name='HAP_Template_RSECE.xlsx'
        )
    else:
        flash('Template nao encontrado', 'error')
        return redirect(url_for('index'))


if __name__ == '__main__':
    print("=" * 60)
    print("HAP 5.1 Web Converter - com Prev/Ref")
    print("=" * 60)
    print()
    print("Abrir no browser: http://localhost:5000")
    print()
    print("Funcionalidades:")
    print("  - E3A -> Excel (exportacao completa)")
    print("  - Aplicar Alteracoes (Prev/Ref)")
    print("  - Excel -> E3A (criar novo)")
    print("  - Download do template")
    print()
    print("Ctrl+C para parar")
    print("=" * 60)

    app.run(debug=False, host='0.0.0.0', port=5000)
