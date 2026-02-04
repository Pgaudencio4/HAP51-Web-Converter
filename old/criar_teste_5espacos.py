"""
Script para criar ficheiro de teste com 5 espaços completos
"""
from openpyxl import Workbook

# Criar workbook de teste
wb = Workbook()

# ============================================
# SHEET ESPACOS
# ============================================
ws = wb.active
ws.title = 'Espacos'

# Cabeçalhos (linha 3)
headers_row3 = [
    'Space Name', 'Floor Area\n(m2)', 'Avg Ceiling Ht\n(m)', 'Building Wt\n(kg/m2)', 'Outdoor Air\n(valor)', 'OA Unit',
    'Occupancy\n(people)', 'Activity Level', 'Sensible\n(W/person)', 'Latent\n(W/person)', 'Schedule',
    'Task Lighting\n(W)', 'General Ltg\n(W)', 'Fixture Type', 'Ballast Mult', 'Schedule',
    'Equipment\n(W/m2)', 'Schedule',
    'Sensible\n(W)', 'Latent\n(W)', 'Sens Sch', 'Lat Sch',
    'Infil Method', 'Design Clg\n(ACH)', 'Design Htg\n(ACH)', 'Energy\n(ACH)',
    'Floor Type', 'Floor Area\n(m2)', 'U-Value\n(W/m2K)', 'Exp Perim\n(m)', 'Edge R\n(m2K/W)', 'Depth\n(m)', 'Bsmt Wall U\n(W/m2K)', 'Wall Ins R\n(m2K/W)', 'Ins Depth\n(m)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
    'Area\n(m2)', 'U-Value\n(W/m2K)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
    'Area\n(m2)', 'U-Value\n(W/m2K)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
]
# Adicionar colunas de paredes (8 x 9 = 72 colunas)
for w in range(8):
    headers_row3.extend(['Exposure', 'Gross Area\n(m2)', 'Wall Type', 'Window 1', 'Win1 Qty', 'Window 2', 'Win2 Qty', 'Door', 'Door Qty'])
# Adicionar colunas de coberturas (4 x 6 = 24 colunas)
for r in range(4):
    headers_row3.extend(['Exposure', 'Gross Area\n(m2)', 'Slope\n(deg)', 'Roof Type', 'Skylight', 'Sky Qty'])

# Escrever cabeçalhos
for col, val in enumerate(headers_row3, 1):
    ws.cell(row=3, column=col, value=val)

# ============================================
# DADOS DOS 5 ESPAÇOS DE TESTE
# ============================================
espacos = [
    {
        'name': 'Escritorio_Sul',
        'area': 100.0, 'height': 3.0, 'weight': 150,
        'oa': 10, 'oa_unit': 'L/s/person',
        'occupancy': 10, 'activity': 'Office Work', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
        'task_light': 0, 'general_light': 1000, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
        'equipment': 15, 'equip_sch': 'Escritorio Equip',
        'misc_sens': 0, 'misc_lat': 0,
        'infil_method': 'Air Change', 'ach_clg': 0.5, 'ach_htg': 0.5, 'ach_energy': 0.5,
        'floor_type': 'Slab Floor On Grade', 'floor_area': 100, 'floor_u': 0.5, 'floor_perim': 40, 'floor_edge_r': 1.0,
        'ceil_area': 100, 'ceil_u': 0.4,
        'walls': [
            {'exp': 'S', 'area': 30, 'type': 'Parede Exterior', 'win1': 'Janela_Sul', 'win1_qty': 2},
            {'exp': 'E', 'area': 25, 'type': 'Parede Exterior', 'win1': 'Janela_Este', 'win1_qty': 1},
        ],
        'roofs': [
            {'exp': 'H', 'area': 100, 'slope': 0, 'type': 'Cobertura Plana', 'sky': 'Claraboia', 'sky_qty': 1},
        ],
    },
    {
        'name': 'Escritorio_Norte',
        'area': 80.0, 'height': 3.0, 'weight': 150,
        'oa': 8, 'oa_unit': 'L/s/person',
        'occupancy': 8, 'activity': 'Office Work', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
        'task_light': 0, 'general_light': 800, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
        'equipment': 15, 'equip_sch': 'Escritorio Equip',
        'misc_sens': 0, 'misc_lat': 0,
        'infil_method': 'Air Change', 'ach_clg': 0.5, 'ach_htg': 0.5, 'ach_energy': 0.5,
        'floor_type': 'Conditioned Space Below', 'floor_area': 80, 'floor_u': 0.5,
        'ceil_area': 80, 'ceil_u': 0.4,
        'walls': [
            {'exp': 'N', 'area': 24, 'type': 'Parede Exterior', 'win1': 'Janela_Norte', 'win1_qty': 3},
            {'exp': 'W', 'area': 20, 'type': 'Parede Exterior', 'win1': 'Janela_Oeste', 'win1_qty': 2},
        ],
        'roofs': [],
    },
    {
        'name': 'Sala_Reunioes',
        'area': 50.0, 'height': 3.0, 'weight': 150,
        'oa': 25, 'oa_unit': 'L/s',
        'occupancy': 12, 'activity': 'Conference', 'sensible': 70, 'latent': 45, 'people_sch': 'Escritorio Ocup',
        'task_light': 0, 'general_light': 600, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
        'equipment': 10, 'equip_sch': 'Escritorio Equip',
        'misc_sens': 500, 'misc_lat': 0, 'misc_sens_sch': 'Escritorio Equip',
        'infil_method': 'Air Change', 'ach_clg': 0.3, 'ach_htg': 0.3, 'ach_energy': 0.3,
        'floor_type': 'Conditioned Space Below', 'floor_area': 50, 'floor_u': 0.5,
        'ceil_area': 50, 'ceil_u': 0.4,
        'walls': [
            {'exp': 'SE', 'area': 21, 'type': 'Parede Exterior', 'win1': 'Janela_SE', 'win1_qty': 2, 'win2': 'Porta_Vidro', 'win2_qty': 1},
        ],
        'roofs': [],
    },
    {
        'name': 'Recepcao',
        'area': 40.0, 'height': 4.0, 'weight': 200,
        'oa': 50, 'oa_unit': 'L/s',
        'occupancy': 5, 'activity': 'Standing/Walking', 'sensible': 80, 'latent': 60, 'people_sch': 'Escritorio Ocup',
        'task_light': 100, 'general_light': 500, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
        'equipment': 5, 'equip_sch': 'Escritorio Equip',
        'misc_sens': 0, 'misc_lat': 0,
        'infil_method': 'Air Change', 'ach_clg': 1.0, 'ach_htg': 1.0, 'ach_energy': 1.0,
        'floor_type': 'Slab Floor On Grade', 'floor_area': 40, 'floor_u': 0.5, 'floor_perim': 26, 'floor_edge_r': 1.0,
        'ceil_area': 40, 'ceil_u': 0.4,
        'walls': [
            {'exp': 'S', 'area': 20, 'type': 'Parede Exterior', 'win1': 'Porta_Entrada', 'win1_qty': 1, 'win2': 'Janela_Sul', 'win2_qty': 1},
            {'exp': 'SW', 'area': 16, 'type': 'Parede Exterior', 'win1': 'Janela_SW', 'win1_qty': 2},
        ],
        'roofs': [],
    },
    {
        'name': 'Cobertura_Total',
        'area': 200.0, 'height': 3.5, 'weight': 180,
        'oa': 20, 'oa_unit': 'L/s/person',
        'occupancy': 15, 'activity': 'Office Work', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
        'task_light': 0, 'general_light': 2000, 'fixture': 'Recessed Vented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
        'equipment': 20, 'equip_sch': 'Escritorio Equip',
        'misc_sens': 1000, 'misc_lat': 200, 'misc_sens_sch': 'Escritorio Equip', 'misc_lat_sch': 'Escritorio Equip',
        'infil_method': 'Air Change', 'ach_clg': 0.5, 'ach_htg': 0.5, 'ach_energy': 0.5,
        'floor_type': 'Conditioned Space Below', 'floor_area': 200, 'floor_u': 0.5,
        'ceil_area': 0, 'ceil_u': 0,
        'walls': [
            {'exp': 'N', 'area': 35, 'type': 'Parede Exterior', 'win1': 'Janela_Norte', 'win1_qty': 4},
            {'exp': 'S', 'area': 35, 'type': 'Parede Exterior', 'win1': 'Janela_Sul', 'win1_qty': 4},
            {'exp': 'E', 'area': 28, 'type': 'Parede Exterior', 'win1': 'Janela_Este', 'win1_qty': 3},
            {'exp': 'W', 'area': 28, 'type': 'Parede Exterior', 'win1': 'Janela_Oeste', 'win1_qty': 3},
        ],
        'roofs': [
            {'exp': 'H', 'area': 120, 'slope': 0, 'type': 'Cobertura Plana', 'sky': 'Claraboia', 'sky_qty': 2},
            {'exp': 'N', 'area': 80, 'slope': 15, 'type': 'Cobertura Inclinada'},
        ],
    },
]

# Escrever dados
for row_idx, esp in enumerate(espacos, 4):
    # GENERAL (1-6)
    ws.cell(row=row_idx, column=1, value=esp['name'])
    ws.cell(row=row_idx, column=2, value=esp['area'])
    ws.cell(row=row_idx, column=3, value=esp['height'])
    ws.cell(row=row_idx, column=4, value=esp['weight'])
    ws.cell(row=row_idx, column=5, value=esp['oa'])
    ws.cell(row=row_idx, column=6, value=esp['oa_unit'])

    # PEOPLE (7-11)
    ws.cell(row=row_idx, column=7, value=esp['occupancy'])
    ws.cell(row=row_idx, column=8, value=esp['activity'])
    ws.cell(row=row_idx, column=9, value=esp['sensible'])
    ws.cell(row=row_idx, column=10, value=esp['latent'])
    ws.cell(row=row_idx, column=11, value=esp['people_sch'])

    # LIGHTING (12-16)
    ws.cell(row=row_idx, column=12, value=esp['task_light'])
    ws.cell(row=row_idx, column=13, value=esp['general_light'])
    ws.cell(row=row_idx, column=14, value=esp['fixture'])
    ws.cell(row=row_idx, column=15, value=esp['ballast'])
    ws.cell(row=row_idx, column=16, value=esp['light_sch'])

    # EQUIPMENT (17-18)
    ws.cell(row=row_idx, column=17, value=esp['equipment'])
    ws.cell(row=row_idx, column=18, value=esp['equip_sch'])

    # MISC (19-22)
    ws.cell(row=row_idx, column=19, value=esp['misc_sens'])
    ws.cell(row=row_idx, column=20, value=esp['misc_lat'])
    ws.cell(row=row_idx, column=21, value=esp.get('misc_sens_sch', ''))
    ws.cell(row=row_idx, column=22, value=esp.get('misc_lat_sch', ''))

    # INFILTRATION (23-26)
    ws.cell(row=row_idx, column=23, value=esp['infil_method'])
    ws.cell(row=row_idx, column=24, value=esp['ach_clg'])
    ws.cell(row=row_idx, column=25, value=esp['ach_htg'])
    ws.cell(row=row_idx, column=26, value=esp['ach_energy'])

    # FLOORS (27-39)
    ws.cell(row=row_idx, column=27, value=esp['floor_type'])
    ws.cell(row=row_idx, column=28, value=esp['floor_area'])
    ws.cell(row=row_idx, column=29, value=esp.get('floor_u', 0.5))
    ws.cell(row=row_idx, column=30, value=esp.get('floor_perim', 0))
    ws.cell(row=row_idx, column=31, value=esp.get('floor_edge_r', 1.0))

    # CEILING PARTITION (40-45)
    ws.cell(row=row_idx, column=40, value=esp.get('ceil_area', 0))
    ws.cell(row=row_idx, column=41, value=esp.get('ceil_u', 0.4))

    # WALLS (52-123) - 8 paredes x 9 colunas
    for w_idx, wall in enumerate(esp.get('walls', [])):
        base_col = 52 + w_idx * 9
        ws.cell(row=row_idx, column=base_col, value=wall.get('exp'))
        ws.cell(row=row_idx, column=base_col+1, value=wall.get('area'))
        ws.cell(row=row_idx, column=base_col+2, value=wall.get('type'))
        ws.cell(row=row_idx, column=base_col+3, value=wall.get('win1'))
        ws.cell(row=row_idx, column=base_col+4, value=wall.get('win1_qty'))
        ws.cell(row=row_idx, column=base_col+5, value=wall.get('win2'))
        ws.cell(row=row_idx, column=base_col+6, value=wall.get('win2_qty'))
        ws.cell(row=row_idx, column=base_col+7, value=wall.get('door'))
        ws.cell(row=row_idx, column=base_col+8, value=wall.get('door_qty'))

    # ROOFS (124-147) - 4 coberturas x 6 colunas
    for r_idx, roof in enumerate(esp.get('roofs', [])):
        base_col = 124 + r_idx * 6
        ws.cell(row=row_idx, column=base_col, value=roof.get('exp'))
        ws.cell(row=row_idx, column=base_col+1, value=roof.get('area'))
        ws.cell(row=row_idx, column=base_col+2, value=roof.get('slope'))
        ws.cell(row=row_idx, column=base_col+3, value=roof.get('type'))
        ws.cell(row=row_idx, column=base_col+4, value=roof.get('sky', ''))
        ws.cell(row=row_idx, column=base_col+5, value=roof.get('sky_qty', 0))

# ============================================
# SHEET TIPOS (vazia - schedules vêm do modelo)
# ============================================
ws_tipos = wb.create_sheet('Tipos')

# ============================================
# SHEET WINDOWS
# ============================================
ws_win = wb.create_sheet('Windows')
ws_win['A1'] = 'WINDOWS'
ws_win['A2'] = 'IDENTIFICAÇÃO'
ws_win['B2'] = 'PROPRIEDADES TÉRMICAS'
ws_win['D2'] = 'DIMENSÕES'
ws_win['A3'] = 'Nome'
ws_win['B3'] = 'U-Value\n(W/m²K)'
ws_win['C3'] = 'SHGC'
ws_win['D3'] = 'Altura\n(m)'
ws_win['E3'] = 'Largura\n(m)'

windows = [
    ('Janela_Sul', 2.5, 0.6, 1.5, 2.0),
    ('Janela_Norte', 2.5, 0.6, 1.5, 2.0),
    ('Janela_Este', 2.5, 0.6, 1.2, 1.5),
    ('Janela_Oeste', 2.5, 0.6, 1.2, 1.5),
    ('Janela_SE', 2.5, 0.6, 1.5, 1.8),
    ('Janela_SW', 2.5, 0.6, 1.5, 1.8),
    ('Porta_Vidro', 3.0, 0.7, 2.2, 1.0),
    ('Porta_Entrada', 3.5, 0.5, 2.2, 1.2),
    ('Claraboia', 3.0, 0.5, 1.0, 1.0),
]
for i, (nome, u, shgc, alt, larg) in enumerate(windows, 4):
    ws_win.cell(row=i, column=1, value=nome)
    ws_win.cell(row=i, column=2, value=u)
    ws_win.cell(row=i, column=3, value=shgc)
    ws_win.cell(row=i, column=4, value=alt)
    ws_win.cell(row=i, column=5, value=larg)

# ============================================
# SHEET WALLS
# ============================================
ws_walls = wb.create_sheet('Walls')
ws_walls['A1'] = 'WALLS'
ws_walls['A2'] = 'IDENTIFICAÇÃO'
ws_walls['B2'] = 'PROPRIEDADES'
ws_walls['A3'] = 'Nome'
ws_walls['B3'] = 'U-Value\n(W/m²K)'
ws_walls['C3'] = 'Peso\n(kg/m²)'
ws_walls['D3'] = 'Espessura\n(m)'

walls = [
    ('Parede Exterior', 0.4, 200, 0.30),
]
for i, (nome, u, peso, esp) in enumerate(walls, 4):
    ws_walls.cell(row=i, column=1, value=nome)
    ws_walls.cell(row=i, column=2, value=u)
    ws_walls.cell(row=i, column=3, value=peso)
    ws_walls.cell(row=i, column=4, value=esp)

# ============================================
# SHEET ROOFS
# ============================================
ws_roofs = wb.create_sheet('Roofs')
ws_roofs['A1'] = 'ROOFS'
ws_roofs['A2'] = 'IDENTIFICAÇÃO'
ws_roofs['B2'] = 'PROPRIEDADES'
ws_roofs['A3'] = 'Nome'
ws_roofs['B3'] = 'U-Value\n(W/m²K)'
ws_roofs['C3'] = 'Peso\n(kg/m²)'
ws_roofs['D3'] = 'Espessura\n(m)'

roofs = [
    ('Cobertura Plana', 0.35, 300, 0.35),
    ('Cobertura Inclinada', 0.40, 250, 0.30),
]
for i, (nome, u, peso, esp) in enumerate(roofs, 4):
    ws_roofs.cell(row=i, column=1, value=nome)
    ws_roofs.cell(row=i, column=2, value=u)
    ws_roofs.cell(row=i, column=3, value=peso)
    ws_roofs.cell(row=i, column=4, value=esp)

# Guardar
output_path = r'C:\Users\pedro\Downloads\Teste_5Espacos.xlsx'
wb.save(output_path)
print(f'Ficheiro criado: {output_path}')
print()
print('=== RESUMO ===')
print('5 Espaços:')
for esp in espacos:
    walls_str = ', '.join([w['exp'] for w in esp.get('walls', [])])
    roofs_str = ', '.join([r['exp'] for r in esp.get('roofs', [])])
    print(f"  - {esp['name']}: {esp['area']}m², Walls=[{walls_str}], Roofs=[{roofs_str}]")
print()
print('9 Janelas, 1 Parede, 2 Coberturas')
