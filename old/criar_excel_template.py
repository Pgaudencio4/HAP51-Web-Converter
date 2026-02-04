"""
Criar Excel template para input de dados HAP 5.1
Cada linha = 1 espaco
Linha 1: Grupos (GENERAL, INTERNALS, etc.)
Linha 2: Headers com unidades
Linha 3: Exemplo preenchido
Linha 4+: Dados
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Cores para cada seccao
COLORS = {
    'General': 'B8CCE4',      # Azul claro
    'Internals': 'C4D79B',    # Verde claro
    'Walls': 'FCD5B4',        # Laranja claro
    'Roofs': 'CCC0DA',        # Roxo claro
    'Infiltration': 'FABF8F', # Laranja
    'Floors': 'D8E4BC',       # Verde lima
    'Partitions': 'E4DFEC',   # Lavanda
}

COLORS_DARK = {
    'General': '4F81BD',      # Azul escuro
    'Internals': '77933C',    # Verde escuro
    'Walls': 'E26B0A',        # Laranja escuro
    'Roofs': '7030A0',        # Roxo escuro
    'Infiltration': 'C65911', # Laranja escuro
    'Floors': '9BBB59',       # Verde lima escuro
    'Partitions': '8064A2',   # Lavanda escuro
}

# Cores alternadas para Walls W1-W8 (tons de laranja/amarelo)
WALL_COLORS = {
    'W1': 'FFE6CC',  # Laranja muito claro
    'W2': 'FFEEDD',  # Peach
    'W3': 'FFF2CC',  # Amarelo claro
    'W4': 'FFE6CC',  # Laranja muito claro
    'W5': 'FFEEDD',  # Peach
    'W6': 'FFF2CC',  # Amarelo claro
    'W7': 'FFE6CC',  # Laranja muito claro
    'W8': 'FFEEDD',  # Peach
}

# Cores alternadas para Roofs R1-R4 (tons de roxo)
ROOF_COLORS = {
    'R1': 'E6D5EC',  # Lavanda claro
    'R2': 'F0E6F5',  # Roxo muito claro
    'R3': 'E6D5EC',  # Lavanda claro
    'R4': 'F0E6F5',  # Roxo muito claro
}

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

thick_left_border = Border(
    left=Side(style='thick', color='000000'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_group(cell, color_key):
    """Estilo para linha de grupo (GENERAL, INTERNALS, etc.)"""
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color=COLORS_DARK[color_key], end_color=COLORS_DARK[color_key], fill_type='solid')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def get_subgroup_color(subgroup):
    """Retorna cor especifica para subgrupos WALL 1-8 ou ROOF 1-4."""
    subgroup_colors = {
        'WALL 1': 'FFE6CC', 'WALL 2': 'FFEEDD', 'WALL 3': 'FFF2CC', 'WALL 4': 'FFE6CC',
        'WALL 5': 'FFEEDD', 'WALL 6': 'FFF2CC', 'WALL 7': 'FFE6CC', 'WALL 8': 'FFEEDD',
        'ROOF 1': 'E6D5EC', 'ROOF 2': 'F0E6F5', 'ROOF 3': 'E6D5EC', 'ROOF 4': 'F0E6F5',
        'CEILING': 'F0E6F5', 'WALL': 'E6D5EC',
        'PEOPLE': 'D5E8D4', 'LIGHTING': 'E2F0D9', 'EQUIPMENT': 'D5E8D4', 'MISC': 'E2F0D9',
    }
    return subgroup_colors.get(subgroup, None)

def is_first_col_of_subgroup(columns, col_idx):
    """Verifica se é a primeira coluna de um subgrupo."""
    if col_idx == 0:
        return True
    current_subgroup = columns[col_idx][1]
    prev_subgroup = columns[col_idx-1][1]
    return current_subgroup != prev_subgroup and current_subgroup != ''

def style_subgroup(cell, color_key, subgroup=''):
    """Estilo para linha de subgrupo (WALL 1, PEOPLE, etc.)."""
    cell.font = Font(bold=True, size=9, color='333333')
    alt_color = get_subgroup_color(subgroup)
    if alt_color:
        # Cor ligeiramente mais escura para subgrupo
        cell.fill = PatternFill(start_color=alt_color, end_color=alt_color, fill_type='solid')
    else:
        cell.fill = PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_header(cell, color_key, subgroup='', is_first=False):
    """Estilo para header com unidade."""
    cell.font = Font(bold=True, size=9)
    alt_color = get_subgroup_color(subgroup)
    if alt_color:
        cell.fill = PatternFill(start_color=alt_color, end_color=alt_color, fill_type='solid')
    else:
        cell.fill = PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')
    if is_first:
        cell.border = thick_left_border
    else:
        cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_data(cell, color_key, subgroup='', is_first=False):
    """Estilo para celula de dados."""
    alt_color = get_subgroup_color(subgroup)
    if alt_color:
        cell.fill = PatternFill(start_color=alt_color, end_color=alt_color, fill_type='solid')
    else:
        cell.fill = PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')
    if is_first:
        cell.border = thick_left_border
    else:
        cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_example(cell, color_key, subgroup='', is_first=False):
    """Estilo para linha de exemplo."""
    cell.font = Font(italic=True, size=9, color='333333')
    alt_color = get_subgroup_color(subgroup)
    if alt_color:
        cell.fill = PatternFill(start_color=alt_color, end_color=alt_color, fill_type='solid')
    else:
        cell.fill = PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')
    if is_first:
        cell.border = thick_left_border
    else:
        cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ============================================================================
# SHEET: Espacos
# ============================================================================
ws = wb.active
ws.title = 'Espacos'

# Freeze linhas 1-3 (grupos + headers + exemplo)
ws.freeze_panes = 'A4'

# Definir colunas: (seccao, subgrupo, header, unidade, largura, exemplo)
# Headers usam exactamente os mesmos termos do HAP 5.1
# Subgrupo permite criar merged cells para WALL 1, WALL 2, PEOPLE, LIGHTING, etc.
columns = [
    # GENERAL (6 colunas) - Tab "General" no HAP
    ('General', '', 'Space Name', '', 25, 'Sala Reunioes'),
    ('General', '', 'Floor Area', 'm2', 10, 50),
    ('General', '', 'Avg Ceiling Ht', 'm', 10, 3.0),
    ('General', '', 'Building Wt', 'kg/m2', 10, 150),
    ('General', '', 'Outdoor Air', 'valor', 10, 10),
    ('General', '', 'OA Unit', '', 12, 'L/s/person'),

    # INTERNALS - People (5 colunas) - Tab "Internals" no HAP
    ('Internals', 'PEOPLE', 'Occupancy', 'people', 10, 10),
    ('Internals', 'PEOPLE', 'Activity Level', '', 14, 'Office Work'),
    ('Internals', 'PEOPLE', 'Sensible', 'W/person', 10, 75),
    ('Internals', 'PEOPLE', 'Latent', 'W/person', 10, 55),
    ('Internals', 'PEOPLE', 'Schedule', '', 12, 'OC_Escritorios'),

    # INTERNALS - Lighting (5 colunas)
    ('Internals', 'LIGHTING', 'Task Lighting', 'W', 10, 0),
    ('Internals', 'LIGHTING', 'General Ltg', 'W', 10, 500),
    ('Internals', 'LIGHTING', 'Fixture Type', '', 16, 'Recessed Unvented'),
    ('Internals', 'LIGHTING', 'Ballast Mult', '', 10, 1.0),
    ('Internals', 'LIGHTING', 'Schedule', '', 12, 'IL_Escritorios'),

    # INTERNALS - Equipment (2 colunas)
    ('Internals', 'EQUIPMENT', 'Equipment', 'W/m2', 10, 15),
    ('Internals', 'EQUIPMENT', 'Schedule', '', 12, 'EQ_Escritorios'),

    # INTERNALS - Miscellaneous (4 colunas)
    ('Internals', 'MISC', 'Sensible', 'W', 10, 0),
    ('Internals', 'MISC', 'Latent', 'W', 10, 0),
    ('Internals', 'MISC', 'Sens Sch', '', 12, ''),
    ('Internals', 'MISC', 'Lat Sch', '', 12, ''),

    # INFILTRATION (4 colunas) - Tab "Infiltration" no HAP
    ('Infiltration', '', 'Infil Method', '', 12, 'Air Change'),
    ('Infiltration', '', 'Design Clg', 'ACH', 10, 0.5),
    ('Infiltration', '', 'Design Htg', 'ACH', 10, 0.5),
    ('Infiltration', '', 'Energy', 'ACH', 10, 0.3),

    # FLOORS (13 colunas) - Tab "Floors" no HAP
    ('Floors', '', 'Floor Type', '', 22, 'Floor Above Uncond Space'),
    ('Floors', '', 'Floor Area', 'm2', 10, 50),
    ('Floors', '', 'U-Value', 'W/m2K', 10, 0.5),
    ('Floors', '', 'Exp Perim', 'm', 10, ''),
    ('Floors', '', 'Edge R', 'm2K/W', 10, ''),
    ('Floors', '', 'Depth', 'm', 8, ''),
    ('Floors', '', 'Bsmt Wall U', 'W/m2K', 10, ''),
    ('Floors', '', 'Wall Ins R', 'm2K/W', 10, ''),
    ('Floors', '', 'Ins Depth', 'm', 8, ''),
    ('Floors', '', 'Unc Max', 'C', 8, 30),
    ('Floors', '', 'Out Max', 'C', 8, 35),
    ('Floors', '', 'Unc Min', 'C', 8, 15),
    ('Floors', '', 'Out Min', 'C', 8, 0),

    # PARTITIONS - Ceiling Partition (6 colunas)
    ('Partitions', 'CEILING', 'Area', 'm2', 10, 50),
    ('Partitions', 'CEILING', 'U-Value', 'W/m2K', 10, 0.8),
    ('Partitions', 'CEILING', 'Unc Max', 'C', 8, 30),
    ('Partitions', 'CEILING', 'Out Max', 'C', 8, 35),
    ('Partitions', 'CEILING', 'Unc Min', 'C', 8, 15),
    ('Partitions', 'CEILING', 'Out Min', 'C', 8, 5),

    # PARTITIONS - Wall Partition (6 colunas)
    ('Partitions', 'WALL', 'Area', 'm2', 10, 20),
    ('Partitions', 'WALL', 'U-Value', 'W/m2K', 10, 1.0),
    ('Partitions', 'WALL', 'Unc Max', 'C', 8, 28),
    ('Partitions', 'WALL', 'Out Max', 'C', 8, 32),
    ('Partitions', 'WALL', 'Unc Min', 'C', 8, 18),
    ('Partitions', 'WALL', 'Out Min', 'C', 8, 10),

    # WALLS - W1 (9 colunas)
    ('Walls', 'WALL 1', 'Exposure', '', 7, 'N'),
    ('Walls', 'WALL 1', 'Gross Area', 'm2', 9, 15),
    ('Walls', 'WALL 1', 'Wall Type', '', 14, 'Paredes Ext'),
    ('Walls', 'WALL 1', 'Window 1', '', 14, 'V0.01'),
    ('Walls', 'WALL 1', 'Win1 Qty', '', 7, 2),
    ('Walls', 'WALL 1', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 1', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 1', 'Door', '', 12, ''),
    ('Walls', 'WALL 1', 'Door Qty', '', 7, ''),

    # WALLS - W2
    ('Walls', 'WALL 2', 'Exposure', '', 7, 'E'),
    ('Walls', 'WALL 2', 'Gross Area', 'm2', 9, 20),
    ('Walls', 'WALL 2', 'Wall Type', '', 14, 'Paredes Ext'),
    ('Walls', 'WALL 2', 'Window 1', '', 14, 'V0.01'),
    ('Walls', 'WALL 2', 'Win1 Qty', '', 7, 3),
    ('Walls', 'WALL 2', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 2', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 2', 'Door', '', 12, 'Porta Ext'),
    ('Walls', 'WALL 2', 'Door Qty', '', 7, 1),

    # WALLS - W3
    ('Walls', 'WALL 3', 'Exposure', '', 7, 'S'),
    ('Walls', 'WALL 3', 'Gross Area', 'm2', 9, 15),
    ('Walls', 'WALL 3', 'Wall Type', '', 14, 'Paredes Ext'),
    ('Walls', 'WALL 3', 'Window 1', '', 14, 'V0.02'),
    ('Walls', 'WALL 3', 'Win1 Qty', '', 7, 2),
    ('Walls', 'WALL 3', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 3', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 3', 'Door', '', 12, ''),
    ('Walls', 'WALL 3', 'Door Qty', '', 7, ''),

    # WALLS - W4
    ('Walls', 'WALL 4', 'Exposure', '', 7, 'W'),
    ('Walls', 'WALL 4', 'Gross Area', 'm2', 9, 20),
    ('Walls', 'WALL 4', 'Wall Type', '', 14, 'Paredes Ext'),
    ('Walls', 'WALL 4', 'Window 1', '', 14, ''),
    ('Walls', 'WALL 4', 'Win1 Qty', '', 7, ''),
    ('Walls', 'WALL 4', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 4', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 4', 'Door', '', 12, ''),
    ('Walls', 'WALL 4', 'Door Qty', '', 7, ''),

    # WALLS - W5
    ('Walls', 'WALL 5', 'Exposure', '', 7, ''),
    ('Walls', 'WALL 5', 'Gross Area', 'm2', 9, ''),
    ('Walls', 'WALL 5', 'Wall Type', '', 14, ''),
    ('Walls', 'WALL 5', 'Window 1', '', 14, ''),
    ('Walls', 'WALL 5', 'Win1 Qty', '', 7, ''),
    ('Walls', 'WALL 5', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 5', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 5', 'Door', '', 12, ''),
    ('Walls', 'WALL 5', 'Door Qty', '', 7, ''),

    # WALLS - W6
    ('Walls', 'WALL 6', 'Exposure', '', 7, ''),
    ('Walls', 'WALL 6', 'Gross Area', 'm2', 9, ''),
    ('Walls', 'WALL 6', 'Wall Type', '', 14, ''),
    ('Walls', 'WALL 6', 'Window 1', '', 14, ''),
    ('Walls', 'WALL 6', 'Win1 Qty', '', 7, ''),
    ('Walls', 'WALL 6', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 6', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 6', 'Door', '', 12, ''),
    ('Walls', 'WALL 6', 'Door Qty', '', 7, ''),

    # WALLS - W7
    ('Walls', 'WALL 7', 'Exposure', '', 7, ''),
    ('Walls', 'WALL 7', 'Gross Area', 'm2', 9, ''),
    ('Walls', 'WALL 7', 'Wall Type', '', 14, ''),
    ('Walls', 'WALL 7', 'Window 1', '', 14, ''),
    ('Walls', 'WALL 7', 'Win1 Qty', '', 7, ''),
    ('Walls', 'WALL 7', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 7', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 7', 'Door', '', 12, ''),
    ('Walls', 'WALL 7', 'Door Qty', '', 7, ''),

    # WALLS - W8
    ('Walls', 'WALL 8', 'Exposure', '', 7, ''),
    ('Walls', 'WALL 8', 'Gross Area', 'm2', 9, ''),
    ('Walls', 'WALL 8', 'Wall Type', '', 14, ''),
    ('Walls', 'WALL 8', 'Window 1', '', 14, ''),
    ('Walls', 'WALL 8', 'Win1 Qty', '', 7, ''),
    ('Walls', 'WALL 8', 'Window 2', '', 14, ''),
    ('Walls', 'WALL 8', 'Win2 Qty', '', 7, ''),
    ('Walls', 'WALL 8', 'Door', '', 12, ''),
    ('Walls', 'WALL 8', 'Door Qty', '', 7, ''),

    # ROOFS - R1 (6 colunas)
    ('Roofs', 'ROOF 1', 'Exposure', '', 7, 'N'),
    ('Roofs', 'ROOF 1', 'Gross Area', 'm2', 9, 50),
    ('Roofs', 'ROOF 1', 'Slope', 'deg', 7, 0),
    ('Roofs', 'ROOF 1', 'Roof Type', '', 14, 'Cobert Plana'),
    ('Roofs', 'ROOF 1', 'Skylight', '', 14, ''),
    ('Roofs', 'ROOF 1', 'Sky Qty', '', 7, ''),

    # ROOFS - R2
    ('Roofs', 'ROOF 2', 'Exposure', '', 7, ''),
    ('Roofs', 'ROOF 2', 'Gross Area', 'm2', 9, ''),
    ('Roofs', 'ROOF 2', 'Slope', 'deg', 7, ''),
    ('Roofs', 'ROOF 2', 'Roof Type', '', 14, ''),
    ('Roofs', 'ROOF 2', 'Skylight', '', 14, ''),
    ('Roofs', 'ROOF 2', 'Sky Qty', '', 7, ''),

    # ROOFS - R3
    ('Roofs', 'ROOF 3', 'Exposure', '', 7, ''),
    ('Roofs', 'ROOF 3', 'Gross Area', 'm2', 9, ''),
    ('Roofs', 'ROOF 3', 'Slope', 'deg', 7, ''),
    ('Roofs', 'ROOF 3', 'Roof Type', '', 14, ''),
    ('Roofs', 'ROOF 3', 'Skylight', '', 14, ''),
    ('Roofs', 'ROOF 3', 'Sky Qty', '', 7, ''),

    # ROOFS - R4
    ('Roofs', 'ROOF 4', 'Exposure', '', 7, ''),
    ('Roofs', 'ROOF 4', 'Gross Area', 'm2', 9, ''),
    ('Roofs', 'ROOF 4', 'Slope', 'deg', 7, ''),
    ('Roofs', 'ROOF 4', 'Roof Type', '', 14, ''),
    ('Roofs', 'ROOF 4', 'Skylight', '', 14, ''),
    ('Roofs', 'ROOF 4', 'Sky Qty', '', 7, ''),
]

# ============================================================================
# SHEET: Espacos
# ============================================================================
ws = wb.active
ws.title = 'Espacos'

# Freeze linhas 1-4 (grupos + subgrupos + headers + exemplo)
ws.freeze_panes = 'A5'

# ============================================================================
# LINHA 1: GRUPOS (GENERAL, INTERNALS, etc.)
# ============================================================================
groups = {}
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    if section not in groups:
        groups[section] = {'start': col, 'end': col}
    else:
        groups[section]['end'] = col

for section, span in groups.items():
    start_col = span['start']
    end_col = span['end']
    if start_col != end_col:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=section.upper())
    style_group(cell, section)

ws.row_dimensions[1].height = 22

# ============================================================================
# LINHA 2: SUBGRUPOS (WALL 1, WALL 2, PEOPLE, LIGHTING, etc.)
# ============================================================================
subgroups = {}
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    key = (section, subgroup)
    if key not in subgroups:
        subgroups[key] = {'start': col, 'end': col, 'section': section, 'subgroup': subgroup}
    else:
        subgroups[key]['end'] = col

for key, span in subgroups.items():
    start_col = span['start']
    end_col = span['end']
    section = span['section']
    subgroup = span['subgroup']
    if start_col != end_col:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    cell = ws.cell(row=2, column=start_col, value=subgroup if subgroup else '')
    style_subgroup(cell, section, subgroup)

ws.row_dimensions[2].height = 20

# ============================================================================
# LINHA 3: HEADERS COM UNIDADES
# ============================================================================
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    if unit:
        text = f"{header}\n({unit})"
    else:
        text = header
    is_first = is_first_col_of_subgroup(columns, col-1)
    cell = ws.cell(row=3, column=col, value=text)
    style_header(cell, section, subgroup, is_first)
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[3].height = 35

# ============================================================================
# LINHA 4: EXEMPLO PREENCHIDO
# ============================================================================
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    is_first = is_first_col_of_subgroup(columns, col-1)
    cell = ws.cell(row=4, column=col, value=example)
    style_example(cell, section, subgroup, is_first)

# ============================================================================
# LINHAS 5-54: DADOS (50 espacos)
# ============================================================================
for row in range(5, 55):
    for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
        is_first = is_first_col_of_subgroup(columns, col-1)
        cell = ws.cell(row=row, column=col, value='')
        style_data(cell, section, subgroup, is_first)

# ============================================================================
# DATA VALIDATIONS (Dropdowns)
# ============================================================================

# OA Unit dropdown - exactamente como no HAP
col_oa_unit = 6
dv_oa = DataValidation(type='list', formula1='"L/s,L/s/m2,L/s/person,%"', allow_blank=True)
ws.add_data_validation(dv_oa)
for row in range(4, 55):
    dv_oa.add(ws.cell(row=row, column=col_oa_unit))

# Activity Level dropdown - exactamente como no HAP
col_activity = 8
dv_activity = DataValidation(type='list', formula1='"Seated at Rest,Office Work,Sedentary Work,Light Bench Work,Medium Work,Heavy Work,Dancing,Athletics"', allow_blank=True)
ws.add_data_validation(dv_activity)
for row in range(4, 55):
    dv_activity.add(ws.cell(row=row, column=col_activity))

# Infiltration Method dropdown - exactamente como no HAP
col_infil = None
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    if header == 'Infil Method':
        col_infil = col
        break
if col_infil:
    dv_mode = DataValidation(type='list', formula1='"Air Change,Crack Method"', allow_blank=True)
    ws.add_data_validation(dv_mode)
    for row in range(4, 55):
        dv_mode.add(ws.cell(row=row, column=col_infil))

# Floor Type dropdown - exactamente como no HAP
col_floor = None
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    if header == 'Floor Type':
        col_floor = col
        break
if col_floor:
    dv_floor = DataValidation(type='list', formula1='"Floor Above Cond Space,Floor Above Uncond Space,Slab Floor On Grade,Slab Floor Below Grade"', allow_blank=True)
    ws.add_data_validation(dv_floor)
    for row in range(4, 55):
        dv_floor.add(ws.cell(row=row, column=col_floor))

# Fixture Type dropdown - exactamente como no HAP
col_fixture = None
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    if header == 'Fixture Type':
        col_fixture = col
        break
if col_fixture:
    dv_fixture = DataValidation(type='list', formula1='"Recessed Unvented,Vented to Return Air,Vented to Supply & Return,Surface Mount/Pendant"', allow_blank=True)
    ws.add_data_validation(dv_fixture)
    for row in range(4, 55):
        dv_fixture.add(ws.cell(row=row, column=col_fixture))

# Exposure dropdown para Walls e Roofs
dv_dir = DataValidation(type='list', formula1='"N,NNE,NE,ENE,E,ESE,SE,SSE,S,SSW,SW,WSW,W,WNW,NW,NNW"', allow_blank=True)
ws.add_data_validation(dv_dir)

dir_cols = []
for col, (section, subgroup, header, unit, width, example) in enumerate(columns, 1):
    if header == 'Exposure':
        dir_cols.append(col)

for col in dir_cols:
    for row in range(4, 55):
        dv_dir.add(ws.cell(row=row, column=col))

print(f'Sheet Espacos criada com {len(columns)} colunas')

# ============================================================================
# SHEET: Tipos (Reference data)
# ============================================================================
ws_tipos = wb.create_sheet('Tipos')

# Estilo para esta sheet
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill_wall = PatternFill(start_color=COLORS_DARK['Walls'], end_color=COLORS_DARK['Walls'], fill_type='solid')
header_fill_roof = PatternFill(start_color=COLORS_DARK['Roofs'], end_color=COLORS_DARK['Roofs'], fill_type='solid')
header_fill_int = PatternFill(start_color=COLORS_DARK['Internals'], end_color=COLORS_DARK['Internals'], fill_type='solid')

# Wall Types
ws_tipos.cell(row=1, column=1, value='WALL TYPES')
ws_tipos.cell(row=1, column=1).font = header_font
ws_tipos.cell(row=1, column=1).fill = header_fill_wall
ws_tipos.cell(row=1, column=2).fill = header_fill_wall
ws_tipos.merge_cells('A1:B1')
ws_tipos.cell(row=2, column=1, value='ID').font = Font(bold=True)
ws_tipos.cell(row=2, column=2, value='Nome').font = Font(bold=True)
ws_tipos.column_dimensions['A'].width = 6
ws_tipos.column_dimensions['B'].width = 35

# Window Types
ws_tipos.cell(row=1, column=4, value='WINDOW TYPES')
ws_tipos.cell(row=1, column=4).font = header_font
ws_tipos.cell(row=1, column=4).fill = header_fill_wall
ws_tipos.cell(row=1, column=5).fill = header_fill_wall
ws_tipos.merge_cells('D1:E1')
ws_tipos.cell(row=2, column=4, value='ID').font = Font(bold=True)
ws_tipos.cell(row=2, column=5, value='Nome').font = Font(bold=True)
ws_tipos.column_dimensions['D'].width = 6
ws_tipos.column_dimensions['E'].width = 35

# Door Types
ws_tipos.cell(row=1, column=7, value='DOOR TYPES')
ws_tipos.cell(row=1, column=7).font = header_font
ws_tipos.cell(row=1, column=7).fill = header_fill_wall
ws_tipos.cell(row=1, column=8).fill = header_fill_wall
ws_tipos.merge_cells('G1:H1')
ws_tipos.cell(row=2, column=7, value='ID').font = Font(bold=True)
ws_tipos.cell(row=2, column=8, value='Nome').font = Font(bold=True)
ws_tipos.column_dimensions['G'].width = 6
ws_tipos.column_dimensions['H'].width = 35

# Roof Types
ws_tipos.cell(row=1, column=10, value='ROOF TYPES')
ws_tipos.cell(row=1, column=10).font = header_font
ws_tipos.cell(row=1, column=10).fill = header_fill_roof
ws_tipos.cell(row=1, column=11).fill = header_fill_roof
ws_tipos.merge_cells('J1:K1')
ws_tipos.cell(row=2, column=10, value='ID').font = Font(bold=True)
ws_tipos.cell(row=2, column=11, value='Nome').font = Font(bold=True)
ws_tipos.column_dimensions['J'].width = 6
ws_tipos.column_dimensions['K'].width = 35

# Schedules
ws_tipos.cell(row=1, column=13, value='SCHEDULES')
ws_tipos.cell(row=1, column=13).font = header_font
ws_tipos.cell(row=1, column=13).fill = header_fill_int
ws_tipos.cell(row=1, column=14).fill = header_fill_int
ws_tipos.merge_cells('M1:N1')
ws_tipos.cell(row=2, column=13, value='ID').font = Font(bold=True)
ws_tipos.cell(row=2, column=14, value='Nome').font = Font(bold=True)
ws_tipos.column_dimensions['M'].width = 6
ws_tipos.column_dimensions['N'].width = 35

# Exemplos de dados
examples = [
    # Wall Types
    [(3, 1, 1), (3, 2, 'Parede Exterior')],
    [(4, 1, 2), (4, 2, 'Parede Interior')],
    # Window Types
    [(3, 4, 1), (3, 5, 'Janela Aluminio')],
    [(4, 4, 2), (4, 5, 'Janela PVC')],
    # Door Types
    [(3, 7, 1), (3, 8, 'Porta Exterior')],
    # Roof Types
    [(3, 10, 1), (3, 11, 'Cobertura Plana')],
    # Schedules
    [(3, 13, 1), (3, 14, 'Ocupacao Escritorio')],
    [(4, 13, 2), (4, 14, 'Iluminacao Escritorio')],
]

for cells in examples:
    for row, col, val in cells:
        ws_tipos.cell(row=row, column=col, value=val)

print('Sheet Tipos criada')

# ============================================================================
# SHEET: Legenda
# ============================================================================
ws_leg = wb.create_sheet('Legenda')
ws_leg.column_dimensions['A'].width = 25
ws_leg.column_dimensions['B'].width = 60

legend = [
    ('HAP 5.1 INPUT TEMPLATE', ''),
    ('', ''),
    ('GENERAL (Azul)', ''),
    ('  Nome', 'Nome do espaco (max 24 caracteres)'),
    ('  Area', 'Area do piso em m2'),
    ('  Altura', 'Altura do tecto em metros'),
    ('  Peso Edif', 'Peso do edificio em kg/m2 (tipico: 70-300)'),
    ('  OA', 'Outdoor Air - caudal de ar novo'),
    ('  OA Unit', 'L/s, L/s/m2, L/s/person, ou %'),
    ('', ''),
    ('INTERNALS (Verde)', ''),
    ('  Pessoas', 'Numero de ocupantes'),
    ('  Actividade', 'Nivel de actividade metabolica'),
    ('  Sensivel/Latente', 'Calor por pessoa em Watts'),
    ('  Luz Tarefa/Geral', 'Potencia de iluminacao em Watts'),
    ('  Ballast', 'Multiplicador do balastro (tipico: 1.0-1.2)'),
    ('  Equip', 'Equipamentos em W/m2'),
    ('  Sch ID', 'ID do Schedule (ver sheet Tipos)'),
    ('', ''),
    ('INFILTRATION (Laranja)', ''),
    ('  Modo', 'When Fan Off ou All Hours'),
    ('  ACH', 'Air Changes per Hour (renovacoes/hora)'),
    ('', ''),
    ('FLOORS (Verde Lima)', ''),
    ('  Tipo Piso', 'Above Cond, Above Uncond, Slab On/Below Grade'),
    ('  U Piso', 'U-value em W/(m2.K)'),
    ('  Perimetro', 'Perimetro exposto em metros (para Slab)'),
    ('  R Aresta', 'R-value isolamento aresta em (m2.K)/W'),
    ('  Profund', 'Profundidade abaixo do solo em metros'),
    ('  T Unc/Amb', 'Temperaturas Max/Min em Celsius'),
    ('', ''),
    ('PARTITIONS (Lavanda)', ''),
    ('  Tipo', 'Ceiling (tecto) ou Wall (parede)'),
    ('  Area', 'Area da particao em m2'),
    ('  U', 'U-value em W/(m2.K)'),
    ('  Temps', 'Temperaturas adjacentes em Celsius'),
    ('', ''),
    ('WALLS W1-W8 (Laranja Claro)', ''),
    ('  Dir', 'Orientacao: N, NE, E, SE, S, SW, W, NW, etc'),
    ('  Area', 'Area bruta da parede em m2'),
    ('  Tipo ID', 'ID do tipo de parede (ver sheet Tipos)'),
    ('  Jan1/Jan2', 'ID e quantidade de janelas'),
    ('  Porta', 'ID e quantidade de portas'),
    ('', ''),
    ('ROOFS R1-R4 (Roxo)', ''),
    ('  Dir', 'Orientacao da cobertura'),
    ('  Area', 'Area da cobertura em m2'),
    ('  Inclin', 'Inclinacao em graus (0 = horizontal)'),
    ('  Tipo ID', 'ID do tipo de cobertura'),
    ('  Clarab', 'Claraboia: usa IDs de Window'),
    ('', ''),
    ('CONVERSOES AUTOMATICAS', ''),
    ('  Areas', 'm2 -> ft2 (x 10.7639)'),
    ('  U-values', 'W/(m2.K) -> BTU/(hr.ft2.F) (/ 5.678)'),
    ('  R-values', '(m2.K)/W -> (hr.ft2.F)/BTU (x 5.678)'),
    ('  Temperaturas', 'C -> F (x 9/5 + 32)'),
    ('  Comprimentos', 'm -> ft (x 3.28084)'),
]

for i, (col1, col2) in enumerate(legend, 1):
    ws_leg.cell(row=i, column=1, value=col1)
    ws_leg.cell(row=i, column=2, value=col2)

    # Styling
    if col1 == 'HAP 5.1 INPUT TEMPLATE':
        ws_leg.cell(row=i, column=1).font = Font(bold=True, size=14)
    elif col1.endswith(')') and not col1.startswith(' '):
        ws_leg.cell(row=i, column=1).font = Font(bold=True, size=11)
    elif col1 == 'CONVERSOES AUTOMATICAS':
        ws_leg.cell(row=i, column=1).font = Font(bold=True, size=11)

print('Sheet Legenda criada')

# ============================================================================
# GUARDAR
# ============================================================================
output_file = r'C:\Users\pedro\Downloads\HAPPXXXX\HAP_Input_Template.xlsx'
wb.save(output_file)
print()
print(f'=== FICHEIRO CRIADO: {output_file} ===')
print(f'Colunas: {len(columns)}')
print('Linha 1: Grupos (GENERAL, INTERNALS, FLOORS, PARTITIONS, WALLS, ROOFS)')
print('Linha 2: Headers com unidades')
print('Linha 3: Exemplo preenchido')
print('Linhas 4-53: 50 espacos disponiveis')
