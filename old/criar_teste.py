import openpyxl
from openpyxl.styles import Font

# Usar o template existente!
wb = openpyxl.load_workbook('HAP_Template_RSECE.xlsx')

# SHEET: Windows - limpar e preencher
ws_win = wb['Windows']
for row in range(4, 20):
    for col in range(1, 6):
        ws_win.cell(row=row, column=col).value = None

windows_data = [
    ['Janela_Duplo_Normal', 2.8, 0.65, 1.2, 1.0],
    ['Janela_Duplo_Solar', 1.8, 0.35, 1.2, 1.0],
    ['Porta_Vidro', 2.5, 0.60, 2.2, 1.5],
    ['Claraboia', 2.2, 0.45, 0.8, 0.8],
]
for row, data in enumerate(windows_data, 4):
    for col, val in enumerate(data, 1):
        ws_win.cell(row=row, column=col, value=val)

# SHEET: Walls - limpar e preencher
ws_wal = wb['Walls']
for row in range(4, 20):
    for col in range(1, 5):
        ws_wal.cell(row=row, column=col).value = None

walls_data = [
    ['Parede_Ext_ETICS', 0.35, 250, 0.35],
    ['Parede_Int_15', 1.20, 150, 0.15],
    ['Parede_Meacao', 0.50, 300, 0.30],
]
for row, data in enumerate(walls_data, 4):
    for col, val in enumerate(data, 1):
        ws_wal.cell(row=row, column=col, value=val)

# SHEET: Roofs - limpar e preencher
ws_rof = wb['Roofs']
for row in range(4, 20):
    for col in range(1, 5):
        ws_rof.cell(row=row, column=col).value = None

roofs_data = [
    ['Cobertura_Plana', 0.40, 350, 0.35],
    ['Cobertura_Inclinada', 0.35, 200, 0.30],
]
for row, data in enumerate(roofs_data, 4):
    for col, val in enumerate(data, 1):
        ws_rof.cell(row=row, column=col, value=val)

# SHEET: Espacos - limpar dados existentes
ws_esp = wb['Espacos']
for row in range(4, 50):
    for col in range(1, 150):
        ws_esp.cell(row=row, column=col).value = None

# DADOS DOS 5 ESPACOS
espacos = [
    {'nome': 'Sala', 'area': 35.0, 'altura': 2.8, 'peso': 250, 'oa': 35, 'oa_unit': 'L/s',
     'ocupacao': 6, 'actividade': 'Seated at Rest', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
     'task_w': 0, 'general_w': 150, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
     'equip_wm2': 5.0, 'equip_sch': 'Escritorio Equip', 'ach_clg': 0.6, 'ach_htg': 0.6, 'ach_energy': 0.6,
     'walls': [
         {'exp': 'N', 'area': 12.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Normal', 'win1_qty': 2},
         {'exp': 'S', 'area': 12.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Porta_Vidro', 'win1_qty': 1},
         {'exp': 'E', 'area': 8.0, 'type': 'Parede_Int_15'},
         {'exp': 'W', 'area': 8.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Solar', 'win1_qty': 1},
     ], 'roofs': []},
    {'nome': 'Quarto1', 'area': 16.0, 'altura': 2.8, 'peso': 250, 'oa': 16, 'oa_unit': 'L/s',
     'ocupacao': 2, 'actividade': 'Seated at Rest', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
     'task_w': 0, 'general_w': 80, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
     'equip_wm2': 2.0, 'equip_sch': 'Escritorio Equip', 'ach_clg': 0.6, 'ach_htg': 0.6, 'ach_energy': 0.6,
     'walls': [
         {'exp': 'N', 'area': 8.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Normal', 'win1_qty': 1},
         {'exp': 'E', 'area': 8.0, 'type': 'Parede_Int_15'},
         {'exp': 'W', 'area': 8.0, 'type': 'Parede_Int_15'},
     ], 'roofs': [{'exp': 'N', 'area': 16.0, 'slope': 0, 'type': 'Cobertura_Plana'}]},
    {'nome': 'Quarto2', 'area': 14.0, 'altura': 2.8, 'peso': 250, 'oa': 14, 'oa_unit': 'L/s',
     'ocupacao': 2, 'actividade': 'Seated at Rest', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
     'task_w': 0, 'general_w': 70, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
     'equip_wm2': 2.0, 'equip_sch': 'Escritorio Equip', 'ach_clg': 0.6, 'ach_htg': 0.6, 'ach_energy': 0.6,
     'walls': [
         {'exp': 'S', 'area': 7.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Solar', 'win1_qty': 1},
         {'exp': 'E', 'area': 8.0, 'type': 'Parede_Int_15'},
         {'exp': 'W', 'area': 8.0, 'type': 'Parede_Ext_ETICS'},
     ], 'roofs': [{'exp': 'S', 'area': 14.0, 'slope': 15, 'type': 'Cobertura_Inclinada'}]},
    {'nome': 'WC', 'area': 5.0, 'altura': 2.5, 'peso': 300, 'oa': 25, 'oa_unit': 'L/s',
     'ocupacao': 1, 'actividade': 'Seated at Rest', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
     'task_w': 0, 'general_w': 40, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
     'equip_wm2': 0, 'equip_sch': 'Escritorio Equip', 'ach_clg': 1.0, 'ach_htg': 1.0, 'ach_energy': 1.0,
     'walls': [
         {'exp': 'N', 'area': 5.0, 'type': 'Parede_Int_15'},
         {'exp': 'E', 'area': 5.0, 'type': 'Parede_Int_15'},
         {'exp': 'W', 'area': 5.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Normal', 'win1_qty': 1},
     ], 'roofs': []},
    {'nome': 'Cozinha', 'area': 12.0, 'altura': 2.8, 'peso': 300, 'oa': 60, 'oa_unit': 'L/s',
     'ocupacao': 2, 'actividade': 'Light Bench Work', 'sensible': 75, 'latent': 55, 'people_sch': 'Escritorio Ocup',
     'task_w': 50, 'general_w': 80, 'fixture': 'Recessed Unvented', 'ballast': 1.0, 'light_sch': 'Escritorio Ilum',
     'equip_wm2': 15.0, 'equip_sch': 'Escritorio Equip', 'ach_clg': 0.8, 'ach_htg': 0.8, 'ach_energy': 0.8,
     'walls': [
         {'exp': 'S', 'area': 8.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Solar', 'win1_qty': 1},
         {'exp': 'E', 'area': 6.0, 'type': 'Parede_Ext_ETICS', 'win1': 'Janela_Duplo_Normal', 'win1_qty': 1},
         {'exp': 'N', 'area': 8.0, 'type': 'Parede_Int_15'},
         {'exp': 'W', 'area': 6.0, 'type': 'Parede_Int_15'},
     ], 'roofs': [{'exp': 'S', 'area': 12.0, 'slope': 0, 'type': 'Cobertura_Plana', 'sky': 'Claraboia', 'sky_qty': 1}]},
]

for row_idx, esp in enumerate(espacos, 4):
    ws_esp.cell(row=row_idx, column=1, value=esp['nome'])
    ws_esp.cell(row=row_idx, column=2, value=esp['area'])
    ws_esp.cell(row=row_idx, column=3, value=esp['altura'])
    ws_esp.cell(row=row_idx, column=4, value=esp['peso'])
    ws_esp.cell(row=row_idx, column=5, value=esp['oa'])
    ws_esp.cell(row=row_idx, column=6, value=esp['oa_unit'])
    ws_esp.cell(row=row_idx, column=7, value=esp['ocupacao'])
    ws_esp.cell(row=row_idx, column=8, value=esp['actividade'])
    ws_esp.cell(row=row_idx, column=9, value=esp['sensible'])
    ws_esp.cell(row=row_idx, column=10, value=esp['latent'])
    ws_esp.cell(row=row_idx, column=11, value=esp['people_sch'])
    ws_esp.cell(row=row_idx, column=12, value=esp['task_w'])
    ws_esp.cell(row=row_idx, column=13, value=esp['general_w'])
    ws_esp.cell(row=row_idx, column=14, value=esp['fixture'])
    ws_esp.cell(row=row_idx, column=15, value=esp['ballast'])
    ws_esp.cell(row=row_idx, column=16, value=esp['light_sch'])
    ws_esp.cell(row=row_idx, column=17, value=esp['equip_wm2'])
    ws_esp.cell(row=row_idx, column=18, value=esp['equip_sch'])
    ws_esp.cell(row=row_idx, column=24, value=esp['ach_clg'])
    ws_esp.cell(row=row_idx, column=25, value=esp['ach_htg'])
    ws_esp.cell(row=row_idx, column=26, value=esp['ach_energy'])

    for w_idx, wall in enumerate(esp.get('walls', [])):
        col_base = 51 + w_idx * 9
        ws_esp.cell(row=row_idx, column=col_base, value=wall.get('exp'))
        ws_esp.cell(row=row_idx, column=col_base + 1, value=wall.get('area'))
        ws_esp.cell(row=row_idx, column=col_base + 2, value=wall.get('type'))
        ws_esp.cell(row=row_idx, column=col_base + 3, value=wall.get('win1'))
        ws_esp.cell(row=row_idx, column=col_base + 4, value=wall.get('win1_qty'))

    for r_idx, roof in enumerate(esp.get('roofs', [])):
        col_base = 123 + r_idx * 6
        ws_esp.cell(row=row_idx, column=col_base, value=roof.get('exp'))
        ws_esp.cell(row=row_idx, column=col_base + 1, value=roof.get('area'))
        ws_esp.cell(row=row_idx, column=col_base + 2, value=roof.get('slope'))
        ws_esp.cell(row=row_idx, column=col_base + 3, value=roof.get('type'))
        ws_esp.cell(row=row_idx, column=col_base + 4, value=roof.get('sky'))
        ws_esp.cell(row=row_idx, column=col_base + 5, value=roof.get('sky_qty'))

wb.save('Edificio_Teste_Template.xlsx')
print('Ficheiro criado: Edificio_Teste_Template.xlsx')
print('(Baseado no HAP_Template_RSECE.xlsx)')
print()
print('=== RESUMO DO EDIFICIO TESTE ===')
print('Windows: 4 tipos')
print('  - Janela_Duplo_Normal: U=2.8, SHGC=0.65')
print('  - Janela_Duplo_Solar: U=1.8, SHGC=0.35')
print('  - Porta_Vidro: U=2.5, SHGC=0.60')
print('  - Claraboia: U=2.2, SHGC=0.45')
print()
print('Walls: 3 tipos')
print('  - Parede_Ext_ETICS: U=0.35')
print('  - Parede_Int_15: U=1.20')
print('  - Parede_Meacao: U=0.50')
print()
print('Roofs: 2 tipos')
print('  - Cobertura_Plana: U=0.40')
print('  - Cobertura_Inclinada: U=0.35')
print()
print('Espacos: 5')
print('  1. Sala: 35m2, 4 paredes, 4 janelas')
print('  2. Quarto1: 16m2, cobertura plana')
print('  3. Quarto2: 14m2, cobertura inclinada')
print('  4. WC: 5m2, 1 janela')
print('  5. Cozinha: 12m2, claraboia, equip=15W/m2')
