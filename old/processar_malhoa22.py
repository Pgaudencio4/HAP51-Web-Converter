"""
Processar HAP_Template_Malhoa22.xlsx e criar ficheiro completo
"""
import openpyxl
import shutil
import math

# Carregar o ficheiro original
wb = openpyxl.load_workbook('HAP_Template_Malhoa22.xlsx')

# ============================================================
# 1. LER WINDOWS DA SHEET SEPARADA
# ============================================================
ws_win = wb['Windows']
windows_by_zone = {}

for row in range(3, ws_win.max_row + 1):
    name = ws_win.cell(row=row, column=1).value
    zone = ws_win.cell(row=row, column=2).value
    if not name or not zone:
        continue

    ori = ws_win.cell(row=row, column=4).value
    u_value = ws_win.cell(row=row, column=5).value or 5.75
    shgc = ws_win.cell(row=row, column=6).value or 0.85
    area = ws_win.cell(row=row, column=7).value or 0

    if zone not in windows_by_zone:
        windows_by_zone[zone] = {}

    # Agrupar por orientação
    if ori not in windows_by_zone[zone]:
        windows_by_zone[zone][ori] = {
            'name': name,
            'u': u_value,
            'shgc': shgc,
            'area': area,
        }
    else:
        # Somar áreas se mesma orientação
        windows_by_zone[zone][ori]['area'] += area

print(f"Janelas processadas: {len(windows_by_zone)} zonas com janelas")

# ============================================================
# 2. LER WALLS DA SHEET SEPARADA
# ============================================================
ws_walls = wb['Walls']
wall_types = {}
for row in range(4, ws_walls.max_row + 1):
    name = ws_walls.cell(row=row, column=1).value
    if not name:
        continue
    wall_types[name] = {
        'u': ws_walls.cell(row=row, column=2).value or 0.5,
        'weight': ws_walls.cell(row=row, column=3).value or 200,
        'thickness': ws_walls.cell(row=row, column=4).value or 0.3,
    }
print(f"Wall types: {list(wall_types.keys())}")

# ============================================================
# 3. LER ROOFS DA SHEET SEPARADA
# ============================================================
ws_roofs = wb['Roofs']
roof_types = {}
for row in range(4, ws_roofs.max_row + 1):
    name = ws_roofs.cell(row=row, column=1).value
    if not name:
        continue
    roof_types[name] = {
        'u': ws_roofs.cell(row=row, column=2).value or 0.4,
        'weight': ws_roofs.cell(row=row, column=3).value or 300,
        'thickness': ws_roofs.cell(row=row, column=4).value or 0.35,
    }
print(f"Roof types: {list(roof_types.keys())}")

# ============================================================
# 4. ACTUALIZAR SHEET ESPACOS COM WALLS E WINDOWS
# ============================================================
ws = wb['Espacos']

# Mapeamento de orientações
ORI_MAP = {
    'N': 'N', 'S': 'S', 'E': 'E', 'W': 'W',
    'NE': 'NE', 'NW': 'NW', 'SE': 'SE', 'SW': 'SW',
    'Norte': 'N', 'Sul': 'S', 'Este': 'E', 'Oeste': 'W',
}

# Nome default para wall type
DEFAULT_WALL = 'Parede Exterior'
DEFAULT_ROOF = 'Cobertura Plana'

# Criar tipos de janelas únicos
unique_windows = {}
for zone, oris in windows_by_zone.items():
    for ori, wdata in oris.items():
        # Usar dimensões aproximadas baseadas na área
        area = wdata['area']
        # Assumir janela standard: altura 1.5m
        height = 1.5
        width = area / height if area > 0 else 1.0

        win_key = f"Win_{wdata['u']:.2f}_{wdata['shgc']:.2f}"
        if win_key not in unique_windows:
            unique_windows[win_key] = {
                'name': win_key,
                'u': wdata['u'],
                'shgc': wdata['shgc'],
                'height': height,
                'width': width,
            }

print(f"Tipos de janelas únicos: {len(unique_windows)}")

# Para cada espaço, criar walls baseadas nas janelas
spaces_updated = 0
for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    # Procurar zona correspondente nas janelas
    zone_name = space_name

    if zone_name in windows_by_zone:
        windows = windows_by_zone[zone_name]

        # Criar uma wall para cada orientação com janelas
        wall_idx = 0
        for ori, wdata in windows.items():
            if wall_idx >= 8:  # Máximo 8 walls
                break

            # Calcular área da parede (assumir parede maior que janela)
            win_area = wdata['area']
            wall_area = win_area * 1.5  # Parede 50% maior que janela

            # Coluna base para esta wall (wall_start = 52, cada wall = 9 colunas)
            col = 52 + wall_idx * 9

            # Escrever dados da wall
            ws.cell(row=row, column=col, value=ORI_MAP.get(ori, ori))  # Exposure
            ws.cell(row=row, column=col+1, value=wall_area)  # Gross Area
            ws.cell(row=row, column=col+2, value=DEFAULT_WALL)  # Wall Type

            # Window 1
            win_type_name = f"Win_{wdata['u']:.2f}_{wdata['shgc']:.2f}"
            ws.cell(row=row, column=col+3, value=win_type_name)  # Window 1
            ws.cell(row=row, column=col+4, value=1)  # Win1 Qty

            wall_idx += 1

        spaces_updated += 1

print(f"Espaços actualizados com walls: {spaces_updated}")

# ============================================================
# 5. ACTUALIZAR SHEET WINDOWS COM FORMATO CORRECTO
# ============================================================
# Limpar e reescrever a sheet Windows no formato esperado pelo conversor
ws_win_new = wb['Windows']

# Limpar dados existentes (manter cabeçalhos)
for row in range(1, ws_win_new.max_row + 1):
    for col in range(1, 9):
        ws_win_new.cell(row=row, column=col, value=None)

# Escrever novos cabeçalhos
ws_win_new.cell(row=1, column=1, value='Nome')
ws_win_new.cell(row=1, column=2, value='U-Value')
ws_win_new.cell(row=1, column=3, value='SHGC')
ws_win_new.cell(row=1, column=4, value='Altura')
ws_win_new.cell(row=1, column=5, value='Largura')

# Escrever tipos de janelas únicos
row = 2
for win_key, wdata in unique_windows.items():
    ws_win_new.cell(row=row, column=1, value=wdata['name'])
    ws_win_new.cell(row=row, column=2, value=wdata['u'])
    ws_win_new.cell(row=row, column=3, value=wdata['shgc'])
    ws_win_new.cell(row=row, column=4, value=wdata['height'])
    ws_win_new.cell(row=row, column=5, value=wdata['width'])
    row += 1

print(f"Sheet Windows actualizada com {len(unique_windows)} tipos")

# ============================================================
# 6. ACTUALIZAR SHEET WALLS COM FORMATO CORRECTO
# ============================================================
ws_walls_new = wb['Walls']

# Limpar e reescrever
for row in range(1, ws_walls_new.max_row + 1):
    for col in range(1, 5):
        ws_walls_new.cell(row=row, column=col, value=None)

ws_walls_new.cell(row=1, column=1, value='Nome')
ws_walls_new.cell(row=1, column=2, value='U-Value')
ws_walls_new.cell(row=1, column=3, value='Peso')
ws_walls_new.cell(row=1, column=4, value='Espessura')

row = 2
for name, wdata in wall_types.items():
    ws_walls_new.cell(row=row, column=1, value=name)
    ws_walls_new.cell(row=row, column=2, value=wdata['u'])
    ws_walls_new.cell(row=row, column=3, value=wdata['weight'])
    ws_walls_new.cell(row=row, column=4, value=wdata['thickness'])
    row += 1

print(f"Sheet Walls actualizada")

# ============================================================
# 7. ACTUALIZAR SHEET ROOFS COM FORMATO CORRECTO
# ============================================================
ws_roofs_new = wb['Roofs']

# Limpar e reescrever
for row in range(1, ws_roofs_new.max_row + 1):
    for col in range(1, 5):
        ws_roofs_new.cell(row=row, column=col, value=None)

ws_roofs_new.cell(row=1, column=1, value='Nome')
ws_roofs_new.cell(row=1, column=2, value='U-Value')
ws_roofs_new.cell(row=1, column=3, value='Peso')
ws_roofs_new.cell(row=1, column=4, value='Espessura')

row = 2
for name, rdata in roof_types.items():
    ws_roofs_new.cell(row=row, column=1, value=name)
    ws_roofs_new.cell(row=row, column=2, value=rdata['u'])
    ws_roofs_new.cell(row=row, column=3, value=rdata['weight'])
    ws_roofs_new.cell(row=row, column=4, value=rdata['thickness'])
    row += 1

print(f"Sheet Roofs actualizada")

# ============================================================
# 8. GUARDAR
# ============================================================
wb.save('Malhoa22_Completo.xlsx')
print("\nFicheiro guardado: Malhoa22_Completo.xlsx")
