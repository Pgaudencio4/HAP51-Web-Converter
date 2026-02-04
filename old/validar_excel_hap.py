"""
Validador de Excel para HAP 5.1
Verifica campo a campo se o ficheiro Excel está correcto para conversão.

Usage:
    python validar_excel_hap.py <ficheiro.xlsx>

Exemplo:
    python validar_excel_hap.py MeuProjeto.xlsx
"""

import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# =============================================================================
# ESPECIFICAÇÃO DO FORMATO HAP
# =============================================================================

SHEETS_OBRIGATORIAS = ['Espacos', 'Windows', 'Walls', 'Roofs', 'Tipos', 'Legenda', 'Schedules_RSECE']

# Valores válidos para campos enumerados
VALORES_VALIDOS = {
    'OA Unit': ['L/s', 'L/s/m2', 'L/s/person', '%'],
    'Activity Level': ['Seated at Rest', 'Office Work', 'Sedentary Work', 'Light Bench Work',
                       'Medium Work', 'Heavy Work', 'Dancing', 'Athletics'],
    'Fixture Type': ['Recessed Unvented', 'Vented to Return Air', 'Vented to Supply & Return',
                     'Surface Mount/Pendant'],
    'Infil Method': ['Air Change'],
    'Floor Type': ['Floor Above Cond Space', 'Floor Above Uncond Space',
                   'Slab Floor On Grade', 'Slab Floor Below Grade'],
    'Exposure': ['N', 'NNE', 'NE', 'ENE', 'E', 'ESE', 'SE', 'SSE',
                 'S', 'SSW', 'SW', 'WSW', 'W', 'WNW', 'NW', 'NNW'],
    'Roof Exposure': ['N', 'NNE', 'NE', 'ENE', 'E', 'ESE', 'SE', 'SSE',
                      'S', 'SSW', 'SW', 'WSW', 'W', 'WNW', 'NW', 'NNW',
                      'H', 'HORIZ', 'HORIZONTAL'],
}

# 82 Schedules RSECE válidos
SCHEDULES_RSECE = [
    'Sample Schedule',
    'Hipermercado Ocup', 'Hipermercado Ilum', 'Hipermercado Equip',
    'Venda Grosso Ocup', 'Venda Grosso Ilum', 'Venda Grosso Equip',
    'Supermercado Ocup', 'Supermercado Ilum', 'Supermercado Equip',
    'Centro Comercial Ocup', 'Centro Comercial Ilum', 'Centro Comercial Equip',
    'Pequena Loja Ocup', 'Pequena Loja Ilum', 'Pequena Loja Equip',
    'Restaurante Ocup', 'Restaurante Ilum', 'Restaurante Equip',
    'Pastelaria Ocup', 'Pastelaria Ilum', 'Pastelaria Equip',
    'Pronto-a-Comer Ocup', 'Pronto-a-Comer Ilum', 'Pronto-a-Comer Equip',
    'Hotel 4-5 Estrelas Ocup', 'Hotel 4-5 Estrelas Ilum', 'Hotel 4-5 Estrelas Equip',
    'Hotel 1-3 Estrelas Ocup', 'Hotel 1-3 Estrelas Ilum', 'Hotel 1-3 Estrelas Equip',
    'Cinema Teatro Ocup', 'Cinema Teatro Ilum', 'Cinema Teatro Equip',
    'Discoteca Ocup', 'Discoteca Ilum', 'Discoteca Equip',
    'Bingo Clube Social Ocup', 'Bingo Clube Social Ilum', 'Bingo Clube Social Equip',
    'Clube Desp Piscina Ocup', 'Clube Desp Piscina Ilum', 'Clube Desp Piscina Equip',
    'Clube Desportivo Ocup', 'Clube Desportivo Ilum', 'Clube Desportivo Equip',
    'Escritorio Ocup', 'Escritorio Ilum', 'Escritorio Equip',
    'Banco Sede Ocup', 'Banco Sede Ilum', 'Banco Sede Equip',
    'Banco Filial Ocup', 'Banco Filial Ilum', 'Banco Filial Equip',
    'Comunicacoes Ocup', 'Comunicacoes Ilum', 'Comunicacoes Equip',
    'Biblioteca Ocup', 'Biblioteca Ilum', 'Biblioteca Equip',
    'Museu Galeria Ocup', 'Museu Galeria Ilum', 'Museu Galeria Equip',
    'Tribunal Camara Ocup', 'Tribunal Camara Ilum', 'Tribunal Camara Equip',
    'Prisao Ocup', 'Prisao Ilum', 'Prisao Equip',
    'Escola Ocup', 'Escola Ilum', 'Escola Equip',
    'Universidade Ocup', 'Universidade Ilum', 'Universidade Equip',
    'Saude Sem Intern Ocup', 'Saude Sem Intern Ilum', 'Saude Sem Intern Equip',
    'Saude Com Intern Ocup', 'Saude Com Intern Ilum', 'Saude Com Intern Equip',
]

# Definição das 147 colunas da sheet Espacos
COLUNAS_ESPACOS = {
    # GENERAL (1-6)
    1: {'nome': 'Space Name', 'tipo': 'string', 'obrigatorio': True, 'max_length': 24},
    2: {'nome': 'Floor Area', 'tipo': 'number', 'obrigatorio': True, 'min': 0},
    3: {'nome': 'Avg Ceiling Ht', 'tipo': 'number', 'obrigatorio': True, 'min': 0},
    4: {'nome': 'Building Wt', 'tipo': 'number', 'obrigatorio': False},
    5: {'nome': 'Outdoor Air', 'tipo': 'number', 'obrigatorio': True},
    6: {'nome': 'OA Unit', 'tipo': 'enum', 'obrigatorio': True, 'valores': 'OA Unit'},

    # PEOPLE (7-11)
    7: {'nome': 'Occupancy', 'tipo': 'number', 'obrigatorio': False, 'min': 0},
    8: {'nome': 'Activity Level', 'tipo': 'enum', 'obrigatorio': False, 'valores': 'Activity Level'},
    9: {'nome': 'Sensible W/person', 'tipo': 'number', 'obrigatorio': False},
    10: {'nome': 'Latent W/person', 'tipo': 'number', 'obrigatorio': False},
    11: {'nome': 'People Schedule', 'tipo': 'schedule', 'obrigatorio': False, 'sufixo': 'Ocup'},

    # LIGHTING (12-16)
    12: {'nome': 'Task Lighting', 'tipo': 'number', 'obrigatorio': False},
    13: {'nome': 'General Lighting', 'tipo': 'number', 'obrigatorio': False},
    14: {'nome': 'Fixture Type', 'tipo': 'enum', 'obrigatorio': False, 'valores': 'Fixture Type'},
    15: {'nome': 'Ballast Mult', 'tipo': 'number', 'obrigatorio': False},
    16: {'nome': 'Light Schedule', 'tipo': 'schedule', 'obrigatorio': False, 'sufixo': 'Ilum'},

    # EQUIPMENT (17-18)
    17: {'nome': 'Equipment W/m2', 'tipo': 'number', 'obrigatorio': False},
    18: {'nome': 'Equipment Schedule', 'tipo': 'schedule', 'obrigatorio': False, 'sufixo': 'Equip'},

    # MISC (19-22)
    19: {'nome': 'Misc Sensible', 'tipo': 'number', 'obrigatorio': False},
    20: {'nome': 'Misc Latent', 'tipo': 'number', 'obrigatorio': False},
    21: {'nome': 'Misc Sens Schedule', 'tipo': 'schedule', 'obrigatorio': False},
    22: {'nome': 'Misc Lat Schedule', 'tipo': 'schedule', 'obrigatorio': False},

    # INFILTRATION (23-26)
    23: {'nome': 'Infil Method', 'tipo': 'enum', 'obrigatorio': False, 'valores': 'Infil Method'},
    24: {'nome': 'Design Clg ACH', 'tipo': 'number', 'obrigatorio': False},
    25: {'nome': 'Design Htg ACH', 'tipo': 'number', 'obrigatorio': False},
    26: {'nome': 'Energy ACH', 'tipo': 'number', 'obrigatorio': False},

    # FLOORS (27-39)
    27: {'nome': 'Floor Type', 'tipo': 'enum', 'obrigatorio': False, 'valores': 'Floor Type'},
    28: {'nome': 'Floor Area', 'tipo': 'number', 'obrigatorio': False},
    29: {'nome': 'Floor U-Value', 'tipo': 'number', 'obrigatorio': False},
    30: {'nome': 'Exp Perim', 'tipo': 'number', 'obrigatorio': False},
    31: {'nome': 'Edge R', 'tipo': 'number', 'obrigatorio': False},
    32: {'nome': 'Depth', 'tipo': 'number', 'obrigatorio': False},
    33: {'nome': 'Bsmt Wall U', 'tipo': 'number', 'obrigatorio': False},
    34: {'nome': 'Wall Ins R', 'tipo': 'number', 'obrigatorio': False},
    35: {'nome': 'Ins Depth', 'tipo': 'number', 'obrigatorio': False},
    36: {'nome': 'Floor Unc Max', 'tipo': 'number', 'obrigatorio': False},
    37: {'nome': 'Floor Out Max', 'tipo': 'number', 'obrigatorio': False},
    38: {'nome': 'Floor Unc Min', 'tipo': 'number', 'obrigatorio': False},
    39: {'nome': 'Floor Out Min', 'tipo': 'number', 'obrigatorio': False},

    # PARTITIONS CEILING (40-45)
    40: {'nome': 'Ceiling Area', 'tipo': 'number', 'obrigatorio': False},
    41: {'nome': 'Ceiling U-Value', 'tipo': 'number', 'obrigatorio': False},
    42: {'nome': 'Ceiling Unc Max', 'tipo': 'number', 'obrigatorio': False},
    43: {'nome': 'Ceiling Out Max', 'tipo': 'number', 'obrigatorio': False},
    44: {'nome': 'Ceiling Unc Min', 'tipo': 'number', 'obrigatorio': False},
    45: {'nome': 'Ceiling Out Min', 'tipo': 'number', 'obrigatorio': False},

    # PARTITIONS WALL (46-51)
    46: {'nome': 'Wall Part Area', 'tipo': 'number', 'obrigatorio': False},
    47: {'nome': 'Wall Part U-Value', 'tipo': 'number', 'obrigatorio': False},
    48: {'nome': 'Wall Part Unc Max', 'tipo': 'number', 'obrigatorio': False},
    49: {'nome': 'Wall Part Out Max', 'tipo': 'number', 'obrigatorio': False},
    50: {'nome': 'Wall Part Unc Min', 'tipo': 'number', 'obrigatorio': False},
    51: {'nome': 'Wall Part Out Min', 'tipo': 'number', 'obrigatorio': False},
}

# Adicionar colunas de WALLS (52-123) - 8 paredes x 9 colunas
for wall_idx in range(8):
    base_col = 52 + (wall_idx * 9)
    wall_num = wall_idx + 1
    COLUNAS_ESPACOS[base_col] = {'nome': f'Wall {wall_num} Exposure', 'tipo': 'enum', 'obrigatorio': False, 'valores': 'Exposure'}
    COLUNAS_ESPACOS[base_col + 1] = {'nome': f'Wall {wall_num} Gross Area', 'tipo': 'number', 'obrigatorio': False, 'min': 0}
    COLUNAS_ESPACOS[base_col + 2] = {'nome': f'Wall {wall_num} Type', 'tipo': 'ref_wall', 'obrigatorio': False}
    COLUNAS_ESPACOS[base_col + 3] = {'nome': f'Wall {wall_num} Window 1', 'tipo': 'ref_window', 'obrigatorio': False}
    COLUNAS_ESPACOS[base_col + 4] = {'nome': f'Wall {wall_num} Win1 Qty', 'tipo': 'number', 'obrigatorio': False, 'min': 0}
    COLUNAS_ESPACOS[base_col + 5] = {'nome': f'Wall {wall_num} Window 2', 'tipo': 'ref_window', 'obrigatorio': False}
    COLUNAS_ESPACOS[base_col + 6] = {'nome': f'Wall {wall_num} Win2 Qty', 'tipo': 'number', 'obrigatorio': False, 'min': 0}
    COLUNAS_ESPACOS[base_col + 7] = {'nome': f'Wall {wall_num} Door', 'tipo': 'ref_door', 'obrigatorio': False}
    COLUNAS_ESPACOS[base_col + 8] = {'nome': f'Wall {wall_num} Door Qty', 'tipo': 'number', 'obrigatorio': False, 'min': 0}

# Adicionar colunas de ROOFS (124-147) - 4 coberturas x 6 colunas
for roof_idx in range(4):
    base_col = 124 + (roof_idx * 6)
    roof_num = roof_idx + 1
    COLUNAS_ESPACOS[base_col] = {'nome': f'Roof {roof_num} Exposure', 'tipo': 'enum', 'obrigatorio': False, 'valores': 'Roof Exposure'}
    COLUNAS_ESPACOS[base_col + 1] = {'nome': f'Roof {roof_num} Gross Area', 'tipo': 'number', 'obrigatorio': False, 'min': 0}
    COLUNAS_ESPACOS[base_col + 2] = {'nome': f'Roof {roof_num} Slope', 'tipo': 'number', 'obrigatorio': False, 'min': 0}
    COLUNAS_ESPACOS[base_col + 3] = {'nome': f'Roof {roof_num} Type', 'tipo': 'ref_roof', 'obrigatorio': False}
    COLUNAS_ESPACOS[base_col + 4] = {'nome': f'Roof {roof_num} Skylight', 'tipo': 'ref_window', 'obrigatorio': False}
    COLUNAS_ESPACOS[base_col + 5] = {'nome': f'Roof {roof_num} Sky Qty', 'tipo': 'number', 'obrigatorio': False, 'min': 0}


# =============================================================================
# CLASSE VALIDADOR
# =============================================================================

class ValidadorHAP:
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = None
        self.erros = []
        self.avisos = []
        self.ok = []
        self.estatisticas = {}

        # Referências carregadas das sheets
        self.windows = set()
        self.walls = set()
        self.roofs = set()
        self.doors = set()

    def log_erro(self, categoria, mensagem, detalhe=None):
        self.erros.append({
            'categoria': categoria,
            'mensagem': mensagem,
            'detalhe': detalhe
        })

    def log_aviso(self, categoria, mensagem, detalhe=None):
        self.avisos.append({
            'categoria': categoria,
            'mensagem': mensagem,
            'detalhe': detalhe
        })

    def log_ok(self, categoria, mensagem):
        self.ok.append({
            'categoria': categoria,
            'mensagem': mensagem
        })

    def carregar_excel(self):
        """Carrega o ficheiro Excel."""
        try:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)
            self.log_ok('FICHEIRO', f'Ficheiro carregado: {os.path.basename(self.filepath)}')
            return True
        except Exception as e:
            self.log_erro('FICHEIRO', f'Erro ao carregar ficheiro', str(e))
            return False

    def validar_sheets(self):
        """Valida se todas as sheets obrigatórias existem."""
        sheets_existentes = self.wb.sheetnames

        for sheet in SHEETS_OBRIGATORIAS:
            if sheet in sheets_existentes:
                self.log_ok('SHEETS', f'Sheet "{sheet}" existe')
            else:
                self.log_erro('SHEETS', f'Sheet "{sheet}" não existe')

        # Verificar sheets extra
        for sheet in sheets_existentes:
            if sheet not in SHEETS_OBRIGATORIAS:
                self.log_aviso('SHEETS', f'Sheet extra encontrada: "{sheet}"')

    def carregar_referencias(self):
        """Carrega os tipos de Windows, Walls, Roofs das respectivas sheets."""

        # Windows
        if 'Windows' in self.wb.sheetnames:
            ws = self.wb['Windows']
            for row in range(4, ws.max_row + 1):
                nome = ws.cell(row=row, column=1).value
                if nome and str(nome).strip():
                    self.windows.add(str(nome).strip())
            self.log_ok('REFERENCIAS', f'{len(self.windows)} tipos de Window carregados')

        # Walls
        if 'Walls' in self.wb.sheetnames:
            ws = self.wb['Walls']
            for row in range(4, ws.max_row + 1):
                nome = ws.cell(row=row, column=1).value
                if nome and str(nome).strip():
                    self.walls.add(str(nome).strip())
            self.log_ok('REFERENCIAS', f'{len(self.walls)} tipos de Wall carregados')

        # Roofs
        if 'Roofs' in self.wb.sheetnames:
            ws = self.wb['Roofs']
            for row in range(4, ws.max_row + 1):
                nome = ws.cell(row=row, column=1).value
                if nome and str(nome).strip():
                    self.roofs.add(str(nome).strip())
            self.log_ok('REFERENCIAS', f'{len(self.roofs)} tipos de Roof carregados')

        # Doors (da sheet Tipos, se existir)
        if 'Tipos' in self.wb.sheetnames:
            ws = self.wb['Tipos']
            for row in range(3, ws.max_row + 1):
                nome = ws.cell(row=row, column=8).value  # Door Types na coluna 8
                if nome and str(nome).strip():
                    self.doors.add(str(nome).strip())
            if self.doors:
                self.log_ok('REFERENCIAS', f'{len(self.doors)} tipos de Door carregados')

    def validar_estrutura_espacos(self):
        """Valida a estrutura da sheet Espacos."""
        if 'Espacos' not in self.wb.sheetnames:
            self.log_erro('ESTRUTURA', 'Sheet Espacos não existe')
            return

        ws = self.wb['Espacos']

        # Verificar número de colunas
        if ws.max_column >= 147:
            self.log_ok('ESTRUTURA', f'Sheet Espacos tem {ws.max_column} colunas (mínimo 147)')
        else:
            self.log_erro('ESTRUTURA', f'Sheet Espacos tem apenas {ws.max_column} colunas (necessário 147)')

        # Verificar linha 1 - Categorias
        categorias_esperadas = {1: 'GENERAL', 7: 'INTERNALS', 23: 'INFILTRATION',
                               27: 'FLOORS', 40: 'PARTITIONS', 52: 'WALLS', 124: 'ROOFS'}
        for col, esperado in categorias_esperadas.items():
            valor = ws.cell(row=1, column=col).value
            if valor == esperado:
                self.log_ok('ESTRUTURA', f'Linha 1, Col {col}: "{esperado}" OK')
            else:
                self.log_erro('ESTRUTURA', f'Linha 1, Col {col}: esperado "{esperado}", encontrado "{valor}"')

        # Verificar linha 3 - Headers principais
        headers_check = {1: 'Space Name', 2: 'Floor Area', 6: 'OA Unit', 11: 'Schedule', 16: 'Schedule'}
        for col, esperado in headers_check.items():
            valor = ws.cell(row=3, column=col).value
            if valor and esperado.lower() in str(valor).lower():
                self.log_ok('ESTRUTURA', f'Header Col {col}: contém "{esperado}"')
            else:
                self.log_aviso('ESTRUTURA', f'Header Col {col}: esperado conter "{esperado}", encontrado "{valor}"')

    def validar_espacos(self):
        """Valida os dados de cada espaço."""
        if 'Espacos' not in self.wb.sheetnames:
            return

        ws = self.wb['Espacos']

        nomes_encontrados = set()
        total_espacos = 0
        campos_preenchidos = {col: 0 for col in COLUNAS_ESPACOS.keys()}
        erros_por_coluna = {col: [] for col in COLUNAS_ESPACOS.keys()}

        # Iterar sobre cada linha de dados (a partir da linha 4)
        for row in range(4, ws.max_row + 1):
            nome = ws.cell(row=row, column=1).value
            if not nome or str(nome).strip() == '':
                continue

            total_espacos += 1
            nome_str = str(nome).strip()

            # Validar cada coluna
            for col, spec in COLUNAS_ESPACOS.items():
                valor = ws.cell(row=row, column=col).value

                # Contabilizar preenchidos
                if valor is not None and str(valor).strip() != '':
                    campos_preenchidos[col] += 1

                # Validar campo
                erro = self._validar_campo(valor, spec, row, col, nome_str)
                if erro:
                    erros_por_coluna[col].append(erro)

            # Verificar nome duplicado
            if nome_str in nomes_encontrados:
                self.log_erro('DADOS', f'Nome duplicado: "{nome_str}"', f'Linha {row}')
            nomes_encontrados.add(nome_str)

            # Verificar tamanho do nome
            if len(nome_str) > 24:
                self.log_erro('DADOS', f'Nome excede 24 caracteres: "{nome_str}" ({len(nome_str)} chars)', f'Linha {row}')

        self.estatisticas['total_espacos'] = total_espacos
        self.estatisticas['campos_preenchidos'] = campos_preenchidos
        self.estatisticas['erros_por_coluna'] = erros_por_coluna

        self.log_ok('DADOS', f'Total de espaços: {total_espacos}')

    def _validar_campo(self, valor, spec, row, col, nome_espaco):
        """Valida um campo individual."""

        # Se vazio e obrigatório
        if (valor is None or str(valor).strip() == '') and spec.get('obrigatorio'):
            return f'Linha {row} ({nome_espaco}): Campo obrigatório vazio'

        # Se vazio e não obrigatório, OK
        if valor is None or str(valor).strip() == '':
            return None

        valor_str = str(valor).strip()
        tipo = spec.get('tipo')

        # Validar tipo number
        if tipo == 'number':
            try:
                num = float(valor)
                if 'min' in spec and num < spec['min']:
                    return f'Linha {row} ({nome_espaco}): Valor {num} abaixo do mínimo {spec["min"]}'
            except:
                return f'Linha {row} ({nome_espaco}): Valor "{valor}" não é número'

        # Validar tipo enum
        elif tipo == 'enum':
            valores_validos = VALORES_VALIDOS.get(spec.get('valores'), [])
            if valor_str not in valores_validos:
                return f'Linha {row} ({nome_espaco}): Valor "{valor_str}" inválido. Válidos: {valores_validos[:3]}...'

        # Validar schedule
        elif tipo == 'schedule':
            if valor_str not in SCHEDULES_RSECE:
                return f'Linha {row} ({nome_espaco}): Schedule "{valor_str}" não existe'
            # Verificar sufixo correcto
            sufixo = spec.get('sufixo')
            if sufixo and not valor_str.endswith(sufixo) and valor_str != 'Sample Schedule':
                return f'Linha {row} ({nome_espaco}): Schedule "{valor_str}" deveria terminar em "{sufixo}"'

        # Validar referência a Window
        elif tipo == 'ref_window':
            if valor_str not in self.windows:
                return f'Linha {row} ({nome_espaco}): Window "{valor_str}" não existe na sheet Windows'

        # Validar referência a Wall
        elif tipo == 'ref_wall':
            if valor_str not in self.walls:
                return f'Linha {row} ({nome_espaco}): Wall Type "{valor_str}" não existe na sheet Walls'

        # Validar referência a Roof
        elif tipo == 'ref_roof':
            if valor_str not in self.roofs:
                return f'Linha {row} ({nome_espaco}): Roof Type "{valor_str}" não existe na sheet Roofs'

        # Validar referência a Door
        elif tipo == 'ref_door':
            if self.doors and valor_str not in self.doors:
                return f'Linha {row} ({nome_espaco}): Door Type "{valor_str}" não existe'

        # Validar tamanho máximo
        if 'max_length' in spec and len(valor_str) > spec['max_length']:
            return f'Linha {row} ({nome_espaco}): Valor excede {spec["max_length"]} caracteres'

        return None

    def validar_sheet_windows(self):
        """Valida a sheet Windows."""
        if 'Windows' not in self.wb.sheetnames:
            self.log_aviso('WINDOWS', 'Sheet Windows não existe')
            return

        ws = self.wb['Windows']
        count = 0

        for row in range(4, ws.max_row + 1):
            nome = ws.cell(row=row, column=1).value
            if not nome:
                continue
            count += 1

            # Verificar campos obrigatórios
            u_value = ws.cell(row=row, column=2).value
            shgc = ws.cell(row=row, column=3).value
            altura = ws.cell(row=row, column=4).value
            largura = ws.cell(row=row, column=5).value

            if u_value is None:
                self.log_erro('WINDOWS', f'U-Value em falta', f'Linha {row}: {nome}')
            if shgc is None:
                self.log_erro('WINDOWS', f'SHGC em falta', f'Linha {row}: {nome}')
            elif shgc and (float(shgc) < 0 or float(shgc) > 1):
                self.log_erro('WINDOWS', f'SHGC deve ser entre 0 e 1', f'Linha {row}: {nome}, valor={shgc}')
            if altura is None:
                self.log_erro('WINDOWS', f'Altura em falta', f'Linha {row}: {nome}')
            if largura is None:
                self.log_erro('WINDOWS', f'Largura em falta', f'Linha {row}: {nome}')

        self.log_ok('WINDOWS', f'{count} tipos de janela validados')

    def validar_sheet_walls(self):
        """Valida a sheet Walls."""
        if 'Walls' not in self.wb.sheetnames:
            self.log_aviso('WALLS', 'Sheet Walls não existe')
            return

        ws = self.wb['Walls']
        count = 0

        for row in range(4, ws.max_row + 1):
            nome = ws.cell(row=row, column=1).value
            if not nome:
                continue
            count += 1

            u_value = ws.cell(row=row, column=2).value
            if u_value is None:
                self.log_erro('WALLS', f'U-Value em falta', f'Linha {row}: {nome}')

        self.log_ok('WALLS', f'{count} tipos de parede validados')

    def validar_sheet_roofs(self):
        """Valida a sheet Roofs."""
        if 'Roofs' not in self.wb.sheetnames:
            self.log_aviso('ROOFS', 'Sheet Roofs não existe')
            return

        ws = self.wb['Roofs']
        count = 0

        for row in range(4, ws.max_row + 1):
            nome = ws.cell(row=row, column=1).value
            if not nome:
                continue
            count += 1

            u_value = ws.cell(row=row, column=2).value
            if u_value is None:
                self.log_erro('ROOFS', f'U-Value em falta', f'Linha {row}: {nome}')

        self.log_ok('ROOFS', f'{count} tipos de cobertura validados')

    def executar(self):
        """Executa todas as validações."""
        print("=" * 70)
        print("VALIDADOR HAP 5.1 - Verificação de Excel")
        print("=" * 70)
        print()

        if not self.carregar_excel():
            return False

        self.validar_sheets()
        self.carregar_referencias()
        self.validar_estrutura_espacos()
        self.validar_espacos()
        self.validar_sheet_windows()
        self.validar_sheet_walls()
        self.validar_sheet_roofs()

        return True

    def gerar_relatorio(self):
        """Gera o relatório de validação."""
        print()
        print("=" * 70)
        print("RELATÓRIO DE VALIDAÇÃO")
        print("=" * 70)
        print()

        # Resumo
        print(f"Ficheiro: {os.path.basename(self.filepath)}")
        print(f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()

        # Estatísticas
        if self.estatisticas:
            print("-" * 70)
            print("ESTATÍSTICAS")
            print("-" * 70)
            print(f"Total de espaços: {self.estatisticas.get('total_espacos', 0)}")
            print()

        # Erros
        print("-" * 70)
        print(f"ERROS: {len(self.erros)}")
        print("-" * 70)
        if self.erros:
            erros_por_cat = {}
            for e in self.erros:
                cat = e['categoria']
                if cat not in erros_por_cat:
                    erros_por_cat[cat] = []
                erros_por_cat[cat].append(e)

            for cat, erros in erros_por_cat.items():
                print(f"\n[{cat}] ({len(erros)} erros)")
                for e in erros[:10]:  # Mostrar máximo 10 por categoria
                    print(f"  ✗ {e['mensagem']}")
                    if e['detalhe']:
                        print(f"    → {e['detalhe']}")
                if len(erros) > 10:
                    print(f"  ... e mais {len(erros) - 10} erros")
        else:
            print("  Nenhum erro encontrado!")

        # Avisos
        print()
        print("-" * 70)
        print(f"AVISOS: {len(self.avisos)}")
        print("-" * 70)
        if self.avisos:
            for a in self.avisos[:20]:
                print(f"  ⚠ [{a['categoria']}] {a['mensagem']}")
                if a['detalhe']:
                    print(f"    → {a['detalhe']}")
            if len(self.avisos) > 20:
                print(f"  ... e mais {len(self.avisos) - 20} avisos")
        else:
            print("  Nenhum aviso")

        # OK
        print()
        print("-" * 70)
        print(f"VERIFICAÇÕES OK: {len(self.ok)}")
        print("-" * 70)
        for o in self.ok:
            print(f"  ✓ [{o['categoria']}] {o['mensagem']}")

        # Campos preenchidos
        if 'campos_preenchidos' in self.estatisticas:
            print()
            print("-" * 70)
            print("CAMPOS OBRIGATÓRIOS - PREENCHIMENTO")
            print("-" * 70)
            total = self.estatisticas.get('total_espacos', 0)
            campos_obrig = {col: spec for col, spec in COLUNAS_ESPACOS.items() if spec.get('obrigatorio')}
            for col, spec in campos_obrig.items():
                preenchidos = self.estatisticas['campos_preenchidos'].get(col, 0)
                status = "✓" if preenchidos == total else "✗"
                print(f"  {status} Col {col:3d} ({spec['nome']}): {preenchidos}/{total}")

        # Conclusão
        print()
        print("=" * 70)
        if len(self.erros) == 0:
            print("✓ FICHEIRO VÁLIDO - Pronto para conversão!")
        else:
            print(f"✗ FICHEIRO COM ERROS - Corrigir {len(self.erros)} erro(s) antes de converter")
        print("=" * 70)

        return len(self.erros) == 0

    def gerar_relatorio_campos(self):
        """Gera relatório detalhado campo a campo."""
        print()
        print("=" * 70)
        print("RELATÓRIO DETALHADO - CAMPO A CAMPO")
        print("=" * 70)

        if 'campos_preenchidos' not in self.estatisticas:
            print("Sem dados de campos")
            return

        total = self.estatisticas.get('total_espacos', 0)

        # Agrupar por categoria
        categorias = {
            'GENERAL': range(1, 7),
            'PEOPLE': range(7, 12),
            'LIGHTING': range(12, 17),
            'EQUIPMENT': range(17, 19),
            'MISC': range(19, 23),
            'INFILTRATION': range(23, 27),
            'FLOORS': range(27, 40),
            'PARTITIONS CEILING': range(40, 46),
            'PARTITIONS WALL': range(46, 52),
            'WALLS': range(52, 124),
            'ROOFS': range(124, 148),
        }

        for cat_nome, cols in categorias.items():
            print()
            print(f"--- {cat_nome} ---")

            for col in cols:
                if col not in COLUNAS_ESPACOS:
                    continue

                spec = COLUNAS_ESPACOS[col]
                preenchidos = self.estatisticas['campos_preenchidos'].get(col, 0)
                erros = len(self.estatisticas['erros_por_coluna'].get(col, []))

                obrig = "OBRIG" if spec.get('obrigatorio') else "     "

                if erros > 0:
                    status = f"✗ {erros} erros"
                elif preenchidos == 0:
                    status = "- vazio"
                elif spec.get('obrigatorio') and preenchidos < total:
                    status = f"⚠ {preenchidos}/{total}"
                else:
                    status = f"✓ {preenchidos}/{total}"

                print(f"  Col {col:3d} [{obrig}] {spec['nome'][:30]:30s} {status}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nErro: Especifique o ficheiro Excel a validar")
        print("\nExemplo:")
        print("  python validar_excel_hap.py MeuProjeto.xlsx")
        sys.exit(1)

    filepath = sys.argv[1]

    if not os.path.exists(filepath):
        print(f"Erro: Ficheiro não encontrado: {filepath}")
        sys.exit(1)

    validador = ValidadorHAP(filepath)

    if validador.executar():
        validador.gerar_relatorio()

        # Perguntar se quer relatório detalhado
        print()
        resposta = input("Mostrar relatório detalhado campo a campo? (s/n): ")
        if resposta.lower() == 's':
            validador.gerar_relatorio_campos()

        sys.exit(0 if len(validador.erros) == 0 else 1)
    else:
        print("Erro ao executar validação")
        sys.exit(1)


if __name__ == '__main__':
    main()
