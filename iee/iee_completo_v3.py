"""
IEE Completo v3 - Folha de cálculo completa para certificação energética SCE

SEQUÊNCIA DE PREENCHIMENTO (ordem lógica):
==========================================
1. Detalhe PREV/REF    - Dados brutos HAP (automático dos CSV)
2. Mensal PREV/REF     - Dados mensais (automático dos CSV)
3. Simulação           - EER/COP dos equipamentos → Energia + Aerotermia
4. Iluminação ENU      - Potência × Horas
5. AQS                 - Pessoas, litros, sistema, solar térmico
6. PV                  - Potência instalada
7. Equipamentos Extra  - Equipamentos não simulados
8. Elevadores          - Carga, viagens, percurso
9. Ventilação Extra    - Sistemas não simulados
10. Bombagem           - Bombas não simuladas
11. Desagregação       - Resume tudo (FÓRMULAS automáticas)
12. Energia Primária   - Área útil → kWhEP (FÓRMULAS)
13. IEE                - Indicadores (FÓRMULAS)
14. Classe             - Resultado final (FÓRMULAS)

Usage:
    python iee_completo_v3.py <pasta_prev> <pasta_ref> [output.xlsx]
"""

import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# ESTILOS
# =============================================================================

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10)

SECTION_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
SECTION_FONT = Font(bold=True, size=10)

TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_FONT = Font(bold=True)

FORMULA_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RESULT_FILL = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
PREV_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
REF_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
REN_FILL = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)

CLASS_COLORS = {
    'A+': '00A651', 'A': '50B848', 'B': 'B5D334',
    'B-': 'FFF200', 'C': 'F7941D', 'D': 'F15A29',
    'E': 'ED1C24', 'F': 'BE1E2D'
}

# =============================================================================
# CONSTANTES SCE
# =============================================================================

FPU = {
    'electricidade': 2.5,
    'gas_natural': 1.0,
}

RIEE_LIMITES = [
    ('A+', 0, 0.25), ('A', 0.26, 0.50), ('B', 0.51, 0.75),
    ('B-', 0.76, 1.00), ('C', 1.01, 1.50), ('D', 1.51, 2.00),
    ('E', 2.01, 2.50), ('F', 2.51, 999),
]

MONTHS = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']


# =============================================================================
# FUNÇÕES LEITURA CSV
# =============================================================================

def read_hap_csv(filepath):
    """Lê um CSV do HAP e retorna dados"""
    with open(filepath, 'r', encoding='latin-1') as f:
        lines = f.readlines()

    if len(lines) < 5:
        return None, {}

    sistema = lines[1].split(';')[0].replace('Monthly Simulation Results for ', '').strip()
    headers = lines[3].strip().split(';')

    totals = {}
    monthly_data = []

    for line in lines[4:]:
        line = line.strip()
        if not line or line.startswith('Month'):
            continue

        values = line.split(';')
        month = values[0] if values else ''

        month_data = {'Month': month}
        for i, val in enumerate(values[1:], 1):
            if i < len(headers) and val:
                col = headers[i]
                try:
                    num_val = int(val)
                    totals[col] = totals.get(col, 0) + num_val
                    month_data[col] = num_val
                except:
                    pass

        if month:
            monthly_data.append(month_data)

    return sistema, {'totals': totals, 'monthly': monthly_data, 'headers': headers}


def load_project_data(project_folder):
    """Carrega todos os dados de um projecto HAP"""
    pattern = os.path.join(project_folder, 'HAP51_Monthly_*.csv')
    csv_files = sorted(glob.glob(pattern))
    systems_data = {}

    for csv_file in csv_files:
        sistema, data = read_hap_csv(csv_file)
        if sistema and data and sistema != 'TODOS':
            systems_data[sistema] = data

    return systems_data


# =============================================================================
# 1-2. FOLHAS DETALHE E MENSAL (Automáticas dos CSV)
# =============================================================================

def create_detalhe_sheet(ws, data, title):
    """Cria folha Detalhe com dados brutos anuais do HAP"""
    sistemas = list(data.keys())

    columns = [
        ('Sistema', 'A', 22),
        ('Lighting (kWh)', 'B', 12),
        ('Equipment (kWh)', 'C', 12),
        ('Central Clg Input', 'D', 14),
        ('Terminal Clg Input', 'E', 14),
        ('Central Htg Input', 'F', 14),
        ('Terminal Htg Input', 'G', 14),
        ('Central Aux Htg', 'H', 12),
        ('Terminal Aux Htg', 'I', 12),
        ('Supply Fan', 'J', 10),
        ('Return Fan', 'K', 10),
        ('Exhaust Fan', 'L', 10),
        ('Vent Fan', 'M', 10),
        ('Clg Coil Load', 'N', 12),
        ('Htg Coil Load', 'O', 12),
    ]

    csv_map = {
        'B': 'Lighting (kWh)',
        'C': 'Electric Equipment (kWh)',
        'D': 'Central Unit Clg Input (kWh)',
        'E': 'Terminal Unit Clg Input (kWh)',
        'F': 'Central Unit Htg Input (kWh)',
        'G': 'Terminal Unit Htg Input (kWh)',
        'H': 'Central Unit Aux. Htg. Input (kWh)',
        'I': 'Terminal Unit Aux. Htg. Input (kWh)',
        'J': 'Supply Fan (kWh)',
        'K': 'Return Fan (kWh)',
        'L': 'Exhaust Fan (kWh)',
        'M': 'Ventilation Fan (kWh)',
        'N': 'Central Cooling Coil Load (kWh)',
        'O': 'Central Heating Coil Load (kWh)',
    }

    # Título
    ws.merge_cells('A1:O1')
    ws.cell(1, 1, value=f'DADOS {title} - Consumos Anuais por Sistema (do HAP)').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 1: Estes dados são extraídos automaticamente dos CSV do HAP').font = Font(italic=True, color='666666')

    # Headers
    for nome, col, width in columns:
        col_num = ord(col) - ord('A') + 1
        cell = ws.cell(3, col_num, value=nome)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[col].width = width

    # Dados
    for i, sistema in enumerate(sistemas):
        row = 4 + i
        ws.cell(row, 1, value=sistema).border = THIN_BORDER

        for col_letter, csv_col in csv_map.items():
            col_num = ord(col_letter) - ord('A') + 1
            value = data[sistema]['totals'].get(csv_col, 0)
            cell = ws.cell(row, col_num, value=value if value > 0 else '')
            cell.border = THIN_BORDER

    # Total
    row_total = 4 + len(sistemas)
    ws.cell(row_total, 1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row_total, 1).fill = TOTAL_FILL

    for col_letter in csv_map.keys():
        col_num = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row_total, col_num, value=f'=SUM({col_letter}4:{col_letter}{row_total-1})')
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    ws.freeze_panes = 'B4'
    return row_total


def create_mensal_sheet(ws, data, title):
    """Cria folha com dados mensais do HAP - TODOS os tipos de consumo + RESUMO"""
    sistemas = list(data.keys())

    ws.merge_cells('A1:O1')
    ws.cell(1, 1, value=f'DADOS {title} - Consumos Mensais (do HAP)').font = Font(bold=True, size=14)

    # Headers
    headers = ['Sistema', 'Tipo'] + MONTHS + ['Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(2, col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # TODOS os tipos de consumo do HAP
    cols_show = [
        ('Lighting (kWh)', 'Iluminação'),
        ('Electric Equipment (kWh)', 'Equipamentos'),
        ('Central Unit Clg Input (kWh)', 'Arref. Central'),
        ('Terminal Unit Clg Input (kWh)', 'Arref. Terminal'),
        ('Central Unit Htg Input (kWh)', 'Aquec. Central'),
        ('Terminal Unit Htg Input (kWh)', 'Aquec. Terminal'),
        ('Central Unit Aux. Htg. Input (kWh)', 'Aux Htg Central'),
        ('Terminal Unit Aux. Htg. Input (kWh)', 'Aux Htg Terminal'),
        ('Supply Fan (kWh)', 'Vent. Insuflação'),
        ('Return Fan (kWh)', 'Vent. Retorno'),
        ('Exhaust Fan (kWh)', 'Vent. Extração'),
        ('Ventilation Fan (kWh)', 'Vent. Mecânica'),
        ('Central Cooling Coil Load (kWh)', 'Carga Arref.'),
        ('Central Heating Coil Load (kWh)', 'Carga Aquec.'),
    ]

    # Guardar linhas por tipo e por sistema para os resumos
    rows_by_type = {nome: [] for _, nome in cols_show}
    rows_by_sistema = {sistema: [] for sistema in sistemas}

    row = 3
    for sistema in sistemas:
        monthly = data[sistema].get('monthly', [])
        if not monthly:
            continue

        for csv_col, nome in cols_show:
            if data[sistema]['totals'].get(csv_col, 0) == 0:
                continue

            ws.cell(row, 1, value=sistema).border = THIN_BORDER
            ws.cell(row, 2, value=nome).border = THIN_BORDER

            for col, month_data in enumerate(monthly[:12], 3):
                val = month_data.get(csv_col, 0)
                ws.cell(row, col, value=val if val > 0 else '').border = THIN_BORDER

            ws.cell(row, 15, value=f'=SUM(C{row}:N{row})').border = THIN_BORDER
            ws.cell(row, 15).fill = FORMULA_FILL

            rows_by_type[nome].append(row)
            rows_by_sistema[sistema].append(row)
            row += 1

    last_detail_row = row - 1

    # Larguras da parte detalhe
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 7

    # =========================================================================
    # RESUMO À DIREITA (coluna R em diante)
    # =========================================================================
    col_resumo_start = 18  # Coluna R

    ws.cell(1, col_resumo_start, value='RESUMO POR TIPO DE CONSUMO').font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=col_resumo_start, end_row=1, end_column=col_resumo_start + 13)

    # Headers do resumo
    headers_resumo = ['Tipo'] + MONTHS + ['TOTAL']
    for col, h in enumerate(headers_resumo):
        cell = ws.cell(2, col_resumo_start + col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Dados do resumo - soma por tipo
    row_resumo = 3
    resumo_rows = {}

    for csv_col, nome in cols_show:
        rows_list = rows_by_type[nome]
        if not rows_list:
            continue

        ws.cell(row_resumo, col_resumo_start, value=nome).border = THIN_BORDER
        ws.cell(row_resumo, col_resumo_start).font = Font(bold=True)

        # Para cada mês (colunas C a N = 3 a 14)
        for month_idx in range(12):
            col_month = col_resumo_start + 1 + month_idx
            detail_col = get_column_letter(3 + month_idx)

            # Soma das linhas deste tipo
            if len(rows_list) == 1:
                formula = f'={detail_col}{rows_list[0]}'
            else:
                sum_parts = '+'.join([f'{detail_col}{r}' for r in rows_list])
                formula = f'={sum_parts}'

            ws.cell(row_resumo, col_month, value=formula).border = THIN_BORDER

        # Total anual
        col_total = col_resumo_start + 13
        start_col_letter = get_column_letter(col_resumo_start + 1)
        end_col_letter = get_column_letter(col_resumo_start + 12)
        ws.cell(row_resumo, col_total, value=f'=SUM({start_col_letter}{row_resumo}:{end_col_letter}{row_resumo})')
        ws.cell(row_resumo, col_total).fill = FORMULA_FILL
        ws.cell(row_resumo, col_total).border = THIN_BORDER
        ws.cell(row_resumo, col_total).font = Font(bold=True)

        resumo_rows[nome] = row_resumo
        row_resumo += 1

    # Linha TOTAL GERAL
    row_total = row_resumo
    ws.cell(row_total, col_resumo_start, value='TOTAL GERAL').font = TOTAL_FONT
    ws.cell(row_total, col_resumo_start).fill = TOTAL_FILL
    ws.cell(row_total, col_resumo_start).border = THIN_BORDER

    for month_idx in range(12):
        col_month = col_resumo_start + 1 + month_idx
        col_letter = get_column_letter(col_month)
        ws.cell(row_total, col_month, value=f'=SUM({col_letter}3:{col_letter}{row_total-1})')
        ws.cell(row_total, col_month).fill = TOTAL_FILL
        ws.cell(row_total, col_month).border = THIN_BORDER
        ws.cell(row_total, col_month).font = TOTAL_FONT

    col_total = col_resumo_start + 13
    start_col_letter = get_column_letter(col_resumo_start + 1)
    end_col_letter = get_column_letter(col_resumo_start + 12)
    ws.cell(row_total, col_total, value=f'=SUM({start_col_letter}{row_total}:{end_col_letter}{row_total})')
    ws.cell(row_total, col_total).fill = TOTAL_FILL
    ws.cell(row_total, col_total).border = THICK_BORDER
    ws.cell(row_total, col_total).font = TOTAL_FONT

    # Larguras do resumo por tipo
    ws.column_dimensions[get_column_letter(col_resumo_start)].width = 16
    for col in range(col_resumo_start + 1, col_resumo_start + 14):
        ws.column_dimensions[get_column_letter(col)].width = 8

    # =========================================================================
    # RESUMO POR SISTEMA (abaixo do resumo por tipo)
    # =========================================================================
    row_sistema_start = row_total + 3  # 2 linhas de espaço

    ws.cell(row_sistema_start, col_resumo_start, value='RESUMO POR SISTEMA').font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row_sistema_start, start_column=col_resumo_start, end_row=row_sistema_start, end_column=col_resumo_start + 13)

    # Headers do resumo por sistema
    row_sistema_start += 1
    for col, h in enumerate(headers_resumo):
        cell = ws.cell(row_sistema_start, col_resumo_start + col, value=h if col == 0 else h)
        if col == 0:
            cell.value = 'Sistema'
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Dados do resumo por sistema
    row_sistema = row_sistema_start + 1
    row_sistema_first = row_sistema

    for sistema in sistemas:
        rows_list = rows_by_sistema[sistema]
        if not rows_list:
            continue

        ws.cell(row_sistema, col_resumo_start, value=sistema).border = THIN_BORDER
        ws.cell(row_sistema, col_resumo_start).font = Font(bold=True)

        # Para cada mês (colunas C a N = 3 a 14)
        for month_idx in range(12):
            col_month = col_resumo_start + 1 + month_idx
            detail_col = get_column_letter(3 + month_idx)

            # Soma das linhas deste sistema
            if len(rows_list) == 1:
                formula = f'={detail_col}{rows_list[0]}'
            else:
                sum_parts = '+'.join([f'{detail_col}{r}' for r in rows_list])
                formula = f'={sum_parts}'

            ws.cell(row_sistema, col_month, value=formula).border = THIN_BORDER

        # Total anual
        col_total = col_resumo_start + 13
        start_col_letter = get_column_letter(col_resumo_start + 1)
        end_col_letter = get_column_letter(col_resumo_start + 12)
        ws.cell(row_sistema, col_total, value=f'=SUM({start_col_letter}{row_sistema}:{end_col_letter}{row_sistema})')
        ws.cell(row_sistema, col_total).fill = FORMULA_FILL
        ws.cell(row_sistema, col_total).border = THIN_BORDER
        ws.cell(row_sistema, col_total).font = Font(bold=True)

        row_sistema += 1

    # Linha TOTAL GERAL por sistema
    ws.cell(row_sistema, col_resumo_start, value='TOTAL GERAL').font = TOTAL_FONT
    ws.cell(row_sistema, col_resumo_start).fill = TOTAL_FILL
    ws.cell(row_sistema, col_resumo_start).border = THIN_BORDER

    for month_idx in range(12):
        col_month = col_resumo_start + 1 + month_idx
        col_letter = get_column_letter(col_month)
        ws.cell(row_sistema, col_month, value=f'=SUM({col_letter}{row_sistema_first}:{col_letter}{row_sistema-1})')
        ws.cell(row_sistema, col_month).fill = TOTAL_FILL
        ws.cell(row_sistema, col_month).border = THIN_BORDER
        ws.cell(row_sistema, col_month).font = TOTAL_FONT

    col_total = col_resumo_start + 13
    start_col_letter = get_column_letter(col_resumo_start + 1)
    end_col_letter = get_column_letter(col_resumo_start + 12)
    ws.cell(row_sistema, col_total, value=f'=SUM({start_col_letter}{row_sistema}:{end_col_letter}{row_sistema})')
    ws.cell(row_sistema, col_total).fill = TOTAL_FILL
    ws.cell(row_sistema, col_total).border = THICK_BORDER
    ws.cell(row_sistema, col_total).font = TOTAL_FONT


# =============================================================================
# 3. FOLHA SIMULAÇÃO (EER/COP → Energia + Aerotermia)
# =============================================================================

def create_simulacao_sheet(ws, data, row_total_detalhe):
    """Cria folha Simulação com cálculo de EER/COP e Aerotermia"""
    sistemas = list(data.keys())
    num_sistemas = len(sistemas)

    ws.merge_cells('A1:P1')
    ws.cell(1, 1, value='SIMULAÇÃO - Cálculo Energia Final e Aerotermia').font = Font(bold=True, size=14)

    ws.cell(2, 1, value='Passo 3: Preencher EER e COP de cada sistema. A energia e aerotermia são calculadas automaticamente.').font = Font(italic=True, color='666666')

    # -------------------------------------------------------------------------
    # Secção 1: NECESSIDADES (do HAP)
    # -------------------------------------------------------------------------
    ws.cell(4, 1, value='NECESSIDADES TÉRMICAS (do HAP)').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL
    ws.merge_cells('A4:F4')

    headers_nec = ['Sistema', 'Cool (kWh)', 'Heat (kWh)', 'Heat+5%', 'Iluminação', 'Equipamentos']
    for col, h in enumerate(headers_nec, 1):
        cell = ws.cell(5, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    row = 6
    for i, sistema in enumerate(sistemas):
        row_det = 4 + i  # linha no Detalhe

        ws.cell(row, 1, value=sistema).border = THIN_BORDER
        # Cool = Central Clg Load (N) + Terminal Clg Load
        ws.cell(row, 2, value=f"=DetalhePREV!N{row_det}").border = THIN_BORDER
        ws.cell(row, 2).fill = FORMULA_FILL
        # Heat = Central Htg Load (O)
        ws.cell(row, 3, value=f"=DetalhePREV!O{row_det}").border = THIN_BORDER
        ws.cell(row, 3).fill = FORMULA_FILL
        # Heat + 5%
        ws.cell(row, 4, value=f'=C{row}*1.05').border = THIN_BORDER
        ws.cell(row, 4).fill = FORMULA_FILL
        # Iluminação
        ws.cell(row, 5, value=f"=DetalhePREV!B{row_det}").border = THIN_BORDER
        ws.cell(row, 5).fill = FORMULA_FILL
        # Equipamentos
        ws.cell(row, 6, value=f"=DetalhePREV!C{row_det}").border = THIN_BORDER
        ws.cell(row, 6).fill = FORMULA_FILL
        row += 1

    row_total_nec = row
    ws.cell(row, 1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        ws.cell(row, col, value=f'=SUM({col_letter}6:{col_letter}{row-1})')
        ws.cell(row, col).font = TOTAL_FONT
        ws.cell(row, col).fill = TOTAL_FILL
        ws.cell(row, col).border = THIN_BORDER

    # -------------------------------------------------------------------------
    # Secção 2: CÁLCULO COM EER/COP
    # -------------------------------------------------------------------------
    row += 2
    row_calc_start = row
    ws.cell(row, 1, value='CÁLCULO ENERGIA FINAL (com EER/COP)').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:L{row}')
    row += 1

    headers_calc = ['Sistema', 'Nec Cool', 'Nec Heat', 'EER', 'COP',
                    'Energia Cool', 'Energia Heat', 'Aero Cool', 'Aero Heat']
    for col, h in enumerate(headers_calc, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        if h in ['EER', 'COP']:
            cell.fill = INPUT_FILL
            cell.font = Font(bold=True, color='000000')  # Preto para ser visível no amarelo
    row += 1

    row_calc_data_start = row
    for i, sistema in enumerate(sistemas):
        row_nec = 6 + i  # linha nas necessidades

        ws.cell(row, 1, value=sistema).border = THIN_BORDER
        # Nec Cool (referência)
        ws.cell(row, 2, value=f'=B{row_nec}').border = THIN_BORDER
        ws.cell(row, 2).fill = FORMULA_FILL
        # Nec Heat (referência)
        ws.cell(row, 3, value=f'=D{row_nec}').border = THIN_BORDER  # Heat+5%
        ws.cell(row, 3).fill = FORMULA_FILL
        # EER (INPUT) - fonte preta para ser visível no fundo amarelo
        ws.cell(row, 4, value=3.0)
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 4).font = Font(color='000000')
        # COP (INPUT) - fonte preta para ser visível no fundo amarelo
        ws.cell(row, 5, value=3.0)
        ws.cell(row, 5).fill = INPUT_FILL
        ws.cell(row, 5).border = THIN_BORDER
        ws.cell(row, 5).font = Font(color='000000')
        # Energia Cool = Nec / EER
        ws.cell(row, 6, value=f'=IF(D{row}>0,B{row}/D{row},0)')
        ws.cell(row, 6).fill = FORMULA_FILL
        ws.cell(row, 6).border = THIN_BORDER
        ws.cell(row, 6).number_format = '0'
        # Energia Heat = Nec / COP
        ws.cell(row, 7, value=f'=IF(E{row}>0,C{row}/E{row},0)')
        ws.cell(row, 7).fill = FORMULA_FILL
        ws.cell(row, 7).border = THIN_BORDER
        ws.cell(row, 7).number_format = '0'
        # Aerotermia Cool = Nec - Energia
        ws.cell(row, 8, value=f'=B{row}-F{row}')
        ws.cell(row, 8).fill = REN_FILL
        ws.cell(row, 8).border = THIN_BORDER
        ws.cell(row, 8).number_format = '0'
        # Aerotermia Heat = Nec - Energia
        ws.cell(row, 9, value=f'=C{row}-G{row}')
        ws.cell(row, 9).fill = REN_FILL
        ws.cell(row, 9).border = THIN_BORDER
        ws.cell(row, 9).number_format = '0'
        row += 1

    row_total_calc = row
    ws.cell(row, 1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    for col in range(2, 10):
        if col in [4, 5]:  # EER/COP não soma
            continue
        col_letter = get_column_letter(col)
        ws.cell(row, col, value=f'=SUM({col_letter}{row_calc_data_start}:{col_letter}{row-1})')
        ws.cell(row, col).font = TOTAL_FONT
        ws.cell(row, col).fill = TOTAL_FILL
        ws.cell(row, col).border = THIN_BORDER

    # -------------------------------------------------------------------------
    # Secção 3: RESUMO PARA DESAGREGAÇÃO
    # -------------------------------------------------------------------------
    row += 2
    ws.cell(row, 1, value='RESUMO PARA DESAGREGAÇÃO').font = Font(bold=True, size=12)
    row += 1

    resumo = [
        ('Iluminação (Simulação)', f'=E{row_total_nec}', 'EE', 'S'),
        ('Equipamentos (Simulação)', f'=F{row_total_nec}', 'EE', 'T'),
        ('Arrefecimento EE', f'=F{row_total_calc}', 'EE', 'S'),
        ('Aquecimento EE', f'=G{row_total_calc}', 'EE', 'S'),
        ('Arrefecimento REN (Aerotermia)', f'=H{row_total_calc}', 'REN', 'REN'),
        ('Aquecimento REN (Aerotermia)', f'=I{row_total_calc}', 'REN', 'REN'),
        ('Ventilação (Simulação)', f"=DetalhePREV!J{row_total_detalhe}+'DetalhePREV'!K{row_total_detalhe}", 'EE', 'S'),
    ]

    row_resumo_start = row
    for nome, formula, fonte, tipo in resumo:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=formula)
        ws.cell(row, 2).fill = RESULT_FILL
        ws.cell(row, 2).border = THICK_BORDER
        ws.cell(row, 2).number_format = '0'
        ws.cell(row, 3, value='kWh/ano')
        ws.cell(row, 4, value=fonte)
        ws.cell(row, 5, value=tipo)
        row += 1

    # Larguras
    ws.column_dimensions['A'].width = 25
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return row_resumo_start, row_total_calc


# =============================================================================
# 4-10. FOLHAS AUXILIARES DE INPUT
# =============================================================================

def create_iluminacao_enu_sheet(ws):
    """Cria folha de Iluminação ENU e Exterior"""
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='ILUMINAÇÃO ENU E EXTERIOR').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 4: Preencher potência e horas de funcionamento').font = Font(italic=True, color='666666')

    # ENU
    ws.cell(4, 1, value='ILUMINAÇÃO ENU (Tipo B)').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL

    headers = ['Espaço', 'Potência [W]', 'Horas/ano', 'Energia [kWh]']
    for col, h in enumerate(headers, 1):
        ws.cell(5, col, value=h).font = SUBHEADER_FONT
        ws.cell(5, col).fill = SUBHEADER_FILL
        ws.cell(5, col).border = THIN_BORDER

    espacos = ['Armazéns', 'Zona Técnica', 'Cozinha ENU', 'Garagem', 'Outros ENU']
    row = 6
    for esp in espacos:
        ws.cell(row, 1, value=esp).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}/1000').fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        row += 1

    row_total_enu = row
    ws.cell(row, 1, value='TOTAL ENU').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D6:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL

    # Exterior
    row += 2
    ws.cell(row, 1, value='ILUMINAÇÃO EXTERIOR').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1

    for col, h in enumerate(headers, 1):
        ws.cell(row, col, value=h).font = SUBHEADER_FONT
        ws.cell(row, col).fill = SUBHEADER_FILL
        ws.cell(row, col).border = THIN_BORDER
    row += 1

    espacos_ext = ['Estacionamento', 'Jardins', 'Fachada', 'Outros']
    row_start_ext = row
    for esp in espacos_ext:
        ws.cell(row, 1, value=esp).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}/1000').fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        row += 1

    row_total_ext = row
    ws.cell(row, 1, value='TOTAL EXTERIOR').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start_ext}:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14

    return row_total_enu, row_total_ext


def create_aqs_sheet(ws):
    """Cria folha AQS"""
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='AQS - Águas Quentes Sanitárias').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 5: Preencher dados de ocupação e sistema').font = Font(italic=True, color='666666')

    # Parâmetros
    ws.cell(4, 1, value='PARÂMETROS').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL

    params = [
        ('Número de pessoas', 5, '', 'pessoas'),
        ('Consumo litros/pessoa/dia', 6, 40, 'L/p/dia'),
        ('Dias por ano', 7, 365, 'dias'),
        ('Temp. água fria', 8, 15, '°C'),
        ('Temp. AQS', 9, 60, '°C'),
    ]

    for nome, row, default, unit in params:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=default if default else None)
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=unit)

    # Cálculo necessidades
    ws.cell(11, 1, value='CÁLCULO NECESSIDADES').font = SECTION_FONT
    ws.cell(11, 1).fill = SECTION_FILL

    ws.cell(12, 1, value='Volume diário [L]')
    ws.cell(12, 2, value='=B5*B6').fill = FORMULA_FILL

    ws.cell(13, 1, value='Qútil [kWh/ano]')
    ws.cell(13, 2, value='=B5*B6*B7*4.186*(B9-B8)/3600').fill = FORMULA_FILL
    ws.cell(13, 2).number_format = '0'
    row_qutil = 13

    # Solar térmico
    ws.cell(15, 1, value='SOLAR TÉRMICO').font = SECTION_FONT
    ws.cell(15, 1).fill = REN_FILL

    ws.cell(16, 1, value='Contribuição ST [kWh/ano]')
    ws.cell(16, 2).fill = INPUT_FILL
    ws.cell(16, 2).border = THICK_BORDER
    ws.cell(16, 3, value='← Do SCE.ER')
    row_st = 16

    ws.cell(17, 1, value='Necessidades após ST')
    ws.cell(17, 2, value=f'=B{row_qutil}-B{row_st}').fill = FORMULA_FILL

    # Sistema apoio
    ws.cell(19, 1, value='SISTEMA DE APOIO').font = SECTION_FONT
    ws.cell(19, 1).fill = SECTION_FILL

    ws.cell(20, 1, value='1 - Bomba de Calor (COP)')
    ws.cell(20, 2, value=3.5).fill = INPUT_FILL
    ws.cell(20, 3, value='Fonte: EE')

    ws.cell(21, 1, value='2 - Caldeira Gás (η)')
    ws.cell(21, 2, value=0.92).fill = INPUT_FILL
    ws.cell(21, 3, value='Fonte: GN')

    ws.cell(22, 1, value='3 - Resistência (η)')
    ws.cell(22, 2, value=1.0).fill = INPUT_FILL
    ws.cell(22, 3, value='Fonte: EE')

    ws.cell(24, 1, value='Sistema escolhido (1, 2 ou 3):')
    ws.cell(24, 2, value=1).fill = INPUT_FILL
    ws.cell(24, 2).border = THICK_BORDER
    row_sel = 24

    # Resultados
    ws.cell(26, 1, value='RESULTADOS').font = SECTION_FONT
    ws.cell(26, 1).fill = RESULT_FILL

    ws.cell(27, 1, value='Rendimento')
    ws.cell(27, 2, value=f'=INDEX(B20:B22,B{row_sel})').fill = FORMULA_FILL

    ws.cell(28, 1, value='Energia Final AQS')
    ws.cell(28, 2, value='=ROUND(B17/B27,0)')
    ws.cell(28, 2).fill = RESULT_FILL
    ws.cell(28, 2).border = THICK_BORDER
    ws.cell(28, 3, value='kWh/ano')
    row_ef = 28

    ws.cell(29, 1, value='Aerotermia (se BC)')
    ws.cell(29, 2, value=f'=IF(B{row_sel}=1,B17-B{row_ef},0)')
    ws.cell(29, 2).fill = REN_FILL
    ws.cell(29, 2).border = THICK_BORDER
    ws.cell(29, 3, value='kWh/ano')

    # Resumo
    ws.cell(31, 1, value='RESUMO PARA DESAGREGAÇÃO').font = Font(bold=True)
    ws.cell(32, 1, value='AQS - EE')
    ws.cell(32, 2, value=f'=IF(OR(B{row_sel}=1,B{row_sel}=3),B{row_ef},0)')
    ws.cell(32, 2).fill = RESULT_FILL
    ws.cell(32, 2).border = THICK_BORDER

    ws.cell(33, 1, value='AQS - GN')
    ws.cell(33, 2, value=f'=IF(B{row_sel}=2,B{row_ef},0)')
    ws.cell(33, 2).fill = RESULT_FILL
    ws.cell(33, 2).border = THICK_BORDER

    ws.cell(34, 1, value='AQS - REN Aero')
    ws.cell(34, 2, value='=B29')
    ws.cell(34, 2).fill = REN_FILL
    ws.cell(34, 2).border = THICK_BORDER

    ws.cell(35, 1, value='AQS - REN Solar')
    ws.cell(35, 2, value=f'=B{row_st}')
    ws.cell(35, 2).fill = REN_FILL
    ws.cell(35, 2).border = THICK_BORDER

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12


def create_pv_sheet(ws):
    """Cria folha PV"""
    ws.merge_cells('A1:D1')
    ws.cell(1, 1, value='PV - Produção Fotovoltaica').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 6: Preencher potência instalada').font = Font(italic=True, color='666666')

    ws.cell(4, 1, value='PARÂMETROS').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL

    ws.cell(5, 1, value='Potência instalada')
    ws.cell(5, 2).fill = INPUT_FILL
    ws.cell(5, 2).border = THICK_BORDER
    ws.cell(5, 3, value='kWp')

    ws.cell(6, 1, value='Horas equivalentes')
    ws.cell(6, 2, value=1400).fill = INPUT_FILL
    ws.cell(6, 3, value='h/ano')

    ws.cell(7, 1, value='Perdas sistema')
    ws.cell(7, 2, value=15).fill = INPUT_FILL
    ws.cell(7, 3, value='%')

    ws.cell(8, 1, value='Factor auto-consumo')
    ws.cell(8, 2, value=0.8).fill = INPUT_FILL

    ws.cell(10, 1, value='RESULTADOS').font = SECTION_FONT
    ws.cell(10, 1).fill = RESULT_FILL

    ws.cell(11, 1, value='Produção bruta')
    ws.cell(11, 2, value='=B5*B6').fill = FORMULA_FILL
    ws.cell(11, 3, value='kWh/ano')

    ws.cell(12, 1, value='Produção líquida')
    ws.cell(12, 2, value='=B5*B6*(1-B7/100)').fill = FORMULA_FILL
    ws.cell(12, 3, value='kWh/ano')

    ws.cell(13, 1, value='Auto-consumo')
    ws.cell(13, 2, value='=B12*B8')
    ws.cell(13, 2).fill = REN_FILL
    ws.cell(13, 2).border = THICK_BORDER
    ws.cell(13, 3, value='kWh/ano')

    ws.cell(15, 1, value='PV para IEE (auto-consumo)')
    ws.cell(15, 2, value='=B13')
    ws.cell(15, 2).fill = RESULT_FILL
    ws.cell(15, 2).border = THICK_BORDER

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def create_equipamentos_extra_sheet(ws):
    """Cria folha Equipamentos Extra"""
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='EQUIPAMENTOS EXTRA - Não simulados').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 7: Preencher equipamentos não incluídos no HAP').font = Font(italic=True, color='666666')

    # EE
    ws.cell(4, 1, value='EQUIPAMENTOS EE (Tipo T)').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL

    headers = ['Equipamento', 'Potência [W]', 'Horas/ano', 'Qtd', 'Energia [kWh]']
    for col, h in enumerate(headers, 1):
        ws.cell(5, col, value=h).font = SUBHEADER_FONT
        ws.cell(5, col).fill = SUBHEADER_FILL
        ws.cell(5, col).border = THIN_BORDER

    eqs = ['Cozinha Industrial', 'Frio Comercial', 'Servidores', 'Outros EE']
    row = 6
    for eq in eqs:
        ws.cell(row, 1, value=eq).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f'=B{row}*C{row}*D{row}/1000').fill = FORMULA_FILL
        ws.cell(row, 5).border = THIN_BORDER
        row += 1

    row_total_ee = row
    ws.cell(row, 1, value='TOTAL EE Extra').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 5, value=f'=SUM(E6:E{row-1})').font = TOTAL_FONT
    ws.cell(row, 5).fill = TOTAL_FILL

    # GN
    row += 2
    ws.cell(row, 1, value='EQUIPAMENTOS GN (Tipo T)').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    row += 1

    ws.cell(row, 1, value='Equipamento').font = SUBHEADER_FONT
    ws.cell(row, 1).fill = SUBHEADER_FILL
    ws.cell(row, 2, value='Consumo [kWh/ano]').font = SUBHEADER_FONT
    ws.cell(row, 2).fill = SUBHEADER_FILL
    row += 1

    row_start_gn = row
    for eq in ['Cozinha (fogões)', 'Outros GN']:
        ws.cell(row, 1, value=eq).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        row += 1

    row_total_gn = row
    ws.cell(row, 1, value='TOTAL GN Extra').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 2, value=f'=SUM(B{row_start_gn}:B{row-1})').font = TOTAL_FONT
    ws.cell(row, 2).fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14

    return row_total_ee, row_total_gn


def create_elevadores_sheet(ws):
    """Cria folha Elevadores"""
    ws.merge_cells('A1:D1')
    ws.cell(1, 1, value='ELEVADORES - Cálculo SCE/RECS').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 8: Preencher dados dos elevadores').font = Font(italic=True, color='666666')
    ws.cell(3, 1, value='Fórmula: Edasc = (1.08 × Qasc × na × sm)/1000 + 100×21').font = Font(italic=True, size=9)

    ws.cell(5, 1, value='ELEVADOR 1').font = SECTION_FONT
    ws.cell(5, 1).fill = SECTION_FILL

    params = [
        ('Qasc - Carga nominal [kg]', 6, ''),
        ('na - Viagens/ano', 7, 750),
        ('p - Paragens/viagem', 8, 1),
        ('lm - Percurso médio [m]', 9, ''),
        ('sm = p × lm [m]', 10, '=B8*B9'),
        ('Edasc [Wh/dia]', 11, '=1.08*B6*B7*B10/1000+100*21'),
        ('Dias activos/ano', 12, 365),
        ('Energia anual [kWh]', 13, '=B11*B12/1000'),
    ]

    for nome, row, val in params:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        if isinstance(val, str) and val.startswith('='):
            ws.cell(row, 2, value=val).fill = FORMULA_FILL
        elif val:
            ws.cell(row, 2, value=val).fill = INPUT_FILL
        else:
            ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER

    ws.cell(15, 1, value='ELEVADOR 2').font = SECTION_FONT
    ws.cell(15, 1).fill = SECTION_FILL

    params2 = [
        ('Qasc - Carga nominal [kg]', 16, ''),
        ('na - Viagens/ano', 17, 750),
        ('p - Paragens/viagem', 18, 0.67),
        ('lm - Percurso médio [m]', 19, ''),
        ('sm = p × lm [m]', 20, '=B18*B19'),
        ('Edasc [Wh/dia]', 21, '=1.08*B16*B17*B20/1000+100*21'),
        ('Dias activos/ano', 22, 365),
        ('Energia anual [kWh]', 23, '=B21*B22/1000'),
    ]

    for nome, row, val in params2:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        if isinstance(val, str) and val.startswith('='):
            ws.cell(row, 2, value=val).fill = FORMULA_FILL
        elif val:
            ws.cell(row, 2, value=val).fill = INPUT_FILL
        else:
            ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER

    ws.cell(25, 1, value='TOTAL ELEVADORES')
    ws.cell(25, 2, value='=B13+B23')
    ws.cell(25, 2).fill = RESULT_FILL
    ws.cell(25, 2).border = THICK_BORDER
    ws.cell(25, 3, value='kWh/ano')

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15


def create_ventilacao_extra_sheet(ws):
    """Cria folha Ventilação Extra"""
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='VENTILAÇÃO EXTRA - Tipo S não simulada').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 9: Preencher ventilação não incluída no HAP').font = Font(italic=True, color='666666')

    ws.cell(4, 1, value='VENTILAÇÃO TIPO S').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL

    headers = ['Sistema', 'Potência [kW]', 'Horas/dia', 'Dias/ano', 'Energia [kWh]']
    for col, h in enumerate(headers, 1):
        ws.cell(5, col, value=h).font = SUBHEADER_FONT
        ws.cell(5, col).fill = SUBHEADER_FILL
        ws.cell(5, col).border = THIN_BORDER

    sistemas = ['Insuflação', 'Extração WC', 'Extração Cozinha', 'Outros']
    row = 6
    for sist in sistemas:
        ws.cell(row, 1, value=sist).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f'=B{row}*C{row}*D{row}').fill = FORMULA_FILL
        ws.cell(row, 5).border = THIN_BORDER
        row += 1

    row_total = row
    ws.cell(row, 1, value='TOTAL Ventilação Extra').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 5, value=f'=SUM(E6:E{row-1})').font = TOTAL_FONT
    ws.cell(row, 5).fill = TOTAL_FILL
    ws.cell(row, 5).border = THICK_BORDER

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    return row_total


def create_bombagem_sheet(ws):
    """Cria folha Bombagem"""
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='BOMBAGEM - Tipo S não simulada').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 10: Preencher bombas não incluídas no HAP').font = Font(italic=True, color='666666')

    ws.cell(4, 1, value='BOMBAS TIPO S').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL

    headers = ['Sistema', 'Potência [kW]', 'Horas/dia', 'Dias/ano', 'Energia [kWh]']
    for col, h in enumerate(headers, 1):
        ws.cell(5, col, value=h).font = SUBHEADER_FONT
        ws.cell(5, col).fill = SUBHEADER_FILL
        ws.cell(5, col).border = THIN_BORDER

    bombas = ['Bombas AQS', 'Bombas Solar', 'Bombas AVAC', 'Outras']
    row = 6
    for bomba in bombas:
        ws.cell(row, 1, value=bomba).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f'=B{row}*C{row}*D{row}').fill = FORMULA_FILL
        ws.cell(row, 5).border = THIN_BORDER
        row += 1

    row_total = row
    ws.cell(row, 1, value='TOTAL Bombagem').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 5, value=f'=SUM(E6:E{row-1})').font = TOTAL_FONT
    ws.cell(row, 5).fill = TOTAL_FILL
    ws.cell(row, 5).border = THICK_BORDER

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    return row_total


# =============================================================================
# 11. DESAGREGAÇÃO CONSUMOS
# =============================================================================

def create_desagregacao_sheet(ws, row_sim_resumo):
    """Cria folha Desagregação - resume todos os consumos"""
    ws.merge_cells('A1:F1')
    ws.cell(1, 1, value='DESAGREGAÇÃO CONSUMOS - Resumo Completo').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 11: Esta folha é calculada automaticamente a partir das anteriores').font = Font(italic=True, color='666666')

    headers = ['Utilização', 'Simulacao', 'Extra', 'TOTAL', 'Fonte', 'Tipo']
    for col, h in enumerate(headers, 1):
        ws.cell(4, col, value=h).font = HEADER_FONT
        ws.cell(4, col).fill = HEADER_FILL
        ws.cell(4, col).border = THIN_BORDER

    # Electricidade
    ws.cell(5, 1, value='ELECTRICIDADE').font = SECTION_FONT
    ws.cell(5, 1).fill = SECTION_FILL
    ws.merge_cells('A5:F5')

    consumos_ee = [
        ('Iluminação (Simulação)', f"=Simulacao!B{row_sim_resumo}", '', 'S'),
        ('IluminacaoENU', '', "=IluminacaoENU!D11", 'S'),
        ('Iluminação Exterior', '', "=IluminacaoENU!D19", 'S'),
        ('Arrefecimento', f"=Simulacao!B{row_sim_resumo+2}", '', 'S'),
        ('Aquecimento', f"=Simulacao!B{row_sim_resumo+3}", '', 'S'),
        ('Ventilação (Simulação)', f"=Simulacao!B{row_sim_resumo+6}", '', 'S'),
        ('Ventilação (Extra)', '', "=VentilacaoExtra!E10", 'S'),
        ('Bombagem', '', "=Bombagem!E10", 'S'),
        ('AQS EE', '', "=AQS!B32", 'S'),
        ('Elevadores', '', "=Elevadores!B25", 'S'),
        ('Equipamentos (Simulação)', f"=Simulacao!B{row_sim_resumo+1}", '', 'T'),
        ('Equipamentos (Extra)', '', "=EquipamentosExtra!E10", 'T'),
    ]

    row = 6
    row_start_ee = row
    for nome, form_sim, form_extra, tipo in consumos_ee:
        ws.cell(row, 1, value=nome).border = THIN_BORDER

        if form_sim:
            ws.cell(row, 2, value=form_sim).fill = FORMULA_FILL
        else:
            ws.cell(row, 2, value=0)
        ws.cell(row, 2).border = THIN_BORDER

        if form_extra:
            ws.cell(row, 3, value=form_extra).fill = FORMULA_FILL
        else:
            ws.cell(row, 3, value=0)
        ws.cell(row, 3).border = THIN_BORDER

        ws.cell(row, 4, value=f'=B{row}+C{row}').fill = TOTAL_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value='EE').border = THIN_BORDER
        ws.cell(row, 6, value=tipo).border = THIN_BORDER
        row += 1
    row_end_ee = row - 1

    ws.cell(row, 1, value='TOTAL EE').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start_ee}:D{row_end_ee})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ee = row
    row += 2

    # Gás Natural
    ws.cell(row, 1, value='GÁS NATURAL').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    consumos_gn = [
        ('AQS GN', "=AQS!B33", 'S'),
        ('Equipamentos GN', "=EquipamentosExtra!B16", 'T'),
    ]

    row_start_gn = row
    for nome, formula, tipo in consumos_gn:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=0).border = THIN_BORDER
        ws.cell(row, 3, value=formula).fill = FORMULA_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}+C{row}').fill = TOTAL_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value='GN').border = THIN_BORDER
        ws.cell(row, 6, value=tipo).border = THIN_BORDER
        row += 1
    row_end_gn = row - 1

    ws.cell(row, 1, value='TOTAL GN').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start_gn}:D{row_end_gn})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_gn = row
    row += 2

    # Renováveis
    ws.cell(row, 1, value='RENOVÁVEIS').font = SECTION_FONT
    ws.cell(row, 1).fill = REN_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    renovaveis = [
        ('Aerotermia Arrefecimento', f"=Simulacao!B{row_sim_resumo+4}"),
        ('Aerotermia Aquecimento', f"=Simulacao!B{row_sim_resumo+5}"),
        ('Aerotermia AQS', "=AQS!B34"),
        ('Solar Térmico', "=AQS!B35"),
        ('Fotovoltaico', "=PV!B15"),
    ]

    row_start_ren = row
    for nome, formula in renovaveis:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=0).border = THIN_BORDER
        ws.cell(row, 3, value=formula).fill = REN_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=C{row}').fill = REN_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value='REN').border = THIN_BORDER
        ws.cell(row, 6, value='REN').border = THIN_BORDER
        row += 1
    row_end_ren = row - 1

    ws.cell(row, 1, value='TOTAL REN').font = TOTAL_FONT
    ws.cell(row, 1).fill = REN_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start_ren}:D{row_end_ren})').font = TOTAL_FONT
    ws.cell(row, 4).fill = REN_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ren = row

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8

    return row_start_ee, row_end_ee, row_total_ee, row_start_gn, row_end_gn, row_total_gn, row_start_ren, row_end_ren, row_total_ren


# =============================================================================
# 12. ENERGIA PRIMÁRIA
# =============================================================================

def create_energia_primaria_sheet(ws, desag_info):
    """Cria folha Energia Primária"""
    row_start_ee, row_end_ee, row_total_ee, row_start_gn, row_end_gn, row_total_gn, row_start_ren, row_end_ren, row_total_ren = desag_info

    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='ENERGIA PRIMÁRIA - Conversão kWh → kWhEP').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 12: Preencher a Área Útil').font = Font(italic=True, color='666666')

    # Área
    ws.cell(4, 1, value='Área Útil [m²]:').font = Font(bold=True)
    ws.cell(4, 2).fill = INPUT_FILL
    ws.cell(4, 2).border = THICK_BORDER
    ws.cell(4, 3, value='← PREENCHER').font = Font(color='FF0000')

    # Factores
    ws.cell(6, 1, value='FACTORES Fpu').font = SECTION_FONT
    ws.cell(6, 1).fill = SECTION_FILL
    ws.cell(7, 1, value='Electricidade')
    ws.cell(7, 2, value=2.5)
    ws.cell(7, 3, value='kWhEP/kWh')
    ws.cell(8, 1, value='Gás Natural')
    ws.cell(8, 2, value=1.0)
    ws.cell(8, 3, value='kWhEP/kWh')

    # Tabela
    ws.cell(10, 1, value='ENERGIA PRIMÁRIA').font = SECTION_FONT
    ws.cell(10, 1).fill = PREV_FILL

    headers = ['Utilização', 'Energia Final', 'Fpu', 'EnergiaPrimaria', 'Tipo']
    for col, h in enumerate(headers, 1):
        ws.cell(11, col, value=h).font = SUBHEADER_FONT
        ws.cell(11, col).fill = SUBHEADER_FILL
        ws.cell(11, col).border = THIN_BORDER

    # EE
    row = 12
    for i in range(row_start_ee, row_end_ee + 1):
        ws.cell(row, 1, value=f"=Desagregacao!A{i}").border = THIN_BORDER
        ws.cell(row, 2, value=f"=Desagregacao!D{i}").fill = FORMULA_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value='=$B$7').border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}').fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f"=Desagregacao!F{i}").border = THIN_BORDER
        row += 1

    # GN
    for i in range(row_start_gn, row_end_gn + 1):
        ws.cell(row, 1, value=f"=Desagregacao!A{i}").border = THIN_BORDER
        ws.cell(row, 2, value=f"=Desagregacao!D{i}").fill = FORMULA_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value='=$B$8').border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}').fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f"=Desagregacao!F{i}").border = THIN_BORDER
        row += 1

    row_total = row
    ws.cell(row, 1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 2, value=f'=SUM(B12:B{row-1})').font = TOTAL_FONT
    ws.cell(row, 2).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D12:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row += 1

    # Total Tipo S
    ws.cell(row, 1, value='TOTAL Tipo S').font = Font(bold=True)
    ws.cell(row, 1).fill = PREV_FILL
    ws.cell(row, 4, value=f'=SUMIF(E12:E{row-2},"S",D12:D{row-2})')
    ws.cell(row, 4).fill = PREV_FILL
    ws.cell(row, 4).font = Font(bold=True)
    ws.cell(row, 4).border = THICK_BORDER
    row_total_s = row
    row += 2

    # Renováveis
    ws.cell(row, 1, value='ENERGIA RENOVÁVEL').font = SECTION_FONT
    ws.cell(row, 1).fill = REN_FILL
    row += 1

    for col, h in enumerate(headers[:4], 1):
        ws.cell(row, col, value=h).font = SUBHEADER_FONT
        ws.cell(row, col).fill = SUBHEADER_FILL
        ws.cell(row, col).border = THIN_BORDER
    row += 1

    row_ren_start = row
    for i in range(row_start_ren, row_end_ren + 1):
        ws.cell(row, 1, value=f"=Desagregacao!A{i}").border = THIN_BORDER
        ws.cell(row, 2, value=f"=Desagregacao!D{i}").fill = REN_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value='=$B$7').border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}').fill = REN_FILL
        ws.cell(row, 4).border = THIN_BORDER
        row += 1

    ws.cell(row, 1, value='TOTAL REN').font = TOTAL_FONT
    ws.cell(row, 1).fill = REN_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_ren_start}:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = REN_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ren = row

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 8

    return row_total_s, row_total_ren


# =============================================================================
# 13. IEE
# =============================================================================

def create_iee_sheet(ws, row_total_s, row_total_ren, row_total_ref):
    """Cria folha IEE

    IEEref,s é calculado automaticamente a partir dos dados REF (DetalheREF):
    - Tipo S = Iluminação + Arrefecimento + Aquecimento + Ventilação
    - IEEref,s = (Consumos Tipo S REF) × Fpu / Área

    Na folha DetalheREF, linha row_total_ref:
    - B = Iluminação
    - D+E = Arrefecimento (Central + Terminal)
    - F+G+H+I = Aquecimento (Central + Terminal + Aux Central + Aux Terminal)
    - J+K+L+M = Ventilação (Supply + Return + Exhaust + Vent)
    """
    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='IEE - Indicadores de Eficiência Energética').font = Font(bold=True, size=14)
    ws.cell(2, 1, value='Passo 13: IEEref,s calculado automaticamente dos dados REF').font = Font(italic=True, color='666666')

    headers = ['Indicador', 'Fórmula', 'Valor', 'Unidade']
    for col, h in enumerate(headers, 1):
        ws.cell(4, col, value=h).font = HEADER_FONT
        ws.cell(4, col).fill = HEADER_FILL
        ws.cell(4, col).border = THIN_BORDER

    # IEEprev,s
    ws.cell(5, 1, value='IEEprev,s').font = Font(bold=True)
    ws.cell(5, 2, value='EP_prev_s / Área')
    ws.cell(5, 3, value=f"=IFERROR(EnergiaPrimaria!D{row_total_s}/EnergiaPrimaria!$B$4,0)")
    ws.cell(5, 3).fill = RESULT_FILL
    ws.cell(5, 3).border = THICK_BORDER
    ws.cell(5, 3).number_format = '0.00'
    ws.cell(5, 4, value='kWhEP/m².ano')

    # IEEref,s (CALCULADO AUTOMATICAMENTE)
    # Tipo S REF = Ilum(B) + Arref(D+E) + Aquec(F+G+H+I) + Vent(J+K+L+M)
    # IEEref,s = Tipo_S_REF × Fpu(2.5) / Área
    ref_tipo_s_formula = f"(DetalheREF!B{row_total_ref}+DetalheREF!D{row_total_ref}+DetalheREF!E{row_total_ref}+DetalheREF!F{row_total_ref}+DetalheREF!G{row_total_ref}+DetalheREF!H{row_total_ref}+DetalheREF!I{row_total_ref}+DetalheREF!J{row_total_ref}+DetalheREF!K{row_total_ref}+DetalheREF!L{row_total_ref}+DetalheREF!M{row_total_ref})"
    iee_ref_formula = f"=IFERROR({ref_tipo_s_formula}*EnergiaPrimaria!$B$7/EnergiaPrimaria!$B$4,0)"

    ws.cell(6, 1, value='IEEref,s').font = Font(bold=True)
    ws.cell(6, 2, value='EP_ref_s / Área')
    ws.cell(6, 3, value=iee_ref_formula)
    ws.cell(6, 3).fill = REF_FILL
    ws.cell(6, 3).border = THICK_BORDER
    ws.cell(6, 3).number_format = '0.00'
    ws.cell(6, 4, value='kWhEP/m².ano')

    # IEEren
    ws.cell(7, 1, value='IEEren').font = Font(bold=True)
    ws.cell(7, 2, value='EP_ren / Área')
    ws.cell(7, 3, value=f"=IFERROR(EnergiaPrimaria!D{row_total_ren}/EnergiaPrimaria!$B$4,0)")
    ws.cell(7, 3).fill = REN_FILL
    ws.cell(7, 3).border = THICK_BORDER
    ws.cell(7, 3).number_format = '0.00'
    ws.cell(7, 4, value='kWhEP/m².ano')

    # RIEE
    ws.cell(9, 1, value='RIEE').font = Font(bold=True, size=16)
    ws.cell(9, 2, value='(IEEprev,s - IEEren) / IEEref,s')
    ws.cell(9, 3, value='=IFERROR((C5-C7)/C6,"-")')
    ws.cell(9, 3).fill = RESULT_FILL
    ws.cell(9, 3).border = THICK_BORDER
    ws.cell(9, 3).font = Font(bold=True, size=16)
    ws.cell(9, 3).number_format = '0.00'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18


# =============================================================================
# 14. CLASSE
# =============================================================================

def create_classe_sheet(ws):
    """Cria folha Classe"""
    ws.merge_cells('A1:D1')
    ws.cell(1, 1, value='CLASSE ENERGÉTICA').font = Font(bold=True, size=16)
    ws.cell(2, 1, value='Passo 14: Resultado final').font = Font(italic=True, color='666666')

    ws.cell(4, 1, value='RIEE:').font = Font(bold=True, size=14)
    ws.cell(4, 2, value="=IEE!C9")
    ws.cell(4, 2).font = Font(bold=True, size=14)
    ws.cell(4, 2).number_format = '0.00'

    ws.cell(6, 1, value='CLASSE:').font = Font(bold=True, size=16)
    classe_formula = '=IF(ISERROR(B4),"-",IF(B4<=0.25,"A+",IF(B4<=0.5,"A",IF(B4<=0.75,"B",IF(B4<=1,"B-",IF(B4<=1.5,"C",IF(B4<=2,"D",IF(B4<=2.5,"E","F"))))))))'
    ws.cell(6, 2, value=classe_formula)
    ws.cell(6, 2).font = Font(bold=True, size=48)
    ws.cell(6, 2).alignment = Alignment(horizontal='center')

    ws.cell(9, 1, value='ESCALA:').font = Font(bold=True)

    row = 10
    for classe, min_val, max_val in RIEE_LIMITES:
        ws.cell(row, 1, value=classe)
        ws.cell(row, 1).fill = PatternFill(start_color=CLASS_COLORS[classe], end_color=CLASS_COLORS[classe], fill_type='solid')
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal='center')
        ws.cell(row, 1).border = THIN_BORDER

        if classe == 'A+':
            ws.cell(row, 2, value='≤ 0.25')
        elif classe == 'F':
            ws.cell(row, 2, value='> 2.50')
        else:
            ws.cell(row, 2, value=f'{min_val:.2f} - {max_val:.2f}')
        ws.cell(row, 2).border = THIN_BORDER
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15


# =============================================================================
# 15. LEGENDA
# =============================================================================

def create_legenda_sheet(ws):
    """Cria folha Legenda"""
    ws.cell(1, 1, value='SEQUÊNCIA DE PREENCHIMENTO').font = Font(bold=True, size=14)

    sequencia = [
        ('1-2', 'Detalhe PREV/REF, Mensal', 'Automático dos CSV', 'Dados brutos do HAP'),
        ('3', 'Simulacao', 'EER e COP', 'Energia final e aerotermia'),
        ('4', 'IluminacaoENU', 'Potência e horas', 'Iluminação não simulada'),
        ('5', 'AQS', 'Pessoas, litros, sistema', 'Águas quentes sanitárias'),
        ('6', 'PV', 'Potência instalada', 'Fotovoltaico'),
        ('7', 'EquipamentosExtra', 'Potência e horas', 'Equipamentos não simulados'),
        ('8', 'Elevadores', 'Carga, viagens', 'Energia elevadores'),
        ('9', 'VentilacaoExtra', 'Potência e horas', 'Ventilação não simulada'),
        ('10', 'Bombagem', 'Potência e horas', 'Bombas não simuladas'),
        ('11', 'Desagregacao', 'Automático', 'Resume tudo'),
        ('12', 'EnergiaPrimaria', 'Área útil', 'Conversão kWh→kWhEP'),
        ('13', 'IEE', 'IEEref,s (se necessário)', 'Indicadores'),
        ('14', 'Classe', 'Automático', 'Resultado final'),
    ]

    headers = ['Passo', 'Folha', 'Preencher', 'Resultado']
    for col, h in enumerate(headers, 1):
        ws.cell(3, col, value=h).font = HEADER_FONT
        ws.cell(3, col).fill = HEADER_FILL
        ws.cell(3, col).border = THIN_BORDER

    row = 4
    for passo, folha, preencher, resultado in sequencia:
        ws.cell(row, 1, value=passo).border = THIN_BORDER
        ws.cell(row, 2, value=folha).border = THIN_BORDER
        ws.cell(row, 3, value=preencher).border = THIN_BORDER
        ws.cell(row, 4, value=resultado).border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row, 1, value='CORES:').font = Font(bold=True)
    row += 1
    ws.cell(row, 1, value='Amarelo').fill = INPUT_FILL
    ws.cell(row, 2, value='= Preencher')
    row += 1
    ws.cell(row, 1, value='Verde claro').fill = FORMULA_FILL
    ws.cell(row, 2, value='= Fórmula automática')
    row += 1
    ws.cell(row, 1, value='Verde escuro').fill = REN_FILL
    ws.cell(row, 2, value='= Energia renovável')
    row += 1
    ws.cell(row, 1, value='Laranja').fill = RESULT_FILL
    ws.cell(row, 2, value='= Resultado')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def create_iee_completo(prev_data, ref_data, output_path):
    """Cria a folha Excel completa"""
    wb = openpyxl.Workbook()

    # 1. Detalhe PREV
    ws = wb.active
    ws.title = 'DetalhePREV'
    ws.sheet_properties.tabColor = 'DDEBF7'
    row_total_prev = create_detalhe_sheet(ws, prev_data, 'PREVISTO')

    # 2. Detalhe REF
    ws = wb.create_sheet('DetalheREF')
    ws.sheet_properties.tabColor = 'FCE4D6'
    row_total_ref = create_detalhe_sheet(ws, ref_data, 'REFERÊNCIA')

    # 3. Mensal PREV
    ws = wb.create_sheet('MensalPREV')
    ws.sheet_properties.tabColor = 'DDEBF7'
    create_mensal_sheet(ws, prev_data, 'PREVISTO')

    # 4. Mensal REF
    ws = wb.create_sheet('MensalREF')
    ws.sheet_properties.tabColor = 'FCE4D6'
    create_mensal_sheet(ws, ref_data, 'REFERÊNCIA')

    # 5. Simulação
    ws = wb.create_sheet('Simulacao')
    ws.sheet_properties.tabColor = 'FF0000'
    row_sim_resumo, row_sim_total = create_simulacao_sheet(ws, prev_data, row_total_prev)

    # 6. Iluminação ENU
    ws = wb.create_sheet('IluminacaoENU')
    ws.sheet_properties.tabColor = '92D050'
    create_iluminacao_enu_sheet(ws)

    # 7. AQS
    ws = wb.create_sheet('AQS')
    ws.sheet_properties.tabColor = 'FFC000'
    create_aqs_sheet(ws)

    # 8. PV
    ws = wb.create_sheet('PV')
    ws.sheet_properties.tabColor = '7030A0'
    create_pv_sheet(ws)

    # 9. Equipamentos Extra
    ws = wb.create_sheet('EquipamentosExtra')
    ws.sheet_properties.tabColor = 'FFC000'
    create_equipamentos_extra_sheet(ws)

    # 10. Elevadores
    ws = wb.create_sheet('Elevadores')
    ws.sheet_properties.tabColor = '00B0F0'
    create_elevadores_sheet(ws)

    # 11. Ventilação Extra
    ws = wb.create_sheet('VentilacaoExtra')
    ws.sheet_properties.tabColor = '7030A0'
    create_ventilacao_extra_sheet(ws)

    # 12. Bombagem
    ws = wb.create_sheet('Bombagem')
    ws.sheet_properties.tabColor = '000000'
    create_bombagem_sheet(ws)

    # 13. Desagregação
    ws = wb.create_sheet('Desagregacao')
    ws.sheet_properties.tabColor = 'FF0000'
    desag_info = create_desagregacao_sheet(ws, row_sim_resumo)

    # 14. Energia Primária
    ws = wb.create_sheet('EnergiaPrimaria')
    ws.sheet_properties.tabColor = 'FF0000'
    row_total_s, row_total_ren = create_energia_primaria_sheet(ws, desag_info)

    # 15. IEE
    ws = wb.create_sheet('IEE')
    ws.sheet_properties.tabColor = 'FF0000'
    create_iee_sheet(ws, row_total_s, row_total_ren, row_total_ref)

    # 16. Classe
    ws = wb.create_sheet('Classe')
    ws.sheet_properties.tabColor = 'FF0000'
    create_classe_sheet(ws)

    # 17. Legenda
    ws = wb.create_sheet('Legenda')
    create_legenda_sheet(ws)

    wb.save(output_path)
    return len(prev_data), len(ref_data)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    prev_folder = sys.argv[1]
    ref_folder = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else 'IEE_Completo_v3.xlsx'

    print(f"A carregar PREV: {prev_folder}")
    prev_data = load_project_data(prev_folder)
    print(f"  → {len(prev_data)} sistemas")

    print(f"A carregar REF: {ref_folder}")
    ref_data = load_project_data(ref_folder)
    print(f"  → {len(ref_data)} sistemas")

    if not prev_data or not ref_data:
        print("ERRO: CSV não encontrados")
        sys.exit(1)

    num_prev, num_ref = create_iee_completo(prev_data, ref_data, output_path)

    print(f"\n{'='*60}")
    print(f"FICHEIRO CRIADO: {output_path}")
    print(f"{'='*60}")
    print("""
SEQUÊNCIA DE PREENCHIMENTO:
  1-2. Detalhe/Mensal  → Automático (CSV)
  3.   Simulação       → Preencher EER/COP
  4.   Iluminação ENU  → Preencher potência/horas
  5.   AQS             → Preencher pessoas, sistema
  6.   PV              → Preencher potência
  7.   Equipamentos    → Preencher equipamentos extra
  8.   Elevadores      → Preencher carga, viagens
  9.   Ventilação      → Preencher ventilação extra
  10.  Bombagem        → Preencher bombas
  11.  Desagregação    → Automático
  12.  Energia Primária→ Preencher ÁREA ÚTIL
  13.  IEE             → Preencher IEEref,s
  14.  Classe          → RESULTADO FINAL
""")


if __name__ == '__main__':
    main()
