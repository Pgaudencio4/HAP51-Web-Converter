"""
Cálculo IEE - Folha completa para certificação energética SCE

Cria uma folha Excel profissional que:
- Recebe dados de duas simulações HAP (PREV e REF)
- Calcula IEEprev, IEEref, IEEren
- Calcula RIEE e determina classe energética
- Todas as células com fórmulas para rastreabilidade

Usage:
    python calculo_iee.py <pasta_prev> <pasta_ref> [output.xlsx]

Exemplo:
    python calculo_iee.py "C:\\Projecto_PREV" "C:\\Projecto_REF" IEE_Calculo.xlsx
"""

import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference

# =============================================================================
# ESTILOS
# =============================================================================

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10)

SECTION_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
SECTION_FONT = Font(bold=True, size=10)

TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_FONT = Font(bold=True)

FORMULA_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RESULT_FILL = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Cores das classes energéticas
CLASS_COLORS = {
    'A+': '00A651', 'A': '50B848', 'B': 'B5D334',
    'B-': 'FFF200', 'C': 'F7941D', 'D': 'F15A29',
    'E': 'ED1C24', 'F': 'BE1E2D'
}

# =============================================================================
# CONSTANTES SCE
# =============================================================================

# Factores de conversão para energia primária (Despacho 15793-D/2013)
FPU = {
    'electricidade': 2.5,  # kWhEP/kWh
    'gas_natural': 1.0,    # kWhEP/kWh
    'gpl': 1.0,            # kWhEP/kWh
    'gasoleo': 1.0,        # kWhEP/kWh
    'biomassa': 1.0,       # kWhEP/kWh
}

# Limites RIEE para classes energéticas
RIEE_LIMITES = [
    ('A+', 0, 0.25),
    ('A', 0.26, 0.50),
    ('B', 0.51, 0.75),
    ('B-', 0.76, 1.00),
    ('C', 1.01, 1.50),
    ('D', 1.51, 2.00),
    ('E', 2.01, 2.50),
    ('F', 2.51, 999),
]

# Colunas do CSV HAP
CSV_COLUMNS = [
    ('Lighting (kWh)', 'Iluminação'),
    ('Electric Equipment (kWh)', 'Equipamentos'),
    ('Central Unit Clg Input (kWh)', 'Arrefecimento Central'),
    ('Terminal Unit Clg Input (kWh)', 'Arrefecimento Terminal'),
    ('Central Unit Htg Input (kWh)', 'Aquecimento Central'),
    ('Terminal Unit Htg Input (kWh)', 'Aquecimento Terminal'),
    ('Central Unit Aux. Htg. Input (kWh)', 'Aquec. Aux. Central'),
    ('Terminal Unit Aux. Htg. Input (kWh)', 'Aquec. Aux. Terminal'),
    ('Supply Fan (kWh)', 'Ventilador Insuflação'),
    ('Return Fan (kWh)', 'Ventilador Retorno'),
    ('Exhaust Fan (kWh)', 'Ventilador Extração'),
    ('Ventilation Fan (kWh)', 'Ventilador Ventilação'),
]


# =============================================================================
# FUNÇÕES
# =============================================================================

def read_hap_csv(filepath):
    """Lê um CSV do HAP e retorna dicionário com totais anuais"""
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if len(lines) < 5:
        return None, {}

    sistema = lines[1].split(';')[0].replace('Monthly Simulation Results for ', '').strip()
    headers = lines[3].strip().split(';')
    totals = {}

    for line in lines[4:]:
        line = line.strip()
        if not line or line.startswith('Month'):
            continue
        values = line.split(';')
        for i, val in enumerate(values[1:], 1):
            if i < len(headers) and val:
                col = headers[i]
                try:
                    totals[col] = totals.get(col, 0) + int(val)
                except:
                    pass

    return sistema, {'totals': totals, 'headers': headers}


def find_hap_csvs(project_folder):
    """Encontra todos os CSVs do HAP numa pasta"""
    pattern = os.path.join(project_folder, 'HAP51_Monthly_*.csv')
    return sorted(glob.glob(pattern))


def load_project_data(project_folder):
    """Carrega todos os dados de um projecto HAP"""
    csv_files = find_hap_csvs(project_folder)
    systems_data = {}

    for csv_file in csv_files:
        sistema, data = read_hap_csv(csv_file)
        if sistema and data and sistema != 'TODOS':
            systems_data[sistema] = data

    return systems_data


def create_iee_excel(prev_data, ref_data, output_path):
    """Cria a folha Excel completa de cálculo IEE"""

    wb = openpyxl.Workbook()

    # =========================================================================
    # FOLHA 1: DADOS PREV (valores dos CSVs)
    # =========================================================================
    ws_prev = wb.active
    ws_prev.title = 'Dados PREV'

    create_data_sheet(ws_prev, prev_data, 'PREVISTO')

    # =========================================================================
    # FOLHA 2: DADOS REF (valores dos CSVs)
    # =========================================================================
    ws_ref = wb.create_sheet('Dados REF')
    create_data_sheet(ws_ref, ref_data, 'REFERÊNCIA')

    # =========================================================================
    # FOLHA 3: CÁLCULO IEE (fórmulas)
    # =========================================================================
    ws_calc = wb.create_sheet('Cálculo IEE')

    num_sistemas_prev = len(prev_data)
    num_sistemas_ref = len(ref_data)

    # Título
    ws_calc.merge_cells('B2:H2')
    ws_calc.cell(2, 2, value='CÁLCULO DO INDICADOR DE EFICIÊNCIA ENERGÉTICA (IEE)').font = Font(bold=True, size=16)

    ws_calc.merge_cells('B3:H3')
    ws_calc.cell(3, 2, value='Sistema de Certificação Energética (SCE) - Portugal').font = Font(italic=True, size=11)

    # Área útil (input)
    ws_calc.cell(5, 2, value='DADOS DO EDIFÍCIO').font = SECTION_FONT
    ws_calc.cell(5, 2).fill = SECTION_FILL

    ws_calc.cell(6, 2, value='Área Útil [m²]')
    ws_calc.cell(6, 3).fill = INPUT_FILL
    ws_calc.cell(6, 3).border = THIN_BORDER
    ws_calc.cell(6, 4, value='← Preencher').font = Font(italic=True, color='FF0000')

    # Factores de conversão
    ws_calc.cell(8, 2, value='FACTORES DE CONVERSÃO (Fpu)').font = SECTION_FONT
    ws_calc.cell(8, 2).fill = SECTION_FILL

    ws_calc.cell(9, 2, value='Electricidade')
    ws_calc.cell(9, 3, value=FPU['electricidade'])
    ws_calc.cell(9, 4, value='kWhEP/kWh')

    ws_calc.cell(10, 2, value='Gás Natural')
    ws_calc.cell(10, 3, value=FPU['gas_natural'])
    ws_calc.cell(10, 4, value='kWhEP/kWh')

    # Consumos PREV
    ws_calc.cell(12, 2, value='CONSUMOS PREVISTO (Energia Final)').font = SECTION_FONT
    ws_calc.cell(12, 2).fill = SECTION_FILL
    ws_calc.merge_cells('B12:D12')

    row = 13
    prev_total_row = 2 + num_sistemas_prev  # linha do TOTAL na folha Dados PREV

    consumos_prev = [
        ('Iluminação Interior', f"='Dados PREV'!B{prev_total_row}", 'S'),
        ('Equipamentos', f"='Dados PREV'!C{prev_total_row}", 'T'),
        ('Arrefecimento', f"='Dados PREV'!D{prev_total_row}+'Dados PREV'!E{prev_total_row}", 'S'),
        ('Aquecimento', f"='Dados PREV'!F{prev_total_row}+'Dados PREV'!G{prev_total_row}", 'S'),
        ('Aquec. Auxiliar', f"='Dados PREV'!H{prev_total_row}+'Dados PREV'!I{prev_total_row}", 'S'),
        ('Ventilação', f"='Dados PREV'!J{prev_total_row}+'Dados PREV'!K{prev_total_row}+'Dados PREV'!L{prev_total_row}+'Dados PREV'!M{prev_total_row}", 'S'),
    ]

    ws_calc.cell(row, 2, value='Utilização').font = Font(bold=True)
    ws_calc.cell(row, 3, value='kWh/ano').font = Font(bold=True)
    ws_calc.cell(row, 4, value='Tipo').font = Font(bold=True)
    ws_calc.cell(row, 5, value='kWhEP/ano').font = Font(bold=True)
    row += 1

    consumo_rows_prev = {}
    for nome, formula, tipo in consumos_prev:
        ws_calc.cell(row, 2, value=nome)
        ws_calc.cell(row, 3, value=formula)
        ws_calc.cell(row, 3).fill = FORMULA_FILL
        ws_calc.cell(row, 4, value=tipo)
        # Energia primária = consumo × Fpu
        ws_calc.cell(row, 5, value=f'=C{row}*$C$9')
        ws_calc.cell(row, 5).fill = FORMULA_FILL
        consumo_rows_prev[nome] = row
        row += 1

    # Campos manuais (não vêm do HAP)
    campos_manuais = [
        ('Iluminação Exterior', 'S'),
        ('Bombagem', 'S'),
        ('AQS (Electricidade)', 'S'),
        ('Elevadores', 'S'),
    ]

    for nome, tipo in campos_manuais:
        ws_calc.cell(row, 2, value=nome)
        ws_calc.cell(row, 3).fill = INPUT_FILL
        ws_calc.cell(row, 3).border = THIN_BORDER
        ws_calc.cell(row, 4, value=tipo)
        ws_calc.cell(row, 5, value=f'=C{row}*$C$9')
        ws_calc.cell(row, 5).fill = FORMULA_FILL
        consumo_rows_prev[nome] = row
        row += 1

    # Gás Natural
    ws_calc.cell(row, 2, value='AQS (Gás Natural)')
    ws_calc.cell(row, 3).fill = INPUT_FILL
    ws_calc.cell(row, 3).border = THIN_BORDER
    ws_calc.cell(row, 4, value='S')
    ws_calc.cell(row, 5, value=f'=C{row}*$C$10')
    ws_calc.cell(row, 5).fill = FORMULA_FILL
    row += 1

    ws_calc.cell(row, 2, value='Equipamentos (Gás Natural)')
    ws_calc.cell(row, 3).fill = INPUT_FILL
    ws_calc.cell(row, 3).border = THIN_BORDER
    ws_calc.cell(row, 4, value='T')
    ws_calc.cell(row, 5, value=f'=C{row}*$C$10')
    ws_calc.cell(row, 5).fill = FORMULA_FILL
    row += 1

    # Total Energia Final e Primária
    row_total_prev = row
    ws_calc.cell(row, 2, value='TOTAL PREV').font = TOTAL_FONT
    ws_calc.cell(row, 2).fill = TOTAL_FILL
    ws_calc.cell(row, 3, value=f'=SUM(C14:C{row-1})')
    ws_calc.cell(row, 3).font = TOTAL_FONT
    ws_calc.cell(row, 3).fill = TOTAL_FILL
    ws_calc.cell(row, 5, value=f'=SUM(E14:E{row-1})')
    ws_calc.cell(row, 5).font = TOTAL_FONT
    ws_calc.cell(row, 5).fill = TOTAL_FILL
    row += 2

    # Consumos REF (estrutura similar)
    ws_calc.cell(row, 2, value='CONSUMOS REFERÊNCIA (Energia Final)').font = SECTION_FONT
    ws_calc.cell(row, 2).fill = SECTION_FILL
    ws_calc.merge_cells(f'B{row}:D{row}')
    row += 1

    ref_total_row = 2 + num_sistemas_ref

    consumos_ref = [
        ('Iluminação Interior', f"='Dados REF'!B{ref_total_row}", 'S'),
        ('Equipamentos', f"='Dados REF'!C{ref_total_row}", 'T'),
        ('Arrefecimento', f"='Dados REF'!D{ref_total_row}+'Dados REF'!E{ref_total_row}", 'S'),
        ('Aquecimento', f"='Dados REF'!F{ref_total_row}+'Dados REF'!G{ref_total_row}", 'S'),
        ('Aquec. Auxiliar', f"='Dados REF'!H{ref_total_row}+'Dados REF'!I{ref_total_row}", 'S'),
        ('Ventilação', f"='Dados REF'!J{ref_total_row}+'Dados REF'!K{ref_total_row}+'Dados REF'!L{ref_total_row}+'Dados REF'!M{ref_total_row}", 'S'),
    ]

    ws_calc.cell(row, 2, value='Utilização').font = Font(bold=True)
    ws_calc.cell(row, 3, value='kWh/ano').font = Font(bold=True)
    ws_calc.cell(row, 4, value='Tipo').font = Font(bold=True)
    ws_calc.cell(row, 5, value='kWhEP/ano').font = Font(bold=True)
    row += 1

    row_start_ref = row
    for nome, formula, tipo in consumos_ref:
        ws_calc.cell(row, 2, value=nome)
        ws_calc.cell(row, 3, value=formula)
        ws_calc.cell(row, 3).fill = FORMULA_FILL
        ws_calc.cell(row, 4, value=tipo)
        ws_calc.cell(row, 5, value=f'=C{row}*$C$9')
        ws_calc.cell(row, 5).fill = FORMULA_FILL
        row += 1

    # Campos manuais REF
    for nome, tipo in campos_manuais:
        ws_calc.cell(row, 2, value=nome)
        ws_calc.cell(row, 3).fill = INPUT_FILL
        ws_calc.cell(row, 3).border = THIN_BORDER
        ws_calc.cell(row, 4, value=tipo)
        ws_calc.cell(row, 5, value=f'=C{row}*$C$9')
        ws_calc.cell(row, 5).fill = FORMULA_FILL
        row += 1

    # Gás Natural REF
    ws_calc.cell(row, 2, value='AQS (Gás Natural)')
    ws_calc.cell(row, 3).fill = INPUT_FILL
    ws_calc.cell(row, 3).border = THIN_BORDER
    ws_calc.cell(row, 4, value='S')
    ws_calc.cell(row, 5, value=f'=C{row}*$C$10')
    ws_calc.cell(row, 5).fill = FORMULA_FILL
    row += 1

    ws_calc.cell(row, 2, value='Equipamentos (Gás Natural)')
    ws_calc.cell(row, 3).fill = INPUT_FILL
    ws_calc.cell(row, 3).border = THIN_BORDER
    ws_calc.cell(row, 4, value='T')
    ws_calc.cell(row, 5, value=f'=C{row}*$C$10')
    ws_calc.cell(row, 5).fill = FORMULA_FILL
    row += 1

    row_total_ref = row
    ws_calc.cell(row, 2, value='TOTAL REF').font = TOTAL_FONT
    ws_calc.cell(row, 2).fill = TOTAL_FILL
    ws_calc.cell(row, 3, value=f'=SUM(C{row_start_ref}:C{row-1})')
    ws_calc.cell(row, 3).font = TOTAL_FONT
    ws_calc.cell(row, 3).fill = TOTAL_FILL
    ws_calc.cell(row, 5, value=f'=SUM(E{row_start_ref}:E{row-1})')
    ws_calc.cell(row, 5).font = TOTAL_FONT
    ws_calc.cell(row, 5).fill = TOTAL_FILL
    row += 2

    # Renováveis
    ws_calc.cell(row, 2, value='ENERGIA RENOVÁVEL').font = SECTION_FONT
    ws_calc.cell(row, 2).fill = SECTION_FILL
    ws_calc.merge_cells(f'B{row}:D{row}')
    row += 1

    ws_calc.cell(row, 2, value='Fonte').font = Font(bold=True)
    ws_calc.cell(row, 3, value='kWh/ano').font = Font(bold=True)
    ws_calc.cell(row, 5, value='kWhEP/ano').font = Font(bold=True)
    row += 1

    row_ren_start = row
    renovaveis = ['Solar Térmico (AQS)', 'Aerotermia (AQS)', 'Aerotermia (Aquecimento)', 'Aerotermia (Arrefecimento)', 'Fotovoltaico (autoconsumo)']
    for ren in renovaveis:
        ws_calc.cell(row, 2, value=ren)
        ws_calc.cell(row, 3).fill = INPUT_FILL
        ws_calc.cell(row, 3).border = THIN_BORDER
        ws_calc.cell(row, 5, value=f'=C{row}*$C$9')
        ws_calc.cell(row, 5).fill = FORMULA_FILL
        row += 1

    row_ren_total = row
    ws_calc.cell(row, 2, value='TOTAL RENOVÁVEL').font = TOTAL_FONT
    ws_calc.cell(row, 2).fill = TOTAL_FILL
    ws_calc.cell(row, 3, value=f'=SUM(C{row_ren_start}:C{row-1})')
    ws_calc.cell(row, 3).font = TOTAL_FONT
    ws_calc.cell(row, 3).fill = TOTAL_FILL
    ws_calc.cell(row, 5, value=f'=SUM(E{row_ren_start}:E{row-1})')
    ws_calc.cell(row, 5).font = TOTAL_FONT
    ws_calc.cell(row, 5).fill = TOTAL_FILL
    row += 2

    # RESULTADOS IEE
    ws_calc.cell(row, 2, value='INDICADORES DE EFICIÊNCIA ENERGÉTICA').font = Font(bold=True, size=14)
    ws_calc.cell(row, 2).fill = HEADER_FILL
    ws_calc.cell(row, 2).font = HEADER_FONT
    ws_calc.merge_cells(f'B{row}:E{row}')
    row += 1

    ws_calc.cell(row, 2, value='Indicador').font = Font(bold=True)
    ws_calc.cell(row, 3, value='Fórmula').font = Font(bold=True)
    ws_calc.cell(row, 4, value='Valor').font = Font(bold=True)
    ws_calc.cell(row, 5, value='Unidade').font = Font(bold=True)
    row += 1

    # IEEprev
    ws_calc.cell(row, 2, value='IEEprev')
    ws_calc.cell(row, 3, value='= Ep_prev / Área')
    ws_calc.cell(row, 4, value=f'=E{row_total_prev}/$C$6')
    ws_calc.cell(row, 4).fill = RESULT_FILL
    ws_calc.cell(row, 5, value='kWhEP/m².ano')
    row_iee_prev = row
    row += 1

    # IEEref
    ws_calc.cell(row, 2, value='IEEref')
    ws_calc.cell(row, 3, value='= Ep_ref / Área')
    ws_calc.cell(row, 4, value=f'=E{row_total_ref}/$C$6')
    ws_calc.cell(row, 4).fill = RESULT_FILL
    ws_calc.cell(row, 5, value='kWhEP/m².ano')
    row_iee_ref = row
    row += 1

    # IEEren
    ws_calc.cell(row, 2, value='IEEren')
    ws_calc.cell(row, 3, value='= Ep_ren / Área')
    ws_calc.cell(row, 4, value=f'=E{row_ren_total}/$C$6')
    ws_calc.cell(row, 4).fill = RESULT_FILL
    ws_calc.cell(row, 5, value='kWhEP/m².ano')
    row_iee_ren = row
    row += 1

    # RIEE
    ws_calc.cell(row, 2, value='RIEE').font = Font(bold=True, size=12)
    ws_calc.cell(row, 3, value='= (IEEprev - IEEren) / IEEref')
    ws_calc.cell(row, 4, value=f'=(D{row_iee_prev}-D{row_iee_ren})/D{row_iee_ref}')
    ws_calc.cell(row, 4).fill = RESULT_FILL
    ws_calc.cell(row, 4).font = Font(bold=True, size=12)
    row_riee = row
    row += 2

    # Classe energética
    ws_calc.cell(row, 2, value='CLASSE ENERGÉTICA').font = Font(bold=True, size=14)
    ws_calc.cell(row, 2).fill = HEADER_FILL
    ws_calc.cell(row, 2).font = HEADER_FONT
    ws_calc.merge_cells(f'B{row}:E{row}')
    row += 1

    # Fórmula para determinar classe
    classe_formula = f'=IF(D{row_riee}<=0.25,"A+",IF(D{row_riee}<=0.5,"A",IF(D{row_riee}<=0.75,"B",IF(D{row_riee}<=1,"B-",IF(D{row_riee}<=1.5,"C",IF(D{row_riee}<=2,"D",IF(D{row_riee}<=2.5,"E","F")))))))'

    ws_calc.cell(row, 2, value='Classe:')
    ws_calc.cell(row, 3, value=classe_formula)
    ws_calc.cell(row, 3).font = Font(bold=True, size=24)
    ws_calc.cell(row, 3).alignment = Alignment(horizontal='center')
    row += 2

    # Tabela de limites RIEE
    ws_calc.cell(row, 2, value='Limites RIEE por Classe:').font = Font(bold=True)
    row += 1
    for classe, min_val, max_val in RIEE_LIMITES:
        ws_calc.cell(row, 2, value=classe)
        ws_calc.cell(row, 2).fill = PatternFill(start_color=CLASS_COLORS[classe], end_color=CLASS_COLORS[classe], fill_type='solid')
        if classe == 'A+':
            ws_calc.cell(row, 3, value=f'≤ {max_val}')
        elif classe == 'F':
            ws_calc.cell(row, 3, value=f'> 2.50')
        else:
            ws_calc.cell(row, 3, value=f'{min_val} - {max_val}')
        row += 1

    # Ajustar larguras
    ws_calc.column_dimensions['B'].width = 25
    ws_calc.column_dimensions['C'].width = 15
    ws_calc.column_dimensions['D'].width = 12
    ws_calc.column_dimensions['E'].width = 15

    # =========================================================================
    # FOLHA 4: LEGENDA
    # =========================================================================
    ws_leg = wb.create_sheet('Legenda')

    ws_leg.cell(2, 2, value='LEGENDA E INSTRUÇÕES').font = Font(bold=True, size=14)

    ws_leg.cell(4, 2, value='CORES DAS CÉLULAS:').font = Font(bold=True)

    ws_leg.cell(5, 2, value='Verde')
    ws_leg.cell(5, 2).fill = FORMULA_FILL
    ws_leg.cell(5, 3, value='= Fórmula (valor calculado automaticamente)')

    ws_leg.cell(6, 2, value='Amarelo')
    ws_leg.cell(6, 2).fill = INPUT_FILL
    ws_leg.cell(6, 3, value='= Input (preencher manualmente)')

    ws_leg.cell(7, 2, value='Laranja')
    ws_leg.cell(7, 2).fill = RESULT_FILL
    ws_leg.cell(7, 3, value='= Resultado final')

    ws_leg.cell(9, 2, value='TIPOS DE CONSUMO:').font = Font(bold=True)
    ws_leg.cell(10, 2, value='Tipo S')
    ws_leg.cell(10, 3, value='= Conta para classificação energética')
    ws_leg.cell(11, 2, value='Tipo T')
    ws_leg.cell(11, 3, value='= NÃO conta para classificação')

    ws_leg.cell(13, 2, value='CONSUMOS TIPO S:').font = Font(bold=True)
    tipo_s = ['Aquecimento e Arrefecimento', 'Ventilação AVAC', 'Bombagem AVAC',
              'AQS', 'Iluminação Interior', 'Iluminação Exterior (desde 2016)',
              'Elevadores (desde 2016)']
    for i, item in enumerate(tipo_s):
        ws_leg.cell(14 + i, 3, value=f'• {item}')

    ws_leg.cell(22, 2, value='CONSUMOS TIPO T:').font = Font(bold=True)
    tipo_t = ['Equipamentos (computadores, etc.)', 'Refrigeração comercial', 'Outros']
    for i, item in enumerate(tipo_t):
        ws_leg.cell(23 + i, 3, value=f'• {item}')

    ws_leg.cell(27, 2, value='FÓRMULAS:').font = Font(bold=True)
    ws_leg.cell(28, 2, value='IEEprev = Σ(Consumos × Fpu) / Área')
    ws_leg.cell(29, 2, value='RIEE = (IEEprev,s - IEEren) / IEEref,s')

    ws_leg.column_dimensions['B'].width = 20
    ws_leg.column_dimensions['C'].width = 50

    # =========================================================================
    # GUARDAR
    # =========================================================================
    wb.save(output_path)
    return len(prev_data), len(ref_data)


def create_data_sheet(ws, data, title):
    """Cria uma folha de dados brutos"""

    sistemas = [(s, d) for s, d in data.items()]

    # Título
    ws.cell(1, 1, value='Sistema').font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 1).border = THIN_BORDER

    for col, (csv_col, nome) in enumerate(CSV_COLUMNS, 2):
        cell = ws.cell(1, col, value=nome)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Dados
    row = 2
    for sistema, sdata in sistemas:
        ws.cell(row, 1, value=sistema).border = THIN_BORDER

        for col, (csv_col, nome) in enumerate(CSV_COLUMNS, 2):
            value = sdata['totals'].get(csv_col, 0)
            cell = ws.cell(row, col, value=value if value > 0 else '')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

        row += 1

    # TOTAL
    ws.cell(row, 1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 1).border = THIN_BORDER

    for col in range(2, len(CSV_COLUMNS) + 2):
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}2:{col_letter}{row-1})'
        cell = ws.cell(row, col, value=formula)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    # Larguras
    ws.column_dimensions['A'].width = 20
    for col in range(2, len(CSV_COLUMNS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = 'B2'


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        print("\nERRO: Necessário indicar as pastas PREV e REF")
        sys.exit(1)

    prev_folder = sys.argv[1]
    ref_folder = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else 'Calculo_IEE.xlsx'

    # Verificar pastas
    if not os.path.isdir(prev_folder):
        print(f"Erro: Pasta PREV não encontrada: {prev_folder}")
        sys.exit(1)

    if not os.path.isdir(ref_folder):
        print(f"Erro: Pasta REF não encontrada: {ref_folder}")
        sys.exit(1)

    # Carregar dados
    print(f"A carregar dados PREV de: {prev_folder}")
    prev_data = load_project_data(prev_folder)
    print(f"  Encontrados {len(prev_data)} sistemas")

    print(f"\nA carregar dados REF de: {ref_folder}")
    ref_data = load_project_data(ref_folder)
    print(f"  Encontrados {len(ref_data)} sistemas")

    if not prev_data:
        print("ERRO: Nenhum CSV encontrado na pasta PREV")
        sys.exit(1)

    if not ref_data:
        print("ERRO: Nenhum CSV encontrado na pasta REF")
        sys.exit(1)

    # Criar Excel
    num_prev, num_ref = create_iee_excel(prev_data, ref_data, output_path)

    print(f"\n{'='*50}")
    print(f"Ficheiro criado: {output_path}")
    print(f"{'='*50}")
    print(f"  - Folha 'Dados PREV': {num_prev} sistemas (valores)")
    print(f"  - Folha 'Dados REF': {num_ref} sistemas (valores)")
    print(f"  - Folha 'Cálculo IEE': Fórmulas completas")
    print(f"  - Folha 'Legenda': Instruções")
    print(f"\nPróximos passos:")
    print(f"  1. Abrir o ficheiro no Excel")
    print(f"  2. Preencher a Área Útil [m²]")
    print(f"  3. Preencher campos amarelos (AQS, Bombagem, etc.)")
    print(f"  4. Ver resultado: RIEE e Classe Energética")


if __name__ == '__main__':
    main()
