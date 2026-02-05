"""
IEE Completo - Folha de cálculo completa para certificação energética SCE

Cria uma folha Excel profissional e completa que:
- Recebe dados de duas simulações HAP (PREV e REF)
- Inclui TODAS as folhas auxiliares de cálculo
- Calcula IEEprev, IEEref, IEEren, RIEE
- Determina classe energética
- TUDO com fórmulas para rastreabilidade total

Folhas criadas:
 1. Detalhe PREV       - Dados brutos HAP (VALORES)
 2. Detalhe REF        - Dados brutos HAP (VALORES)
 3. Mensal PREV        - Dados mensais (VALORES)
 4. Mensal REF         - Dados mensais (VALORES)
 5. Iluminação ENU     - Cálculo iluminação ENU e exterior (INPUT + FÓRMULAS)
 6. AQS                - Cálculo águas quentes sanitárias (INPUT + FÓRMULAS)
 7. PV                 - Produção fotovoltaica (INPUT + FÓRMULAS)
 8. Equipamentos Extra - Equipamentos não simulados (INPUT + FÓRMULAS)
 9. Elevadores         - Cálculo energia elevadores (INPUT + FÓRMULAS)
10. Ventilação Extra   - Ventilação tipo S adicional (INPUT + FÓRMULAS)
11. Bombagem           - Bombas AQS e AVAC (INPUT + FÓRMULAS)
12. Consumos Tipologia - Resumo por categoria (FÓRMULAS)
13. Energia Primária   - Conversão kWh→kWhEP (FÓRMULAS)
14. Renováveis         - Contribuição renovável (FÓRMULAS)
15. IEE                - Indicadores (FÓRMULAS)
16. Classe             - Resultado final (FÓRMULAS)
17. Legenda            - Instruções e referências

Usage:
    python iee_completo.py <pasta_prev> <pasta_ref> [output.xlsx]
"""

import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# ESTILOS
# =============================================================================

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10)

SECTION_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
SECTION_FONT = Font(bold=True, size=10)

TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_FONT = Font(bold=True)

FORMULA_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RESULT_FILL = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
PREV_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
REF_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

# Cores especiais para folhas
TAB_GREEN = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
TAB_PURPLE = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
TAB_ORANGE = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

THICK_BORDER = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)

CLASS_COLORS = {
    'A+': '00A651', 'A': '50B848', 'B': 'B5D334',
    'B-': 'FFF200', 'C': 'F7941D', 'D': 'F15A29',
    'E': 'ED1C24', 'F': 'BE1E2D'
}

# =============================================================================
# CONSTANTES SCE
# =============================================================================

FPU = {
    'electricidade': 2.5,
    'gas_natural': 1.0,
}

RIEE_LIMITES = [
    ('A+', 0, 0.25),
    ('A', 0.26, 0.50),
    ('B', 0.51, 0.75),
    ('B-', 0.76, 1.00),
    ('C', 1.01, 1.50),
    ('D', 1.51, 2.00),
    ('E', 2.01, 2.50),
    ('F', 2.51, 999),
]

DETALHE_COLUMNS = [
    ('Lighting (kWh)', 'Lighting', 'B'),
    ('Electric Equipment (kWh)', 'Electric Equipment', 'C'),
    ('Central Unit Clg Input (kWh)', 'Central Clg Input', 'D'),
    ('Terminal Unit Clg Input (kWh)', 'Terminal Clg Input', 'E'),
    ('Central Unit Htg Input (kWh)', 'Central Htg Input', 'F'),
    ('Terminal Unit Htg Input (kWh)', 'Terminal Htg Input', 'G'),
    ('Central Unit Aux. Htg. Input (kWh)', 'Central Aux Htg', 'H'),
    ('Terminal Unit Aux. Htg. Input (kWh)', 'Terminal Aux Htg', 'I'),
    ('Supply Fan (kWh)', 'Supply Fan', 'J'),
    ('Return Fan (kWh)', 'Return Fan', 'K'),
    ('Exhaust Fan (kWh)', 'Exhaust Fan', 'L'),
    ('Ventilation Fan (kWh)', 'Ventilation Fan', 'M'),
    ('Central Cooling Coil Load (kWh)', 'Central Clg Load', 'N'),
    ('Central Heating Coil Load (kWh)', 'Central Htg Load', 'O'),
    ('Terminal Cooling Coil Load (kWh)', 'Terminal Clg Load', 'P'),
    ('Terminal Heating Coil Load (kWh)', 'Terminal Htg Load', 'Q'),
]

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTHS_FULL = ['January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']


# =============================================================================
# FUNÇÕES LEITURA CSV
# =============================================================================

def read_hap_csv(filepath):
    """Lê um CSV do HAP e retorna dicionário com totais anuais e dados mensais"""
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if len(lines) < 5:
        return None, {}

    sistema = lines[1].split(';')[0].replace('Monthly Simulation Results for ', '').strip()
    headers = lines[3].strip().split(';')

    totals = {}
    monthly_data = []

    for line in lines[4:]:
        line = line.strip()
        if not line or line.startswith('Month'):
            continue

        values = line.split(';')
        month = values[0] if values else ''

        month_data = {'Month': month}
        for i, val in enumerate(values[1:], 1):
            if i < len(headers) and val:
                col = headers[i]
                try:
                    num_val = int(val)
                    totals[col] = totals.get(col, 0) + num_val
                    month_data[col] = num_val
                except:
                    pass

        if month:
            monthly_data.append(month_data)

    return sistema, {'totals': totals, 'monthly': monthly_data, 'headers': headers}


def find_hap_csvs(project_folder):
    """Encontra todos os CSVs do HAP numa pasta"""
    pattern = os.path.join(project_folder, 'HAP51_Monthly_*.csv')
    return sorted(glob.glob(pattern))


def load_project_data(project_folder):
    """Carrega todos os dados de um projecto HAP"""
    csv_files = find_hap_csvs(project_folder)
    systems_data = {}

    for csv_file in csv_files:
        sistema, data = read_hap_csv(csv_file)
        if sistema and data and sistema != 'TODOS':
            systems_data[sistema] = data

    return systems_data


# =============================================================================
# FOLHAS DETALHE E MENSAL
# =============================================================================

def create_detalhe_sheet(ws, data, title):
    """Cria folha Detalhe com dados brutos anuais"""
    sistemas = list(data.keys())
    num_sistemas = len(sistemas)

    ws.merge_cells('A1:Q1')
    ws.cell(1, 1, value=f'DADOS {title} - Consumos Anuais por Sistema').font = Font(bold=True, size=14)

    ws.cell(2, 1, value='Sistema').font = HEADER_FONT
    ws.cell(2, 1).fill = HEADER_FILL
    ws.cell(2, 1).border = THIN_BORDER

    for csv_col, nome, col_letter in DETALHE_COLUMNS:
        col_num = ord(col_letter) - ord('A') + 1
        cell = ws.cell(2, col_num, value=nome)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, sistema in enumerate(sistemas):
        row = 3 + i
        ws.cell(row, 1, value=sistema).border = THIN_BORDER

        for csv_col, nome, col_letter in DETALHE_COLUMNS:
            col_num = ord(col_letter) - ord('A') + 1
            value = data[sistema]['totals'].get(csv_col, 0)
            cell = ws.cell(row, col_num, value=value if value > 0 else '')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

    row_total = 3 + num_sistemas
    ws.cell(row_total, 1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row_total, 1).fill = TOTAL_FILL
    ws.cell(row_total, 1).border = THIN_BORDER

    for csv_col, nome, col_letter in DETALHE_COLUMNS:
        col_num = ord(col_letter) - ord('A') + 1
        formula = f'=SUM({col_letter}3:{col_letter}{row_total-1})'
        cell = ws.cell(row_total, col_num, value=formula)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    ws.column_dimensions['A'].width = 22
    for csv_col, nome, col_letter in DETALHE_COLUMNS:
        ws.column_dimensions[col_letter].width = 12

    ws.freeze_panes = 'B3'
    return row_total


def create_mensal_sheet(ws, data, title):
    """Cria folha com dados mensais"""
    sistemas = list(data.keys())

    ws.merge_cells('A1:O1')
    ws.cell(1, 1, value=f'DADOS {title} - Consumos Mensais').font = Font(bold=True, size=14)

    ws.cell(2, 1, value='Sistema').font = HEADER_FONT
    ws.cell(2, 1).fill = HEADER_FILL
    ws.cell(2, 2, value='Tipo').font = HEADER_FONT
    ws.cell(2, 2).fill = HEADER_FILL

    for col, month in enumerate(MONTHS, 3):
        cell = ws.cell(2, col, value=month)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    ws.cell(2, 15, value='Total').font = HEADER_FONT
    ws.cell(2, 15).fill = HEADER_FILL

    cols_mensal = [
        ('Lighting (kWh)', 'Lighting'),
        ('Electric Equipment (kWh)', 'Equipment'),
        ('Central Unit Clg Input (kWh)', 'Cooling'),
        ('Central Unit Htg Input (kWh)', 'Heating'),
        ('Supply Fan (kWh)', 'Supply Fan'),
    ]

    row = 3
    for sistema in sistemas:
        monthly = data[sistema].get('monthly', [])
        if not monthly:
            continue

        for csv_col, nome in cols_mensal:
            if data[sistema]['totals'].get(csv_col, 0) == 0:
                continue

            ws.cell(row, 1, value=sistema).border = THIN_BORDER
            ws.cell(row, 2, value=nome).border = THIN_BORDER

            for col, month_full in enumerate(MONTHS_FULL, 3):
                for m_data in monthly:
                    if m_data.get('Month', '').startswith(month_full[:3]):
                        val = m_data.get(csv_col, 0)
                        cell = ws.cell(row, col, value=val if val > 0 else '')
                        cell.border = THIN_BORDER
                        break

            cell = ws.cell(row, 15, value=f'=SUM(C{row}:N{row})')
            cell.border = THIN_BORDER
            cell.fill = FORMULA_FILL
            row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 8

    ws.freeze_panes = 'C3'


# =============================================================================
# FOLHAS AUXILIARES DE CÁLCULO
# =============================================================================

def create_iluminacao_enu_sheet(ws):
    """Cria folha de Iluminação ENU e Exterior"""

    # Título
    ws.merge_cells('A1:H1')
    ws.cell(1, 1, value='ILUMINAÇÃO - Espaços Não Úteis (ENU) e Exterior').font = Font(bold=True, size=14)

    # Descrição
    ws.cell(2, 1, value='Iluminação não incluída na simulação HAP (armazéns, zonas técnicas, exterior)').font = Font(italic=True)

    # Secção ENU
    ws.cell(4, 1, value='ILUMINAÇÃO ENU (Tipo B)').font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL
    ws.merge_cells('A4:D4')

    headers = ['Espaço', 'Potência [W]', 'Horas/ano', 'Energia [kWh/ano]']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(5, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    espacos_enu = ['Armazéns', 'Zona Técnica', 'Cozinha ENU', 'Garagem', 'Outros ENU']
    row = 6
    for espaco in espacos_enu:
        ws.cell(row, 1, value=espaco).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}/1000')
        ws.cell(row, 4).fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        row += 1

    row_total_enu = row
    ws.cell(row, 1, value='TOTAL ENU').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 1).border = THIN_BORDER
    ws.cell(row, 4, value=f'=SUM(D6:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THIN_BORDER

    # Secção Exterior
    row += 2
    ws.cell(row, 1, value='ILUMINAÇÃO EXTERIOR').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    espacos_ext = ['Estacionamento', 'Jardins', 'Fachada', 'Sinalização', 'Outros Exterior']
    row_start_ext = row
    for espaco in espacos_ext:
        ws.cell(row, 1, value=espaco).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}/1000')
        ws.cell(row, 4).fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        row += 1

    row_total_ext = row
    ws.cell(row, 1, value='TOTAL EXTERIOR').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 1).border = THIN_BORDER
    ws.cell(row, 4, value=f'=SUM(D{row_start_ext}:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THIN_BORDER

    # Resumo
    row += 2
    ws.cell(row, 1, value='RESUMO ILUMINAÇÃO').font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, value='Iluminação ENU (Tipo B)')
    ws.cell(row, 4, value=f'=D{row_total_enu}')
    ws.cell(row, 4).fill = RESULT_FILL
    ws.cell(row, 4).border = THICK_BORDER
    ws.cell(row, 5, value='kWh/ano')
    row += 1
    ws.cell(row, 1, value='Iluminação Exterior')
    ws.cell(row, 4, value=f'=D{row_total_ext}')
    ws.cell(row, 4).fill = RESULT_FILL
    ws.cell(row, 4).border = THICK_BORDER
    ws.cell(row, 5, value='kWh/ano')

    # Nota
    row += 2
    ws.cell(row, 1, value='Nota: Iluminação exterior conta como Tipo S desde 2016').font = Font(italic=True, color='666666')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10

    return row_total_enu, row_total_ext


def create_aqs_sheet(ws):
    """Cria folha de cálculo AQS (Águas Quentes Sanitárias)"""

    ws.merge_cells('A1:J1')
    ws.cell(1, 1, value='AQS - Águas Quentes Sanitárias').font = Font(bold=True, size=14)

    # Parâmetros gerais
    ws.cell(3, 1, value='PARÂMETROS GERAIS').font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    ws.merge_cells('A3:D3')

    params = [
        ('Número de pessoas (ocupação)', 'B5', '', 'pessoas'),
        ('Consumo litros/pessoa/dia', 'B6', '40', 'L/pessoa/dia'),
        ('Número de dias/ano', 'B7', '365', 'dias'),
        ('Temperatura água fria', 'B8', '15', '°C'),
        ('Temperatura AQS', 'B9', '60', '°C'),
    ]

    row = 5
    for nome, cell_ref, default, unit in params:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        if default:
            ws.cell(row, 2, value=float(default) if '.' in default else int(default))
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=unit)
        row += 1

    # Cálculo necessidades
    row += 1
    ws.cell(row, 1, value='CÁLCULO NECESSIDADES AQS').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row, 1, value='Volume diário')
    ws.cell(row, 2, value='=B5*B6')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 3, value='L/dia')
    row += 1

    ws.cell(row, 1, value='Volume anual')
    ws.cell(row, 2, value='=B5*B6*B7')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 3, value='L/ano')
    row += 1

    ws.cell(row, 1, value='Qútil (energia útil)')
    ws.cell(row, 2, value='=B5*B6*B7*4.186*(B9-B8)/3600')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_qutil = row
    row += 1

    # Sistema de produção
    row += 1
    ws.cell(row, 1, value='SISTEMA DE PRODUÇÃO AQS').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row, 1, value='Contribuição Solar Térmico')
    ws.cell(row, 2).fill = INPUT_FILL
    ws.cell(row, 2).border = THIN_BORDER
    ws.cell(row, 3, value='kWh/ano')
    ws.cell(row, 4, value='← Do SCE.ER ou simulação')
    row_solar = row
    row += 1

    ws.cell(row, 1, value='Necessidades após ST')
    ws.cell(row, 2, value=f'=B{row_qutil}-B{row_solar}')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 3, value='kWh/ano')
    row_apos_st = row
    row += 1

    row += 1
    ws.cell(row, 1, value='Sistema de apoio:').font = Font(bold=True)
    row += 1

    sistemas = [
        ('Bomba de Calor (COP)', 'B' + str(row), '3.5', 'EE'),
        ('Caldeira a Gás (η)', 'B' + str(row+1), '0.92', 'GN'),
        ('Resistência Elétrica (η)', 'B' + str(row+2), '1.0', 'EE'),
    ]

    row_sistema_start = row
    for nome, cell, default, fonte in sistemas:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=float(default))
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value='%' if float(default) < 1 else '')
        ws.cell(row, 4, value=f'Fonte: {fonte}')
        row += 1

    row += 1
    ws.cell(row, 1, value='Selecionar sistema (1, 2 ou 3):')
    ws.cell(row, 2, value=1)
    ws.cell(row, 2).fill = INPUT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    row_selecao = row
    row += 1

    # Resultados
    row += 1
    ws.cell(row, 1, value='RESULTADOS AQS').font = SECTION_FONT
    ws.cell(row, 1).fill = RESULT_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row, 1, value='Rendimento selecionado')
    ws.cell(row, 2, value=f'=INDEX(B{row_sistema_start}:B{row_sistema_start+2},B{row_selecao})')
    ws.cell(row, 2).fill = FORMULA_FILL
    row_rend = row
    row += 1

    ws.cell(row, 1, value='Energia Final AQS')
    ws.cell(row, 2, value=f'=ROUND(B{row_apos_st}/B{row_rend},0)')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_ef = row
    row += 1

    ws.cell(row, 1, value='Energia Renovável (Aerotermia)')
    ws.cell(row, 2, value=f'=IF(B{row_selecao}=1,B{row_apos_st}-B{row_ef},0)')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_ren_aero = row
    row += 1

    # Resumo final
    row += 2
    ws.cell(row, 1, value='RESUMO PARA DESAGREGAÇÃO').font = Font(bold=True, size=12)
    row += 1

    resumo = [
        ('AQS - Electricidade (EE)', f'=IF(OR(B{row_selecao}=1,B{row_selecao}=3),B{row_ef},0)'),
        ('AQS - Gás Natural (GN)', f'=IF(B{row_selecao}=2,B{row_ef},0)'),
        ('AQS - Renovável Aerotermia', f'=B{row_ren_aero}'),
        ('AQS - Renovável Solar Térmico', f'=B{row_solar}'),
    ]

    row_resumo_start = row
    for nome, formula in resumo:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=formula)
        ws.cell(row, 2).fill = RESULT_FILL
        ws.cell(row, 2).border = THICK_BORDER
        ws.cell(row, 2).number_format = '0'
        ws.cell(row, 3, value='kWh/ano')
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25

    return row_resumo_start


def create_pv_sheet(ws):
    """Cria folha de Produção Fotovoltaica"""

    ws.merge_cells('A1:F1')
    ws.cell(1, 1, value='PV - Produção Fotovoltaica').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='PARÂMETROS DO SISTEMA PV').font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    ws.merge_cells('A3:D3')

    params = [
        ('Potência instalada', '', 'kWp'),
        ('Horas equivalentes (HSP)', '1400', 'h/ano'),
        ('Perdas sistema (%)', '15', '%'),
        ('Factor de auto-consumo', '0.8', ''),
    ]

    row = 5
    for nome, default, unit in params:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        if default:
            ws.cell(row, 2, value=float(default) if '.' in default else int(default))
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=unit)
        row += 1

    row += 1
    ws.cell(row, 1, value='CÁLCULO PRODUÇÃO').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws.cell(row, 1, value='Produção bruta')
    ws.cell(row, 2, value='=B5*B6')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row += 1

    ws.cell(row, 1, value='Produção líquida')
    ws.cell(row, 2, value='=B5*B6*(1-B7/100)')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_prod = row
    row += 1

    ws.cell(row, 1, value='Auto-consumo')
    ws.cell(row, 2, value=f'=B{row_prod}*B8')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_auto = row
    row += 1

    ws.cell(row, 1, value='Exportado para rede')
    ws.cell(row, 2, value=f'=B{row_prod}-B{row_auto}')
    ws.cell(row, 2).fill = FORMULA_FILL
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row += 2

    ws.cell(row, 1, value='RESULTADO PARA IEE').font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, value='PV Auto-consumo (Renovável)')
    ws.cell(row, 2, value=f'=B{row_auto}')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_resultado = row

    ws.cell(row+2, 1, value='Nota: Só o auto-consumo conta para o IEE').font = Font(italic=True, color='666666')

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12

    return row_resultado


def create_equipamentos_extra_sheet(ws):
    """Cria folha de Equipamentos Extra"""

    ws.merge_cells('A1:F1')
    ws.cell(1, 1, value='EQUIPAMENTOS EXTRA - Não incluídos na simulação').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='EQUIPAMENTOS ELÉCTRICOS (Tipo T)').font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    ws.merge_cells('A3:E3')

    headers = ['Equipamento', 'Potência [W]', 'Horas/ano', 'Quantidade', 'Energia [kWh/ano]']
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    equipamentos_ee = ['Cozinha Industrial', 'Frio Comercial', 'Computadores Extra',
                       'Servidores', 'Outros EE']
    row = 5
    for eq in equipamentos_ee:
        ws.cell(row, 1, value=eq).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f'=B{row}*C{row}*D{row}/1000')
        ws.cell(row, 5).fill = FORMULA_FILL
        ws.cell(row, 5).border = THIN_BORDER
        row += 1

    row_total_ee = row
    ws.cell(row, 1, value='TOTAL EE Extra').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 5, value=f'=SUM(E5:E{row-1})').font = TOTAL_FONT
    ws.cell(row, 5).fill = TOTAL_FILL
    ws.cell(row, 5).border = THIN_BORDER

    row += 2
    ws.cell(row, 1, value='EQUIPAMENTOS A GÁS NATURAL (Tipo T)').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    headers_gn = ['Equipamento', 'Consumo [kWh/ano]', '', '', '']
    for col, h in enumerate(headers_gn, 1):
        if h:
            cell = ws.cell(row, col, value=h)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER
    row += 1

    equipamentos_gn = ['Cozinha (fogões)', 'Outros GN']
    row_start_gn = row
    for eq in equipamentos_gn:
        ws.cell(row, 1, value=eq).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        row += 1

    row_total_gn = row
    ws.cell(row, 1, value='TOTAL GN Extra').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 2, value=f'=SUM(B{row_start_gn}:B{row-1})').font = TOTAL_FONT
    ws.cell(row, 2).fill = TOTAL_FILL
    ws.cell(row, 2).border = THIN_BORDER

    row += 2
    ws.cell(row, 1, value='RESUMO EQUIPAMENTOS EXTRA').font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, value='Equipamentos EE (Tipo T)')
    ws.cell(row, 2, value=f'=E{row_total_ee}')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 3, value='kWh/ano')
    row += 1
    ws.cell(row, 1, value='Equipamentos GN (Tipo T)')
    ws.cell(row, 2, value=f'=B{row_total_gn}')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 3, value='kWh/ano')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    return row_total_ee, row_total_gn


def create_elevadores_sheet(ws):
    """Cria folha de cálculo Elevadores (fórmula SCE)"""

    ws.merge_cells('A1:H1')
    ws.cell(1, 1, value='ELEVADORES - Cálculo segundo SCE/RECS').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='Fórmula: Edasc = (1.08 × Qasc × na × sm) / 1000 + 100 × 21').font = Font(italic=True)
    ws.cell(4, 1, value='Energia anual = Edasc × dias_ano / 1000').font = Font(italic=True)

    # Elevador 1
    ws.cell(6, 1, value='ELEVADOR 1').font = SECTION_FONT
    ws.cell(6, 1).fill = SECTION_FILL
    ws.merge_cells('A6:D6')

    params_elev = [
        ('Qasc - Carga nominal', '', 'kg'),
        ('na - Viagens/ano', '750', 'viagens'),
        ('p - Paragens médias/viagem', '1', ''),
        ('lm - Percurso médio', '', 'm'),
        ('sm = p × lm', '=B9*B10', 'm'),
        ('Edasc (energia diária)', '=1.08*B7*B8*B11/1000+100*21', 'Wh/dia'),
        ('daasc - Dias activos/ano', '365', 'dias'),
        ('Energia anual', '=B12*B13/1000', 'kWh/ano'),
    ]

    row = 7
    for nome, formula, unit in params_elev:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        if formula.startswith('='):
            ws.cell(row, 2, value=formula)
            ws.cell(row, 2).fill = FORMULA_FILL
        elif formula:
            ws.cell(row, 2, value=int(formula) if formula.isdigit() else float(formula))
            ws.cell(row, 2).fill = INPUT_FILL
        else:
            ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=unit)
        row += 1
    row_elev1 = row - 1

    # Elevador 2
    row += 1
    ws.cell(row, 1, value='ELEVADOR 2').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    params_elev2 = [
        ('Qasc - Carga nominal', '', 'kg'),
        ('na - Viagens/ano', '750', 'viagens'),
        ('p - Paragens médias/viagem', '0.67', ''),
        ('lm - Percurso médio', '', 'm'),
        ('sm = p × lm', f'=B{row+2}*B{row+3}', 'm'),
        ('Edasc (energia diária)', f'=1.08*B{row}*B{row+1}*B{row+4}/1000+100*21', 'Wh/dia'),
        ('daasc - Dias activos/ano', '365', 'dias'),
        ('Energia anual', f'=B{row+5}*B{row+6}/1000', 'kWh/ano'),
    ]

    row_start_e2 = row
    for nome, formula, unit in params_elev2:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        if formula.startswith('='):
            ws.cell(row, 2, value=formula)
            ws.cell(row, 2).fill = FORMULA_FILL
        elif formula:
            ws.cell(row, 2, value=int(formula) if formula.isdigit() else float(formula))
            ws.cell(row, 2).fill = INPUT_FILL
        else:
            ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=unit)
        row += 1
    row_elev2 = row - 1

    # Total
    row += 2
    ws.cell(row, 1, value='TOTAL ELEVADORES').font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, value='Energia Total Elevadores')
    ws.cell(row, 2, value=f'=B{row_elev1}+B{row_elev2}')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_total = row

    # Notas
    row += 2
    ws.cell(row, 1, value='LEGENDA:').font = Font(bold=True)
    row += 1
    ws.cell(row, 1, value='Qasc = Carga nominal do ascensor [kg]').font = Font(size=9)
    row += 1
    ws.cell(row, 1, value='na = Número de viagens por ano (default: 750 para baixo uso)').font = Font(size=9)
    row += 1
    ws.cell(row, 1, value='p = Número de paragens médio por viagem').font = Font(size=9)
    row += 1
    ws.cell(row, 1, value='lm = Percurso médio de viagem [m]').font = Font(size=9)
    row += 1
    ws.cell(row, 1, value='sm = Distância percorrida por viagem = p × lm [m]').font = Font(size=9)
    row += 1
    ws.cell(row, 1, value='21 = Potência em standby estimada [W]').font = Font(size=9)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12

    return row_total


def create_ventilacao_extra_sheet(ws):
    """Cria folha de Ventilação Extra Tipo S"""

    ws.merge_cells('A1:F1')
    ws.cell(1, 1, value='VENTILAÇÃO EXTRA - Sistemas Tipo S não simulados').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='VENTILAÇÃO INSUFLAÇÃO/EXTRAÇÃO (Tipo S)').font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    ws.merge_cells('A3:E3')

    headers = ['Sistema', 'Potência [kW]', 'Horas/dia', 'Dias/ano', 'Energia [kWh/ano]']
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    sistemas = ['Insuflação IS', 'Extração WC', 'Extração Cozinha',
                'Ventilação Industrial', 'Outros Vent S']
    row = 5
    for sist in sistemas:
        ws.cell(row, 1, value=sist).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f'=B{row}*C{row}*D{row}')
        ws.cell(row, 5).fill = FORMULA_FILL
        ws.cell(row, 5).border = THIN_BORDER
        row += 1

    row_total = row
    ws.cell(row, 1, value='TOTAL Ventilação Extra').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 5, value=f'=SUM(E5:E{row-1})').font = TOTAL_FONT
    ws.cell(row, 5).fill = TOTAL_FILL
    ws.cell(row, 5).border = THIN_BORDER

    row += 2
    ws.cell(row, 1, value='RESULTADO').font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, value='Ventilação Extra (Tipo S)')
    ws.cell(row, 2, value=f'=E{row_total}')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_resultado = row

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    return row_resultado


def create_bombagem_sheet(ws):
    """Cria folha de Bombagem"""

    ws.merge_cells('A1:F1')
    ws.cell(1, 1, value='BOMBAGEM - Bombas não incluídas na simulação').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='BOMBAS AQS E OUTRAS (Tipo S)').font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    ws.merge_cells('A3:E3')

    headers = ['Sistema', 'Potência [kW]', 'Horas/dia', 'Dias/ano', 'Energia [kWh/ano]']
    row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    bombas = ['Bombas AQS circulação', 'Bombas AQS solar', 'Bombas piscina',
              'Bombas AVAC extra', 'Outras bombas']
    row = 5
    for bomba in bombas:
        ws.cell(row, 1, value=bomba).border = THIN_BORDER
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3).fill = INPUT_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4).fill = INPUT_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=f'=B{row}*C{row}*D{row}')
        ws.cell(row, 5).fill = FORMULA_FILL
        ws.cell(row, 5).border = THIN_BORDER
        row += 1

    row_total = row
    ws.cell(row, 1, value='TOTAL Bombagem').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 5, value=f'=SUM(E5:E{row-1})').font = TOTAL_FONT
    ws.cell(row, 5).fill = TOTAL_FILL
    ws.cell(row, 5).border = THIN_BORDER

    row += 2
    ws.cell(row, 1, value='RESULTADO').font = Font(bold=True, size=12)
    row += 1
    ws.cell(row, 1, value='Bombagem Total (Tipo S)')
    ws.cell(row, 2, value=f'=E{row_total}')
    ws.cell(row, 2).fill = RESULT_FILL
    ws.cell(row, 2).border = THICK_BORDER
    ws.cell(row, 2).number_format = '0'
    ws.cell(row, 3, value='kWh/ano')
    row_resultado = row

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    return row_resultado


def create_consumos_tipologia_sheet(ws, row_total_prev, row_total_ref):
    """Cria folha de Consumos por Tipologia - resume tudo"""

    ws.merge_cells('A1:H1')
    ws.cell(1, 1, value='CONSUMOS POR TIPOLOGIA - Desagregação Completa').font = Font(bold=True, size=14)

    # Headers
    headers = ['Utilização', 'Simulação HAP', 'Extra', 'TOTAL', 'Tipo', 'Fonte']
    row = 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # ELECTRICIDADE
    row = 4
    ws.cell(row, 1, value='ELECTRICIDADE').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    consumos_ee = [
        ('Iluminação Interior', f"='Detalhe PREV'!B{row_total_prev}", '', 'S'),
        ('Iluminação ENU', '', "='Iluminação ENU'!D11", 'S'),
        ('Iluminação Exterior', '', "='Iluminação ENU'!D19", 'S'),
        ('Aquecimento', f"='Detalhe PREV'!F{row_total_prev}+'Detalhe PREV'!G{row_total_prev}+'Detalhe PREV'!H{row_total_prev}+'Detalhe PREV'!I{row_total_prev}", '', 'S'),
        ('Arrefecimento', f"='Detalhe PREV'!D{row_total_prev}+'Detalhe PREV'!E{row_total_prev}", '', 'S'),
        ('Ventilação (simulada)', f"='Detalhe PREV'!J{row_total_prev}+'Detalhe PREV'!K{row_total_prev}+'Detalhe PREV'!L{row_total_prev}+'Detalhe PREV'!M{row_total_prev}", '', 'S'),
        ('Ventilação (extra)', '', "='Ventilação Extra'!B13", 'S'),
        ('Bombagem', '', "='Bombagem'!B13", 'S'),
        ('AQS (EE)', '', "='AQS'!B37", 'S'),
        ('Elevadores', '', "='Elevadores'!B30", 'S'),
        ('Equipamentos (simulados)', f"='Detalhe PREV'!C{row_total_prev}", '', 'T'),
        ('Equipamentos (extra)', '', "='Equipamentos Extra'!E10", 'T'),
    ]

    row_start_ee = row
    for nome, formula_hap, formula_extra, tipo in consumos_ee:
        ws.cell(row, 1, value=nome).border = THIN_BORDER

        if formula_hap:
            ws.cell(row, 2, value=formula_hap)
            ws.cell(row, 2).fill = FORMULA_FILL
        else:
            ws.cell(row, 2, value=0)
        ws.cell(row, 2).border = THIN_BORDER

        if formula_extra:
            ws.cell(row, 3, value=formula_extra)
            ws.cell(row, 3).fill = FORMULA_FILL
        else:
            ws.cell(row, 3, value=0)
        ws.cell(row, 3).border = THIN_BORDER

        ws.cell(row, 4, value=f'=B{row}+C{row}')
        ws.cell(row, 4).fill = TOTAL_FILL
        ws.cell(row, 4).border = THIN_BORDER

        ws.cell(row, 5, value=tipo).border = THIN_BORDER
        ws.cell(row, 6, value='EE').border = THIN_BORDER
        row += 1
    row_end_ee = row - 1

    # Total EE
    ws.cell(row, 1, value='TOTAL ELECTRICIDADE').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start_ee}:D{row_end_ee})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ee = row
    row += 2

    # GÁS NATURAL
    ws.cell(row, 1, value='GÁS NATURAL').font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    consumos_gn = [
        ('AQS (GN)', '', "='AQS'!B38", 'S'),
        ('Equipamentos (GN)', '', "='Equipamentos Extra'!B16", 'T'),
    ]

    row_start_gn = row
    for nome, formula_hap, formula_extra, tipo in consumos_gn:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=0).border = THIN_BORDER

        if formula_extra:
            ws.cell(row, 3, value=formula_extra)
            ws.cell(row, 3).fill = FORMULA_FILL
        else:
            ws.cell(row, 3, value=0)
        ws.cell(row, 3).border = THIN_BORDER

        ws.cell(row, 4, value=f'=B{row}+C{row}')
        ws.cell(row, 4).fill = TOTAL_FILL
        ws.cell(row, 4).border = THIN_BORDER

        ws.cell(row, 5, value=tipo).border = THIN_BORDER
        ws.cell(row, 6, value='GN').border = THIN_BORDER
        row += 1
    row_end_gn = row - 1

    # Total GN
    ws.cell(row, 1, value='TOTAL GÁS NATURAL').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start_gn}:D{row_end_gn})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_gn = row
    row += 2

    # RENOVÁVEIS
    ws.cell(row, 1, value='RENOVÁVEIS').font = SECTION_FONT
    ws.cell(row, 1).fill = TAB_GREEN
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    renovaveis = [
        ('Solar Térmico (AQS)', "='AQS'!B40"),
        ('Aerotermia (AQS)', "='AQS'!B39"),
        ('Fotovoltaico (auto-consumo)', "='PV'!B17"),
    ]

    row_start_ren = row
    for nome, formula in renovaveis:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=0).border = THIN_BORDER
        ws.cell(row, 3, value=formula)
        ws.cell(row, 3).fill = FORMULA_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.cell(row, 4, value=f'=C{row}')
        ws.cell(row, 4).fill = TOTAL_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value='REN').border = THIN_BORDER
        ws.cell(row, 6, value='REN').border = THIN_BORDER
        row += 1
    row_end_ren = row - 1

    # Total REN
    ws.cell(row, 1, value='TOTAL RENOVÁVEIS').font = TOTAL_FONT
    ws.cell(row, 1).fill = TAB_GREEN
    ws.cell(row, 4, value=f'=SUM(D{row_start_ren}:D{row_end_ren})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TAB_GREEN
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ren = row

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8

    return row_total_ee, row_total_gn, row_total_ren, row_start_ee, row_end_ee, row_start_gn, row_end_gn


# =============================================================================
# ENERGIA PRIMÁRIA E IEE
# =============================================================================

def create_energia_primaria_sheet(ws, consumos_info):
    """Cria folha de conversão para Energia Primária"""

    row_total_ee, row_total_gn, row_total_ren, row_start_ee, row_end_ee, row_start_gn, row_end_gn = consumos_info

    ws.merge_cells('A1:G1')
    ws.cell(1, 1, value='ENERGIA PRIMÁRIA - Conversão kWh → kWhEP').font = Font(bold=True, size=14)

    # Área útil
    ws.cell(3, 1, value='Área Útil [m²]:').font = Font(bold=True)
    ws.cell(3, 2).fill = INPUT_FILL
    ws.cell(3, 2).border = THICK_BORDER
    ws.cell(3, 3, value='← PREENCHER').font = Font(italic=True, color='FF0000')

    # Factores Fpu
    ws.cell(5, 1, value='FACTORES DE CONVERSÃO (Fpu)').font = SECTION_FONT
    ws.cell(5, 1).fill = SECTION_FILL

    ws.cell(6, 1, value='Electricidade')
    ws.cell(6, 2, value=FPU['electricidade'])
    ws.cell(6, 3, value='kWhEP/kWh')

    ws.cell(7, 1, value='Gás Natural')
    ws.cell(7, 2, value=FPU['gas_natural'])
    ws.cell(7, 3, value='kWhEP/kWh')

    # Tabela PREV
    ws.cell(9, 1, value='ENERGIA PRIMÁRIA - PREVISTO').font = SECTION_FONT
    ws.cell(9, 1).fill = PREV_FILL
    ws.merge_cells('A9:E9')

    headers = ['Utilização', 'Energia Final (kWh)', 'Fpu', 'Energia Primária (kWhEP)', 'Tipo']
    row = 10
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    # Consumos de Consumos por Tipologia
    consumos = [
        ('Iluminação Interior', f"='Consumos Tipologia'!D{row_start_ee}", '$B$6', 'S'),
        ('Iluminação ENU', f"='Consumos Tipologia'!D{row_start_ee+1}", '$B$6', 'S'),
        ('Iluminação Exterior', f"='Consumos Tipologia'!D{row_start_ee+2}", '$B$6', 'S'),
        ('Aquecimento', f"='Consumos Tipologia'!D{row_start_ee+3}", '$B$6', 'S'),
        ('Arrefecimento', f"='Consumos Tipologia'!D{row_start_ee+4}", '$B$6', 'S'),
        ('Ventilação (simulada)', f"='Consumos Tipologia'!D{row_start_ee+5}", '$B$6', 'S'),
        ('Ventilação (extra)', f"='Consumos Tipologia'!D{row_start_ee+6}", '$B$6', 'S'),
        ('Bombagem', f"='Consumos Tipologia'!D{row_start_ee+7}", '$B$6', 'S'),
        ('AQS (EE)', f"='Consumos Tipologia'!D{row_start_ee+8}", '$B$6', 'S'),
        ('Elevadores', f"='Consumos Tipologia'!D{row_start_ee+9}", '$B$6', 'S'),
        ('Equipamentos (EE)', f"='Consumos Tipologia'!D{row_start_ee+10}+'Consumos Tipologia'!D{row_start_ee+11}", '$B$6', 'T'),
        ('AQS (GN)', f"='Consumos Tipologia'!D{row_start_gn}", '$B$7', 'S'),
        ('Equipamentos (GN)', f"='Consumos Tipologia'!D{row_start_gn+1}", '$B$7', 'T'),
    ]

    row_start = row
    for nome, formula, fpu_ref, tipo in consumos:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=formula)
        ws.cell(row, 2).fill = FORMULA_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=f'={fpu_ref}').border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}')
        ws.cell(row, 4).fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value=tipo).border = THIN_BORDER
        row += 1
    row_end = row - 1

    # Total PREV
    ws.cell(row, 1, value='TOTAL PREV').font = TOTAL_FONT
    ws.cell(row, 1).fill = TOTAL_FILL
    ws.cell(row, 2, value=f'=SUM(B{row_start}:B{row_end})').font = TOTAL_FONT
    ws.cell(row, 2).fill = TOTAL_FILL
    ws.cell(row, 4, value=f'=SUM(D{row_start}:D{row_end})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TOTAL_FILL
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ep = row
    row += 1

    # Total Tipo S PREV
    ws.cell(row, 1, value='TOTAL Tipo S PREV').font = Font(bold=True)
    ws.cell(row, 1).fill = PREV_FILL
    # Soma só tipo S (exclui equipamentos)
    tipo_s_rows = [row_start + i for i in range(10)] + [row_start + 11]  # todos menos equipamentos EE e GN
    formula_s = '+'.join([f'D{r}' for r in tipo_s_rows])
    ws.cell(row, 4, value=f'={formula_s}')
    ws.cell(row, 4).fill = PREV_FILL
    ws.cell(row, 4).font = Font(bold=True)
    ws.cell(row, 4).border = THICK_BORDER
    row_total_s_prev = row
    row += 2

    # Renováveis
    ws.cell(row, 1, value='ENERGIA RENOVÁVEL').font = SECTION_FONT
    ws.cell(row, 1).fill = TAB_GREEN
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    renovaveis = [
        ('Solar Térmico', f"='Consumos Tipologia'!D{row_end_gn+3}", '$B$6'),
        ('Aerotermia', f"='Consumos Tipologia'!D{row_end_gn+4}", '$B$6'),
        ('Fotovoltaico', f"='Consumos Tipologia'!D{row_end_gn+5}", '$B$6'),
    ]

    row_start_ren = row
    for nome, formula, fpu_ref in renovaveis:
        ws.cell(row, 1, value=nome).border = THIN_BORDER
        ws.cell(row, 2, value=formula)
        ws.cell(row, 2).fill = FORMULA_FILL
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, value=f'={fpu_ref}').border = THIN_BORDER
        ws.cell(row, 4, value=f'=B{row}*C{row}')
        ws.cell(row, 4).fill = FORMULA_FILL
        ws.cell(row, 4).border = THIN_BORDER
        ws.cell(row, 5, value='REN').border = THIN_BORDER
        row += 1

    ws.cell(row, 1, value='TOTAL RENOVÁVEL').font = TOTAL_FONT
    ws.cell(row, 1).fill = TAB_GREEN
    ws.cell(row, 4, value=f'=SUM(D{row_start_ren}:D{row-1})').font = TOTAL_FONT
    ws.cell(row, 4).fill = TAB_GREEN
    ws.cell(row, 4).border = THICK_BORDER
    row_total_ren = row

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 8

    return row_total_s_prev, row_total_ren


def create_iee_sheet(ws, row_total_s_prev, row_total_ren):
    """Cria folha de cálculo IEE"""

    ws.merge_cells('A1:E1')
    ws.cell(1, 1, value='INDICADORES DE EFICIÊNCIA ENERGÉTICA (IEE)').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='Indicador').font = HEADER_FONT
    ws.cell(3, 1).fill = HEADER_FILL
    ws.cell(3, 2, value='Fórmula').font = HEADER_FONT
    ws.cell(3, 2).fill = HEADER_FILL
    ws.cell(3, 3, value='Valor').font = HEADER_FONT
    ws.cell(3, 3).fill = HEADER_FILL
    ws.cell(3, 4, value='Unidade').font = HEADER_FONT
    ws.cell(3, 4).fill = HEADER_FILL

    # IEEprev,s
    ws.cell(4, 1, value='IEEprev,s').font = Font(bold=True)
    ws.cell(4, 2, value='= EP_prev_s / Área')
    ws.cell(4, 3, value=f"='Energia Primária'!D{row_total_s_prev}/'Energia Primária'!$B$3")
    ws.cell(4, 3).fill = RESULT_FILL
    ws.cell(4, 3).border = THICK_BORDER
    ws.cell(4, 3).number_format = '0.00'
    ws.cell(4, 4, value='kWhEP/m².ano')

    # IEEref,s (input manual)
    ws.cell(5, 1, value='IEEref,s').font = Font(bold=True)
    ws.cell(5, 2, value='= Do cálculo REF')
    ws.cell(5, 3).fill = INPUT_FILL
    ws.cell(5, 3).border = THICK_BORDER
    ws.cell(5, 3).number_format = '0.00'
    ws.cell(5, 4, value='kWhEP/m².ano ← PREENCHER')

    # IEEren
    ws.cell(6, 1, value='IEEren').font = Font(bold=True)
    ws.cell(6, 2, value='= EP_ren / Área')
    ws.cell(6, 3, value=f"='Energia Primária'!D{row_total_ren}/'Energia Primária'!$B$3")
    ws.cell(6, 3).fill = RESULT_FILL
    ws.cell(6, 3).border = THICK_BORDER
    ws.cell(6, 3).number_format = '0.00'
    ws.cell(6, 4, value='kWhEP/m².ano')

    # RIEE
    ws.cell(8, 1, value='RIEE').font = Font(bold=True, size=14)
    ws.cell(8, 2, value='= (IEEprev,s - IEEren) / IEEref,s')
    ws.cell(8, 3, value='=(C4-C6)/C5')
    ws.cell(8, 3).fill = RESULT_FILL
    ws.cell(8, 3).border = THICK_BORDER
    ws.cell(8, 3).font = Font(bold=True, size=14)
    ws.cell(8, 3).number_format = '0.00'

    # Nota
    ws.cell(10, 1, value='Nota: IEEref,s deve ser calculado com simulação REF ou valores de referência SCE').font = Font(italic=True, color='666666')

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 22


def create_classe_sheet(ws):
    """Cria folha de Classe Energética"""

    ws.merge_cells('A1:F1')
    ws.cell(1, 1, value='CLASSE ENERGÉTICA').font = Font(bold=True, size=16)

    ws.cell(3, 1, value='RIEE:').font = Font(bold=True, size=14)
    ws.cell(3, 2, value="='IEE'!C8")
    ws.cell(3, 2).font = Font(bold=True, size=14)
    ws.cell(3, 2).number_format = '0.00'

    ws.cell(5, 1, value='CLASSE:').font = Font(bold=True, size=16)

    classe_formula = '=IF(B3<=0.25,"A+",IF(B3<=0.5,"A",IF(B3<=0.75,"B",IF(B3<=1,"B-",IF(B3<=1.5,"C",IF(B3<=2,"D",IF(B3<=2.5,"E","F")))))))'
    ws.cell(5, 2, value=classe_formula)
    ws.cell(5, 2).font = Font(bold=True, size=36)
    ws.cell(5, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B5:C6')

    ws.cell(8, 1, value='Limites RIEE por Classe:').font = Font(bold=True)

    row = 9
    for classe, min_val, max_val in RIEE_LIMITES:
        ws.cell(row, 1, value=classe)
        ws.cell(row, 1).fill = PatternFill(start_color=CLASS_COLORS[classe], end_color=CLASS_COLORS[classe], fill_type='solid')
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal='center')
        ws.cell(row, 1).border = THIN_BORDER

        if classe == 'A+':
            ws.cell(row, 2, value='≤ 0.25')
        elif classe == 'F':
            ws.cell(row, 2, value='> 2.50')
        else:
            ws.cell(row, 2, value=f'{min_val:.2f} - {max_val:.2f}')
        ws.cell(row, 2).border = THIN_BORDER
        row += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_legenda_sheet(ws):
    """Cria folha de Legenda"""

    ws.cell(1, 1, value='LEGENDA E INSTRUÇÕES').font = Font(bold=True, size=14)

    ws.cell(3, 1, value='CORES DAS CÉLULAS:').font = Font(bold=True)

    ws.cell(4, 1, value='Amarelo')
    ws.cell(4, 1).fill = INPUT_FILL
    ws.cell(4, 2, value='= INPUT - Preencher manualmente')

    ws.cell(5, 1, value='Verde claro')
    ws.cell(5, 1).fill = FORMULA_FILL
    ws.cell(5, 2, value='= FÓRMULA - Calculado automaticamente')

    ws.cell(6, 1, value='Laranja')
    ws.cell(6, 1).fill = RESULT_FILL
    ws.cell(6, 2, value='= RESULTADO - Valor final')

    ws.cell(7, 1, value='Azul claro')
    ws.cell(7, 1).fill = PREV_FILL
    ws.cell(7, 2, value='= Dados PREVISTO')

    ws.cell(8, 1, value='Rosa claro')
    ws.cell(8, 1).fill = REF_FILL
    ws.cell(8, 2, value='= Dados REFERÊNCIA')

    ws.cell(10, 1, value='FOLHAS DO FICHEIRO:').font = Font(bold=True)

    folhas = [
        ('Detalhe PREV/REF', 'Dados brutos da simulação HAP'),
        ('Mensal PREV/REF', 'Dados mensais por sistema'),
        ('Iluminação ENU', 'Cálculo iluminação ENU e exterior'),
        ('AQS', 'Cálculo águas quentes sanitárias'),
        ('PV', 'Produção fotovoltaica'),
        ('Equipamentos Extra', 'Equipamentos não simulados'),
        ('Elevadores', 'Cálculo energia elevadores (RECS)'),
        ('Ventilação Extra', 'Ventilação tipo S adicional'),
        ('Bombagem', 'Bombas AQS e AVAC'),
        ('Consumos Tipologia', 'Desagregação completa'),
        ('Energia Primária', 'Conversão kWh → kWhEP'),
        ('IEE', 'Indicadores de eficiência'),
        ('Classe', 'Classe energética final'),
    ]

    row = 11
    for folha, desc in folhas:
        ws.cell(row, 1, value=folha)
        ws.cell(row, 2, value=desc)
        row += 1

    row += 1
    ws.cell(row, 1, value='TIPOS DE CONSUMO:').font = Font(bold=True)
    row += 1
    ws.cell(row, 1, value='Tipo S')
    ws.cell(row, 2, value='= CONTA para classificação (AVAC, AQS, Iluminação, Elevadores)')
    row += 1
    ws.cell(row, 1, value='Tipo T')
    ws.cell(row, 2, value='= NÃO conta para classificação (Equipamentos)')
    row += 1
    ws.cell(row, 1, value='REN')
    ws.cell(row, 2, value='= Energia renovável (deduzida ao IEEprev)')

    row += 2
    ws.cell(row, 1, value='FÓRMULAS SCE:').font = Font(bold=True)
    row += 1
    ws.cell(row, 1, value='IEEprev,s = Σ(Consumos_S × Fpu) / Área')
    row += 1
    ws.cell(row, 1, value='IEEren = Σ(Produção_Renovável × Fpu) / Área')
    row += 1
    ws.cell(row, 1, value='RIEE = (IEEprev,s - IEEren) / IEEref,s')

    row += 2
    ws.cell(row, 1, value='FACTORES Fpu:').font = Font(bold=True)
    row += 1
    ws.cell(row, 1, value='Electricidade: 2.5 kWhEP/kWh')
    row += 1
    ws.cell(row, 1, value='Gás Natural: 1.0 kWhEP/kWh')

    row += 2
    ws.cell(row, 1, value='REFERÊNCIAS:').font = Font(bold=True)
    row += 1
    ws.cell(row, 2, value='• Decreto-Lei 101-D/2020')
    row += 1
    ws.cell(row, 2, value='• Despacho 15793-D/2013 (factores Fpu)')
    row += 1
    ws.cell(row, 2, value='• Manual SCE - DGEG/ADENE')
    row += 1
    ws.cell(row, 2, value='• Guia SCE - Indicadores RECS')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 55


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def create_iee_completo(prev_data, ref_data, output_path):
    """Cria a folha Excel completa"""

    wb = openpyxl.Workbook()

    num_sistemas_prev = len(prev_data)
    num_sistemas_ref = len(ref_data)

    # 1. Detalhe PREV
    ws_prev = wb.active
    ws_prev.title = 'Detalhe PREV'
    ws_prev.sheet_properties.tabColor = 'DDEBF7'
    row_total_prev = create_detalhe_sheet(ws_prev, prev_data, 'PREVISTO')

    # 2. Detalhe REF
    ws_ref = wb.create_sheet('Detalhe REF')
    ws_ref.sheet_properties.tabColor = 'FCE4D6'
    row_total_ref = create_detalhe_sheet(ws_ref, ref_data, 'REFERÊNCIA')

    # 3. Mensal PREV
    ws_mensal_prev = wb.create_sheet('Mensal PREV')
    ws_mensal_prev.sheet_properties.tabColor = 'DDEBF7'
    create_mensal_sheet(ws_mensal_prev, prev_data, 'PREVISTO')

    # 4. Mensal REF
    ws_mensal_ref = wb.create_sheet('Mensal REF')
    ws_mensal_ref.sheet_properties.tabColor = 'FCE4D6'
    create_mensal_sheet(ws_mensal_ref, ref_data, 'REFERÊNCIA')

    # 5. Iluminação ENU
    ws_ilum = wb.create_sheet('Iluminação ENU')
    ws_ilum.sheet_properties.tabColor = '92D050'
    create_iluminacao_enu_sheet(ws_ilum)

    # 6. AQS
    ws_aqs = wb.create_sheet('AQS')
    ws_aqs.sheet_properties.tabColor = 'FFC000'
    create_aqs_sheet(ws_aqs)

    # 7. PV
    ws_pv = wb.create_sheet('PV')
    ws_pv.sheet_properties.tabColor = '7030A0'
    create_pv_sheet(ws_pv)

    # 8. Equipamentos Extra
    ws_eq = wb.create_sheet('Equipamentos Extra')
    ws_eq.sheet_properties.tabColor = 'FFC000'
    create_equipamentos_extra_sheet(ws_eq)

    # 9. Elevadores
    ws_elev = wb.create_sheet('Elevadores')
    ws_elev.sheet_properties.tabColor = '00B0F0'
    create_elevadores_sheet(ws_elev)

    # 10. Ventilação Extra
    ws_vent = wb.create_sheet('Ventilação Extra')
    ws_vent.sheet_properties.tabColor = '7030A0'
    create_ventilacao_extra_sheet(ws_vent)

    # 11. Bombagem
    ws_bomb = wb.create_sheet('Bombagem')
    ws_bomb.sheet_properties.tabColor = '000000'
    create_bombagem_sheet(ws_bomb)

    # 12. Consumos por Tipologia
    ws_cons = wb.create_sheet('Consumos Tipologia')
    ws_cons.sheet_properties.tabColor = 'FF0000'
    consumos_info = create_consumos_tipologia_sheet(ws_cons, row_total_prev, row_total_ref)

    # 13. Energia Primária
    ws_ep = wb.create_sheet('Energia Primária')
    ws_ep.sheet_properties.tabColor = 'FF0000'
    row_total_s_prev, row_total_ren = create_energia_primaria_sheet(ws_ep, consumos_info)

    # 14. IEE
    ws_iee = wb.create_sheet('IEE')
    ws_iee.sheet_properties.tabColor = 'FF0000'
    create_iee_sheet(ws_iee, row_total_s_prev, row_total_ren)

    # 15. Classe
    ws_classe = wb.create_sheet('Classe')
    ws_classe.sheet_properties.tabColor = 'FF0000'
    create_classe_sheet(ws_classe)

    # 16. Legenda
    ws_leg = wb.create_sheet('Legenda')
    create_legenda_sheet(ws_leg)

    wb.save(output_path)

    return num_sistemas_prev, num_sistemas_ref


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        print("\nERRO: Necessário indicar as pastas PREV e REF")
        sys.exit(1)

    prev_folder = sys.argv[1]
    ref_folder = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else 'IEE_Completo.xlsx'

    if not os.path.isdir(prev_folder):
        print(f"Erro: Pasta PREV não encontrada: {prev_folder}")
        sys.exit(1)

    if not os.path.isdir(ref_folder):
        print(f"Erro: Pasta REF não encontrada: {ref_folder}")
        sys.exit(1)

    print(f"A carregar dados PREV de: {prev_folder}")
    prev_data = load_project_data(prev_folder)
    print(f"  Encontrados {len(prev_data)} sistemas")

    print(f"\nA carregar dados REF de: {ref_folder}")
    ref_data = load_project_data(ref_folder)
    print(f"  Encontrados {len(ref_data)} sistemas")

    if not prev_data:
        print("ERRO: Nenhum CSV encontrado na pasta PREV")
        sys.exit(1)

    if not ref_data:
        print("ERRO: Nenhum CSV encontrado na pasta REF")
        sys.exit(1)

    num_prev, num_ref = create_iee_completo(prev_data, ref_data, output_path)

    print(f"\n{'='*60}")
    print(f"Ficheiro criado: {output_path}")
    print(f"{'='*60}")
    print(f"\nFOLHAS CRIADAS:")
    print(f"  1. Detalhe PREV       - {num_prev} sistemas (VALORES)")
    print(f"  2. Detalhe REF        - {num_ref} sistemas (VALORES)")
    print(f"  3. Mensal PREV        - Dados mensais (VALORES)")
    print(f"  4. Mensal REF         - Dados mensais (VALORES)")
    print(f"  5. Iluminação ENU     - Iluminação ENU e exterior (INPUT)")
    print(f"  6. AQS                - Águas quentes sanitárias (INPUT)")
    print(f"  7. PV                 - Fotovoltaico (INPUT)")
    print(f"  8. Equipamentos Extra - Equipamentos não simulados (INPUT)")
    print(f"  9. Elevadores         - Cálculo RECS (INPUT)")
    print(f" 10. Ventilação Extra   - Ventilação tipo S (INPUT)")
    print(f" 11. Bombagem           - Bombas (INPUT)")
    print(f" 12. Consumos Tipologia - Desagregação (FÓRMULAS)")
    print(f" 13. Energia Primária   - Conversão kWh→kWhEP (FÓRMULAS)")
    print(f" 14. IEE                - Indicadores (FÓRMULAS)")
    print(f" 15. Classe             - Classe energética (FÓRMULAS)")
    print(f" 16. Legenda            - Instruções")
    print(f"\n{'='*60}")
    print(f"PRÓXIMOS PASSOS:")
    print(f"  1. Abrir o ficheiro no Excel")
    print(f"  2. Preencher células AMARELAS nas folhas auxiliares:")
    print(f"     - Iluminação ENU, AQS, PV, Equipamentos, Elevadores, etc.")
    print(f"  3. Ir a 'Energia Primária' e preencher Área Útil [m²]")
    print(f"  4. Ir a 'IEE' e preencher IEEref,s (do cálculo REF)")
    print(f"  5. Ver resultado na folha 'Classe'")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
