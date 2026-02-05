"""
HAP to IEE - Extrai dados de simulação do HAP para cálculo de IEE

Lê os ficheiros CSV exportados pelo HAP (System Simulation Report)
e cria uma folha Excel organizada com os consumos por sistema.

Estrutura:
- Detalhe: Dados brutos extraídos dos CSVs (valores)
- Resumo: Fórmulas que apontam para Detalhe
- IEEprev: Fórmulas para cálculo do IEE

Usage:
    python hap_to_iee.py <pasta_projecto_hap> [output.xlsx]

Exemplo:
    python hap_to_iee.py "C:\\E20-II\\Projects\\CasaAlecrim2025_Ref" resultado_simulacao.xlsx
"""

import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# ESTILOS
# =============================================================================

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
TOTAL_FONT = Font(bold=True)
FORMULA_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# =============================================================================
# COLUNAS CSV
# =============================================================================

# Ordem das colunas na folha Detalhe (nomes originais do CSV, sem "(kWh)")
DETALHE_COLUMNS = [
    'Lighting',
    'Electric Equipment',
    'Central Unit Clg Input',
    'Terminal Unit Clg Input',
    'Central Unit Htg Input',
    'Terminal Unit Htg Input',
    'Central Unit Aux. Htg. Input',
    'Terminal Unit Aux. Htg. Input',
    'Supply Fan',
    'Return Fan',
    'Exhaust Fan',
    'Ventilation Fan',
    'Central Cooling Coil Load',
    'Central Heating Coil Load',
    'Terminal Cooling Coil Load',
    'Terminal Heating Coil Load',
]

# Mapeamento para Resumo: (nome, colunas do Detalhe a somar)
# As letras referem-se às colunas na folha Detalhe (B=Lighting, C=Equipment, etc.)
RESUMO_MAPPING = [
    ('Lighting', ['B']),
    ('Electric Equipment', ['C']),
    ('Cooling Input', ['D', 'E']),      # Central + Terminal Clg Input
    ('Heating Input', ['F', 'G']),      # Central + Terminal Htg Input
    ('Aux. Htg. Input', ['H', 'I']),    # Central + Terminal Aux Htg
    ('Fans', ['J', 'K', 'L', 'M']),     # Supply + Return + Exhaust + Ventilation
]


# =============================================================================
# FUNÇÕES
# =============================================================================

def read_hap_csv(filepath):
    """Lê um CSV do HAP e retorna dicionário com totais anuais e dados mensais"""

    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if len(lines) < 5:
        return None, {}

    # Nome do sistema (linha 2)
    sistema = lines[1].split(';')[0].replace('Monthly Simulation Results for ', '').strip()

    # Headers (linha 4, índice 3)
    headers = lines[3].strip().split(';')

    # Calcular totais e guardar dados mensais
    totals = {}
    monthly_data = []

    for line in lines[4:]:
        line = line.strip()
        if not line or line.startswith('Month'):
            continue

        values = line.split(';')
        month = values[0] if values else ''

        month_data = {'Month': month}
        for i, val in enumerate(values[1:], 1):
            if i < len(headers) and val:
                col = headers[i]
                try:
                    num_val = int(val)
                    totals[col] = totals.get(col, 0) + num_val
                    month_data[col] = num_val
                except:
                    pass

        if month:
            monthly_data.append(month_data)

    return sistema, {'totals': totals, 'monthly': monthly_data, 'headers': headers}


def find_hap_csvs(project_folder):
    """Encontra todos os CSVs do HAP numa pasta"""
    pattern = os.path.join(project_folder, 'HAP51_Monthly_*.csv')
    return sorted(glob.glob(pattern))


def create_excel(systems_data, output_path):
    """Cria Excel com os dados dos sistemas"""

    wb = openpyxl.Workbook()

    # Filtrar sistemas (remover TODOS)
    sistemas = [(s, d) for s, d in systems_data.items() if s != 'TODOS']
    num_sistemas = len(sistemas)

    # =========================================================================
    # FOLHA 1: DETALHE (dados brutos dos CSVs - VALORES)
    # =========================================================================
    ws_detalhe = wb.active
    ws_detalhe.title = 'Detalhe'

    # Headers na linha 1
    ws_detalhe.cell(1, 1, value='Sistema').font = HEADER_FONT
    ws_detalhe.cell(1, 1).fill = HEADER_FILL
    ws_detalhe.cell(1, 1).border = THIN_BORDER

    for col, col_name in enumerate(DETALHE_COLUMNS, 2):
        cell = ws_detalhe.cell(1, col, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Dados dos sistemas (a partir da linha 2)
    row = 2
    for sistema, data in sistemas:
        ws_detalhe.cell(row, 1, value=sistema).border = THIN_BORDER

        for col, col_name in enumerate(DETALHE_COLUMNS, 2):
            csv_col = col_name + ' (kWh)'
            value = data['totals'].get(csv_col, 0)
            cell = ws_detalhe.cell(row, col, value=value if value > 0 else '')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')

        row += 1

    # Linha TOTAL com fórmulas SUM
    row_total = row
    ws_detalhe.cell(row_total, 1, value='TOTAL').font = TOTAL_FONT
    ws_detalhe.cell(row_total, 1).fill = TOTAL_FILL
    ws_detalhe.cell(row_total, 1).border = THIN_BORDER

    for col in range(2, len(DETALHE_COLUMNS) + 2):
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}2:{col_letter}{row_total - 1})'
        cell = ws_detalhe.cell(row_total, col, value=formula)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

    # Ajustar larguras
    ws_detalhe.column_dimensions['A'].width = 20
    for col in range(2, len(DETALHE_COLUMNS) + 2):
        ws_detalhe.column_dimensions[get_column_letter(col)].width = 12

    ws_detalhe.freeze_panes = 'B2'

    # =========================================================================
    # FOLHA 2: RESUMO (fórmulas que apontam para Detalhe)
    # =========================================================================
    ws_resumo = wb.create_sheet('Resumo')

    # Título
    ws_resumo.cell(1, 1, value='RESUMO SIMULAÇÃO HAP').font = Font(bold=True, size=14)
    ws_resumo.cell(2, 1, value=f'Sistemas: {num_sistemas}')

    # Headers na linha 4
    ws_resumo.cell(4, 1, value='Sistema').font = HEADER_FONT
    ws_resumo.cell(4, 1).fill = HEADER_FILL
    ws_resumo.cell(4, 1).border = THIN_BORDER

    for col, (col_name, _) in enumerate(RESUMO_MAPPING, 2):
        cell = ws_resumo.cell(4, col, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Coluna TOTAL
    total_col = len(RESUMO_MAPPING) + 2
    cell = ws_resumo.cell(4, total_col, value='TOTAL')
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

    # Dados com FÓRMULAS
    for i, (sistema, _) in enumerate(sistemas):
        row = 5 + i
        detalhe_row = 2 + i  # Linha correspondente no Detalhe

        # Nome do sistema (fórmula para Detalhe)
        ws_resumo.cell(row, 1, value=f"=Detalhe!A{detalhe_row}").border = THIN_BORDER

        # Colunas agregadas (fórmulas)
        sum_cols = []
        for col, (col_name, detalhe_cols) in enumerate(RESUMO_MAPPING, 2):
            if len(detalhe_cols) == 1:
                formula = f"=Detalhe!{detalhe_cols[0]}{detalhe_row}"
            else:
                refs = '+'.join([f"Detalhe!{c}{detalhe_row}" for c in detalhe_cols])
                formula = f"={refs}"

            cell = ws_resumo.cell(row, col, value=formula)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')
            cell.fill = FORMULA_FILL
            sum_cols.append(get_column_letter(col))

        # Total da linha (soma das colunas desta linha)
        refs = '+'.join([f"{c}{row}" for c in sum_cols])
        cell = ws_resumo.cell(row, total_col, value=f"={refs}")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        cell.fill = FORMULA_FILL

    # Linha TOTAL
    row_total = 5 + num_sistemas
    ws_resumo.cell(row_total, 1, value='TOTAL').font = TOTAL_FONT
    ws_resumo.cell(row_total, 1).fill = TOTAL_FILL
    ws_resumo.cell(row_total, 1).border = THIN_BORDER

    for col in range(2, total_col + 1):
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}5:{col_letter}{row_total - 1})'
        cell = ws_resumo.cell(row_total, col, value=formula)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

    # Ajustar larguras
    ws_resumo.column_dimensions['A'].width = 20
    for col in range(2, total_col + 1):
        ws_resumo.column_dimensions[get_column_letter(col)].width = 14

    # =========================================================================
    # FOLHA 3: MENSAL (dados mensais - VALORES)
    # =========================================================================
    ws_mensal = wb.create_sheet('Mensal')

    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']

    # Colunas principais para mostrar mensalmente
    mensal_cols = [
        ('Lighting', 'Lighting (kWh)'),
        ('Electric Equipment', 'Electric Equipment (kWh)'),
        ('Central Unit Clg Input', 'Central Unit Clg Input (kWh)'),
        ('Central Unit Htg Input', 'Central Unit Htg Input (kWh)'),
        ('Supply Fan', 'Supply Fan (kWh)'),
        ('Return Fan', 'Return Fan (kWh)'),
    ]

    # Headers
    ws_mensal.cell(1, 1, value='Sistema').font = HEADER_FONT
    ws_mensal.cell(1, 1).fill = HEADER_FILL
    ws_mensal.cell(1, 1).border = THIN_BORDER

    ws_mensal.cell(1, 2, value='Coluna').font = HEADER_FONT
    ws_mensal.cell(1, 2).fill = HEADER_FILL
    ws_mensal.cell(1, 2).border = THIN_BORDER

    for col, month in enumerate(months, 3):
        cell = ws_mensal.cell(1, col, value=month[:3])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    ws_mensal.cell(1, 15, value='Total').font = HEADER_FONT
    ws_mensal.cell(1, 15).fill = HEADER_FILL
    ws_mensal.cell(1, 15).border = THIN_BORDER

    # Dados mensais
    row = 2
    for sistema, data in sistemas:
        monthly = data.get('monthly', [])
        if not monthly:
            continue

        # Para cada coluna com dados
        for col_name, csv_col in mensal_cols:
            if data['totals'].get(csv_col, 0) == 0:
                continue

            ws_mensal.cell(row, 1, value=sistema).border = THIN_BORDER
            ws_mensal.cell(row, 2, value=col_name).border = THIN_BORDER

            total = 0
            for col, month in enumerate(months, 3):
                for m_data in monthly:
                    if m_data.get('Month', '').startswith(month[:3]):
                        val = m_data.get(csv_col, 0)
                        cell = ws_mensal.cell(row, col, value=val if val > 0 else '')
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(horizontal='right')
                        total += val
                        break

            # Total com fórmula SUM
            formula = f'=SUM(C{row}:N{row})'
            cell = ws_mensal.cell(row, 15, value=formula)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')
            cell.fill = FORMULA_FILL
            row += 1

    # Ajustar larguras
    ws_mensal.column_dimensions['A'].width = 18
    ws_mensal.column_dimensions['B'].width = 20
    for col in range(3, 16):
        ws_mensal.column_dimensions[get_column_letter(col)].width = 8

    ws_mensal.freeze_panes = 'C2'

    # =========================================================================
    # FOLHA 4: IEEprev (fórmulas para cálculo IEE)
    # =========================================================================
    ws_iee = wb.create_sheet('IEEprev')

    # Referências para a linha TOTAL do Detalhe
    detalhe_total_row = row_total  # mesma linha que no Resumo

    # Área útil (a preencher manualmente)
    ws_iee.cell(3, 3, value='Área Útil [m2]')
    ws_iee.cell(3, 4, value='').font = Font(bold=True, color='FF0000')
    ws_iee.cell(3, 4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Secção Electricidade [kWh]
    ws_iee.cell(4, 2, value='Electricidade\n[kWh]').font = Font(bold=True)
    ws_iee.cell(4, 2).alignment = Alignment(wrap_text=True)

    # Mapeamento IEE -> Detalhe (coluna na folha Detalhe para a linha TOTAL)
    # Detalhe: B=Lighting, C=Equipment, D=CentralClgInput, E=TerminalClgInput,
    #          F=CentralHtgInput, G=TerminalHtgInput, H=CentralAuxHtg, I=TerminalAuxHtg,
    #          J=SupplyFan, K=ReturnFan, L=ExhaustFan, M=VentilationFan
    detalhe_row_total_num = 2 + num_sistemas  # linha do TOTAL no Detalhe

    iee_rows = [
        ('Aquecimento', f"=Detalhe!F{detalhe_row_total_num}+Detalhe!G{detalhe_row_total_num}"),  # Central+Terminal Htg Input
        ('Arrefecimento', f"=Detalhe!D{detalhe_row_total_num}+Detalhe!E{detalhe_row_total_num}"),  # Central+Terminal Clg Input
        ('Iluminação Interior', f"=Detalhe!B{detalhe_row_total_num}"),  # Lighting
        ('Iluminação Exterior', ''),  # Manual
        ('Ventilação', f"=Detalhe!J{detalhe_row_total_num}+Detalhe!K{detalhe_row_total_num}+Detalhe!L{detalhe_row_total_num}+Detalhe!M{detalhe_row_total_num}"),  # Fans
        ('Bombagem', ''),  # Manual
        ('AQS', ''),  # Manual
        ('Elevador', ''),  # Manual
        ('Equipamentos e Outros', f"=Detalhe!C{detalhe_row_total_num}"),  # Electric Equipment
        ('Aquecimento Piscina', ''),  # Manual
    ]

    for i, (label, formula) in enumerate(iee_rows):
        row_num = 4 + i
        ws_iee.cell(row_num, 3, value=label)
        if formula:
            ws_iee.cell(row_num, 4, value=formula)
            ws_iee.cell(row_num, 4).fill = FORMULA_FILL
        else:
            ws_iee.cell(row_num, 4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Secção Gás Natural [kWh]
    ws_iee.cell(14, 2, value='Gás Natural\n[kWh]').font = Font(bold=True)
    ws_iee.cell(14, 2).alignment = Alignment(wrap_text=True)

    gn_rows = ['Aquecimento', 'AQS', 'Aquecimento Água Piscina', 'Equipamentos']
    for i, label in enumerate(gn_rows):
        row_num = 14 + i
        ws_iee.cell(row_num, 3, value=label)
        ws_iee.cell(row_num, 4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Secção Renovável [kWh]
    ws_iee.cell(18, 2, value='Renovável\n[kWh]').font = Font(bold=True)
    ws_iee.cell(18, 2).alignment = Alignment(wrap_text=True)

    ren_rows = ['Aquecimento', 'Arrefecimento', 'AQS Aero', 'AQS Solar', 'PV']
    for i, label in enumerate(ren_rows):
        row_num = 18 + i
        ws_iee.cell(row_num, 3, value=label)
        ws_iee.cell(row_num, 4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Legenda
    ws_iee.cell(24, 2, value='Legenda:')
    ws_iee.cell(24, 3, value='Verde = Fórmula (vem do Detalhe)')
    ws_iee.cell(24, 3).fill = FORMULA_FILL
    ws_iee.cell(25, 3, value='Amarelo = Preencher manualmente')
    ws_iee.cell(25, 3).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Totais EE (com fórmulas)
    ws_iee.cell(27, 3, value='Total EE (simulação)')
    ws_iee.cell(27, 4, value='=D4+D5+D6+D8+D12')  # Aquec+Arref+Ilum+Vent+Equip
    ws_iee.cell(27, 4).fill = FORMULA_FILL

    # Ajustar larguras
    ws_iee.column_dimensions['B'].width = 15
    ws_iee.column_dimensions['C'].width = 25
    ws_iee.column_dimensions['D'].width = 20

    # =========================================================================
    # GUARDAR
    # =========================================================================
    wb.save(output_path)
    return num_sistemas


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    project_folder = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'Simulacao_HAP.xlsx'

    # Verificar pasta
    if not os.path.isdir(project_folder):
        print(f"Erro: Pasta não encontrada: {project_folder}")
        sys.exit(1)

    # Encontrar CSVs
    csv_files = find_hap_csvs(project_folder)

    if not csv_files:
        print(f"Erro: Nenhum ficheiro HAP51_Monthly_*.csv encontrado em {project_folder}")
        print("Exporta o 'System Simulation Report' como CSV no HAP primeiro.")
        sys.exit(1)

    print(f"Encontrados {len(csv_files)} ficheiros CSV")

    # Ler dados
    systems_data = {}
    for csv_file in csv_files:
        sistema, data = read_hap_csv(csv_file)
        if sistema and data:
            systems_data[sistema] = data
            print(f"  Lido: {sistema}")

    # Criar Excel
    num_sistemas = create_excel(systems_data, output_path)
    print(f"\nFicheiro criado: {output_path}")
    print(f"  - {num_sistemas} sistemas")
    print(f"  - Folha 'Detalhe': Dados brutos dos CSVs (VALORES)")
    print(f"  - Folha 'Resumo': Agregações (FÓRMULAS -> Detalhe)")
    print(f"  - Folha 'Mensal': Dados mensais por sistema (VALORES)")
    print(f"  - Folha 'IEEprev': Estrutura IEE (FÓRMULAS -> Detalhe)")


if __name__ == '__main__':
    main()
