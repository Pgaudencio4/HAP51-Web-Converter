"""
Comparador de Excel HAP - Compara Excel original com extraído do E3A

Compara TODOS os 147 campos do template.

Usage:
    python comparar_excels.py <original.xlsx> <extraido.xlsx> [output.xlsx]

Exemplo:
    python comparar_excels.py Malhoa22_Input.xlsx Malhoa22_Extraido.xlsx Malhoa22_Comparacao.xlsx
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Cores
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

def normalize_value(v):
    """Normaliza valor para comparação"""
    if v is None or v == '':
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == '' or v.lower() in ('none', 'n/a', '-'):
            return None
        # Tentar converter para número
        try:
            f = float(v)
            return round(f, 2) if f != 0 else None
        except:
            return v.lower()
    if isinstance(v, (int, float)):
        return round(float(v), 2) if v != 0 else None
    return v

def compare_values(v1, v2, tolerance=0.1):
    """Compara dois valores com tolerância"""
    n1 = normalize_value(v1)
    n2 = normalize_value(v2)

    # Ambos vazios = match
    if n1 is None and n2 is None:
        return True, 'empty'

    # Um vazio, outro não
    if n1 is None or n2 is None:
        return False, 'missing'

    # Strings
    if isinstance(n1, str) or isinstance(n2, str):
        # Comparação aproximada de strings (schedule names podem ter variações)
        s1 = str(n1).lower().replace(' ', '')
        s2 = str(n2).lower().replace(' ', '')
        if s1 == s2:
            return True, 'match'
        # Verificar se é substring (ex: schedule com/sem texto extra)
        if s1 in s2 or s2 in s1:
            return True, 'partial'
        return False, 'diff'

    # Números - comparar com tolerância relativa
    if abs(n1 - n2) < tolerance:
        return True, 'match'
    if n1 != 0 and abs(n1 - n2) / abs(n1) < 0.05:  # 5% tolerance
        return True, 'close'
    return False, 'diff'

def read_excel_data(filepath):
    """Lê dados do Excel HAP"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Espacos']

    # Headers da linha 3
    headers = [ws.cell(3, col).value or '' for col in range(1, 148)]

    # Dados (linha 4+)
    data = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()
        row_data = [ws.cell(row, col).value for col in range(1, 148)]
        data[name] = row_data

    return headers, data

def create_comparison_excel(headers, data1, data2, output_path):
    """Cria Excel com comparação lado a lado"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparacao'

    # === Cabeçalhos ===
    # Linha 1: Categorias
    categories = [
        (1, 'GENERAL'), (7, 'INTERNALS'), (23, 'INFILTRATION'),
        (27, 'FLOORS'), (40, 'PARTITIONS'), (52, 'WALLS'), (124, 'ROOFS')
    ]

    ws.cell(1, 1, value='SPACE')
    ws.cell(1, 2, value='SOURCE')

    for cat_col, cat_name in categories:
        # Original columns start at 3
        ws.cell(1, cat_col + 2, value=cat_name)
        c = ws.cell(1, cat_col + 2)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    # Linha 2: Sub-categorias
    subcategories = [
        (7, 'PEOPLE'), (12, 'LIGHTING'), (17, 'EQUIPMENT'), (19, 'MISC'),
        (40, 'CEILING'), (46, 'WALL'),
        (52, 'WALL 1'), (61, 'WALL 2'), (70, 'WALL 3'), (79, 'WALL 4'),
        (88, 'WALL 5'), (97, 'WALL 6'), (106, 'WALL 7'), (115, 'WALL 8'),
        (124, 'ROOF 1'), (130, 'ROOF 2'), (136, 'ROOF 3'), (142, 'ROOF 4')
    ]
    for col, name in subcategories:
        c = ws.cell(2, col + 2, value=name)
        c.fill = SUBHEADER_FILL
        c.font = Font(bold=True)

    # Linha 3: Nomes dos campos
    ws.cell(3, 1, value='Space Name')
    ws.cell(3, 2, value='Fonte')
    for i, h in enumerate(headers):
        c = ws.cell(3, i + 3, value=h)
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(wrap_text=True, horizontal='center')

    # === Dados ===
    row = 4

    # Estatísticas
    total_fields = 0
    matching_fields = 0
    diff_fields = 0

    # Todos os espaços (união)
    all_spaces = sorted(set(list(data1.keys()) + list(data2.keys())))

    for space_name in all_spaces:
        orig = data1.get(space_name)
        extr = data2.get(space_name)

        if orig is None:
            # Só existe no extraído
            ws.cell(row, 1, value=space_name)
            ws.cell(row, 2, value='EXTRAIDO')
            for i, v in enumerate(extr):
                ws.cell(row, i + 3, value=v)
            ws.cell(row, 1).fill = YELLOW_FILL
            row += 1
            continue

        if extr is None:
            # Só existe no original
            ws.cell(row, 1, value=space_name)
            ws.cell(row, 2, value='ORIGINAL')
            for i, v in enumerate(orig):
                ws.cell(row, i + 3, value=v)
            ws.cell(row, 1).fill = YELLOW_FILL
            row += 1
            continue

        # Existe em ambos - comparar
        # Linha do original
        ws.cell(row, 1, value=space_name)
        ws.cell(row, 2, value='ORIGINAL')
        for i, v in enumerate(orig):
            ws.cell(row, i + 3, value=v)
        row += 1

        # Linha do extraído
        ws.cell(row, 1, value=space_name)
        ws.cell(row, 2, value='EXTRAIDO')
        for i, v in enumerate(extr):
            c = ws.cell(row, i + 3, value=v)

            # Comparar com original
            match, status = compare_values(orig[i], v)
            total_fields += 1

            if status == 'empty':
                pass  # Ambos vazios
            elif match:
                matching_fields += 1
                c.fill = GREEN_FILL
            else:
                diff_fields += 1
                c.fill = RED_FILL
        row += 1

        # Linha de diferenças
        ws.cell(row, 1, value='')
        ws.cell(row, 2, value='DIFF')
        for i in range(len(orig)):
            match, status = compare_values(orig[i], extr[i])
            if not match and status != 'empty':
                o = normalize_value(orig[i])
                e = normalize_value(extr[i])
                diff_text = f'{o} -> {e}'
                c = ws.cell(row, i + 3, value=diff_text)
                c.fill = RED_FILL
                c.font = Font(bold=True, size=9)
        row += 1

        # Linha vazia
        row += 1

    # === Sumário ===
    ws_sum = wb.create_sheet('Sumario')
    ws_sum.cell(1, 1, value='SUMÁRIO DA COMPARAÇÃO')
    ws_sum.cell(1, 1).font = Font(bold=True, size=14)

    ws_sum.cell(3, 1, value='Total de espaços no original:')
    ws_sum.cell(3, 2, value=len(data1))

    ws_sum.cell(4, 1, value='Total de espaços no extraído:')
    ws_sum.cell(4, 2, value=len(data2))

    ws_sum.cell(5, 1, value='Espaços em comum:')
    ws_sum.cell(5, 2, value=len(set(data1.keys()) & set(data2.keys())))

    ws_sum.cell(7, 1, value='Total de campos comparados:')
    ws_sum.cell(7, 2, value=total_fields)

    ws_sum.cell(8, 1, value='Campos iguais:')
    ws_sum.cell(8, 2, value=matching_fields)
    ws_sum.cell(8, 2).fill = GREEN_FILL

    ws_sum.cell(9, 1, value='Campos diferentes:')
    ws_sum.cell(9, 2, value=diff_fields)
    ws_sum.cell(9, 2).fill = RED_FILL

    match_pct = (matching_fields / total_fields * 100) if total_fields > 0 else 0
    ws_sum.cell(10, 1, value='Percentagem de match:')
    ws_sum.cell(10, 2, value=f'{match_pct:.1f}%')

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 15

    # Ajustar colunas da comparação
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    for col in range(3, 150):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    # Freeze panes
    ws.freeze_panes = 'C4'

    wb.save(output_path)

    print(f'\n=== SUMÁRIO ===')
    print(f'Espaços original:  {len(data1)}')
    print(f'Espaços extraído:  {len(data2)}')
    print(f'Campos comparados: {total_fields}')
    print(f'Campos iguais:     {matching_fields} ({match_pct:.1f}%)')
    print(f'Campos diferentes: {diff_fields}')

    return wb

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    orig_file = sys.argv[1]
    extr_file = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else 'Comparacao.xlsx'

    if not os.path.exists(orig_file):
        print(f"Erro: Ficheiro '{orig_file}' não encontrado!")
        sys.exit(1)
    if not os.path.exists(extr_file):
        print(f"Erro: Ficheiro '{extr_file}' não encontrado!")
        sys.exit(1)

    print(f'Original: {orig_file}')
    print(f'Extraído: {extr_file}')

    # Ler dados
    headers1, data1 = read_excel_data(orig_file)
    headers2, data2 = read_excel_data(extr_file)

    print(f'\nLidos {len(data1)} espaços do original')
    print(f'Lidos {len(data2)} espaços do extraído')

    # Criar comparação
    create_comparison_excel(headers1, data1, data2, output_file)

    print(f'\nFicheiro criado: {output_file}')

if __name__ == '__main__':
    main()
