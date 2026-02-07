"""
Editor E3A - Modifica ficheiros HAP existentes

Suporta edição de:
- Espaços (147 campos)
- Windows (U-Value, SHGC, Altura, Largura)
- Walls (U-Value, Espessura, Massa)
- Roofs (U-Value, Espessura, Massa)

Fluxo:
1. Extrair E3A para Excel: python editor_e3a.py extrair <ficheiro.E3A> <output.xlsx>
2. Editar valores na coluna REF do Excel
3. Aplicar alterações: python editor_e3a.py aplicar <original.E3A> <editor.xlsx> <output.E3A>
"""

import sys
import os
import struct
import zipfile
import tempfile
import re
import math
import subprocess
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# Constantes
SPACE_RECORD_SIZE = 682
WALL_BLOCK_SIZE = 34
WALL_BLOCK_START = 72
ROOF_BLOCK_SIZE = 24
ROOF_BLOCK_START = 344
WINDOW_RECORD_SIZE = 555
WALL_ASSEMBLY_SIZE = 3187
ROOF_ASSEMBLY_SIZE = 3187

# Constantes para assemblies (layers)
ASSEMBLY_SIZE = 3187
LAYER_SIZE = 281
LAYER_START = 377  # Offset onde começa o primeiro layer

# R-Values fixos das superfícies (em IP: ft²·°F·hr/BTU)
R_INSIDE_IP = 0.68    # ~0.12 SI
R_OUTSIDE_IP = 0.33   # ~0.06 SI

# Cores
PREV_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
REF_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
CHECK_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# OA encoding - exact closed-form formula (2026-02-05)
# HAP 5.1 uses a piecewise-linear "fast_exp2" approximation:
#   y = Y0 * fast_exp2(k * (x - 4))
#   k=4 for x<4 (base 16), k=2 for x>=4 (base 4)
#   Y0 = 512 CFM in L/s = 241.637...
_OA_Y0 = 512.0 * (28.316846592 / 60.0)

def _fast_exp2(t):
    n = math.floor(t)
    f = t - n
    return (2.0 ** n) * (1.0 + f)

def _fast_log2(v):
    n = math.floor(math.log2(v))
    f = v / (2.0 ** n) - 1.0
    if f < 0:
        n -= 1
        f = v / (2.0 ** n) - 1.0
    if f >= 1.0:
        n += 1
        f = v / (2.0 ** n) - 1.0
    return n + f

# =============================================================================
# CONVERSÕES SI -> Imperial
# =============================================================================

def m2_to_ft2(m2):
    return m2 * 10.7639 if m2 else 0

def m_to_ft(m):
    return m * 3.28084 if m else 0

def kg_m2_to_lb_ft2(kg):
    return kg * 0.204816 if kg else 0

def c_to_f(c):
    return c * 1.8 + 32 if c is not None else 32

def w_to_btu(w):
    return w * 3.412 if w else 0

def w_m2_to_w_ft2(w):
    return w / 10.764 if w else 0

def u_si_to_ip(u):
    return u / 5.678 if u else 0

def r_si_to_ip(r):
    return r * 5.678 if r else 0

def encode_oa(user_value):
    """Codifica valor OA do utilizador para formato interno HAP.
    Usa a formula piecewise fast_exp2 (identica ao excel_to_hap.py)."""
    if user_value <= 0:
        return 0
    try:
        v = float(user_value) / _OA_Y0
        t = _fast_log2(v)
        if t < 0:
            return t / 4.0 + 4.0
        else:
            return t / 2.0 + 4.0
    except:
        return 0


# =============================================================================
# ASSEMBLIES (LAYERS)
# =============================================================================

def fill_assembly_layers(data, offset, u_value_si, weight_si, absorptivity=0.9):
    """
    Preenche as layers de um Wall/Roof assembly para obter o U-Value e Weight desejados.

    HAP calcula o U-Value a partir das layers, não do valor escrito no offset 269.
    Por isso é necessário criar layers com o R-Value correcto.

    Args:
        data: bytearray do ficheiro DAT
        offset: offset do início do assembly
        u_value_si: U-Value desejado em W/m²K
        weight_si: Weight desejado em kg/m²
        absorptivity: Absorptivity (0-1), default 0.9
    """
    # Converter para IP
    u_value_ip = u_value_si / 5.678  # W/m²K -> BTU/hr·ft²·°F
    weight_ip = weight_si * 0.2048   # kg/m² -> lb/ft²

    # Calcular R-Value do material
    r_total_ip = 1.0 / u_value_ip if u_value_ip > 0 else 10.0
    r_material_ip = r_total_ip - R_INSIDE_IP - R_OUTSIDE_IP
    if r_material_ip < 0.01:
        r_material_ip = 0.01  # Mínimo

    # Usar thickness fixo de 0.1 ft (~30mm) e calcular density para dar o weight
    thickness_ft = 0.1
    density_lb_ft3 = weight_ip / thickness_ft if thickness_ft > 0 else 100.0

    # Absorptivity (offset 255)
    struct.pack_into('<f', data, offset + 255, absorptivity)

    # Surface Color = 2 (Dark) - offset 259
    data[offset + 259] = 2

    # Layer 0: Inside surface resistance (offset 377)
    layer0 = offset + LAYER_START
    inside_name = b'Inside surface resistance' + b' ' * (255 - 25)
    data[layer0:layer0+255] = inside_name
    struct.pack_into('<f', data, layer0 + 257, 0.0)        # Thickness
    struct.pack_into('<f', data, layer0 + 261, 0.0)        # Conductivity
    struct.pack_into('<f', data, layer0 + 265, 0.0)        # Density
    struct.pack_into('<f', data, layer0 + 269, 0.0)        # Specific Heat
    struct.pack_into('<f', data, layer0 + 273, R_INSIDE_IP) # R-Value
    struct.pack_into('<f', data, layer0 + 277, 0.0)        # Weight

    # Layer 1: Material principal (offset 377 + 281 = 658)
    layer1 = offset + LAYER_START + LAYER_SIZE
    material_name = b'Insulation' + b' ' * (255 - 10)
    data[layer1:layer1+255] = material_name
    struct.pack_into('<f', data, layer1 + 257, thickness_ft)      # Thickness
    struct.pack_into('<f', data, layer1 + 261, 0.02)              # Conductivity
    struct.pack_into('<f', data, layer1 + 265, density_lb_ft3)    # Density
    struct.pack_into('<f', data, layer1 + 269, 0.2)               # Specific Heat
    struct.pack_into('<f', data, layer1 + 273, r_material_ip)     # R-Value
    struct.pack_into('<f', data, layer1 + 277, weight_ip)         # Weight

    # Layer 2: Outside surface resistance (offset 377 + 281*2 = 939)
    layer2 = offset + LAYER_START + LAYER_SIZE * 2
    outside_name = b'Outside surface resistance' + b' ' * (255 - 26)
    data[layer2:layer2+255] = outside_name
    struct.pack_into('<f', data, layer2 + 257, 0.0)         # Thickness
    struct.pack_into('<f', data, layer2 + 261, 0.0)         # Conductivity
    struct.pack_into('<f', data, layer2 + 265, 0.0)         # Density
    struct.pack_into('<f', data, layer2 + 269, 0.0)         # Specific Heat
    struct.pack_into('<f', data, layer2 + 273, R_OUTSIDE_IP) # R-Value
    struct.pack_into('<f', data, layer2 + 277, 0.0)         # Weight

    # Limpar layers 3+ (preencher com nulls para o HAP não os detectar)
    for layer_idx in range(3, 9):
        layer_off = offset + LAYER_START + LAYER_SIZE * layer_idx
        if layer_off + LAYER_SIZE <= offset + ASSEMBLY_SIZE:
            data[layer_off:layer_off+255] = b'\x00' * 255
            for i in range(255, LAYER_SIZE):
                data[layer_off + i] = 0


# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

def extract_name(raw_bytes, max_len=24):
    """Extrai nome de bytes - usa apenas primeiros 24 bytes como o extractor"""
    try:
        name_str = raw_bytes[:max_len].decode('latin-1')
        name_str = name_str.split('\x00')[0]
        return name_str.strip()
    except:
        return ''


# =============================================================================
# EXTRACÇÃO PARA EXCEL
# =============================================================================

def extract_for_editing(e3a_path, output_xlsx):
    """Extrai E3A para Excel com formato PREV/REF"""

    print(f"A extrair {e3a_path} para edição...")

    extractor_path = os.path.join(os.path.dirname(__file__), '..', 'extractor', 'hap_extractor.py')
    temp_xlsx = output_xlsx.replace('.xlsx', '_temp.xlsx')

    result = subprocess.run(
        [sys.executable, extractor_path, e3a_path, temp_xlsx],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"Erro no extractor: {result.stderr}")
        return
    print(result.stdout)

    template_path = os.path.join(os.path.dirname(__file__), '..', 'comparador', 'Template_Comparacao.xlsx')

    if not os.path.exists(template_path):
        print(f"Template não encontrado: {template_path}")
        return

    wb_template = openpyxl.load_workbook(template_path)
    wb_data = openpyxl.load_workbook(temp_xlsx)

    # Processar Espacos
    ws = wb_template['Comparacao']
    ws_data = wb_data['Espacos']

    row_out = 4
    for row_in in range(4, ws_data.max_row + 1):
        space_name = ws_data.cell(row_in, 1).value
        if not space_name or str(space_name).strip() == '':
            continue

        out_col = 1
        for in_col in range(1, 148):
            value = ws_data.cell(row_in, in_col).value
            ws.cell(row_out, out_col, value=value).fill = PREV_FILL
            ws.cell(row_out, out_col + 1).fill = REF_FILL
            ws.cell(row_out, out_col + 2).fill = CHECK_FILL
            out_col += 3
        row_out += 1

    # Processar Windows, Walls, Roofs
    for sheet_name in ['Windows', 'Walls', 'Roofs']:
        if sheet_name in wb_data.sheetnames and sheet_name in wb_template.sheetnames:
            ws_t = wb_template[sheet_name]
            ws_d = wb_data[sheet_name]
            row_out = 4
            for row_in in range(4, ws_d.max_row + 1):
                if not ws_d.cell(row_in, 1).value:
                    continue
                out_col = 1
                for in_col in range(1, ws_d.max_column + 1):
                    if out_col > ws_t.max_column:
                        break
                    value = ws_d.cell(row_in, in_col).value
                    ws_t.cell(row_out, out_col, value=value).fill = PREV_FILL
                    ws_t.cell(row_out, out_col + 1).fill = REF_FILL
                    ws_t.cell(row_out, out_col + 2).fill = CHECK_FILL
                    out_col += 3
                row_out += 1

    wb_template.save(output_xlsx)
    if os.path.exists(temp_xlsx):
        os.remove(temp_xlsx)

    print(f"\nFicheiro criado: {output_xlsx}")
    print(f"  - Coluna PREV (amarelo): valores actuais")
    print(f"  - Coluna REF (verde): preencher com novos valores")


# =============================================================================
# APLICAR ALTERAÇÕES
# =============================================================================

def apply_changes(e3a_path, editor_xlsx, output_path):
    """Aplica alterações do Excel ao E3A"""

    print(f"A aplicar alterações de {editor_xlsx}...")

    wb = openpyxl.load_workbook(editor_xlsx, data_only=True)

    # Ler ficheiros do E3A
    with zipfile.ZipFile(e3a_path, 'r') as zf:
        spc_data = bytearray(zf.read('HAP51SPC.DAT'))
        win_data = bytearray(zf.read('HAP51WIN.DAT')) if 'HAP51WIN.DAT' in zf.namelist() else None
        wal_data = bytearray(zf.read('HAP51WAL.DAT')) if 'HAP51WAL.DAT' in zf.namelist() else None
        rof_data = bytearray(zf.read('HAP51ROF.DAT')) if 'HAP51ROF.DAT' in zf.namelist() else None
        inx_data = bytearray(zf.read('HAP51INX.MDB')) if 'HAP51INX.MDB' in zf.namelist() else None

    total_changes = 0

    # =========================================================================
    # MAPEAR WALL ASSEMBLIES (nome -> indice)
    # =========================================================================
    wall_name_to_idx = {}
    if wal_data:
        num_walls = len(wal_data) // WALL_ASSEMBLY_SIZE
        for i in range(num_walls):
            offset = i * WALL_ASSEMBLY_SIZE
            name = wal_data[offset:offset+100].split(b'\x00')[0].decode('latin-1').strip()
            wall_name_to_idx[name] = i

    # Lista para guardar alteracoes de wall type para actualizar Space_Wall_Links
    wall_type_changes = []  # [(space_index, wall_idx), ...]

    # =========================================================================
    # MAPEAR ROOF ASSEMBLIES (nome -> indice)
    # =========================================================================
    roof_name_to_idx = {}
    if rof_data:
        num_roofs = len(rof_data) // ROOF_ASSEMBLY_SIZE
        for i in range(num_roofs):
            offset = i * ROOF_ASSEMBLY_SIZE
            name = rof_data[offset:offset+100].split(b'\x00')[0].decode('latin-1').strip()
            roof_name_to_idx[name] = i

    # Lista para guardar alteracoes de roof type
    roof_type_changes = []  # [(space_name, space_idx, roof_idx), ...]

    # =========================================================================
    # ESPAÇOS
    # =========================================================================
    ws = wb['Comparacao']

    # Mapear nomes para offsets e indices
    space_offsets = {}
    space_indices = {}  # nome -> indice (para Space_Wall_Links)
    num_spaces = len(spc_data) // SPACE_RECORD_SIZE
    for i in range(num_spaces):
        offset = i * SPACE_RECORD_SIZE
        name = extract_name(spc_data[offset:offset+100])
        if name and not name.startswith('Default'):
            space_offsets[name] = offset
            space_indices[name] = i  # indice 0-based, SpaceIndex usa 1-based (i+1)

    # Mapeamento de campos (field_index 1-147) -> (offset, type, conversion)
    FIELD_MAP = {
        # GENERAL
        1: (0, 's24', None),           # Space Name
        2: (24, 'f', m2_to_ft2),       # Floor Area
        3: (28, 'f', m_to_ft),         # Ceiling Height
        4: (32, 'f', kg_m2_to_lb_ft2), # Building Weight
        5: (46, 'f', encode_oa),       # Outdoor Air (valor) - codificação especial
        # 6: OA Unit - código

        # PEOPLE
        7: (580, 'f', None),           # Occupancy
        # 8: Activity Level - código
        9: (586, 'f', w_to_btu),       # Sensible
        10: (590, 'f', w_to_btu),      # Latent
        # 11: Schedule

        # LIGHTING
        12: (600, 'f', None),          # Task Lighting
        13: (606, 'f', None),          # General Lighting
        # 14: Fixture Type
        15: (610, 'f', None),          # Ballast Mult
        # 16: Schedule

        # EQUIPMENT
        17: (656, 'f', w_m2_to_w_ft2), # Equipment W/m2
        # 18: Schedule

        # MISC
        19: (632, 'f', w_to_btu),      # Sensible
        20: (636, 'f', w_to_btu),      # Latent
        # 21, 22: Schedules

        # INFILTRATION
        # 23: Method
        24: (556, 'f', None),          # Design Clg ACH
        25: (562, 'f', None),          # Design Htg ACH
        26: (568, 'f', None),          # Energy ACH

        # FLOORS
        # 27: Floor Type
        28: (494, 'f', m2_to_ft2),     # Floor Area
        29: (498, 'f', u_si_to_ip),    # U-Value
        30: (502, 'f', m_to_ft),       # Exp Perim
        31: (506, 'f', r_si_to_ip),    # Edge R
        32: (510, 'f', m_to_ft),       # Depth
        33: (514, 'f', u_si_to_ip),    # Bsmt Wall U
        34: (518, 'f', r_si_to_ip),    # Wall Ins R
        35: (522, 'f', m_to_ft),       # Ins Depth
        36: (526, 'f', c_to_f),        # Unc Max
        37: (530, 'f', c_to_f),        # Out Max
        38: (534, 'f', c_to_f),        # Unc Min
        39: (538, 'f', c_to_f),        # Out Min

        # PARTITIONS - CEILING
        40: (442, 'f', m2_to_ft2),     # Area
        41: (446, 'f', u_si_to_ip),    # U-Value
        42: (450, 'f', c_to_f),        # Unc Max
        43: (454, 'f', c_to_f),        # Out Max
        44: (458, 'f', c_to_f),        # Unc Min
        45: (462, 'f', c_to_f),        # Out Min

        # PARTITIONS - WALL
        46: (468, 'f', m2_to_ft2),     # Area
        47: (472, 'f', u_si_to_ip),    # U-Value
        48: (476, 'f', c_to_f),        # Unc Max
        49: (480, 'f', c_to_f),        # Out Max
        50: (484, 'f', c_to_f),        # Unc Min
        51: (488, 'f', c_to_f),        # Out Min
    }

    # WALLS (52-123) - 8 walls x 9 campos cada
    # Estrutura do Wall Block (34 bytes):
    #   +0-1: Exposure/Direction (H)
    #   +2-5: Gross Area (f)
    #   +6-7: Wall Type ID (H) <- ESTE é o campo do Wall Type!
    #   +8-9: Window 1 ID, etc.
    # Campos Excel: Exposure(52), Gross Area(53), Wall Type(54), ...
    for w in range(8):
        base_field = 52 + w * 9
        wall_offset = WALL_BLOCK_START + w * WALL_BLOCK_SIZE
        FIELD_MAP[base_field + 1] = (wall_offset + 2, 'f', m2_to_ft2)  # Gross Area (campo 53, 62, ...)
        FIELD_MAP[base_field + 2] = (wall_offset + 6, 'wall_idx', None)  # Wall Type no offset +6 (campo 54, 63, ...)

    # ROOFS (124-147) - 4 roofs x 6 campos
    # Estrutura Roof Block (24 bytes): +0=Exposure, +4=Area, +8=Roof Type ID
    for r in range(4):
        base_field = 124 + r * 6
        roof_offset = ROOF_BLOCK_START + r * ROOF_BLOCK_SIZE
        FIELD_MAP[base_field + 1] = (roof_offset + 4, 'f', m2_to_ft2)  # Gross Area (campo 125, 131, ...)
        FIELD_MAP[base_field + 2] = (roof_offset + 8, 'roof_idx', None)  # Roof Type no offset +8 (campo 126, 132, ...)

    spc_changes_for_inx = []

    changes = 0
    for row in range(4, ws.max_row + 1):
        space_name = ws.cell(row, 1).value  # Nome está no PREV
        if not space_name or space_name not in space_offsets:
            continue

        space_offset = space_offsets[space_name]
        spc_inx_vals = {'name': space_name, 'area': None, 'occ': None}

        for col_prev in range(1, ws.max_column + 1, 3):
            col_ref = col_prev + 1
            ref_value = ws.cell(row, col_ref).value

            if ref_value is None or ref_value == '':
                continue

            field_idx = (col_prev - 1) // 3 + 1
            if field_idx not in FIELD_MAP:
                continue

            offset, ftype, conv = FIELD_MAP[field_idx]
            abs_offset = space_offset + offset

            try:
                if ftype == 'f':
                    val = float(ref_value)
                    val_converted = conv(val) if conv else val
                    spc_data[abs_offset:abs_offset+4] = struct.pack('<f', val_converted)
                    changes += 1

                    # Infiltration: garantir flag ACH mode (2) nos offsets 554/560/566
                    if field_idx in (24, 25, 26):
                        flag_offsets = {24: 554, 25: 560, 26: 566}
                        flag_abs = space_offset + flag_offsets[field_idx]
                        struct.pack_into('<H', spc_data, flag_abs, 2)

                    # Guardar para INX
                    if field_idx == 2:  # Floor Area
                        spc_inx_vals['area'] = val_converted
                    elif field_idx == 7:  # Occupancy
                        spc_inx_vals['occ'] = val  # Occupancy nao tem conversao
                elif ftype.startswith('s'):
                    str_len = int(ftype[1:])
                    encoded = str(ref_value).encode('latin-1')[:str_len-1]
                    encoded = encoded + b'\x00' * (str_len - len(encoded))
                    spc_data[abs_offset:abs_offset+str_len] = encoded
                    changes += 1
                elif ftype == 'wall_idx':
                    # Mapear nome da wall assembly para indice
                    wall_name = str(ref_value).strip()
                    if wall_name in wall_name_to_idx:
                        wall_idx = wall_name_to_idx[wall_name]
                        # Wall Type ID está no offset +6 do wall block (2 bytes, little-endian)
                        struct.pack_into('<H', spc_data, abs_offset, wall_idx)
                        changes += 1
                        # Guardar para actualizar Space_Wall_Links
                        space_idx_1based = space_indices[space_name]
                        wall_type_changes.append((space_name, space_idx_1based, wall_idx))
                elif ftype == 'roof_idx':
                    # Mapear nome do roof assembly para indice
                    roof_name = str(ref_value).strip()
                    if roof_name in roof_name_to_idx:
                        roof_idx = roof_name_to_idx[roof_name]
                        # Roof Type ID está no offset +8 do roof block (2 bytes, little-endian)
                        struct.pack_into('<H', spc_data, abs_offset, roof_idx)
                        changes += 1
                        # Guardar para actualizar Space_Roof_Links (se existir)
                        space_idx_1based = space_indices[space_name]
                        roof_type_changes.append((space_name, space_idx_1based, roof_idx))
            except:
                pass

        if spc_inx_vals['area'] is not None or spc_inx_vals['occ'] is not None:
            spc_changes_for_inx.append(spc_inx_vals)

    print(f"  Espaços: {changes} campos")
    total_changes += changes

    # Actualizar INX para Espacos
    if inx_data and spc_changes_for_inx:
        spc_inx_changes = 0
        for spc_vals in spc_changes_for_inx:
            name_bytes = spc_vals['name'].encode('latin-1')
            idx = inx_data.find(name_bytes)
            if idx != -1:
                if spc_vals['area'] is not None:
                    inx_data[idx-14:idx-10] = struct.pack('<f', spc_vals['area'])
                    spc_inx_changes += 1
                if spc_vals['occ'] is not None:
                    inx_data[idx-10:idx-6] = struct.pack('<f', spc_vals['occ'])
                    spc_inx_changes += 1
        print(f"  Espaços INX (display): {spc_inx_changes} campos")

    # =========================================================================
    # WINDOWS
    # =========================================================================
    win_changes_for_inx = []  # Guardar alteracoes para actualizar INX

    if win_data and 'Windows' in wb.sheetnames:
        ws_win = wb['Windows']

        # Mapear nomes para offsets
        win_offsets = {}
        num_windows = len(win_data) // WINDOW_RECORD_SIZE
        for i in range(num_windows):
            offset = i * WINDOW_RECORD_SIZE
            name = extract_name(win_data[offset:offset+100])
            if name:
                win_offsets[name] = offset

        # Campos: Nome(1), U-Value(2), SHGC(3), Altura(4), Largura(5)
        WIN_MAP = {
            2: (269, 'f', u_si_to_ip),    # U-Value
            3: (273, 'f', None),          # SHGC (offset 273, confirmado)
            4: (257, 'f', m_to_ft),       # Altura
            5: (261, 'f', m_to_ft),       # Largura
        }

        changes = 0
        for row in range(4, ws_win.max_row + 1):
            win_name = ws_win.cell(row, 1).value
            if not win_name or win_name not in win_offsets:
                continue

            win_offset = win_offsets[win_name]

            # Guardar valores para INX (todos os 4 campos)
            win_inx_vals = {'name': win_name, 'u': None, 'shgc': None, 'h': None, 'w': None}

            for col_prev in range(1, 16, 3):  # 5 campos x 3 = 15 colunas
                col_ref = col_prev + 1
                ref_value = ws_win.cell(row, col_ref).value

                if ref_value is None or ref_value == '':
                    continue

                field_idx = (col_prev - 1) // 3 + 1
                if field_idx not in WIN_MAP:
                    continue

                offset, ftype, conv = WIN_MAP[field_idx]
                abs_offset = win_offset + offset

                try:
                    val = float(ref_value)
                    val_converted = conv(val) if conv else val
                    win_data[abs_offset:abs_offset+4] = struct.pack('<f', val_converted)
                    changes += 1

                    # Guardar para INX (valores ja convertidos para IP)
                    if field_idx == 2:
                        win_inx_vals['u'] = val_converted
                    elif field_idx == 3:
                        win_inx_vals['shgc'] = val_converted
                    elif field_idx == 4:
                        win_inx_vals['h'] = val_converted
                    elif field_idx == 5:
                        win_inx_vals['w'] = val_converted
                except:
                    pass

            # Se houve alteracoes, guardar para INX
            if any([win_inx_vals['u'], win_inx_vals['shgc'], win_inx_vals['h'], win_inx_vals['w']]):
                win_changes_for_inx.append(win_inx_vals)

        print(f"  Windows: {changes} campos")
        total_changes += changes

    # =========================================================================
    # ACTUALIZAR HAP51INX.MDB (display values para Windows)
    # =========================================================================
    inx_changes = 0
    if inx_data and win_changes_for_inx:
        for win_vals in win_changes_for_inx:
            win_name = win_vals['name']
            # Procurar nome da janela no INX
            name_bytes = win_name.encode('latin-1')
            idx = inx_data.find(name_bytes)

            if idx == -1:
                continue

            # Estrutura: U-Value(-18), SHGC(-14), Height(-10), Width(-6) antes do nome
            if win_vals['u'] is not None:
                inx_data[idx-18:idx-14] = struct.pack('<f', win_vals['u'])
                inx_changes += 1
            if win_vals['shgc'] is not None:
                inx_data[idx-14:idx-10] = struct.pack('<f', win_vals['shgc'])
                inx_changes += 1
            if win_vals['h'] is not None:
                inx_data[idx-10:idx-6] = struct.pack('<f', win_vals['h'])
                inx_changes += 1
            if win_vals['w'] is not None:
                inx_data[idx-6:idx-2] = struct.pack('<f', win_vals['w'])
                inx_changes += 1

        print(f"  Windows INX (display): {inx_changes} campos")

    # =========================================================================
    # WALLS (Assemblies) - Usa fill_assembly_layers para U-Value correcto
    # =========================================================================
    wal_changes_for_inx = []

    if wal_data and 'Walls' in wb.sheetnames:
        ws_wal = wb['Walls']

        # Mapear nomes para offsets
        wal_offsets = {}
        num_walls = len(wal_data) // WALL_ASSEMBLY_SIZE
        for i in range(num_walls):
            offset = i * WALL_ASSEMBLY_SIZE
            name = extract_name(wal_data[offset:offset+255])
            if name:
                wal_offsets[name] = offset

        # Campos: Nome(1), U-Value(2), Espessura(3), Massa(4), Absorptivity(5)
        # Nota: Espessura é ignorada (calculada automaticamente nas layers)

        changes = 0
        for row in range(4, ws_wal.max_row + 1):
            wal_name = ws_wal.cell(row, 1).value
            if not wal_name or wal_name not in wal_offsets:
                continue

            wal_offset = wal_offsets[wal_name]

            # Recolher valores REF para esta wall
            u_value_si = None
            weight_si = None
            absorptivity = 0.9  # default

            # Campo 2: U-Value (coluna REF = 5)
            ref_u = ws_wal.cell(row, 5).value
            if ref_u is not None and ref_u != '':
                try:
                    u_value_si = float(ref_u)
                except:
                    pass

            # Campo 4: Massa (coluna REF = 11)
            ref_massa = ws_wal.cell(row, 11).value
            if ref_massa is not None and ref_massa != '':
                try:
                    weight_si = float(ref_massa)
                except:
                    pass

            # Campo 5: Absorptivity (coluna REF = 14) - se existir
            if ws_wal.max_column >= 14:
                ref_abs = ws_wal.cell(row, 14).value
                if ref_abs is not None and ref_abs != '':
                    try:
                        absorptivity = float(ref_abs)
                    except:
                        pass

            # Se temos U-Value ou Weight, preencher layers
            if u_value_si is not None or weight_si is not None:
                # Usar valores existentes se não foram especificados
                if u_value_si is None:
                    # Ler U-Value actual do offset 269 (display value)
                    u_ip = struct.unpack_from('<f', wal_data, wal_offset + 269)[0]
                    u_value_si = u_ip * 5.678 if u_ip > 0 else 1.0

                if weight_si is None:
                    # Ler Weight actual do offset 273
                    w_ip = struct.unpack_from('<f', wal_data, wal_offset + 273)[0]
                    weight_si = w_ip / 0.2048 if w_ip > 0 else 50.0

                # Preencher layers com valores calculados
                fill_assembly_layers(wal_data, wal_offset, u_value_si, weight_si, absorptivity)
                changes += 1

                # Guardar U-Value para INX (em IP)
                u_ip_for_inx = u_value_si / 5.678
                wal_changes_for_inx.append({'name': wal_name, 'u': u_ip_for_inx})

        print(f"  Walls: {changes} assemblies actualizados (layers)")
        total_changes += changes

    # Actualizar INX para Walls
    if inx_data and wal_changes_for_inx:
        wal_inx_changes = 0
        for wal_vals in wal_changes_for_inx:
            name_bytes = wal_vals['name'].encode('latin-1')
            idx = inx_data.find(name_bytes)
            if idx != -1 and wal_vals['u'] is not None:
                inx_data[idx-14:idx-10] = struct.pack('<f', wal_vals['u'])
                wal_inx_changes += 1
        print(f"  Walls INX (display): {wal_inx_changes} campos")

    # =========================================================================
    # ROOFS (Assemblies) - Usa fill_assembly_layers para U-Value correcto
    # =========================================================================
    rof_changes_for_inx = []

    if rof_data and 'Roofs' in wb.sheetnames:
        ws_rof = wb['Roofs']

        # Mapear nomes para offsets
        rof_offsets = {}
        num_roofs = len(rof_data) // ROOF_ASSEMBLY_SIZE
        for i in range(num_roofs):
            offset = i * ROOF_ASSEMBLY_SIZE
            name = extract_name(rof_data[offset:offset+255])
            if name:
                rof_offsets[name] = offset

        # Campos: Nome(1), U-Value(2), Espessura(3), Massa(4), Absorptivity(5)
        # Nota: Espessura é ignorada (calculada automaticamente nas layers)

        changes = 0
        for row in range(4, ws_rof.max_row + 1):
            rof_name = ws_rof.cell(row, 1).value
            if not rof_name or rof_name not in rof_offsets:
                continue

            rof_offset = rof_offsets[rof_name]

            # Recolher valores REF para este roof
            u_value_si = None
            weight_si = None
            absorptivity = 0.9  # default

            # Campo 2: U-Value (coluna REF = 5)
            ref_u = ws_rof.cell(row, 5).value
            if ref_u is not None and ref_u != '':
                try:
                    u_value_si = float(ref_u)
                except:
                    pass

            # Campo 4: Massa (coluna REF = 11)
            ref_massa = ws_rof.cell(row, 11).value
            if ref_massa is not None and ref_massa != '':
                try:
                    weight_si = float(ref_massa)
                except:
                    pass

            # Campo 5: Absorptivity (coluna REF = 14) - se existir
            if ws_rof.max_column >= 14:
                ref_abs = ws_rof.cell(row, 14).value
                if ref_abs is not None and ref_abs != '':
                    try:
                        absorptivity = float(ref_abs)
                    except:
                        pass

            # Se temos U-Value ou Weight, preencher layers
            if u_value_si is not None or weight_si is not None:
                # Usar valores existentes se não foram especificados
                if u_value_si is None:
                    # Ler U-Value actual do offset 269 (display value)
                    u_ip = struct.unpack_from('<f', rof_data, rof_offset + 269)[0]
                    u_value_si = u_ip * 5.678 if u_ip > 0 else 1.0

                if weight_si is None:
                    # Ler Weight actual do offset 273
                    w_ip = struct.unpack_from('<f', rof_data, rof_offset + 273)[0]
                    weight_si = w_ip / 0.2048 if w_ip > 0 else 50.0

                # Preencher layers com valores calculados
                fill_assembly_layers(rof_data, rof_offset, u_value_si, weight_si, absorptivity)
                changes += 1

                # Guardar U-Value para INX (em IP)
                u_ip_for_inx = u_value_si / 5.678
                rof_changes_for_inx.append({'name': rof_name, 'u': u_ip_for_inx})

        print(f"  Roofs: {changes} assemblies actualizados (layers)")
        total_changes += changes

    # Actualizar INX para Roofs
    if inx_data and rof_changes_for_inx:
        rof_inx_changes = 0
        for rof_vals in rof_changes_for_inx:
            name_bytes = rof_vals['name'].encode('latin-1')
            idx = inx_data.find(name_bytes)
            if idx != -1 and rof_vals['u'] is not None:
                inx_data[idx-14:idx-10] = struct.pack('<f', rof_vals['u'])
                rof_inx_changes += 1
        print(f"  Roofs INX (display): {rof_inx_changes} campos")

    # =========================================================================
    # GUARDAR E3A
    # =========================================================================
    if total_changes == 0:
        print("\nNenhuma alteração encontrada na coluna REF")
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(e3a_path, 'r') as zf:
            zf.extractall(tmpdir)

        # Guardar ficheiros modificados
        with open(os.path.join(tmpdir, 'HAP51SPC.DAT'), 'wb') as f:
            f.write(spc_data)
        if win_data:
            with open(os.path.join(tmpdir, 'HAP51WIN.DAT'), 'wb') as f:
                f.write(win_data)
        if wal_data:
            with open(os.path.join(tmpdir, 'HAP51WAL.DAT'), 'wb') as f:
                f.write(wal_data)
        if rof_data:
            with open(os.path.join(tmpdir, 'HAP51ROF.DAT'), 'wb') as f:
                f.write(rof_data)
        if inx_data:
            with open(os.path.join(tmpdir, 'HAP51INX.MDB'), 'wb') as f:
                f.write(inx_data)

        # =====================================================================
        # ACTUALIZAR Space_Wall_Links no HAP51INX.MDB (para Wall Type aparecer no HAP)
        # =====================================================================
        if wall_type_changes:
            try:
                import pyodbc
                inx_path = os.path.join(tmpdir, 'HAP51INX.MDB')
                conn = pyodbc.connect(f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={inx_path}')
                cursor = conn.cursor()

                # Obter mapeamento de space names para SpaceIndex.nIndex
                cursor.execute('SELECT nIndex, szName FROM SpaceIndex')
                space_name_to_inx_id = {row[1]: row[0] for row in cursor.fetchall()}

                # Obter mapeamento de wall indices do DAT para WallIndex.nIndex
                # WallIndex tem nIndex que corresponde ao indice no DAT
                cursor.execute('SELECT nIndex, szName FROM WallIndex')
                wall_inx = {row[1]: row[0] for row in cursor.fetchall()}

                wall_links_updated = 0
                for space_name, space_dat_idx, wall_dat_idx in wall_type_changes:
                    # Obter Space_ID do SpaceIndex
                    if space_name not in space_name_to_inx_id:
                        continue
                    space_id = space_name_to_inx_id[space_name]

                    # O wall_dat_idx e o indice no ficheiro DAT (0-based)
                    # Mas WallIndex so tem as walls "validas" com nIndex = indice no DAT
                    # Então precisamos usar wall_dat_idx directamente como Wall_ID

                    # Primeiro remover link existente se houver
                    cursor.execute('DELETE FROM Space_Wall_Links WHERE Space_ID = ?', (space_id,))

                    # Inserir novo link (Wall_ID = indice no DAT)
                    cursor.execute('INSERT INTO Space_Wall_Links (Space_ID, Wall_ID) VALUES (?, ?)',
                                   (space_id, wall_dat_idx))
                    wall_links_updated += 1

                conn.commit()
                conn.close()
                print(f"  Space_Wall_Links: {wall_links_updated} links actualizados")
            except Exception as e:
                print(f"  AVISO: Não foi possível actualizar Space_Wall_Links: {e}")

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(tmpdir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, tmpdir)
                    zf.write(file_path, arc_name)

    print(f"\n  TOTAL: {total_changes} campos actualizados")
    print(f"  Ficheiro criado: {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    cmd = sys.argv[1].lower()

    if cmd == 'extrair':
        if len(sys.argv) < 4:
            print("Uso: python editor_e3a.py extrair <ficheiro.E3A> <output.xlsx>")
            sys.exit(1)
        extract_for_editing(sys.argv[2], sys.argv[3])

    elif cmd == 'aplicar':
        if len(sys.argv) < 5:
            print("Uso: python editor_e3a.py aplicar <original.E3A> <editor.xlsx> <output.E3A>")
            sys.exit(1)
        apply_changes(sys.argv[2], sys.argv[3], sys.argv[4])

    else:
        print(f"Comando desconhecido: {cmd}")
        sys.exit(1)


if __name__ == '__main__':
    main()
