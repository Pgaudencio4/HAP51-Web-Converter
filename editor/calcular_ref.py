"""
Calcular valores REF (Referência RECS) para folha de comparação/edição

Preenche automaticamente as colunas REF com base nos valores de referência
do RECS (Regulamento de Desempenho Energético dos Edifícios de Comércio e Serviços).

Usage:
    python calcular_ref.py <ficheiro_editor.xlsx> [output.xlsx]

Exemplo:
    python calcular_ref.py Malhoa22_EDITOR.xlsx Malhoa22_EDITOR_REF.xlsx

Se não especificar output, cria ficheiro com sufixo _REF.

===============================================================================
FÓRMULAS UTILIZADAS
===============================================================================

1. ILUMINAÇÃO REF
-----------------
Potência REF (W) = DPI_ref × (Iluminância / 100) × Área

Onde:
    DPI_ref = 2.5 W/m²/100lux (valor de referência RECS)
    Iluminância = nível de lux do espaço (depende do tipo)

Níveis de Iluminância por Tipo de Espaço (EN 12464-1):
    | Tipo               | Lux | W/m² REF |
    |--------------------|-----|----------|
    | Escritório/Area    | 500 |    12.5  |
    | Copa/Cozinha       | 300 |     7.5  |
    | IS/WC              | 200 |     5.0  |
    | Zonas técnicas     | 200 |     5.0  |
    | Escadas            | 150 |    3.75  |
    | Circulação         | 100 |     2.5  |
    | Estacionamento     |  75 |   1.875  |

2. CAUDAIS AR NOVO REF
----------------------
Caudal REF = máx(Qocup, Qedif)

Onde:
    Qocup = nº ocupantes × caudal/pessoa
    Qedif = área × caudal/m²

Valores RECS por Actividade Metabólica:
    | Actividade           | met  | Caudal/pessoa |
    |----------------------|------|---------------|
    | Sedentária           | 1.2  | 24 m³/h = 6.67 L/s |
    | Moderada             | 1.75 | 35 m³/h = 9.72 L/s |
    | Alta                 | 5.0  | 98 m³/h = 27.2 L/s |

Valores RECS por Carga Poluente:
    | Carga Poluente              | Caudal/m²           |
    |-----------------------------|---------------------|
    | Sem emissão poluentes       | 3 m³/h/m² = 0.833 L/s/m² |
    | Com emissão poluentes       | 5 m³/h/m² = 1.389 L/s/m² |

Por defeito usa: Sedentária (6.67 L/s/pessoa) + Sem poluentes (0.833 L/s/m²)

===============================================================================
"""

import openpyxl
import sys
import os

# ============================================================================
# CONSTANTES RECS
# ============================================================================

DPI_REF = 2.5  # W/m²/100lux

# Iluminância por tipo de espaço (lux)
LUX_MAP = {
    'area': 500,           # Escritórios/Áreas de trabalho
    'escritorio': 500,     # Escritórios
    'office': 500,         # Escritórios (inglês)
    'sala': 500,           # Salas
    'reuniao': 500,        # Salas de reunião
    'recepcao': 300,       # Recepção
    'circulacao': 100,     # Corredores/Circulação
    'corredor': 100,       # Corredores
    'hall': 100,           # Hall
    'caixa_de_escada': 150, # Escadas
    'escada': 150,         # Escadas
    'is': 200,             # Instalações sanitárias
    'wc': 200,             # WC
    'balneario': 200,      # Balneários
    'zonas_tecnicas': 200, # Zonas técnicas
    'tecnica': 200,        # Zonas técnicas
    'arrumos': 100,        # Arrumos
    'arquivo': 200,        # Arquivo
    'estacionamento': 75,  # Estacionamento
    'garagem': 75,         # Garagem
    'parking': 75,         # Parking
    'copa': 300,           # Copa
    'cozinha': 500,        # Cozinha
    'refeitorio': 200,     # Refeitório
    'restaurante': 200,    # Restaurante
    'bar': 200,            # Bar
    'loja': 300,           # Loja/Comércio
    'comercio': 300,       # Comércio
    'quarto': 300,         # Quarto (hotel)
    'ginasio': 300,        # Ginásio
    'piscina': 300,        # Piscina
}

# Caudais RECS (em m³/h - unidade base para cálculo)
CAUDAL_PESSOA_M3H = {
    'sedentaria': 24,      # m³/h/pessoa
    'moderada': 35,        # m³/h/pessoa
    'alta': 98,            # m³/h/pessoa
}

CAUDAL_M2_M3H = {
    'sem_poluentes': 3,    # m³/h/m²
    'com_poluentes': 5,    # m³/h/m²
}

# Eficácia de ventilação REF
EFICACIA_COM_OCUPACAO = 0.8    # Espaços com ocupação (escritórios, áreas, etc.)
EFICACIA_SEM_OCUPACAO = 1.0    # Espaços sem ocupação (escadas, circulação, IS, etc.)

# ============================================================================
# ESTRUTURA DA FOLHA DE COMPARAÇÃO
# ============================================================================

# Colunas da folha Comparacao (1-indexed)
COLS = {
    'NAME': 1,          # A - Space Name (PREV)
    'AREA': 4,          # D - Floor Area (m2) (PREV)
    'OA_PREV': 13,      # M - Outdoor Air (valor) (PREV)
    'OA_REF': 14,       # N - Outdoor Air (valor) (REF)
    'OCUP': 19,         # S - Occupancy (people) (PREV)
    'TASK_PREV': 34,    # AH - Task Lighting (W) (PREV)
    'TASK_REF': 35,     # AI - Task Lighting (W) (REF)
    'GEN_PREV': 37,     # AK - General Ltg (W) (PREV)
    'GEN_REF': 38,      # AL - General Ltg (W) (REF)
}

# ============================================================================
# FUNÇÕES
# ============================================================================

def get_lux_from_name(name):
    """
    Determina iluminância (lux) baseado no nome do espaço.

    Procura palavras-chave no nome para identificar o tipo de espaço.
    Se não encontrar, assume escritório (500 lux).
    """
    if not name:
        return 500

    name_lower = str(name).lower()

    for key, lux in LUX_MAP.items():
        if key in name_lower:
            return lux

    return 500  # default: escritório


def calcular_ref(filepath, output_path=None, actividade='sedentaria', poluentes=False):
    """
    Preenche colunas REF na folha de comparação/edição.

    Args:
        filepath: Caminho para o ficheiro Excel
        output_path: Caminho para output (opcional)
        actividade: 'sedentaria', 'moderada' ou 'alta'
        poluentes: True se espaços têm emissão de poluentes

    Returns:
        Caminho do ficheiro guardado
    """

    if output_path is None:
        base, ext = os.path.splitext(filepath)
        output_path = f"{base}_REF{ext}"

    print(f"Input:  {filepath}")
    print(f"Output: {output_path}")
    print()

    # Parâmetros RECS (em m³/h)
    caudal_pessoa_m3h = CAUDAL_PESSOA_M3H.get(actividade, CAUDAL_PESSOA_M3H['sedentaria'])
    caudal_m2_m3h = CAUDAL_M2_M3H['com_poluentes'] if poluentes else CAUDAL_M2_M3H['sem_poluentes']

    print(f"Parâmetros RECS:")
    print(f"  Actividade:  {actividade} ({caudal_pessoa_m3h} m³/h/pessoa)")
    print(f"  Poluentes:   {'Sim' if poluentes else 'Não'} ({caudal_m2_m3h} m³/h/m²)")
    print(f"  Eficácia REF: 0.8 (com ocupação) / 1.0 (sem ocupação)")
    print(f"  DPI ref:     {DPI_REF} W/m²/100lux")
    print()

    # Abrir ficheiro
    wb = openpyxl.load_workbook(filepath)

    if 'Comparacao' not in wb.sheetnames:
        print("ERRO: Folha 'Comparacao' não encontrada!")
        print(f"Folhas disponíveis: {wb.sheetnames}")
        return None

    ws = wb['Comparacao']

    # Header
    print(f"{'Espaço':<30} | {'Área':>7} | {'Ocup':>4} | {'Efic':>4} | {'Lux':>3} | {'OA REF':>7} | {'Ltg REF':>8}")
    print("-" * 90)

    count = 0

    # Processar linhas (dados começam na linha 4)
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, COLS['NAME']).value
        if not name:
            continue

        # Ler valores
        area = ws.cell(row, COLS['AREA']).value
        ocup = ws.cell(row, COLS['OCUP']).value

        area_val = float(area) if area else 0
        ocup_val = int(ocup) if ocup else 0

        # Determinar iluminância
        lux = get_lux_from_name(name)

        # Determinar se espaço tem ocupação (para escolher eficácia)
        # Espaços com ocupação: Area, Escritorio, Sala, Copa, etc. -> eficácia 0.8
        # Espaços sem ocupação: Escadas, Circulação, IS, Zonas técnicas, Estacionamento -> eficácia 1.0
        tem_ocupacao = ocup_val > 0 or any(x in str(name).lower() for x in ['area', 'escritorio', 'sala', 'copa', 'quarto', 'gab'])
        eficacia = EFICACIA_COM_OCUPACAO if tem_ocupacao else EFICACIA_SEM_OCUPACAO

        # Calcular valores REF
        # Caudal mínimo RECS: máx(Qocup, Qedif) em m³/h
        q_ocup_m3h = ocup_val * caudal_pessoa_m3h
        q_edif_m3h = area_val * caudal_m2_m3h
        q_min_m3h = max(q_ocup_m3h, q_edif_m3h)

        # Caudal REF = Caudal mínimo / eficácia, convertido para L/s
        # Espaços com ocupação: divide por 0.8
        # Espaços sem ocupação: divide por 1.0 (não altera)
        q_ref_m3h = q_min_m3h / eficacia
        oa_ref_ls = q_ref_m3h / 3.6

        # Iluminação: DPI × (lux/100) × área
        wm2_ref = DPI_REF * (lux / 100)
        ltg_ref = wm2_ref * area_val

        # Escrever valores numéricos (arredondados) para compatibilidade com editor_e3a.py
        # OA REF em L/s (arredondado para cima)
        import math
        oa_ref_rounded = math.ceil(oa_ref_ls)
        ws.cell(row, COLS['OA_REF']).value = oa_ref_rounded

        # Task Lighting REF: 0
        ws.cell(row, COLS['TASK_REF']).value = 0

        # General Lighting REF: DPI × (lux/100) × área (arredondado)
        ltg_ref_rounded = round(ltg_ref)
        ws.cell(row, COLS['GEN_REF']).value = ltg_ref_rounded

        count += 1
        efic_str = "0.8" if tem_ocupacao else "1.0"
        print(f"{str(name)[:30]:<30} | {area_val:>7.1f} | {ocup_val:>4} | {efic_str:>4} | {lux:>3} | {oa_ref_ls:>7.0f} | {ltg_ref:>8.0f}")

    # Guardar
    wb.save(output_path)

    print()
    print(f"Total: {count} espaços processados")
    print(f"Guardado em: {output_path}")
    print()
    print("=" * 70)
    print("FÓRMULAS INSERIDAS:")
    print("=" * 70)
    print()
    print("Coluna N (OA REF) em L/s:")
    print(f"  Espaços COM ocupação: =ROUNDUP(MAX(S{{row}}*{caudal_pessoa_m3h}, D{{row}}*{caudal_m2_m3h})/0.8/3.6, 0)")
    print(f"  Espaços SEM ocupação: =ROUNDUP(MAX(S{{row}}*{caudal_pessoa_m3h}, D{{row}}*{caudal_m2_m3h})/1.0/3.6, 0)")
    print()
    print("  Onde:")
    print(f"    - S{{row}}*{caudal_pessoa_m3h} = Qocup (ocupantes × {caudal_pessoa_m3h} m³/h/pessoa)")
    print(f"    - D{{row}}*{caudal_m2_m3h} = Qedif (área × {caudal_m2_m3h} m³/h/m²)")
    print(f"    - /0.8 = eficácia REF para espaços com ocupação")
    print(f"    - /1.0 = eficácia REF para espaços sem ocupação")
    print(f"    - /3.6 = conversão m³/h para L/s")
    print()
    print("Coluna AI (Task Ltg REF):")
    print("  0")
    print()
    print("Coluna AL (Gen Ltg REF):")
    print("  =ROUND(2.5*(lux/100)*D{row}, 0)")
    print("  Onde lux é determinado pelo tipo de espaço")

    return output_path


# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUsage:")
        print("  python calcular_ref.py <ficheiro.xlsx> [output.xlsx]")
        print()
        print("Opções avançadas (editar script):")
        print("  actividade: 'sedentaria' (default), 'moderada', 'alta'")
        print("  poluentes:  False (default), True")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_file):
        print(f"ERRO: Ficheiro não encontrado: {input_file}")
        sys.exit(1)

    calcular_ref(input_file, output_file)
