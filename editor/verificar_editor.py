"""
Verificador de ficheiro EDITOR - Lista TODAS as colunas REF preenchidas

Usage:
    python verificar_editor.py <ficheiro_editor.xlsx>
"""

import sys
import openpyxl


def verificar_excel(filepath):
    """Verifica todas as colunas REF preenchidas no Excel EDITOR"""

    print(f"A verificar: {filepath}")
    print("=" * 70)

    wb = openpyxl.load_workbook(filepath)

    total_campos = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"\n{'='*70}")
        print(f"FOLHA: {sheet_name}")
        print(f"{'='*70}")

        campos_encontrados = []

        # Percorrer todas as colunas REF (2, 5, 8, 11, ...)
        for col in range(2, ws.max_column + 1, 3):
            # Contar valores preenchidos
            valores = []
            for row in range(4, ws.max_row + 1):
                val = ws.cell(row, col).value
                if val is not None and str(val).strip() != '':
                    valores.append(val)

            if valores:
                # Obter nome do campo (coluna PREV anterior)
                header_prev = ws.cell(3, col - 1).value or ''
                header_prev = header_prev.replace(' (PREV)', '').replace(' (REF)', '')

                field_idx = (col - 2) // 3 + 1

                # Valores unicos
                unique_vals = list(set(str(v) for v in valores))

                campos_encontrados.append({
                    'idx': field_idx,
                    'col': col,
                    'nome': header_prev,
                    'total': len(valores),
                    'valores': unique_vals[:5]
                })

        if campos_encontrados:
            print(f"\nCampos REF preenchidos: {len(campos_encontrados)}")
            print("-" * 70)

            for campo in campos_encontrados:
                print(f"\n  Campo {campo['idx']:3d} | Coluna {campo['col']:3d} | {campo['nome']}")
                print(f"            | Preenchidos: {campo['total']}")
                print(f"            | Valores: {campo['valores']}")
                total_campos += campo['total']
        else:
            print("\n  (Nenhum campo REF preenchido)")

    print(f"\n{'='*70}")
    print(f"RESUMO TOTAL: {total_campos} valores REF a aplicar")
    print(f"{'='*70}")

    return total_campos


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    verificar_excel(sys.argv[1])
