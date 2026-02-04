"""
HAP 5.1 Extractor - Extrai TODOS os dados de um ficheiro E3A para Excel

Usa o mesmo formato do template HAP_Template_RSECE.xlsx com 147 colunas.

Usage:
    python hap_extractor.py <input.E3A> [output.xlsx]

Exemplo:
    python hap_extractor.py Malhoa22_ComSistemas.E3A Malhoa22_Extraido.xlsx
"""

import zipfile
import struct
import sys
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# =============================================================================
# CONSTANTES
# =============================================================================

SPACE_RECORD_SIZE = 682
WALL_BLOCK_SIZE = 34
WALL_BLOCK_START = 72
ROOF_BLOCK_SIZE = 24
ROOF_BLOCK_START = 344
SCHEDULE_RECORD_SIZE = 792
WALL_ASSEMBLY_SIZE = 3187
ROOF_ASSEMBLY_SIZE = 3187
WINDOW_RECORD_SIZE = 555

# Códigos de direcção (inverso)
DIRECTION_NAMES = {
    0: '', 1: 'N', 2: 'NNE', 3: 'NE', 4: 'ENE',
    5: 'E', 6: 'ESE', 7: 'SE', 8: 'SSE',
    9: 'S', 10: 'SSW', 11: 'SW', 12: 'WSW',
    13: 'W', 14: 'WNW', 15: 'NW', 16: 'NNW',
    17: 'H'  # Horizontal
}

ACTIVITY_NAMES = {
    0: 'Seated at Rest', 1: 'Seated at Rest', 2: 'Office Work',
    3: 'Office Work', 4: 'Sedentary Work', 5: 'Medium Work',
    6: 'Heavy Work', 7: 'Dancing', 8: 'Athletics'
}

FIXTURE_NAMES = {
    0: 'Recessed Unvented', 1: 'Vented to Return Air',
    2: 'Vented to Supply & Return', 3: 'Surface Mount/Pendant'
}

FLOOR_TYPE_NAMES = {
    0: '', 1: 'Floor Above Cond Space', 2: 'Floor Above Uncond Space',
    3: 'Slab Floor On Grade', 4: 'Slab Floor Below Grade'
}

OA_UNIT_NAMES = {
    0: '', 1: 'L/s', 2: 'L/s/m2', 3: 'L/s/person', 4: '%'
}

# OA decoding constants
OA_A = 0.00470356
OA_B = 2.71147770

import math

# =============================================================================
# CONVERSÕES (Imperial para SI)
# =============================================================================

def ft2_to_m2(ft2):
    return ft2 * 0.092903 if ft2 else 0

def ft_to_m(ft):
    return ft * 0.3048 if ft else 0

def lb_ft2_to_kg_m2(lb):
    return lb * 4.8824 if lb else 0

def f_to_c(f):
    return (f - 32) / 1.8 if f else 0

def btu_to_w(btu):
    return btu / 3.412 if btu else 0

def w_ft2_to_w_m2(w):
    return w * 10.764 if w else 0

def u_ip_to_si(u):
    """BTU/(hr·ft²·°F) para W/(m²·K)"""
    return u * 5.678 if u else 0

def r_ip_to_si(r):
    """(hr·ft²·°F)/BTU para (m²·K)/W"""
    return r / 5.678 if r else 0

def decode_oa(internal_float):
    """Descodifica valor interno OA para valor do utilizador"""
    if internal_float <= 0:
        return 0
    try:
        return OA_A * math.exp(OA_B * internal_float)
    except:
        return 0

def clean_string(s):
    """Remove caracteres nulos e de controlo de uma string"""
    if not s:
        return ''
    # Remove caracteres de controlo (0x00-0x1F exceto espaços)
    s = ''.join(c if (ord(c) >= 32 or c in '\t\n\r') else '' for c in s)
    return s.strip()

# =============================================================================
# LEITURA DE DADOS DO E3A
# =============================================================================

def read_e3a(filepath):
    """Lê um ficheiro E3A e retorna os dados em formato raw"""
    with zipfile.ZipFile(filepath, 'r') as z:
        files = {name: z.read(name) for name in z.namelist()}
    return files

def extract_spaces(spc_data):
    """Extrai todos os espaços do HAP51SPC.DAT"""
    num_spaces = len(spc_data) // SPACE_RECORD_SIZE
    spaces = []

    for i in range(num_spaces):
        offset = i * SPACE_RECORD_SIZE
        record = spc_data[offset:offset + SPACE_RECORD_SIZE]

        space = extract_space_record(record, i)
        spaces.append(space)

    return spaces

def extract_space_record(data, index):
    """Extrai todos os campos de um registo de espaço (682 bytes)"""
    space = {'_index': index}

    # === GENERAL (cols 1-6) ===
    # 0-23: Nome (24 bytes)
    raw_name = data[0:24]
    try:
        space['name'] = raw_name.split(b'\x00')[0].decode('latin-1').strip()
    except:
        space['name'] = f'Space_{index}'

    # 24-27: Floor Area (ft²)
    area_ft2 = struct.unpack('<f', data[24:28])[0]
    space['area_m2'] = round(ft2_to_m2(area_ft2), 2)

    # 28-31: Ceiling Height (ft)
    height_ft = struct.unpack('<f', data[28:32])[0]
    space['height_m'] = round(ft_to_m(height_ft), 2)

    # 32-35: Building Weight (lb/ft²)
    weight_lb = struct.unpack('<f', data[32:36])[0]
    space['weight_kg_m2'] = round(lb_ft2_to_kg_m2(weight_lb), 1)

    # 46-49: OA Internal Value
    oa_internal = struct.unpack('<f', data[46:50])[0]
    space['oa_value'] = round(decode_oa(oa_internal), 2)

    # 50-51: OA Unit Code
    oa_unit = struct.unpack('<H', data[50:52])[0]
    space['oa_unit'] = OA_UNIT_NAMES.get(oa_unit, '')

    # === INTERNALS - PEOPLE (cols 7-11) ===
    # 580-583: Occupancy
    occupancy = struct.unpack('<f', data[580:584])[0]
    space['occupancy'] = round(occupancy, 1) if occupancy else 0

    # 584-585: Activity Level ID
    activity_id = struct.unpack('<H', data[584:586])[0]
    space['activity'] = ACTIVITY_NAMES.get(activity_id, 'Office Work')

    # 586-589: Sensible Heat (BTU/hr per person)
    sensible_btu = struct.unpack('<f', data[586:590])[0]
    space['sensible_w'] = round(btu_to_w(sensible_btu), 0)

    # 590-593: Latent Heat (BTU/hr per person)
    latent_btu = struct.unpack('<f', data[590:594])[0]
    space['latent_w'] = round(btu_to_w(latent_btu), 0)

    # 594-595: People Schedule ID
    space['people_schedule_id'] = struct.unpack('<H', data[594:596])[0]

    # === INTERNALS - LIGHTING (cols 12-16) ===
    # 600-603: Task Lighting (W)
    task_light = struct.unpack('<f', data[600:604])[0]
    space['task_light_w'] = round(task_light, 0)

    # 604-605: Fixture Type ID
    fixture_id = struct.unpack('<H', data[604:606])[0]
    space['fixture_type'] = FIXTURE_NAMES.get(fixture_id, 'Recessed Unvented')

    # 606-609: General Lighting (W)
    gen_light = struct.unpack('<f', data[606:610])[0]
    space['gen_light_w'] = round(gen_light, 0)

    # 610-613: Ballast Multiplier
    ballast = struct.unpack('<f', data[610:614])[0]
    space['ballast'] = round(ballast, 2) if ballast else 1.0

    # 616-617: Lighting Schedule ID (CONFIRMADO - offset 616, não 614!)
    space['light_schedule_id'] = struct.unpack('<H', data[616:618])[0]

    # === INTERNALS - EQUIPMENT (cols 17-18) ===
    # 656-659: Equipment (W/ft²)
    equip_w_ft2 = struct.unpack('<f', data[656:660])[0]
    space['equip_w_m2'] = round(w_ft2_to_w_m2(equip_w_ft2), 2)

    # 660-661: Equipment Schedule ID
    space['equip_schedule_id'] = struct.unpack('<H', data[660:662])[0]

    # === INTERNALS - MISC (cols 19-22) ===
    # 632-635: Misc Sensible (BTU/hr)
    misc_sens_btu = struct.unpack('<f', data[632:636])[0]
    space['misc_sensible_w'] = round(btu_to_w(misc_sens_btu), 0)

    # 636-639: Misc Latent (BTU/hr)
    misc_lat_btu = struct.unpack('<f', data[636:640])[0]
    space['misc_latent_w'] = round(btu_to_w(misc_lat_btu), 0)

    # 640-641: Misc Sensible Schedule
    space['misc_sens_schedule_id'] = struct.unpack('<H', data[640:642])[0]

    # 644-645: Misc Latent Schedule
    space['misc_lat_schedule_id'] = struct.unpack('<H', data[644:646])[0]

    # === INFILTRATION (cols 23-26) ===
    # O método de infiltração está codificado na coluna do Excel como "Air Change" ou similar
    # No HAP, os valores ACH estão nos offsets 556, 562, 568 (CONFIRMADO)
    # O offset 492 parece ter outro significado

    # Inferir método do facto de termos valores ACH
    space['infil_method'] = 'Air Change'  # Default quando temos ACH values

    # 556-559: Design Cooling ACH (CONFIRMADO)
    space['design_clg_ach'] = round(struct.unpack('<f', data[556:560])[0], 2)

    # 562-565: Design Heating ACH (CONFIRMADO)
    space['design_htg_ach'] = round(struct.unpack('<f', data[562:566])[0], 2)

    # 568-571: Energy ACH (CONFIRMADO)
    space['energy_ach'] = round(struct.unpack('<f', data[568:572])[0], 2)

    # === FLOORS (cols 27-39) ===
    # FLOOR structure: offsets 492-542 (CONFIRMADO em excel_to_hap.py)
    # 492-493: Floor Type ID
    floor_type_id = struct.unpack('<H', data[492:494])[0]
    space['floor_type'] = FLOOR_TYPE_NAMES.get(floor_type_id, '')

    # 494-497: Floor Area (ft²)
    floor_area_ft2 = struct.unpack('<f', data[494:498])[0]
    space['floor_area_m2'] = round(ft2_to_m2(floor_area_ft2), 2)

    # 498-501: Floor U-Value (BTU/(hr·ft²·°F))
    floor_u_ip = struct.unpack('<f', data[498:502])[0]
    space['floor_u_value'] = round(u_ip_to_si(floor_u_ip), 3)

    # 502-505: Exposed Perimeter (ft)
    exp_perim_ft = struct.unpack('<f', data[502:506])[0]
    space['floor_exp_perim'] = round(ft_to_m(exp_perim_ft), 2)

    # 506-509: Edge Insulation R-Value
    edge_r_ip = struct.unpack('<f', data[506:510])[0]
    space['floor_edge_r'] = round(r_ip_to_si(edge_r_ip), 3)

    # 510-513: Depth below grade (ft)
    depth_ft = struct.unpack('<f', data[510:514])[0]
    space['floor_depth'] = round(ft_to_m(depth_ft), 2)

    # 514-517: Basement Wall U-Value
    bsmt_u_ip = struct.unpack('<f', data[514:518])[0]
    space['floor_bsmt_wall_u'] = round(u_ip_to_si(bsmt_u_ip), 3)

    # 518-521: Wall Insulation R-Value
    wall_ins_r_ip = struct.unpack('<f', data[518:522])[0]
    space['floor_wall_ins_r'] = round(r_ip_to_si(wall_ins_r_ip), 3)

    # 522-525: Insulation Depth (ft)
    ins_depth_ft = struct.unpack('<f', data[522:526])[0]
    space['floor_ins_depth'] = round(ft_to_m(ins_depth_ft), 2)

    # 526-541: Floor temps (Unc Max, Out Max, Unc Min, Out Min)
    space['floor_unc_max'] = round(f_to_c(struct.unpack('<f', data[526:530])[0]), 1)
    space['floor_out_max'] = round(f_to_c(struct.unpack('<f', data[530:534])[0]), 1)
    space['floor_unc_min'] = round(f_to_c(struct.unpack('<f', data[534:538])[0]), 1)
    space['floor_out_min'] = round(f_to_c(struct.unpack('<f', data[538:542])[0]), 1)

    # === PARTITIONS (cols 40-51) ===
    # PARTITIONS - Ceiling: offsets 440-466 (CONFIRMADO em excel_to_hap.py)
    # 440-441: Type (1=Ceiling, 2=Wall)
    # 442-445: Area (ft²)
    space['ceil_area_m2'] = round(ft2_to_m2(struct.unpack('<f', data[442:446])[0]), 2)
    # 446-449: U-Value
    space['ceil_u_value'] = round(u_ip_to_si(struct.unpack('<f', data[446:450])[0]), 3)
    # 450-453: Unc Max, 454-457: Out Max, 458-461: Unc Min, 462-465: Out Min
    space['ceil_unc_max'] = round(f_to_c(struct.unpack('<f', data[450:454])[0]), 1)
    space['ceil_out_max'] = round(f_to_c(struct.unpack('<f', data[454:458])[0]), 1)
    space['ceil_unc_min'] = round(f_to_c(struct.unpack('<f', data[458:462])[0]), 1)
    space['ceil_out_min'] = round(f_to_c(struct.unpack('<f', data[462:466])[0]), 1)

    # PARTITIONS - Wall: offsets 466-492 (CONFIRMADO em excel_to_hap.py)
    # 466-467: Type (2=Wall)
    # 468-471: Area (ft²)
    space['wall_part_area_m2'] = round(ft2_to_m2(struct.unpack('<f', data[468:472])[0]), 2)
    # 472-475: U-Value
    space['wall_part_u_value'] = round(u_ip_to_si(struct.unpack('<f', data[472:476])[0]), 3)
    # 476-479: Unc Max, 480-483: Out Max, 484-487: Unc Min, 488-491: Out Min
    space['wall_part_unc_max'] = round(f_to_c(struct.unpack('<f', data[476:480])[0]), 1)
    space['wall_part_out_max'] = round(f_to_c(struct.unpack('<f', data[480:484])[0]), 1)
    space['wall_part_unc_min'] = round(f_to_c(struct.unpack('<f', data[484:488])[0]), 1)
    space['wall_part_out_min'] = round(f_to_c(struct.unpack('<f', data[488:492])[0]), 1)

    # === WALLS (cols 52-123) - 8 walls x 9 campos ===
    space['walls'] = []
    for w in range(8):
        wall_offset = WALL_BLOCK_START + (w * WALL_BLOCK_SIZE)
        wall = extract_wall_block(data, wall_offset)
        space['walls'].append(wall)

    # === ROOFS (cols 124-147) - 4 roofs x 6 campos ===
    space['roofs'] = []
    for r in range(4):
        roof_offset = ROOF_BLOCK_START + (r * ROOF_BLOCK_SIZE)
        roof = extract_roof_block(data, roof_offset)
        space['roofs'].append(roof)

    return space

def extract_wall_block(data, offset):
    """Extrai um bloco de parede (34 bytes)"""
    wall = {}

    # +0: Exposure Code
    exp_code = struct.unpack('<H', data[offset:offset+2])[0]
    wall['exposure'] = DIRECTION_NAMES.get(exp_code, '')

    # +2: Gross Wall Area (ft²)
    area_ft2 = struct.unpack('<f', data[offset+2:offset+6])[0]
    wall['area_m2'] = round(ft2_to_m2(area_ft2), 2)

    # +6: Wall Type ID
    wall['wall_type_id'] = struct.unpack('<H', data[offset+6:offset+8])[0]

    # +8: Window 1 Type ID
    wall['window1_type_id'] = struct.unpack('<H', data[offset+8:offset+10])[0]

    # +12: Window 1 Quantity
    wall['window1_qty'] = struct.unpack('<H', data[offset+12:offset+14])[0]

    # +14: Window 2 Type ID
    wall['window2_type_id'] = struct.unpack('<H', data[offset+14:offset+16])[0]

    # +18: Window 2 Quantity
    wall['window2_qty'] = struct.unpack('<H', data[offset+18:offset+20])[0]

    # +20: Door Type ID
    wall['door_type_id'] = struct.unpack('<H', data[offset+20:offset+22])[0]

    # +22: Door Quantity
    wall['door_qty'] = struct.unpack('<H', data[offset+22:offset+24])[0]

    return wall

def extract_roof_block(data, offset):
    """Extrai um bloco de cobertura (24 bytes)"""
    roof = {}

    # +0: Exposure Code
    exp_code = struct.unpack('<H', data[offset:offset+2])[0]
    roof['exposure'] = DIRECTION_NAMES.get(exp_code, '')

    # +2: Slope (degrees)
    roof['slope'] = struct.unpack('<H', data[offset+2:offset+4])[0]

    # +4: Gross Area (ft²)
    area_ft2 = struct.unpack('<f', data[offset+4:offset+8])[0]
    roof['area_m2'] = round(ft2_to_m2(area_ft2), 2)

    # +8: Roof Type ID
    roof['roof_type_id'] = struct.unpack('<H', data[offset+8:offset+10])[0]

    # +10: Skylight Type ID
    roof['skylight_type_id'] = struct.unpack('<H', data[offset+10:offset+12])[0]

    # +12: Skylight Quantity
    roof['skylight_qty'] = struct.unpack('<H', data[offset+12:offset+14])[0]

    return roof

def extract_schedules(sch_data):
    """Extrai nomes dos schedules"""
    num_schedules = len(sch_data) // SCHEDULE_RECORD_SIZE
    schedules = []

    for i in range(num_schedules):
        offset = i * SCHEDULE_RECORD_SIZE
        # Nome: bytes 0-79
        raw_name = sch_data[offset:offset+80]
        try:
            name = raw_name.split(b'\x00')[0].decode('latin-1').strip()
            name = clean_string(name)
        except:
            name = f'Schedule_{i}'
        schedules.append(name)

    return schedules

def extract_walls_assemblies(wal_data):
    """Extrai dados dos wall assemblies"""
    num_walls = len(wal_data) // WALL_ASSEMBLY_SIZE
    walls = []
    walls_detail = []  # Lista de dicionários com detalhes

    for i in range(num_walls):
        offset = i * WALL_ASSEMBLY_SIZE
        raw_name = wal_data[offset:offset+255]
        try:
            name = raw_name.split(b'\x00')[0].decode('latin-1').strip()
            name = clean_string(name)
        except:
            name = f'Wall_{i}'
        walls.append(name)

        # Extrair mais detalhes
        try:
            # U-value (offset 269, IP -> SI)
            u_value_ip = struct.unpack('<f', wal_data[offset+269:offset+273])[0]
            u_value_si = round(u_ip_to_si(u_value_ip), 3) if 0 < u_value_ip < 5 else 0

            # Espessura (offset 261, ft -> m)
            thickness_ft = struct.unpack('<f', wal_data[offset+261:offset+265])[0]
            thickness_m = round(thickness_ft * 0.3048, 3) if 0 < thickness_ft < 10 else 0

            # Massa (offset 273, lb/ft² -> kg/m²)
            mass_lb = struct.unpack('<f', wal_data[offset+273:offset+277])[0]
            mass_kg = round(mass_lb * 4.88243, 1) if 0 < mass_lb < 500 else 0
        except:
            u_value_si = 0
            thickness_m = 0
            mass_kg = 0

        walls_detail.append({
            'name': name,
            'u_value': u_value_si,
            'thickness': thickness_m,
            'mass': mass_kg,
        })

    return walls, walls_detail

def extract_roofs_assemblies(rof_data):
    """Extrai dados dos roof assemblies"""
    num_roofs = len(rof_data) // ROOF_ASSEMBLY_SIZE
    roofs = []
    roofs_detail = []

    for i in range(num_roofs):
        offset = i * ROOF_ASSEMBLY_SIZE
        raw_name = rof_data[offset:offset+255]
        try:
            name = raw_name.split(b'\x00')[0].decode('latin-1').strip()
            name = clean_string(name)
        except:
            name = f'Roof_{i}'
        roofs.append(name)

        # Extrair mais detalhes (mesmos offsets que walls)
        try:
            # U-value (offset 269, IP -> SI)
            u_value_ip = struct.unpack('<f', rof_data[offset+269:offset+273])[0]
            u_value_si = round(u_ip_to_si(u_value_ip), 3) if 0 < u_value_ip < 5 else 0

            # Espessura (offset 261, ft -> m)
            thickness_ft = struct.unpack('<f', rof_data[offset+261:offset+265])[0]
            thickness_m = round(thickness_ft * 0.3048, 3) if 0 < thickness_ft < 10 else 0

            # Massa (offset 273, lb/ft² -> kg/m²)
            mass_lb = struct.unpack('<f', rof_data[offset+273:offset+277])[0]
            mass_kg = round(mass_lb * 4.88243, 1) if 0 < mass_lb < 1000 else 0
        except:
            u_value_si = 0
            thickness_m = 0
            mass_kg = 0

        roofs_detail.append({
            'name': name,
            'u_value': u_value_si,
            'thickness': thickness_m,
            'mass': mass_kg,
        })

    return roofs, roofs_detail

def extract_windows(win_data):
    """Extrai dados das windows"""
    num_windows = len(win_data) // WINDOW_RECORD_SIZE
    windows = []
    windows_detail = []

    for i in range(num_windows):
        offset = i * WINDOW_RECORD_SIZE
        raw_name = win_data[offset:offset+100]
        try:
            name = raw_name.split(b'\x00')[0].decode('latin-1').strip()
            name = clean_string(name)
        except:
            name = f'Window_{i}'
        windows.append(name)

        # Extrair mais detalhes - offsets correctos descobertos por análise hex
        try:
            # Dimensões (em ft) - offset 257 e 261
            height_ft = struct.unpack('<f', win_data[offset+257:offset+261])[0]
            width_ft = struct.unpack('<f', win_data[offset+261:offset+265])[0]
            height_m = round(ft_to_m(height_ft), 2) if height_ft > 0 else 0
            width_m = round(ft_to_m(width_ft), 2) if width_ft > 0 else 0

            # U-value IP - offset 269
            u_value_ip = struct.unpack('<f', win_data[offset+269:offset+273])[0]
            u_value_si = round(u_ip_to_si(u_value_ip), 3) if 0 < u_value_ip < 5 else 0

            # SHGC - offset 433
            shgc = struct.unpack('<f', win_data[offset+433:offset+437])[0]
            shgc = round(shgc, 3) if 0 < shgc < 1 else 0
        except:
            u_value_si = 0
            shgc = 0
            height_m = 0
            width_m = 0

        windows_detail.append({
            'name': name,
            'u_value': u_value_si,
            'shgc': shgc,
            'height': height_m,
            'width': width_m,
        })

    return windows, windows_detail

# =============================================================================
# ESCRITA DO EXCEL
# =============================================================================

def create_excel(spaces, schedules, walls, roofs, windows, output_path):
    """Cria Excel com todos os dados extraídos no formato do template"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Espacos'

    # Estilos
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # === LINHA 1: Categorias principais ===
    categories = [
        (1, 'GENERAL'), (7, 'INTERNALS'), (23, 'INFILTRATION'),
        (27, 'FLOORS'), (40, 'PARTITIONS'), (52, 'WALLS'), (124, 'ROOFS')
    ]
    for col, name in categories:
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    # === LINHA 2: Sub-categorias ===
    subcategories = [
        (7, 'PEOPLE'), (12, 'LIGHTING'), (17, 'EQUIPMENT'), (19, 'MISC'),
        (40, 'CEILING'), (46, 'WALL'),
        (52, 'WALL 1'), (61, 'WALL 2'), (70, 'WALL 3'), (79, 'WALL 4'),
        (88, 'WALL 5'), (97, 'WALL 6'), (106, 'WALL 7'), (115, 'WALL 8'),
        (124, 'ROOF 1'), (130, 'ROOF 2'), (136, 'ROOF 3'), (142, 'ROOF 4')
    ]
    for col, name in subcategories:
        cell = ws.cell(row=2, column=col, value=name)
        cell.fill = subheader_fill
        cell.font = Font(bold=True)

    # === LINHA 3: Headers dos campos (147 colunas) ===
    headers = [
        # GENERAL (1-6)
        'Space Name', 'Floor Area\n(m2)', 'Avg Ceiling Ht\n(m)', 'Building Wt\n(kg/m2)',
        'Outdoor Air\n(valor)', 'OA Unit',
        # PEOPLE (7-11)
        'Occupancy\n(people)', 'Activity Level', 'Sensible\n(W/person)', 'Latent\n(W/person)', 'Schedule',
        # LIGHTING (12-16)
        'Task Lighting\n(W)', 'General Ltg\n(W)', 'Fixture Type', 'Ballast Mult', 'Schedule',
        # EQUIPMENT (17-18)
        'Equipment\n(W/m2)', 'Schedule',
        # MISC (19-22)
        'Sensible\n(W)', 'Latent\n(W)', 'Sens Sch', 'Lat Sch',
        # INFILTRATION (23-26)
        'Infil Method', 'Design Clg\n(ACH)', 'Design Htg\n(ACH)', 'Energy\n(ACH)',
        # FLOORS (27-39) - 13 campos
        'Floor Type', 'Floor Area\n(m2)', 'U-Value\n(W/m2K)', 'Exp Perim\n(m)',
        'Edge R\n(m2K/W)', 'Depth\n(m)', 'Bsmt Wall U\n(W/m2K)', 'Wall Ins R\n(m2K/W)',
        'Ins Depth\n(m)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
        # PARTITIONS - CEILING (40-45) - 6 campos
        'Area\n(m2)', 'U-Value\n(W/m2K)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
        # PARTITIONS - WALL (46-51) - 6 campos
        'Area\n(m2)', 'U-Value\n(W/m2K)', 'Unc Max\n(C)', 'Out Max\n(C)', 'Unc Min\n(C)', 'Out Min\n(C)',
    ]

    # WALLS (52-123) - 8 walls x 9 campos = 72 colunas
    wall_headers = ['Exposure', 'Gross Area\n(m2)', 'Wall Type', 'Window 1', 'Win1 Qty',
                    'Window 2', 'Win2 Qty', 'Door', 'Door Qty']
    for w in range(8):
        headers.extend(wall_headers)

    # ROOFS (124-147) - 4 roofs x 6 campos = 24 colunas
    roof_headers = ['Exposure', 'Gross Area\n(m2)', 'Slope\n(deg)', 'Roof Type', 'Skylight', 'Sky Qty']
    for r in range(4):
        headers.extend(roof_headers)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = subheader_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = thin_border

    # === LINHAS 4+: Dados dos espaços ===
    row_idx = 4
    for space in spaces:
        if space['_index'] == 0:  # Skip Default Space
            continue

        col = 1

        # GENERAL (1-6)
        ws.cell(row=row_idx, column=col, value=space['name']); col += 1
        ws.cell(row=row_idx, column=col, value=space['area_m2']); col += 1
        ws.cell(row=row_idx, column=col, value=space['height_m']); col += 1
        ws.cell(row=row_idx, column=col, value=space['weight_kg_m2']); col += 1
        ws.cell(row=row_idx, column=col, value=space['oa_value']); col += 1
        ws.cell(row=row_idx, column=col, value=space['oa_unit']); col += 1

        # PEOPLE (7-11)
        ws.cell(row=row_idx, column=col, value=space['occupancy']); col += 1
        ws.cell(row=row_idx, column=col, value=space['activity']); col += 1
        ws.cell(row=row_idx, column=col, value=space['sensible_w']); col += 1
        ws.cell(row=row_idx, column=col, value=space['latent_w']); col += 1
        sch_name = schedules[space['people_schedule_id']] if space['people_schedule_id'] < len(schedules) else ''
        ws.cell(row=row_idx, column=col, value=sch_name); col += 1

        # LIGHTING (12-16)
        ws.cell(row=row_idx, column=col, value=space['task_light_w']); col += 1
        ws.cell(row=row_idx, column=col, value=space['gen_light_w']); col += 1
        ws.cell(row=row_idx, column=col, value=space['fixture_type']); col += 1
        ws.cell(row=row_idx, column=col, value=space['ballast']); col += 1
        sch_name = schedules[space['light_schedule_id']] if space['light_schedule_id'] < len(schedules) else ''
        ws.cell(row=row_idx, column=col, value=sch_name); col += 1

        # EQUIPMENT (17-18)
        ws.cell(row=row_idx, column=col, value=space['equip_w_m2']); col += 1
        sch_name = schedules[space['equip_schedule_id']] if space['equip_schedule_id'] < len(schedules) else ''
        ws.cell(row=row_idx, column=col, value=sch_name); col += 1

        # MISC (19-22)
        ws.cell(row=row_idx, column=col, value=space['misc_sensible_w']); col += 1
        ws.cell(row=row_idx, column=col, value=space['misc_latent_w']); col += 1
        sch_name = schedules[space['misc_sens_schedule_id']] if space['misc_sens_schedule_id'] < len(schedules) else ''
        ws.cell(row=row_idx, column=col, value=sch_name); col += 1
        sch_name = schedules[space['misc_lat_schedule_id']] if space['misc_lat_schedule_id'] < len(schedules) else ''
        ws.cell(row=row_idx, column=col, value=sch_name); col += 1

        # INFILTRATION (23-26)
        ws.cell(row=row_idx, column=col, value=space['infil_method']); col += 1
        ws.cell(row=row_idx, column=col, value=space['design_clg_ach']); col += 1
        ws.cell(row=row_idx, column=col, value=space['design_htg_ach']); col += 1
        ws.cell(row=row_idx, column=col, value=space['energy_ach']); col += 1

        # FLOORS (27-39)
        ws.cell(row=row_idx, column=col, value=space['floor_type']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_area_m2']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_u_value']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_exp_perim']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_edge_r']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_depth']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_bsmt_wall_u']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_wall_ins_r']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_ins_depth']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_unc_max']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_out_max']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_unc_min']); col += 1
        ws.cell(row=row_idx, column=col, value=space['floor_out_min']); col += 1

        # PARTITIONS - CEILING (40-45)
        ws.cell(row=row_idx, column=col, value=space['ceil_area_m2']); col += 1
        ws.cell(row=row_idx, column=col, value=space['ceil_u_value']); col += 1
        ws.cell(row=row_idx, column=col, value=space['ceil_unc_max']); col += 1
        ws.cell(row=row_idx, column=col, value=space['ceil_out_max']); col += 1
        ws.cell(row=row_idx, column=col, value=space['ceil_unc_min']); col += 1
        ws.cell(row=row_idx, column=col, value=space['ceil_out_min']); col += 1

        # PARTITIONS - WALL (46-51)
        ws.cell(row=row_idx, column=col, value=space['wall_part_area_m2']); col += 1
        ws.cell(row=row_idx, column=col, value=space['wall_part_u_value']); col += 1
        ws.cell(row=row_idx, column=col, value=space['wall_part_unc_max']); col += 1
        ws.cell(row=row_idx, column=col, value=space['wall_part_out_max']); col += 1
        ws.cell(row=row_idx, column=col, value=space['wall_part_unc_min']); col += 1
        ws.cell(row=row_idx, column=col, value=space['wall_part_out_min']); col += 1

        # WALLS (52-123) - 8 walls x 9 campos
        for wall in space['walls']:
            ws.cell(row=row_idx, column=col, value=wall['exposure']); col += 1
            ws.cell(row=row_idx, column=col, value=wall['area_m2']); col += 1
            wall_name = walls[wall['wall_type_id']] if wall['wall_type_id'] < len(walls) else ''
            ws.cell(row=row_idx, column=col, value=wall_name); col += 1
            win1_name = windows[wall['window1_type_id']] if wall['window1_type_id'] < len(windows) else ''
            ws.cell(row=row_idx, column=col, value=win1_name); col += 1
            ws.cell(row=row_idx, column=col, value=wall['window1_qty'] if wall['window1_qty'] else ''); col += 1
            win2_name = windows[wall['window2_type_id']] if wall['window2_type_id'] < len(windows) else ''
            ws.cell(row=row_idx, column=col, value=win2_name); col += 1
            ws.cell(row=row_idx, column=col, value=wall['window2_qty'] if wall['window2_qty'] else ''); col += 1
            ws.cell(row=row_idx, column=col, value=''); col += 1  # Door name (not extracted)
            ws.cell(row=row_idx, column=col, value=wall['door_qty'] if wall['door_qty'] else ''); col += 1

        # ROOFS (124-147) - 4 roofs x 6 campos
        for roof in space['roofs']:
            ws.cell(row=row_idx, column=col, value=roof['exposure']); col += 1
            ws.cell(row=row_idx, column=col, value=roof['area_m2']); col += 1
            ws.cell(row=row_idx, column=col, value=roof['slope'] if roof['slope'] else ''); col += 1
            roof_name = roofs[roof['roof_type_id']] if roof['roof_type_id'] < len(roofs) else ''
            ws.cell(row=row_idx, column=col, value=roof_name); col += 1
            ws.cell(row=row_idx, column=col, value=''); col += 1  # Skylight name
            ws.cell(row=row_idx, column=col, value=roof['skylight_qty'] if roof['skylight_qty'] else ''); col += 1

        row_idx += 1  # Próxima linha

    # Ajustar largura das colunas
    for col in range(1, 148):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    ws.column_dimensions['A'].width = 20  # Space Name

    # Freeze panes
    ws.freeze_panes = 'B4'

    wb.save(output_path)
    return wb

def create_windows_sheet(wb, windows_detail):
    """Cria folha Windows"""
    ws = wb.create_sheet('Windows')

    # Headers
    ws.cell(1, 1, value='WINDOWS')
    ws.cell(2, 1, value='IDENTIFICAÇÃO')
    ws.cell(2, 2, value='PROPRIEDADES TÉRMICAS')
    ws.cell(2, 4, value='DIMENSÕES')

    ws.cell(3, 1, value='Nome')
    ws.cell(3, 2, value='U-Value\n(W/m²K)')
    ws.cell(3, 3, value='SHGC')
    ws.cell(3, 4, value='Altura\n(m)')
    ws.cell(3, 5, value='Largura\n(m)')

    # Dados
    for i, win in enumerate(windows_detail, 4):
        if win['name'] and not win['name'].startswith('Sample'):
            ws.cell(i, 1, value=win['name'])
            ws.cell(i, 2, value=win['u_value'] if win['u_value'] else '')
            ws.cell(i, 3, value=win['shgc'] if win['shgc'] else '')
            ws.cell(i, 4, value=win['height'] if win['height'] else '')
            ws.cell(i, 5, value=win['width'] if win['width'] else '')

    # Ajustar colunas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 12

def create_walls_sheet(wb, walls_detail):
    """Cria folha Walls"""
    ws = wb.create_sheet('Walls')

    # Headers
    ws.cell(1, 1, value='WALLS')
    ws.cell(2, 1, value='IDENTIFICAÇÃO')
    ws.cell(2, 2, value='PROPRIEDADES TÉRMICAS')
    ws.cell(2, 3, value='DIMENSÕES')
    ws.cell(2, 4, value='PROPRIEDADES FÍSICAS')

    ws.cell(3, 1, value='Nome')
    ws.cell(3, 2, value='U-Value\n(W/m²K)')
    ws.cell(3, 3, value='Espessura\n(m)')
    ws.cell(3, 4, value='Massa\n(kg/m²)')

    # Dados
    row = 4
    for wall in walls_detail:
        if wall['name'] and not wall['name'].startswith('Sample') and not wall['name'].startswith('Default'):
            ws.cell(row, 1, value=wall['name'])
            ws.cell(row, 2, value=wall['u_value'] if wall['u_value'] else '')
            ws.cell(row, 3, value=wall['thickness'] if wall.get('thickness') else '')
            ws.cell(row, 4, value=wall['mass'] if wall.get('mass') else '')
            row += 1

    # Ajustar colunas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 12

def create_roofs_sheet(wb, roofs_detail):
    """Cria folha Roofs"""
    ws = wb.create_sheet('Roofs')

    # Headers
    ws.cell(1, 1, value='ROOFS')
    ws.cell(2, 1, value='IDENTIFICAÇÃO')
    ws.cell(2, 2, value='PROPRIEDADES TÉRMICAS')
    ws.cell(2, 3, value='DIMENSÕES')
    ws.cell(2, 4, value='PROPRIEDADES FÍSICAS')

    ws.cell(3, 1, value='Nome')
    ws.cell(3, 2, value='U-Value\n(W/m²K)')
    ws.cell(3, 3, value='Espessura\n(m)')
    ws.cell(3, 4, value='Massa\n(kg/m²)')

    # Dados
    row = 4
    for roof in roofs_detail:
        if roof['name'] and not roof['name'].startswith('Sample') and not roof['name'].startswith('Default'):
            ws.cell(row, 1, value=roof['name'])
            ws.cell(row, 2, value=roof['u_value'] if roof['u_value'] else '')
            ws.cell(row, 3, value=roof['thickness'] if roof.get('thickness') else '')
            ws.cell(row, 4, value=roof['mass'] if roof.get('mass') else '')
            row += 1

    # Ajustar colunas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 12

# =============================================================================
# MAIN
# =============================================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace('.E3A', '_Extraido.xlsx').replace('.e3a', '_Extraido.xlsx')

    if not os.path.exists(input_file):
        print(f"Erro: Ficheiro '{input_file}' nao encontrado!")
        sys.exit(1)

    print(f"A extrair dados de: {input_file}")

    # Ler E3A
    files = read_e3a(input_file)

    # Extrair dados
    spaces = extract_spaces(files.get('HAP51SPC.DAT', b''))
    schedules = extract_schedules(files.get('HAP51SCH.DAT', b''))
    walls, walls_detail = extract_walls_assemblies(files.get('HAP51WAL.DAT', b''))
    roofs, roofs_detail = extract_roofs_assemblies(files.get('HAP51ROF.DAT', b''))
    windows, windows_detail = extract_windows(files.get('HAP51WIN.DAT', b''))

    print(f"  Espacos: {len(spaces)} (incluindo Default Space)")
    print(f"  Schedules: {len(schedules)}")
    print(f"  Wall Assemblies: {len(walls)}")
    print(f"  Roof Assemblies: {len(roofs)}")
    print(f"  Windows: {len(windows)}")

    # Criar Excel
    wb = create_excel(spaces, schedules, walls, roofs, windows, output_file)

    # Adicionar folhas Windows, Walls, Roofs
    create_windows_sheet(wb, windows_detail)
    create_walls_sheet(wb, walls_detail)
    create_roofs_sheet(wb, roofs_detail)

    wb.save(output_file)

    print(f"\nFicheiro criado: {output_file}")
    print("Folhas: Espacos, Windows, Walls, Roofs")

if __name__ == '__main__':
    main()
