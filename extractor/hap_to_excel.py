"""
HAP to Excel Exporter
=====================
Export space data from HAP 5.1 .E3A file to Excel for review/modification.

Usage:
    python hap_to_excel.py <input_e3a> [output_xlsx]
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from hap_library import HAPProject, DIRECTIONS


def export_hap_to_excel(hap_path: str, excel_path: str = None):
    """Export HAP project to Excel."""
    if not excel_path:
        excel_path = Path(hap_path).stem + '_export.xlsx'

    print(f"Opening HAP file: {hap_path}")
    project = HAPProject.open(hap_path)
    print(f"Found {len(project.spaces)} spaces")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Espacos'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = [
        'Nome', 'Area m2', 'Altura m', 'Peso kg/m2',
        'OA Valor', 'OA Unidade',
        'Arref C', 'Arref HR%', 'Aquec C', 'Aquec HR%', 'SHR',
        'Ocupacao', 'Sens W/pes', 'Lat W/pes',
        'Ilum W', 'Tipo Lum', 'Balastro', 'Task W',
        'Equip W/m2',
        'Misc Sens W', 'Misc Lat W',
        'Infil Modo', 'Infil ACH'
    ]

    # Add wall headers
    for direction in DIRECTIONS:
        headers.extend([
            f'{direction} Par.Tipo', f'{direction} Area m2',
            f'{direction} Jan.Tipo', f'{direction} Jan.Qtd'
        ])

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Write data
    for row, space in enumerate(project.spaces, 2):
        col = 1

        # Basic info
        ws.cell(row=row, column=col, value=space.name); col += 1
        ws.cell(row=row, column=col, value=round(space.floor_area_m2, 1)); col += 1
        ws.cell(row=row, column=col, value=round(space.ceiling_height_m, 2)); col += 1
        ws.cell(row=row, column=col, value=round(space.building_weight_kg_m2, 0)); col += 1

        # OA
        ws.cell(row=row, column=col, value=round(space.oa_value, 1)); col += 1
        ws.cell(row=row, column=col, value=space.oa_unit); col += 1

        # Thermostat
        ws.cell(row=row, column=col, value=round(space.cooling_setpoint_c, 1)); col += 1
        ws.cell(row=row, column=col, value=round(space.cooling_rh, 0)); col += 1
        ws.cell(row=row, column=col, value=round(space.heating_setpoint_c, 1)); col += 1
        ws.cell(row=row, column=col, value=round(space.heating_rh, 0)); col += 1
        ws.cell(row=row, column=col, value=round(space.sensible_heat_ratio, 2)); col += 1

        # People
        ws.cell(row=row, column=col, value=round(space.occupancy, 0)); col += 1
        ws.cell(row=row, column=col, value=round(space.sensible_heat_w, 0)); col += 1
        ws.cell(row=row, column=col, value=round(space.latent_heat_w, 0)); col += 1

        # Lighting
        ws.cell(row=row, column=col, value=round(space.overhead_lighting_w, 0)); col += 1
        ws.cell(row=row, column=col, value=space.fixture_type_id); col += 1
        ws.cell(row=row, column=col, value=round(space.ballast_multiplier, 2)); col += 1
        ws.cell(row=row, column=col, value=round(space.task_lighting_w, 0)); col += 1

        # Equipment
        ws.cell(row=row, column=col, value=round(space.equipment_w_m2, 1)); col += 1

        # Misc
        ws.cell(row=row, column=col, value=round(space.misc_sensible_btu_hr / 3.412, 0)); col += 1
        ws.cell(row=row, column=col, value=round(space.misc_latent_btu_hr / 3.412, 0)); col += 1

        # Infiltration
        ws.cell(row=row, column=col, value=space.infiltration.mode); col += 1
        ws.cell(row=row, column=col, value=round(space.infiltration.design_cooling_ach, 2)); col += 1

        # Walls
        for direction in DIRECTIONS:
            wall = space.walls.get(direction)
            if wall and (wall.wall_type_id > 0 or wall.wall_area_ft2 > 0):
                ws.cell(row=row, column=col, value=wall.wall_type_id); col += 1
                ws.cell(row=row, column=col, value=round(wall.wall_area_m2, 1)); col += 1
                ws.cell(row=row, column=col, value=wall.window1_type_id if wall.window1_type_id > 0 else ''); col += 1
                ws.cell(row=row, column=col, value=wall.window1_quantity if wall.window1_quantity > 0 else ''); col += 1
            else:
                col += 4

        # Apply borders
        for c in range(1, col):
            ws.cell(row=row, column=c).border = thin_border

    # Freeze header row
    ws.freeze_panes = 'B2'

    # Save
    wb.save(excel_path)
    print(f"Exported to: {excel_path}")

    # Summary
    print(f"\nSummary:")
    print(f"  Spaces: {len(project.spaces)}")
    print(f"  Total area: {sum(s.floor_area_m2 for s in project.spaces):.1f} m²")
    print(f"  Total occupancy: {sum(s.occupancy for s in project.spaces):.0f} people")

    return excel_path


def main():
    if len(sys.argv) < 2:
        print("HAP to Excel Exporter")
        print("="*50)
        print("\nUsage:")
        print(f"  python {sys.argv[0]} <input_e3a> [output_xlsx]")
        print("\nExample:")
        print(f"  python {sys.argv[0]} projeto.E3A")
        print(f"  python {sys.argv[0]} projeto.E3A dados_exportados.xlsx")
        sys.exit(1)

    hap_path = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not Path(hap_path).exists():
        print(f"Error: HAP file not found: {hap_path}")
        sys.exit(1)

    export_hap_to_excel(hap_path, excel_path)


if __name__ == '__main__':
    main()
