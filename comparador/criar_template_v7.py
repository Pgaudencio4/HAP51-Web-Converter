"""
Criar Template v7 COMPLETO com:
- Folha Comparacao (Espacos) com 147 campos x 3 colunas
- Folha Windows
- Folha Walls
- Folha Roofs

Todas com a mesma formatação: categorias com cores, F1/F2/? e bordas
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Bordas
THIN = Side(style='thin', color='000000')
THICK = Side(style='thick', color='000000')

# Cores
WHITE_FONT = Font(bold=True, color='FFFFFF')
BLACK_FONT = Font(bold=True)

# Cores para Ficheiro 1 vs Ficheiro 2
FILE1_FILL = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')  # Azul
FILE2_FILL = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')  # Laranja
CHECK_FILL = PatternFill(start_color='E2E2E2', end_color='E2E2E2', fill_type='solid')  # Cinza

# Cores por CATEGORIA
CATEGORY_COLORS = {
    'GENERAL': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
    'INTERNALS': PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),
    'INFILTRATION': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
    'FLOORS': PatternFill(start_color='9E480E', end_color='9E480E', fill_type='solid'),
    'PARTITIONS': PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),
    'WALLS': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),
    'ROOFS': PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid'),
    'WINDOWS': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
}

# Cores claras para subcategorias
SUBCATEGORY_COLORS = {
    'GENERAL': PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid'),
    'INTERNALS': PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid'),
    'INFILTRATION': PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid'),
    'FLOORS': PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid'),
    'PARTITIONS': PatternFill(start_color='CDA4DE', end_color='CDA4DE', fill_type='solid'),
    'WALLS': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
    'ROOFS': PatternFill(start_color='9DC3E6', end_color='9DC3E6', fill_type='solid'),
    'WINDOWS': PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid'),
}

def create_simple_comparison_sheet(wb, sheet_name, title, title_fill, fields):
    """
    Cria uma folha de comparação formatada para Windows/Walls/Roofs.

    fields: lista de tuplos (nome_campo, subcategoria)
    """
    ws = wb.create_sheet(sheet_name)

    num_fields = len(fields)
    total_cols = num_fields * 3

    # Agrupar por subcategoria mantendo ordem
    subcats_ordered = []
    subcats_seen = set()
    for field_name, subcat in fields:
        if subcat not in subcats_seen:
            subcats_ordered.append(subcat)
            subcats_seen.add(subcat)

    # Calcular ranges das subcategorias
    subcat_ranges = {}
    col = 1
    for field_name, subcat in fields:
        if subcat not in subcat_ranges:
            subcat_ranges[subcat] = {'start': col, 'end': col + 2}
        else:
            subcat_ranges[subcat]['end'] = col + 2
        col += 3

    # Verificar se há subcategorias ou só uma
    has_multiple_subcats = len(subcats_ordered) > 1

    if has_multiple_subcats:
        # Linha 1: Título (merged horizontalmente)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(1, 1, value=title)
        c.fill = title_fill
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Linha 2: Subcategorias
        for subcat in subcats_ordered:
            rng = subcat_ranges[subcat]
            if rng['end'] > rng['start']:
                ws.merge_cells(start_row=2, start_column=rng['start'], end_row=2, end_column=rng['end'])
            c = ws.cell(2, rng['start'], value=subcat)
            subcat_fill = SUBCATEGORY_COLORS.get(title, SUBCATEGORY_COLORS['GENERAL'])
            c.fill = subcat_fill
            c.font = BLACK_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')
            # Preencher células merged
            for col_idx in range(rng['start'], rng['end'] + 1):
                ws.cell(2, col_idx).fill = subcat_fill
    else:
        # Sem subcategorias: merge vertical linhas 1-2
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=total_cols)
        c = ws.cell(1, 1, value=title)
        c.fill = title_fill
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx in range(1, total_cols + 1):
            ws.cell(1, col_idx).fill = title_fill
            ws.cell(2, col_idx).fill = title_fill

    # Linha 3: Headers (F1, F2, ?)
    col = 1
    for field_name, subcat in fields:
        c1 = ws.cell(3, col, value=f"{field_name} (F1)")
        c1.fill = FILE1_FILL
        c1.font = Font(bold=True, size=9)
        c1.alignment = Alignment(horizontal='center', wrap_text=True)

        c2 = ws.cell(3, col + 1, value=f"{field_name} (F2)")
        c2.fill = FILE2_FILL
        c2.font = Font(bold=True, size=9)
        c2.alignment = Alignment(horizontal='center', wrap_text=True)

        c3 = ws.cell(3, col + 2, value="?")
        c3.fill = CHECK_FILL
        c3.font = BLACK_FONT
        c3.alignment = Alignment(horizontal='center')

        col += 3

    # Bordas grossas à volta de cada subcategoria
    for subcat in subcats_ordered:
        rng = subcat_ranges[subcat]
        start_row = 2 if has_multiple_subcats else 1
        for r in range(start_row, 4):  # Até linha 3 (headers)
            for col_idx in range(rng['start'], rng['end'] + 1):
                cell = ws.cell(r, col_idx)
                left = THICK if col_idx == rng['start'] else THIN
                right = THICK if col_idx == rng['end'] else THIN
                top = THICK if r == start_row else THIN
                bottom = THICK if r == 3 else THIN
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Ajustar larguras
    for col_idx in range(1, total_cols + 1):
        letter = get_column_letter(col_idx)
        if (col_idx - 1) % 3 == 2:  # Coluna ?
            ws.column_dimensions[letter].width = 5
        else:
            ws.column_dimensions[letter].width = 15

    ws.freeze_panes = 'A4'
    return ws

def create_espacos_sheet(wb):
    """
    Cria a folha Comparacao para Espacos com os 147 campos organizados por categoria.
    """
    ws = wb.create_sheet('Comparacao')

    # Definição das categorias e subcategorias (147 campos)
    # Formato: (header_name, category, subcategory)
    fields = [
        # GENERAL (1-6)
        ('Space Name', 'GENERAL', ''),
        ('Space Type', 'GENERAL', ''),
        ('Area', 'GENERAL', ''),
        ('Height', 'GENERAL', ''),
        ('Floor Number', 'GENERAL', ''),
        ('Multiplier', 'GENERAL', ''),

        # INTERNALS - PEOPLE (7-11)
        ('People Activity Level', 'INTERNALS', 'PEOPLE'),
        ('Occupants', 'INTERNALS', 'PEOPLE'),
        ('Sensible', 'INTERNALS', 'PEOPLE'),
        ('Latent', 'INTERNALS', 'PEOPLE'),
        ('People Schedule', 'INTERNALS', 'PEOPLE'),

        # INTERNALS - LIGHTING (12-16)
        ('Lighting W/m2', 'INTERNALS', 'LIGHTING'),
        ('Ballast Multiplier', 'INTERNALS', 'LIGHTING'),
        ('Sensible Rad', 'INTERNALS', 'LIGHTING'),
        ('Sensible Conv', 'INTERNALS', 'LIGHTING'),
        ('Light Schedule', 'INTERNALS', 'LIGHTING'),

        # INTERNALS - EQUIPMENT (17-18)
        ('Equipment W/m2', 'INTERNALS', 'EQUIPMENT'),
        ('Equipment Schedule', 'INTERNALS', 'EQUIPMENT'),

        # INTERNALS - MISC (19-22)
        ('Misc Load', 'INTERNALS', 'MISC'),
        ('Misc Sensible', 'INTERNALS', 'MISC'),
        ('Misc Latent', 'INTERNALS', 'MISC'),
        ('Misc Schedule', 'INTERNALS', 'MISC'),

        # INFILTRATION (23-26)
        ('ACH Heating', 'INFILTRATION', ''),
        ('ACH Cooling', 'INFILTRATION', ''),
        ('ACH Ventilation', 'INFILTRATION', ''),
        ('Infil Schedule', 'INFILTRATION', ''),

        # FLOORS (27-39) - 13 campos
        ('Edge R 1', 'FLOORS', ''),
        ('Floor Length 1', 'FLOORS', ''),
        ('Floor Parcel 1', 'FLOORS', ''),
        ('Edge R 2', 'FLOORS', ''),
        ('Floor Length 2', 'FLOORS', ''),
        ('Floor Parcel 2', 'FLOORS', ''),
        ('Edge R 3', 'FLOORS', ''),
        ('Floor Length 3', 'FLOORS', ''),
        ('Floor Parcel 3', 'FLOORS', ''),
        ('Edge R 4', 'FLOORS', ''),
        ('Floor Length 4', 'FLOORS', ''),
        ('Floor Parcel 4', 'FLOORS', ''),
        ('Floor Area Total', 'FLOORS', ''),

        # PARTITIONS - CEILING (40-45)
        ('Ceiling U-Value', 'PARTITIONS', 'CEILING'),
        ('Ceiling Area', 'PARTITIONS', 'CEILING'),
        ('Ceiling Temp', 'PARTITIONS', 'CEILING'),
        ('Ceiling U-Value 2', 'PARTITIONS', 'CEILING'),
        ('Ceiling Area 2', 'PARTITIONS', 'CEILING'),
        ('Ceiling Temp 2', 'PARTITIONS', 'CEILING'),

        # PARTITIONS - WALL (46-51)
        ('Part Wall U-Value', 'PARTITIONS', 'WALL'),
        ('Part Wall Area', 'PARTITIONS', 'WALL'),
        ('Part Wall Temp', 'PARTITIONS', 'WALL'),
        ('Part Wall U-Value 2', 'PARTITIONS', 'WALL'),
        ('Part Wall Area 2', 'PARTITIONS', 'WALL'),
        ('Part Wall Temp 2', 'PARTITIONS', 'WALL'),

        # WALLS 1-8 (52-123) - 9 campos cada
    ]

    # Adicionar WALLS 1-8
    for wall_num in range(1, 9):
        wall_fields = [
            (f'Wall{wall_num} Assembly', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Orientation', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Tilt', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Gross Area', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Window', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Win Area', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Win Height', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Shading', 'WALLS', f'WALL {wall_num}'),
            (f'Wall{wall_num} Shade Depth', 'WALLS', f'WALL {wall_num}'),
        ]
        fields.extend(wall_fields)

    # ROOFS 1-4 (124-147) - 6 campos cada
    for roof_num in range(1, 5):
        roof_fields = [
            (f'Roof{roof_num} Assembly', 'ROOFS', f'ROOF {roof_num}'),
            (f'Roof{roof_num} Orientation', 'ROOFS', f'ROOF {roof_num}'),
            (f'Roof{roof_num} Tilt', 'ROOFS', f'ROOF {roof_num}'),
            (f'Roof{roof_num} Area', 'ROOFS', f'ROOF {roof_num}'),
            (f'Roof{roof_num} Skylight', 'ROOFS', f'ROOF {roof_num}'),
            (f'Roof{roof_num} Sky Area', 'ROOFS', f'ROOF {roof_num}'),
        ]
        fields.extend(roof_fields)

    # Calcular ranges de categorias e subcategorias
    cat_ranges = []
    sub_ranges = []

    current_cat = None
    cat_start = 1
    current_sub = None
    sub_start = 1

    out_col = 1
    for i, (header, cat, sub) in enumerate(fields):
        # Mudança de categoria
        if cat != current_cat:
            if current_cat is not None:
                has_subs = any(f[2] for f in fields if f[1] == current_cat)
                cat_ranges.append((cat_start, out_col - 1, current_cat, has_subs))
            current_cat = cat
            cat_start = out_col

        # Mudança de subcategoria
        if sub != current_sub:
            if current_sub is not None and current_sub != '':
                sub_ranges.append((sub_start, out_col - 1, current_sub, fields[i-1][1]))
            current_sub = sub
            sub_start = out_col

        out_col += 3

    # Fechar última categoria e subcategoria
    has_subs = any(f[2] for f in fields if f[1] == current_cat)
    cat_ranges.append((cat_start, out_col - 1, current_cat, has_subs))
    if current_sub and current_sub != '':
        sub_ranges.append((sub_start, out_col - 1, current_sub, fields[-1][1]))

    total_cols = (len(fields)) * 3

    # LINHA 1 e 2: Categorias
    for start, end, cat_name, has_subs in cat_ranges:
        cat_fill = CATEGORY_COLORS.get(cat_name, CATEGORY_COLORS['GENERAL'])

        if has_subs:
            # Merge só linha 1
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            c = ws.cell(1, start, value=cat_name)
            c.fill = cat_fill
            c.font = WHITE_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')
            for col in range(start, end + 1):
                ws.cell(1, col).fill = cat_fill
        else:
            # Merge linhas 1-2
            ws.merge_cells(start_row=1, start_column=start, end_row=2, end_column=end)
            c = ws.cell(1, start, value=cat_name)
            c.fill = cat_fill
            c.font = WHITE_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')
            for col in range(start, end + 1):
                ws.cell(1, col).fill = cat_fill
                ws.cell(2, col).fill = cat_fill

    # LINHA 2: Subcategorias
    for start, end, sub_name, parent_cat in sub_ranges:
        sub_fill = SUBCATEGORY_COLORS.get(parent_cat, SUBCATEGORY_COLORS['GENERAL'])
        if end > start:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        c = ws.cell(2, start, value=sub_name)
        c.fill = sub_fill
        c.font = BLACK_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(start, end + 1):
            ws.cell(2, col).fill = sub_fill

    # LINHA 3: Headers (F1, F2, ?)
    out_col = 1
    for header, cat, sub in fields:
        c1 = ws.cell(3, out_col, value=f'{header} (F1)')
        c1.fill = FILE1_FILL
        c1.font = Font(bold=True, size=9)
        c1.alignment = Alignment(horizontal='center', wrap_text=True)

        c2 = ws.cell(3, out_col + 1, value=f'{header} (F2)')
        c2.fill = FILE2_FILL
        c2.font = Font(bold=True, size=9)
        c2.alignment = Alignment(horizontal='center', wrap_text=True)

        c3 = ws.cell(3, out_col + 2, value='?')
        c3.fill = CHECK_FILL
        c3.font = BLACK_FONT
        c3.alignment = Alignment(horizontal='center')

        out_col += 3

    # Bordas grossas
    all_sections = []
    for start, end, sub_name, parent_cat in sub_ranges:
        all_sections.append((start, end, 2))  # Começa na linha 2
    for start, end, cat_name, has_subs in cat_ranges:
        if not has_subs:
            all_sections.append((start, end, 1))  # Começa na linha 1

    for start_col, end_col, start_row in all_sections:
        for r in range(start_row, 4):  # Até linha 3
            for c in range(start_col, end_col + 1):
                cell = ws.cell(r, c)
                left = THICK if c == start_col else THIN
                right = THICK if c == end_col else THIN
                top = THICK if r == start_row else THIN
                bottom = THICK if r == 3 else THIN
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Ajustar colunas
    for col in range(1, total_cols + 1):
        letter = get_column_letter(col)
        if (col - 1) % 3 == 2:
            ws.column_dimensions[letter].width = 5
        else:
            ws.column_dimensions[letter].width = 12

    ws.freeze_panes = 'D4'
    return ws


def main():
    output_path = r'\\100.77.204.117\Programas\HAPPXXXX\Template_Comparacao_v7.xlsx'

    print("Criando Template v7 completo...")
    wb = openpyxl.Workbook()

    # Remover sheet default
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 1. Criar folha Comparacao (Espacos)
    print("Criando folha Comparacao (Espacos - 147 campos)...")
    create_espacos_sheet(wb)

    # 2. Campos para Windows
    windows_fields = [
        ('Nome', 'IDENTIFICAÇÃO'),
        ('U-Value', 'PROPRIEDADES TÉRMICAS'),
        ('SHGC', 'PROPRIEDADES TÉRMICAS'),
        ('Altura', 'DIMENSÕES'),
        ('Largura', 'DIMENSÕES'),
    ]

    # 3. Campos para Walls
    walls_fields = [
        ('Nome', 'IDENTIFICAÇÃO'),
        ('U-Value', 'PROPRIEDADES TÉRMICAS'),
        ('Espessura', 'DIMENSÕES'),
        ('Massa', 'PROPRIEDADES FÍSICAS'),
    ]

    # 4. Campos para Roofs
    roofs_fields = [
        ('Nome', 'IDENTIFICAÇÃO'),
        ('U-Value', 'PROPRIEDADES TÉRMICAS'),
        ('Espessura', 'DIMENSÕES'),
        ('Massa', 'PROPRIEDADES FÍSICAS'),
    ]

    # Criar folhas
    print("Criando folha Windows...")
    create_simple_comparison_sheet(wb, 'Windows', 'WINDOWS', CATEGORY_COLORS['WINDOWS'], windows_fields)

    print("Criando folha Walls...")
    create_simple_comparison_sheet(wb, 'Walls', 'WALLS', CATEGORY_COLORS['WALLS'], walls_fields)

    print("Criando folha Roofs...")
    create_simple_comparison_sheet(wb, 'Roofs', 'ROOFS', CATEGORY_COLORS['ROOFS'], roofs_fields)

    # Guardar
    wb.save(output_path)
    print(f"\n=== Template v7 criado com sucesso! ===")
    print(f"Ficheiro: {output_path}")
    print(f"Folhas: Comparacao, Windows, Walls, Roofs")

if __name__ == '__main__':
    main()
