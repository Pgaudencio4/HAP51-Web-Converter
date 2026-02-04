"""
Comparador Excel HAP - Usa template formatado

Lê um template com a formatação desejada e preenche com dados de comparação.
Mantém a formatação do template (cores, bordas) e preenche os dados.

Usage:
    python comparar_com_template.py <template.xlsx> <ficheiro1.xlsx> <ficheiro2.xlsx> [output.xlsx]
"""

import sys
import os
import re
import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Cores para resultados
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Cores para dados (mais claras que headers)
FILE1_DATA = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # Azul claro
FILE2_DATA = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')  # Laranja claro
CHECK_DATA = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # Cinza claro

# Bordas
THIN = Side(style='thin', color='000000')
THICK = Side(style='thick', color='000000')

def normalize_value(v):
    """Normaliza valor para comparação"""
    if v is None or v == '':
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == '' or v.lower() in ('none', 'n/a', '-'):
            return None
        try:
            f = float(v)
            return round(f, 2) if f != 0 else None
        except:
            return v.lower()
    if isinstance(v, (int, float)):
        return round(float(v), 2) if v != 0 else None
    return v

def compare_values(v1, v2, tolerance=0.5):
    """Compara dois valores"""
    n1 = normalize_value(v1)
    n2 = normalize_value(v2)

    if n1 is None and n2 is None:
        return True, ''
    if n1 is None:
        return False, 'F2'
    if n2 is None:
        return False, 'F1'

    if isinstance(n1, str) or isinstance(n2, str):
        s1 = str(n1).lower().replace(' ', '')
        s2 = str(n2).lower().replace(' ', '')
        if s1 == s2:
            return True, 'OK'
        if s1 in s2 or s2 in s1:
            return True, '~'
        return False, 'DIFF'

    if abs(n1 - n2) < tolerance:
        return True, 'OK'
    if n1 != 0 and abs(n1 - n2) / abs(n1) < 0.05:
        return True, '~'
    return False, 'DIFF'

def get_short_name(filepath):
    """Extrai nome curto do ficheiro"""
    basename = os.path.basename(filepath)
    name = os.path.splitext(basename)[0]
    name = re.sub(r'_extraido$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'_export.*$', '', name, flags=re.IGNORECASE)
    if len(name) > 20:
        name = name[:17] + '...'
    return name

def read_excel_data(filepath):
    """Lê dados do Excel HAP - folha Espacos"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Espacos']

    # Dados (linha 4+)
    data = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()
        row_data = [ws.cell(row, col).value for col in range(1, 148)]
        data[name] = row_data

    return data


def read_sheet_data(filepath, sheet_name, data_start_row=4):
    """Lê dados de uma folha específica (Windows, Walls, Roofs)"""
    try:
        wb = openpyxl.load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]

        # Dados (linha data_start_row+)
        data = {}
        for row in range(data_start_row, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or str(name).strip() == '':
                continue
            name = str(name).strip()
            row_data = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
            data[name] = row_data

        return data
    except:
        return {}

def copy_cell_style(source_cell, target_cell):
    """Copia o estilo de uma célula para outra"""
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format

def detect_section_borders(ws, max_col):
    """Detecta onde estão as bordas grossas (secções) na linha 3"""
    sections = []
    start_col = 1

    for col in range(1, max_col + 2):
        if col > max_col:
            # Fechar última secção
            if start_col <= max_col:
                sections.append((start_col, max_col))
            break

        cell = ws.cell(3, col)
        if cell.border and cell.border.left and cell.border.left.style == 'thick':
            if col > start_col:
                sections.append((start_col, col - 1))
            start_col = col

    return sections


def create_comparison_from_template(template_path, file1_path, file2_path, output_path):
    """Cria comparação usando template formatado"""

    # Carregar template
    print(f"A carregar template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Comparacao']

    # Ler dados dos ficheiros
    print(f"A ler dados de: {file1_path}")
    data1 = read_excel_data(file1_path)
    print(f"A ler dados de: {file2_path}")
    data2 = read_excel_data(file2_path)

    name1 = get_short_name(file1_path)
    name2 = get_short_name(file2_path)

    print(f"\nEspaços em {name1}: {len(data1)}")
    print(f"Espaços em {name2}: {len(data2)}")

    max_col = 147 * 3  # 441 colunas

    # Detectar secções (bordas grossas)
    sections = detect_section_borders(ws, max_col)

    # Actualizar headers na linha 3 com nomes dos ficheiros
    for col in range(1, max_col + 1):
        cell = ws.cell(3, col)
        if cell.value:
            val = str(cell.value)
            val = re.sub(r'\([^)]+\)$', '', val).strip()
            col_type = (col - 1) % 3
            if col_type == 0:
                cell.value = f"{val} ({name1})"
            elif col_type == 1:
                cell.value = f"{val} ({name2})"

    # Estatísticas
    stats = {'ok': 0, 'diff': 0, 'f1_only': 0, 'f2_only': 0, 'both_empty': 0}

    # Escrever dados
    all_spaces = sorted(set(data1.keys()) | set(data2.keys()))
    row = 4
    last_row = 4 + len([s for s in all_spaces if s in data1 and s in data2]) - 1

    for space_name in all_spaces:
        d1 = data1.get(space_name)
        d2 = data2.get(space_name)

        if d1 is None or d2 is None:
            continue

        out_col = 1
        for in_col in range(1, 148):
            v1 = d1[in_col - 1]
            v2 = d2[in_col - 1]

            # Coluna Ficheiro 1 (azul claro)
            c1 = ws.cell(row, out_col, value=v1)
            c1.fill = FILE1_DATA

            # Coluna Ficheiro 2 (laranja claro)
            c2 = ws.cell(row, out_col + 1, value=v2)
            c2.fill = FILE2_DATA

            # Coluna Check
            match, status = compare_values(v1, v2)
            c3 = ws.cell(row, out_col + 2, value=status)
            c3.alignment = Alignment(horizontal='center')

            if status == '':
                c3.fill = CHECK_DATA
                stats['both_empty'] += 1
            elif match:
                c3.fill = GREEN_FILL
                stats['ok'] += 1
            else:
                c3.fill = RED_FILL
                if status == 'F1':
                    stats['f1_only'] += 1
                elif status == 'F2':
                    stats['f2_only'] += 1
                else:
                    stats['diff'] += 1

            out_col += 3

        row += 1

    last_row = row - 1

    # Aplicar bordas grossas às secções (dados)
    for start_col, end_col in sections:
        for r in range(4, last_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(r, c)
                left = THICK if c == start_col else THIN
                right = THICK if c == end_col else THIN
                top = THIN
                bottom = THICK if r == last_row else THIN
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Comparar folhas Windows, Walls, Roofs
    for sheet_name in ['Windows', 'Walls', 'Roofs']:
        if sheet_name in wb.sheetnames:
            compare_simple_sheet(wb, sheet_name, file1_path, file2_path, name1, name2)

    # Guardar
    wb.save(output_path)

    # Mostrar estatísticas
    print(f'\n=== SUMÁRIO ESPACOS ===')
    print(f'Espaços:        {len(all_spaces)}')
    print(f'Campos OK:      {stats["ok"]}')
    print(f'Diferentes:     {stats["diff"]}')
    print(f'Só em {name1}:  {stats["f1_only"]}')
    print(f'Só em {name2}:  {stats["f2_only"]}')
    print(f'Ambos vazios:   {stats["both_empty"]}')

    total = stats['ok'] + stats['diff']
    if total > 0:
        print(f'\nMatch: {stats["ok"] / total * 100:.1f}%')

    return wb


def compare_simple_sheet(wb, sheet_name, file1_path, file2_path, name1, name2):
    """Compara uma folha simples (Windows, Walls, Roofs)"""
    ws = wb[sheet_name]

    # Ler dados dos ficheiros
    data1 = read_sheet_data(file1_path, sheet_name)
    data2 = read_sheet_data(file2_path, sheet_name)

    if not data1 and not data2:
        print(f'\n{sheet_name}: Sem dados em ambos os ficheiros')
        return

    print(f'\n{sheet_name}: {len(data1)} em {name1}, {len(data2)} em {name2}')

    # Determinar número de colunas de dados (excluindo F1/F2/?)
    # No template v7, cada campo tem 3 colunas
    max_template_col = ws.max_column
    num_fields = max_template_col // 3

    # Actualizar headers na linha 3 com nomes dos ficheiros
    for col in range(1, max_template_col + 1):
        cell = ws.cell(3, col)
        if cell.value:
            val = str(cell.value)
            val = re.sub(r'\([^)]+\)$', '', val).strip()
            col_type = (col - 1) % 3
            if col_type == 0:
                cell.value = f"{val} ({name1})"
            elif col_type == 1:
                cell.value = f"{val} ({name2})"

    # Todos os itens (união)
    all_items = sorted(set(data1.keys()) | set(data2.keys()))

    # Escrever dados a partir da linha 4
    row = 4
    stats = {'ok': 0, 'diff': 0}

    for item_name in all_items:
        d1 = data1.get(item_name, [])
        d2 = data2.get(item_name, [])

        # Normalizar tamanho
        max_len = max(len(d1), len(d2), num_fields)
        while len(d1) < max_len:
            d1.append(None)
        while len(d2) < max_len:
            d2.append(None)

        out_col = 1
        for i in range(num_fields):
            v1 = d1[i] if i < len(d1) else None
            v2 = d2[i] if i < len(d2) else None

            # Coluna F1
            c1 = ws.cell(row, out_col, value=v1)
            c1.fill = FILE1_DATA

            # Coluna F2
            c2 = ws.cell(row, out_col + 1, value=v2)
            c2.fill = FILE2_DATA

            # Coluna Check
            match, status = compare_values(v1, v2)
            c3 = ws.cell(row, out_col + 2, value=status)
            c3.alignment = Alignment(horizontal='center')

            if status == '':
                c3.fill = CHECK_DATA
            elif match:
                c3.fill = GREEN_FILL
                stats['ok'] += 1
            else:
                c3.fill = RED_FILL
                stats['diff'] += 1

            out_col += 3

        row += 1

    print(f'  OK: {stats["ok"]}, DIFF: {stats["diff"]}')

def main():
    if len(sys.argv) < 4:
        print(__doc__)
        print("\nExemplo:")
        print("  python comparar_com_template.py template.xlsx Prev.xlsx Ref.xlsx output.xlsx")
        sys.exit(1)

    template_path = sys.argv[1]
    file1_path = sys.argv[2]
    file2_path = sys.argv[3]
    output_path = sys.argv[4] if len(sys.argv) > 4 else 'Comparacao_Output.xlsx'

    for f in [template_path, file1_path, file2_path]:
        if not os.path.exists(f):
            print(f"Erro: '{f}' não encontrado!")
            sys.exit(1)

    create_comparison_from_template(template_path, file1_path, file2_path, output_path)
    print(f'\nFicheiro criado: {output_path}')

if __name__ == '__main__':
    main()
