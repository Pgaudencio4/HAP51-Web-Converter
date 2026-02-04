"""
Comparador Excel HAP - Formato lado a lado

Para cada campo do template, cria 3 colunas:
  - Valor Ficheiro 1
  - Valor Ficheiro 2
  - OK/DIFF (verificação)

Usage:
    python comparar_lado_a_lado.py <ficheiro1.xlsx> <ficheiro2.xlsx> [output.xlsx]
"""

import sys
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

# Bordas
THIN_BORDER = Side(style='thin', color='000000')
THICK_BORDER = Side(style='thick', color='000000')

# Cores
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

# Cores para Ficheiro 1 vs Ficheiro 2 (bem distintas)
FILE1_FILL = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')  # Azul
FILE2_FILL = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')  # Laranja/Salmão
CHECK_FILL = PatternFill(start_color='E2E2E2', end_color='E2E2E2', fill_type='solid')  # Cinza

# Cores por CATEGORIA (cada uma diferente)
CATEGORY_COLORS = {
    'GENERAL': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),      # Azul escuro
    'INTERNALS': PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid'),    # Verde
    'INFILTRATION': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'), # Amarelo/Laranja
    'FLOORS': PatternFill(start_color='9E480E', end_color='9E480E', fill_type='solid'),       # Castanho
    'PARTITIONS': PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),   # Roxo
    'WALLS': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),        # Vermelho escuro
    'ROOFS': PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid'),        # Azul claro
}

# Cores claras para subcategorias (baseadas nas categorias)
SUBCATEGORY_COLORS = {
    'GENERAL': PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid'),
    'INTERNALS': PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid'),
    'INFILTRATION': PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid'),
    'FLOORS': PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid'),
    'PARTITIONS': PatternFill(start_color='CDA4DE', end_color='CDA4DE', fill_type='solid'),
    'WALLS': PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'),
    'ROOFS': PatternFill(start_color='9DC3E6', end_color='9DC3E6', fill_type='solid'),
}

def normalize_value(v):
    """Normaliza valor para comparação"""
    if v is None or v == '':
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == '' or v.lower() in ('none', 'n/a', '-'):
            return None
        try:
            f = float(v)
            return round(f, 2) if f != 0 else None
        except:
            return v.lower()
    if isinstance(v, (int, float)):
        return round(float(v), 2) if v != 0 else None
    return v

def compare_values(v1, v2, tolerance=0.5):
    """Compara dois valores"""
    n1 = normalize_value(v1)
    n2 = normalize_value(v2)

    # Ambos vazios = OK
    if n1 is None and n2 is None:
        return True, ''

    # Um vazio, outro não
    if n1 is None:
        return False, 'F2'  # Só no ficheiro 2
    if n2 is None:
        return False, 'F1'  # Só no ficheiro 1

    # Strings
    if isinstance(n1, str) or isinstance(n2, str):
        s1 = str(n1).lower().replace(' ', '')
        s2 = str(n2).lower().replace(' ', '')
        if s1 == s2:
            return True, 'OK'
        if s1 in s2 or s2 in s1:
            return True, '~'  # Match parcial
        return False, 'DIFF'

    # Números
    if abs(n1 - n2) < tolerance:
        return True, 'OK'
    if n1 != 0 and abs(n1 - n2) / abs(n1) < 0.05:
        return True, '~'
    return False, 'DIFF'

def read_excel_data(filepath):
    """Lê dados do Excel HAP"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Espacos']

    # Categorias (linha 1)
    categories = {}
    current_cat = ''
    for col in range(1, 148):
        val = ws.cell(1, col).value
        if val:
            current_cat = val
        categories[col] = current_cat

    # Subcategorias (linha 2)
    subcategories = {}
    current_sub = ''
    for col in range(1, 148):
        val = ws.cell(2, col).value
        if val:
            current_sub = val
        subcategories[col] = current_sub

    # Headers (linha 3)
    headers = [ws.cell(3, col).value or '' for col in range(1, 148)]

    # Dados (linha 4+)
    data = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or str(name).strip() == '':
            continue
        name = str(name).strip()
        row_data = [ws.cell(row, col).value for col in range(1, 148)]
        data[name] = row_data

    return categories, subcategories, headers, data

def get_short_name(filepath):
    """Extrai nome curto do ficheiro para usar como label"""
    basename = os.path.basename(filepath)
    # Remover extensão
    name = os.path.splitext(basename)[0]
    # Remover sufixos comuns
    name = re.sub(r'_extraido$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'_export.*$', '', name, flags=re.IGNORECASE)
    # Se nome muito longo, abreviar
    if len(name) > 25:
        name = name[:22] + '...'
    return name

def create_comparison_excel(categories, subcategories, headers, data1, data2, output_path, name1='Ficheiro1', name2='Ficheiro2'):
    """Cria Excel com comparação lado a lado"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparacao'

    # Estrutura: para cada campo original, 3 colunas (File1, File2, Check)
    # Total: 147 campos * 3 = 441 colunas

    # === Calcular ranges para categorias ===
    cat_ranges = []  # [(start_out_col, end_out_col, cat_name, has_subcategories), ...]
    current_cat = None
    start_in_col = 1

    for in_col in range(1, 149):  # 149 para fechar última
        cat = categories.get(in_col, '') if in_col < 148 else ''
        if cat != current_cat:
            if current_cat is not None:
                # Verificar se esta categoria tem subcategorias
                has_subs = False
                for c in range(start_in_col, in_col):
                    if subcategories.get(c, ''):
                        has_subs = True
                        break
                start_out = (start_in_col - 1) * 3 + 1
                end_out = (in_col - 1) * 3 if in_col <= 147 else 147 * 3
                cat_ranges.append((start_out, end_out, current_cat, has_subs))
            current_cat = cat
            start_in_col = in_col

    # === Calcular ranges para subcategorias ===
    sub_ranges = []  # [(start_out_col, end_out_col, sub_name, parent_cat), ...]
    current_sub = None
    start_in_col = 1

    for in_col in range(1, 149):
        sub = subcategories.get(in_col, '') if in_col < 148 else ''
        if sub != current_sub:
            if current_sub is not None and current_sub != '':
                start_out = (start_in_col - 1) * 3 + 1
                end_out = (in_col - 1) * 3
                parent_cat = categories.get(start_in_col, '')
                sub_ranges.append((start_out, end_out, current_sub, parent_cat))
            current_sub = sub
            start_in_col = in_col

    # === LINHA 1 e 2: Categorias e Subcategorias ===
    for start, end, cat_name, has_subs in cat_ranges:
        if cat_name:
            cat_fill = CATEGORY_COLORS.get(cat_name, CATEGORY_COLORS['GENERAL'])

            if has_subs:
                # Categoria COM subcategorias: só merge horizontal na linha 1
                c = ws.cell(1, start, value=cat_name)
                c.fill = cat_fill
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal='center', vertical='center')
                if end > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                    for col in range(start, end + 1):
                        ws.cell(1, col).fill = cat_fill
            else:
                # Categoria SEM subcategorias: merge VERTICAL (linhas 1 e 2) e horizontal
                c = ws.cell(1, start, value=cat_name)
                c.fill = cat_fill
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal='center', vertical='center')
                # Merge vertical (linhas 1-2) e horizontal
                ws.merge_cells(start_row=1, start_column=start, end_row=2, end_column=end)
                # Aplicar cor a todas as células merged
                for col in range(start, end + 1):
                    ws.cell(1, col).fill = cat_fill
                    ws.cell(2, col).fill = cat_fill

    # === LINHA 2: Subcategorias ===
    for start, end, sub_name, parent_cat in sub_ranges:
        if sub_name:
            sub_fill = SUBCATEGORY_COLORS.get(parent_cat, SUBCATEGORY_COLORS['GENERAL'])
            c = ws.cell(2, start, value=sub_name)
            c.fill = sub_fill
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
            if end > start:
                ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
                for col in range(start, end + 1):
                    ws.cell(2, col).fill = sub_fill

    # === LINHA 3: Headers (File1 | File2 | OK?) ===
    out_col = 1
    for in_col in range(1, 148):
        h = headers[in_col - 1].replace('\n', ' ') if headers[in_col - 1] else f'Col{in_col}'

        # Coluna Ficheiro 1 (Azul)
        c = ws.cell(3, out_col, value=f'{h} ({name1})')
        c.fill = FILE1_FILL
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(wrap_text=True, horizontal='center')

        # Coluna Ficheiro 2 (Laranja/Salmão)
        c = ws.cell(3, out_col + 1, value=f'{h} ({name2})')
        c.fill = FILE2_FILL
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(wrap_text=True, horizontal='center')

        # Coluna Check (Cinza)
        c = ws.cell(3, out_col + 2, value='?')
        c.fill = CHECK_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')

        out_col += 3

    # === LINHA 4+: Dados ===
    row = 4
    stats = {'ok': 0, 'diff': 0, 'orig_only': 0, 'e3a_only': 0, 'both_empty': 0}

    # Cores claras para dados (mais subtis que headers)
    FILE1_DATA = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # Azul muito claro
    FILE2_DATA = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')  # Laranja muito claro

    all_spaces = sorted(set(data1.keys()) | set(data2.keys()))

    for space_name in all_spaces:
        orig = data1.get(space_name)
        extr = data2.get(space_name)

        if orig is None or extr is None:
            continue  # Skip espaços que não existem em ambos

        out_col = 1
        for in_col in range(1, 148):
            v1 = orig[in_col - 1]
            v2 = extr[in_col - 1]

            # Coluna Ficheiro 1 (com cor de fundo)
            c1 = ws.cell(row, out_col, value=v1)
            c1.fill = FILE1_DATA

            # Coluna Ficheiro 2 (com cor de fundo)
            c2 = ws.cell(row, out_col + 1, value=v2)
            c2.fill = FILE2_DATA

            # Coluna Check
            match, status = compare_values(v1, v2)
            c = ws.cell(row, out_col + 2, value=status)
            c.alignment = Alignment(horizontal='center')

            if status == '':
                stats['both_empty'] += 1
            elif match:
                c.fill = GREEN_FILL
                stats['ok'] += 1
            else:
                c.fill = RED_FILL
                if status == 'F1':
                    stats['orig_only'] += 1
                elif status == 'F2':
                    stats['e3a_only'] += 1
                else:
                    stats['diff'] += 1

            out_col += 3

        row += 1

    # === Aplicar BORDAS GROSSAS à volta de cada secção ===
    last_row = row - 1  # Última linha de dados

    # Calcular TODAS as secções que precisam de bordas
    all_section_ranges = []

    # Adicionar subcategorias (começam na linha 2)
    for start, end, sub_name, parent_cat in sub_ranges:
        if sub_name:
            all_section_ranges.append((start, end, sub_name, 2))  # start_row = 2

    # Adicionar categorias SEM subcategorias (começam na linha 1, ocupam linhas 1-2)
    for start, end, cat_name, has_subs in cat_ranges:
        if cat_name and not has_subs:
            all_section_ranges.append((start, end, cat_name, 1))  # start_row = 1

    # Aplicar bordas grossas a TODAS as secções
    for start_col, end_col, section_name, start_row in all_section_ranges:
        if not section_name:
            continue

        # Borda à volta de toda a secção (desde start_row até última linha de dados)
        for r in range(start_row, last_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(r, c)

                # Determinar quais lados precisam de borda grossa
                left = THICK_BORDER if c == start_col else THIN_BORDER
                right = THICK_BORDER if c == end_col else THIN_BORDER
                top = THICK_BORDER if r == start_row else THIN_BORDER
                bottom = THICK_BORDER if r == last_row else THIN_BORDER

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # === Ajustar largura das colunas ===
    for col in range(1, out_col):
        col_letter = get_column_letter(col)
        if (col - 1) % 3 == 2:  # Colunas de check
            ws.column_dimensions[col_letter].width = 5
        else:
            ws.column_dimensions[col_letter].width = 12

    # Primeira coluna (Space Name) mais larga
    ws.column_dimensions['A'].width = 22

    # Freeze panes
    ws.freeze_panes = 'D4'

    # === Folha de Sumário ===
    ws_sum = wb.create_sheet('Sumario')
    ws_sum.cell(1, 1, value='SUMÁRIO DA COMPARAÇÃO')
    ws_sum.cell(1, 1).font = Font(bold=True, size=14)

    ws_sum.cell(3, 1, value='Espaços comparados:')
    ws_sum.cell(3, 2, value=len(all_spaces))

    ws_sum.cell(5, 1, value='Campos OK:')
    ws_sum.cell(5, 2, value=stats['ok'])
    ws_sum.cell(5, 2).fill = GREEN_FILL

    ws_sum.cell(6, 1, value='Campos diferentes:')
    ws_sum.cell(6, 2, value=stats['diff'])
    ws_sum.cell(6, 2).fill = RED_FILL

    ws_sum.cell(7, 1, value=f'Só em {name1}:')
    ws_sum.cell(7, 2, value=stats['orig_only'])

    ws_sum.cell(8, 1, value=f'Só em {name2}:')
    ws_sum.cell(8, 2, value=stats['e3a_only'])

    ws_sum.cell(9, 1, value='Ambos vazios:')
    ws_sum.cell(9, 2, value=stats['both_empty'])

    total_com_dados = stats['ok'] + stats['diff']
    if total_com_dados > 0:
        pct = stats['ok'] / total_com_dados * 100
        ws_sum.cell(11, 1, value='Match % (campos com dados):')
        ws_sum.cell(11, 2, value=f'{pct:.1f}%')
        ws_sum.cell(11, 2).font = Font(bold=True, size=12)

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 15

    # === Legenda ===
    ws_sum.cell(14, 1, value='LEGENDA:')
    ws_sum.cell(14, 1).font = Font(bold=True)

    ws_sum.cell(15, 1, value='OK')
    ws_sum.cell(15, 1).fill = GREEN_FILL
    ws_sum.cell(15, 2, value='Valores iguais')

    ws_sum.cell(16, 1, value='~')
    ws_sum.cell(16, 1).fill = GREEN_FILL
    ws_sum.cell(16, 2, value='Match parcial (substring)')

    ws_sum.cell(17, 1, value='DIFF')
    ws_sum.cell(17, 1).fill = RED_FILL
    ws_sum.cell(17, 2, value='Valores diferentes')

    ws_sum.cell(18, 1, value='F1')
    ws_sum.cell(18, 1).fill = RED_FILL
    ws_sum.cell(18, 2, value=f'Só existe em {name1}')

    ws_sum.cell(19, 1, value='F2')
    ws_sum.cell(19, 1).fill = RED_FILL
    ws_sum.cell(19, 2, value=f'Só existe em {name2}')

    ws_sum.cell(20, 1, value='(vazio)')
    ws_sum.cell(20, 2, value='Ambos vazios')

    wb.save(output_path)

    print(f'\n=== SUMÁRIO ===')
    print(f'Espaços:        {len(all_spaces)}')
    print(f'Campos OK:      {stats["ok"]}')
    print(f'Diferentes:     {stats["diff"]}')
    print(f'Só original:    {stats["orig_only"]}')
    print(f'Só E3A:         {stats["e3a_only"]}')
    print(f'Ambos vazios:   {stats["both_empty"]}')
    if total_com_dados > 0:
        print(f'\nMatch: {stats["ok"] / total_com_dados * 100:.1f}%')

    return wb

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    orig_file = sys.argv[1]
    extr_file = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else 'Comparacao_LadoALado.xlsx'

    if not os.path.exists(orig_file):
        print(f"Erro: '{orig_file}' não encontrado!")
        sys.exit(1)
    if not os.path.exists(extr_file):
        print(f"Erro: '{extr_file}' não encontrado!")
        sys.exit(1)

    print(f'Original: {orig_file}')
    print(f'E3A:      {extr_file}')

    # Ler dados
    cat1, sub1, hdr1, data1 = read_excel_data(orig_file)
    cat2, sub2, hdr2, data2 = read_excel_data(extr_file)

    # Extrair nomes curtos dos ficheiros
    name1 = get_short_name(orig_file)
    name2 = get_short_name(extr_file)

    print(f'\nEspaços em {name1}: {len(data1)}')
    print(f'Espaços em {name2}: {len(data2)}')

    # Criar comparação
    create_comparison_excel(cat1, sub1, hdr1, data1, data2, output_file, name1, name2)

    print(f'\nFicheiro criado: {output_file}')

if __name__ == '__main__':
    main()
