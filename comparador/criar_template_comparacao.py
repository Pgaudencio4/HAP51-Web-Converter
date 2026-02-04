"""
Cria template de comparação baseado no formato do conversor (HAP_Template_RSECE.xlsx)

O template gerado tem para cada campo original 3 colunas:
- PREV (F1) - valores do primeiro ficheiro
- REF (F2) - valores do segundo ficheiro
- ? - indicador de diferença (OK, DIFF, F1, F2)

Usage:
    python criar_template_comparacao.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Cores
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
PREV_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Amarelo
REF_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')   # Verde
CHECK_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid') # Cinza

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def criar_template():
    """Cria template de comparação a partir do template do conversor"""

    # Carregar template do conversor
    conversor_path = os.path.join(os.path.dirname(__file__), '..', 'conversor', 'templates', 'HAP_Template_RSECE.xlsx')
    wb_conv = openpyxl.load_workbook(conversor_path)
    ws_conv = wb_conv['Espacos']

    # Criar novo workbook
    wb = openpyxl.Workbook()

    # =========================================================================
    # FOLHA COMPARAÇÃO (Espaços)
    # =========================================================================
    ws = wb.active
    ws.title = 'Comparacao'

    # Ler headers do conversor (linha 3)
    conv_headers = []
    for col in range(1, 148):
        val = ws_conv.cell(3, col).value
        conv_headers.append(val if val else '')

    # Ler categorias do conversor (linha 1)
    conv_categories = {}
    for col in range(1, 148):
        val = ws_conv.cell(1, col).value
        if val:
            conv_categories[col] = val

    # Ler subcategorias do conversor (linha 2)
    conv_subcategories = {}
    for col in range(1, 148):
        val = ws_conv.cell(2, col).value
        if val:
            conv_subcategories[col] = val

    # === LINHA 1: Categorias (expandidas para 3 colunas cada) ===
    for orig_col, cat_name in conv_categories.items():
        new_col = (orig_col - 1) * 3 + 1
        cell = ws.cell(1, new_col, value=cat_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # === LINHA 2: Subcategorias (expandidas para 3 colunas cada) ===
    for orig_col, subcat_name in conv_subcategories.items():
        new_col = (orig_col - 1) * 3 + 1
        cell = ws.cell(2, new_col, value=subcat_name)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)

    # === LINHA 3: Headers (3 colunas por campo: PREV, REF, ?) ===
    for i, header in enumerate(conv_headers):
        base_col = i * 3 + 1
        header_clean = header.replace('\n', ' ').strip() if header else ''

        # Coluna PREV (F1)
        c1 = ws.cell(3, base_col, value=f"{header_clean} (PREV)")
        c1.fill = PREV_FILL
        c1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        c1.border = THIN_BORDER

        # Coluna REF (F2)
        c2 = ws.cell(3, base_col + 1, value=f"{header_clean} (REF)")
        c2.fill = REF_FILL
        c2.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        c2.border = THIN_BORDER

        # Coluna ?
        c3 = ws.cell(3, base_col + 2, value="?")
        c3.fill = CHECK_FILL
        c3.alignment = Alignment(horizontal='center', vertical='center')
        c3.border = THIN_BORDER

    # Ajustar largura das colunas
    for col in range(1, 147 * 3 + 1):
        col_type = (col - 1) % 3
        if col_type == 2:  # Coluna ?
            ws.column_dimensions[get_column_letter(col)].width = 6
        else:
            ws.column_dimensions[get_column_letter(col)].width = 12

    # Congelar painéis
    ws.freeze_panes = 'D4'

    # =========================================================================
    # FOLHA WINDOWS
    # =========================================================================
    ws_win = wb.create_sheet('Windows')
    win_headers = ['Nome', 'U-Value (W/m2K)', 'SHGC', 'Altura (m)', 'Largura (m)']

    for i, header in enumerate(win_headers):
        base_col = i * 3 + 1

        c1 = ws_win.cell(3, base_col, value=f"{header} (PREV)")
        c1.fill = PREV_FILL
        c1.border = THIN_BORDER

        c2 = ws_win.cell(3, base_col + 1, value=f"{header} (REF)")
        c2.fill = REF_FILL
        c2.border = THIN_BORDER

        c3 = ws_win.cell(3, base_col + 2, value="?")
        c3.fill = CHECK_FILL
        c3.border = THIN_BORDER

    ws_win.freeze_panes = 'D4'

    # =========================================================================
    # FOLHA WALLS
    # =========================================================================
    ws_wal = wb.create_sheet('Walls')
    wal_headers = ['Nome', 'U-Value (W/m2K)', 'Espessura (m)', 'Massa (kg/m2)']

    for i, header in enumerate(wal_headers):
        base_col = i * 3 + 1

        c1 = ws_wal.cell(3, base_col, value=f"{header} (PREV)")
        c1.fill = PREV_FILL
        c1.border = THIN_BORDER

        c2 = ws_wal.cell(3, base_col + 1, value=f"{header} (REF)")
        c2.fill = REF_FILL
        c2.border = THIN_BORDER

        c3 = ws_wal.cell(3, base_col + 2, value="?")
        c3.fill = CHECK_FILL
        c3.border = THIN_BORDER

    ws_wal.freeze_panes = 'D4'

    # =========================================================================
    # FOLHA ROOFS
    # =========================================================================
    ws_rof = wb.create_sheet('Roofs')
    rof_headers = ['Nome', 'U-Value (W/m2K)', 'Espessura (m)', 'Massa (kg/m2)']

    for i, header in enumerate(rof_headers):
        base_col = i * 3 + 1

        c1 = ws_rof.cell(3, base_col, value=f"{header} (PREV)")
        c1.fill = PREV_FILL
        c1.border = THIN_BORDER

        c2 = ws_rof.cell(3, base_col + 1, value=f"{header} (REF)")
        c2.fill = REF_FILL
        c2.border = THIN_BORDER

        c3 = ws_rof.cell(3, base_col + 2, value="?")
        c3.fill = CHECK_FILL
        c3.border = THIN_BORDER

    ws_rof.freeze_panes = 'D4'

    # Guardar
    output_path = os.path.join(os.path.dirname(__file__), 'Template_Comparacao.xlsx')
    wb.save(output_path)

    print(f"Template criado: {output_path}")
    print(f"  - Folha 'Comparacao': {147} campos x 3 colunas = {147*3} colunas")
    print(f"  - Folha 'Windows': {len(win_headers)} campos")
    print(f"  - Folha 'Walls': {len(wal_headers)} campos")
    print(f"  - Folha 'Roofs': {len(rof_headers)} campos")

    return output_path


if __name__ == '__main__':
    criar_template()
