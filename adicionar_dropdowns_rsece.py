"""
Adicionar Dropdowns RSECE ao Excel existente
=============================================

Adiciona dropdowns com os schedules RSECE nas colunas de schedules
que já existem no Excel (colunas 11, 16, 18, 21, 22).

Usage:
    python adicionar_dropdowns_rsece.py <input.xlsx> [output.xlsx]

Exemplo:
    python adicionar_dropdowns_rsece.py HAP_Exemplo_5Espacos.xlsx
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Importar lista de schedules
from criar_perfis_rsece import PERFIS_RSECE


def get_all_schedules():
    """Retorna lista de todos os schedules disponíveis."""
    schedules = ['Sample Schedule']
    for tipo_key in PERFIS_RSECE.keys():
        schedules.append(f'{tipo_key} Ocup')
        schedules.append(f'{tipo_key} Ilum')
        schedules.append(f'{tipo_key} Equip')
    return schedules


def adicionar_dropdowns(input_file: str, output_file: str = None):
    """Adiciona dropdowns de schedules RSECE ao Excel."""

    if output_file is None:
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_RSECE{ext}"

    print(f"A carregar: {input_file}")
    wb = load_workbook(input_file)
    ws = wb['Espacos']

    # Lista de schedules
    schedules = get_all_schedules()
    schedule_str = ','.join(schedules)
    print(f"Schedules disponíveis: {len(schedules)}")

    # Colunas de schedules no formato original:
    # Col 11: People Schedule
    # Col 16: Lighting Schedule
    # Col 18: Equipment Schedule
    # Col 21: Misc Sens Schedule
    # Col 22: Misc Lat Schedule
    schedule_columns = [11, 16, 18, 21, 22]

    # Encontrar última linha com dados
    max_row = ws.max_row
    print(f"Linhas de dados: 4 a {max_row}")

    # Adicionar dropdown a cada coluna de schedule
    for col in schedule_columns:
        col_letter = get_column_letter(col)
        header = ws.cell(row=3, column=col).value
        print(f"  Coluna {col} ({col_letter}): {header}")

        # Criar validação
        dv = DataValidation(
            type="list",
            formula1=f'"{schedule_str}"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Schedule Inválido",
            error="Selecione um schedule da lista RSECE.",
        )

        # Aplicar da linha 4 até max_row + 50 (para futuras linhas)
        dv.add(f'{col_letter}4:{col_letter}{max_row + 50}')
        ws.add_data_validation(dv)

    # Criar/actualizar sheet com lista de schedules para referência
    if 'Schedules_RSECE' in wb.sheetnames:
        del wb['Schedules_RSECE']

    ws_sch = wb.create_sheet('Schedules_RSECE')
    ws_sch.cell(row=1, column=1, value='Schedules RSECE Disponíveis')
    ws_sch.cell(row=1, column=2, value='Tipologia')

    for i, nome in enumerate(schedules, start=2):
        ws_sch.cell(row=i, column=1, value=nome)
        # Extrair tipologia
        if nome == 'Sample Schedule':
            ws_sch.cell(row=i, column=2, value='Default')
        else:
            tipo = nome.rsplit(' ', 1)[0]  # Remove "Ocup", "Ilum", "Equip"
            ws_sch.cell(row=i, column=2, value=tipo)

    ws_sch.column_dimensions['A'].width = 30
    ws_sch.column_dimensions['B'].width = 25

    # Guardar
    wb.save(output_file)
    print()
    print(f"Ficheiro guardado: {output_file}")
    print()
    print("Dropdowns adicionados nas colunas:")
    for col in schedule_columns:
        header = ws.cell(row=3, column=col).value
        if header:
            header_clean = str(header).replace('\n', ' ')[:30]
            print(f"  Col {col}: {header_clean}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_file):
        print(f"Erro: Ficheiro não encontrado: {input_file}")
        return

    adicionar_dropdowns(input_file, output_file)


if __name__ == '__main__':
    main()
