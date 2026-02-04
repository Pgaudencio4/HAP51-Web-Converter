"""
Processar HAP_Template_Malhoa22.xlsx - Versão Final
Usa TODOS os dados das sheets: Espacos, Windows, Walls, Roofs
"""
import openpyxl
import shutil
import math

print("="*60)
print("PROCESSADOR MALHOA22 - VERSÃO FINAL")
print("="*60)

# Carregar o ficheiro original
print("\n1. Carregando ficheiro...")
wb = openpyxl.load_workbook('HAP_Template_Malhoa22.xlsx')

# ============================================================
# 2. LER SHEET WALLS (tipos de parede)
# ============================================================
print("\n2. Lendo tipos de paredes (sheet Walls)...")
ws_walls_sheet = wb['Walls']
wall_types = {}
for row in range(4, ws_walls_sheet.max_row + 1):
    name = ws_walls_sheet.cell(row=row, column=1).value
    if name:
        wall_types[name] = {
            'u': ws_walls_sheet.cell(row=row, column=2).value or 0.5,
            'weight': ws_walls_sheet.cell(row=row, column=3).value or 200,
            'thickness': ws_walls_sheet.cell(row=row, column=4).value or 0.3,
        }
        print(f"   Wall Type: {name} (U={wall_types[name]['u']})")

DEFAULT_WALL_TYPE = list(wall_types.keys())[0] if wall_types else 'Parede Exterior'

# ============================================================
# 3. LER SHEET ROOFS (tipos de cobertura)
# ============================================================
print("\n3. Lendo tipos de coberturas (sheet Roofs)...")
ws_roofs_sheet = wb['Roofs']
roof_types = {}
for row in range(4, ws_roofs_sheet.max_row + 1):
    name = ws_roofs_sheet.cell(row=row, column=1).value
    if name:
        roof_types[name] = {
            'u': ws_roofs_sheet.cell(row=row, column=2).value or 0.4,
            'weight': ws_roofs_sheet.cell(row=row, column=3).value or 300,
            'thickness': ws_roofs_sheet.cell(row=row, column=4).value or 0.35,
        }
        print(f"   Roof Type: {name} (U={roof_types[name]['u']})")

DEFAULT_ROOF_TYPE = list(roof_types.keys())[0] if roof_types else 'Cobertura Plana'

# ============================================================
# 4. LER SHEET WINDOWS (tipos de janela com área e qty)
# ============================================================
print("\n4. Lendo janelas (sheet Windows)...")
ws_win_sheet = wb['Windows']
windows_data = {}
for row in range(3, ws_win_sheet.max_row + 1):
    name = ws_win_sheet.cell(row=row, column=1).value
    if name:
        windows_data[name] = {
            'zone': ws_win_sheet.cell(row=row, column=2).value,
            'piso': ws_win_sheet.cell(row=row, column=3).value,
            'ori': ws_win_sheet.cell(row=row, column=4).value,
            'u': ws_win_sheet.cell(row=row, column=5).value or 5.75,
            'shgc': ws_win_sheet.cell(row=row, column=6).value or 0.85,
            'area': ws_win_sheet.cell(row=row, column=7).value or 0,
            'qty': ws_win_sheet.cell(row=row, column=8).value or 1,
        }

print(f"   Total de janelas: {len(windows_data)}")

# Criar tipos de janelas únicos baseados em U-value e SHGC
unique_window_types = {}
for name, wdata in windows_data.items():
    # Calcular dimensões da janela baseado na área e qty
    total_area = wdata['area']
    qty = wdata['qty'] or 1
    area_per_window = total_area / qty if qty > 0 else total_area

    # Assumir altura de 1.5m e calcular largura
    height = 1.5
    width = area_per_window / height if area_per_window > 0 else 1.0

    # Criar tipo único
    u_val = round(wdata['u'], 2)
    shgc_val = round(wdata['shgc'], 2)

    # Usar o próprio nome da janela como tipo (cada janela é um tipo)
    unique_window_types[name] = {
        'name': name,
        'u': u_val,
        'shgc': shgc_val,
        'height': height,
        'width': round(width, 2),
        'area': round(area_per_window, 2),
    }

print(f"   Tipos de janelas criados: {len(unique_window_types)}")

# ============================================================
# 5. PROCESSAR SHEET ESPACOS - PREENCHER WALL TYPES
# ============================================================
print("\n5. Processando espaços (sheet Espacos)...")
ws = wb['Espacos']

espacos_processados = 0
walls_preenchidas = 0
roofs_preenchidas = 0

for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    espacos_processados += 1

    # Processar 8 walls
    for w in range(8):
        col = 52 + w * 9
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value

        if exp and area:
            # Preencher Wall Type se estiver vazio
            wall_type = ws.cell(row=row, column=col+2).value
            if not wall_type:
                ws.cell(row=row, column=col+2, value=DEFAULT_WALL_TYPE)
            walls_preenchidas += 1

    # Processar 4 roofs (colunas 124+)
    for r in range(4):
        col = 124 + r * 6
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value

        if exp and area:
            # Preencher Roof Type se estiver vazio
            roof_type = ws.cell(row=row, column=col+3).value
            if not roof_type:
                ws.cell(row=row, column=col+3, value=DEFAULT_ROOF_TYPE)
            roofs_preenchidas += 1

print(f"   Espaços processados: {espacos_processados}")
print(f"   Walls preenchidas: {walls_preenchidas}")
print(f"   Roofs preenchidas: {roofs_preenchidas}")

# ============================================================
# 6. RECRIAR SHEET WINDOWS NO FORMATO DO CONVERSOR
# ============================================================
print("\n6. Recriando sheet Windows no formato do conversor...")

# Apagar sheet existente
if 'Windows' in wb.sheetnames:
    del wb['Windows']
ws_win_new = wb.create_sheet('Windows')

# Cabeçalhos (formato esperado pelo conversor)
ws_win_new.cell(row=1, column=1, value='Nome')
ws_win_new.cell(row=1, column=2, value='U-Value')
ws_win_new.cell(row=1, column=3, value='SHGC')
ws_win_new.cell(row=1, column=4, value='Altura')
ws_win_new.cell(row=1, column=5, value='Largura')

# Dados
row = 2
for name, wdata in unique_window_types.items():
    ws_win_new.cell(row=row, column=1, value=wdata['name'])
    ws_win_new.cell(row=row, column=2, value=wdata['u'])
    ws_win_new.cell(row=row, column=3, value=wdata['shgc'])
    ws_win_new.cell(row=row, column=4, value=wdata['height'])
    ws_win_new.cell(row=row, column=5, value=wdata['width'])
    row += 1

print(f"   Janelas escritas: {len(unique_window_types)}")

# ============================================================
# 7. RECRIAR SHEET WALLS NO FORMATO DO CONVERSOR
# ============================================================
print("\n7. Recriando sheet Walls no formato do conversor...")

if 'Walls' in wb.sheetnames:
    del wb['Walls']
ws_walls_new = wb.create_sheet('Walls')

ws_walls_new.cell(row=1, column=1, value='Nome')
ws_walls_new.cell(row=1, column=2, value='U-Value')
ws_walls_new.cell(row=1, column=3, value='Peso')
ws_walls_new.cell(row=1, column=4, value='Espessura')

row = 2
for name, wdata in wall_types.items():
    ws_walls_new.cell(row=row, column=1, value=name)
    ws_walls_new.cell(row=row, column=2, value=wdata['u'])
    ws_walls_new.cell(row=row, column=3, value=wdata['weight'])
    ws_walls_new.cell(row=row, column=4, value=wdata['thickness'])
    row += 1

print(f"   Wall types escritos: {len(wall_types)}")

# ============================================================
# 8. RECRIAR SHEET ROOFS NO FORMATO DO CONVERSOR
# ============================================================
print("\n8. Recriando sheet Roofs no formato do conversor...")

if 'Roofs' in wb.sheetnames:
    del wb['Roofs']
ws_roofs_new = wb.create_sheet('Roofs')

ws_roofs_new.cell(row=1, column=1, value='Nome')
ws_roofs_new.cell(row=1, column=2, value='U-Value')
ws_roofs_new.cell(row=1, column=3, value='Peso')
ws_roofs_new.cell(row=1, column=4, value='Espessura')

row = 2
for name, rdata in roof_types.items():
    ws_roofs_new.cell(row=row, column=1, value=name)
    ws_roofs_new.cell(row=row, column=2, value=rdata['u'])
    ws_roofs_new.cell(row=row, column=3, value=rdata['weight'])
    ws_roofs_new.cell(row=row, column=4, value=rdata['thickness'])
    row += 1

print(f"   Roof types escritos: {len(roof_types)}")

# ============================================================
# 9. GUARDAR FICHEIRO
# ============================================================
output_file = 'Malhoa22_Final.xlsx'
print(f"\n9. Guardando ficheiro: {output_file}")
wb.save(output_file)

print("\n" + "="*60)
print("PROCESSAMENTO COMPLETO!")
print("="*60)
print(f"\nFicheiro criado: {output_file}")
print(f"\nResumo:")
print(f"  - Espaços: {espacos_processados}")
print(f"  - Walls com dados: {walls_preenchidas}")
print(f"  - Roofs com dados: {roofs_preenchidas}")
print(f"  - Tipos de janelas: {len(unique_window_types)}")
print(f"  - Tipos de paredes: {len(wall_types)}")
print(f"  - Tipos de coberturas: {len(roof_types)}")
print(f"\nPróximo passo:")
print(f"  python excel_to_hap.py {output_file} Modelo_RSECE.E3A Malhoa22.E3A")
