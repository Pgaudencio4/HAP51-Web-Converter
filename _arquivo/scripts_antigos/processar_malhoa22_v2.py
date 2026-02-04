"""
Processar HAP_Template_Malhoa22.xlsx e criar ficheiro completo
Versão 2 - Criar paredes e janelas genéricas para espaços com envolvente
"""
import openpyxl
import shutil
import copy

# Carregar o ficheiro original
print("Carregando ficheiro...")
wb = openpyxl.load_workbook('HAP_Template_Malhoa22.xlsx')

ws = wb['Espacos']

# ============================================================
# 1. DEFINIR TIPOS DE PAREDES E JANELAS
# ============================================================

# Tipo de parede exterior
WALL_TYPE = 'Parede Exterior'
WALL_U = 0.5  # W/m2K
WALL_WEIGHT = 200  # kg/m2
WALL_THICKNESS = 0.3  # m

# Tipo de janela
WIN_TYPE = 'Janela Simples'
WIN_U = 5.75  # W/m2K
WIN_SHGC = 0.85
WIN_HEIGHT = 1.5  # m
WIN_WIDTH = 2.0  # m

# Tipo de cobertura
ROOF_TYPE = 'Cobertura Plana'
ROOF_U = 0.4  # W/m2K
ROOF_WEIGHT = 300  # kg/m2
ROOF_THICKNESS = 0.35  # m

# ============================================================
# 2. IDENTIFICAR ESPAÇOS COM ENVOLVENTE EXTERIOR
# ============================================================
# Espaços que tipicamente têm paredes exteriores
ESPACOS_COM_ENVOLVENTE = ['Area', 'Escritorio', 'Cozinha']

# Mapeamento piso -> orientações típicas
# Assumindo edifício rectangular com todas as orientações
ORIENTACOES = ['N', 'S', 'E', 'W']

print("Processando espaços...")
espacos_actualizados = 0

for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    area = ws.cell(row=row, column=2).value or 0
    height = ws.cell(row=row, column=3).value or 2.6

    # Verificar se é um espaço com envolvente
    tem_envolvente = any(tipo in space_name for tipo in ESPACOS_COM_ENVOLVENTE)

    if tem_envolvente and area > 10:  # Só para espaços significativos
        # Calcular perímetro aproximado (assumindo quadrado)
        import math
        perimetro = 4 * math.sqrt(area)

        # Distribuir paredes por 4 orientações
        parede_comprimento = perimetro / 4
        parede_area = parede_comprimento * height

        # Área de janela (30% da parede para escritórios)
        ratio_janela = 0.3 if 'Area' in space_name or 'Escritorio' in space_name else 0.15
        janela_area = parede_area * ratio_janela

        # Número de janelas (arredondar para cima)
        win_area_unit = WIN_HEIGHT * WIN_WIDTH
        num_janelas = max(1, int(janela_area / win_area_unit))

        # Criar 4 paredes (uma para cada orientação)
        for i, ori in enumerate(ORIENTACOES):
            col = 52 + i * 9  # Coluna base para wall i

            # Escrever dados da wall
            ws.cell(row=row, column=col, value=ori)  # Exposure
            ws.cell(row=row, column=col+1, value=round(parede_area, 2))  # Gross Area
            ws.cell(row=row, column=col+2, value=WALL_TYPE)  # Wall Type
            ws.cell(row=row, column=col+3, value=WIN_TYPE)  # Window 1
            ws.cell(row=row, column=col+4, value=num_janelas)  # Win1 Qty

        espacos_actualizados += 1

print(f"Espaços actualizados com paredes: {espacos_actualizados}")

# ============================================================
# 3. ADICIONAR COBERTURAS AO ÚLTIMO PISO
# ============================================================
# Identificar espaços do último piso (S12)
print("Adicionando coberturas ao último piso...")
coberturas_adicionadas = 0

for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    # Verificar se é piso 12 (S12) ou similar topo
    if space_name.startswith('S12') or space_name.startswith('R11') or space_name.startswith('M06'):
        area = ws.cell(row=row, column=2).value or 0

        if area > 0:
            # Coluna 124 = primeira cobertura
            ws.cell(row=row, column=124, value='N')  # Exposure (horizontal = N)
            ws.cell(row=row, column=125, value=round(area, 2))  # Gross Area
            ws.cell(row=row, column=126, value=0)  # Slope (plana)
            ws.cell(row=row, column=127, value=ROOF_TYPE)  # Roof Type

            coberturas_adicionadas += 1

print(f"Coberturas adicionadas: {coberturas_adicionadas}")

# ============================================================
# 4. CRIAR/ACTUALIZAR SHEET WINDOWS
# ============================================================
print("Actualizando sheet Windows...")

# Remover sheet existente e criar nova
if 'Windows' in wb.sheetnames:
    del wb['Windows']
ws_win = wb.create_sheet('Windows')

# Cabeçalhos
ws_win.cell(row=1, column=1, value='Nome')
ws_win.cell(row=1, column=2, value='U-Value')
ws_win.cell(row=1, column=3, value='SHGC')
ws_win.cell(row=1, column=4, value='Altura')
ws_win.cell(row=1, column=5, value='Largura')

# Dados
ws_win.cell(row=2, column=1, value=WIN_TYPE)
ws_win.cell(row=2, column=2, value=WIN_U)
ws_win.cell(row=2, column=3, value=WIN_SHGC)
ws_win.cell(row=2, column=4, value=WIN_HEIGHT)
ws_win.cell(row=2, column=5, value=WIN_WIDTH)

# ============================================================
# 5. CRIAR/ACTUALIZAR SHEET WALLS
# ============================================================
print("Actualizando sheet Walls...")

if 'Walls' in wb.sheetnames:
    del wb['Walls']
ws_walls = wb.create_sheet('Walls')

ws_walls.cell(row=1, column=1, value='Nome')
ws_walls.cell(row=1, column=2, value='U-Value')
ws_walls.cell(row=1, column=3, value='Peso')
ws_walls.cell(row=1, column=4, value='Espessura')

ws_walls.cell(row=2, column=1, value=WALL_TYPE)
ws_walls.cell(row=2, column=2, value=WALL_U)
ws_walls.cell(row=2, column=3, value=WALL_WEIGHT)
ws_walls.cell(row=2, column=4, value=WALL_THICKNESS)

# ============================================================
# 6. CRIAR/ACTUALIZAR SHEET ROOFS
# ============================================================
print("Actualizando sheet Roofs...")

if 'Roofs' in wb.sheetnames:
    del wb['Roofs']
ws_roofs = wb.create_sheet('Roofs')

ws_roofs.cell(row=1, column=1, value='Nome')
ws_roofs.cell(row=1, column=2, value='U-Value')
ws_roofs.cell(row=1, column=3, value='Peso')
ws_roofs.cell(row=1, column=4, value='Espessura')

ws_roofs.cell(row=2, column=1, value=ROOF_TYPE)
ws_roofs.cell(row=2, column=2, value=ROOF_U)
ws_roofs.cell(row=2, column=3, value=ROOF_WEIGHT)
ws_roofs.cell(row=2, column=4, value=ROOF_THICKNESS)

# ============================================================
# 7. GUARDAR
# ============================================================
output_file = 'Malhoa22_Completo.xlsx'
wb.save(output_file)
print(f"\nFicheiro guardado: {output_file}")
print("\nAgora pode executar o conversor:")
print(f"  python excel_to_hap.py {output_file} Modelo_RSECE.E3A Malhoa22_Final.E3A")
