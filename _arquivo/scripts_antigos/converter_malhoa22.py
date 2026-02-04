"""
CONVERSOR COMPLETO: HAP_Template_Malhoa22.xlsx -> Malhoa22.E3A

Analisa e processa TODAS as sheets do ficheiro Excel:
- Espacos: 115 espaços com dados completos
- Windows: 71 tipos de janelas com U, SHGC, Área, Qtd
- Walls: 1 tipo (Parede Exterior)
- Roofs: 1 tipo (Cobertura Plana)
- Tipos: Mapeamento de IDs

O ficheiro Excel tem:
- Colunas 1-6: GENERAL (nome, área, altura, peso, OA)
- Colunas 7-11: PEOPLE (ocupação, actividade, calor, schedule)
- Colunas 12-16: LIGHTING (task, general, fixture, ballast, schedule)
- Colunas 17-18: EQUIPMENT (W/m2, schedule)
- Colunas 19-22: MISC
- Colunas 23-26: INFILTRATION (método, ACH clg/htg/energy)
- Colunas 27-39: FLOORS (tipo, área, U, etc.)
- Colunas 40-45: PARTITIONS CEILING
- Colunas 46-51: PARTITIONS WALL
- Colunas 52-123: WALLS (8 walls x 9 campos)
- Colunas 124-147: ROOFS (4 roofs x 6 campos)
"""

import openpyxl
import shutil
import os

print("="*70)
print("CONVERSOR MALHOA22 - ANÁLISE E CONVERSÃO COMPLETA")
print("="*70)

# ============================================================
# 1. CARREGAR FICHEIRO ORIGINAL
# ============================================================
print("\n[1] Carregando HAP_Template_Malhoa22.xlsx...")
wb = openpyxl.load_workbook('HAP_Template_Malhoa22.xlsx')
print(f"    Sheets: {wb.sheetnames}")

# ============================================================
# 2. LER E REGISTAR SHEET WALLS
# ============================================================
print("\n[2] Analisando sheet Walls...")
ws_walls = wb['Walls']
wall_types = {}
for row in range(4, ws_walls.max_row + 1):
    name = ws_walls.cell(row=row, column=1).value
    if name:
        wall_types[name] = {
            'u': ws_walls.cell(row=row, column=2).value or 0.5,
            'weight': ws_walls.cell(row=row, column=3).value or 200,
            'thickness': ws_walls.cell(row=row, column=4).value or 0.3,
        }
print(f"    Wall types encontrados: {len(wall_types)}")
for name, data in wall_types.items():
    print(f"      - {name}: U={data['u']}, Peso={data['weight']}, Esp={data['thickness']}")

DEFAULT_WALL_TYPE = list(wall_types.keys())[0] if wall_types else 'Parede Exterior'

# ============================================================
# 3. LER E REGISTAR SHEET ROOFS
# ============================================================
print("\n[3] Analisando sheet Roofs...")
ws_roofs = wb['Roofs']
roof_types = {}
for row in range(4, ws_roofs.max_row + 1):
    name = ws_roofs.cell(row=row, column=1).value
    if name:
        roof_types[name] = {
            'u': ws_roofs.cell(row=row, column=2).value or 0.4,
            'weight': ws_roofs.cell(row=row, column=3).value or 300,
            'thickness': ws_roofs.cell(row=row, column=4).value or 0.35,
        }
print(f"    Roof types encontrados: {len(roof_types)}")
for name, data in roof_types.items():
    print(f"      - {name}: U={data['u']}, Peso={data['weight']}, Esp={data['thickness']}")

DEFAULT_ROOF_TYPE = list(roof_types.keys())[0] if roof_types else 'Cobertura Plana'

# ============================================================
# 4. LER E REGISTAR SHEET WINDOWS
# ============================================================
print("\n[4] Analisando sheet Windows...")
ws_windows = wb['Windows']
windows_data = {}
for row in range(3, ws_windows.max_row + 1):
    name = ws_windows.cell(row=row, column=1).value
    if name:
        area_total = ws_windows.cell(row=row, column=7).value or 0
        qty = ws_windows.cell(row=row, column=8).value or 1
        area_per_win = area_total / qty if qty > 0 else area_total

        # Calcular dimensões: assumir altura 1.5m
        height = 1.5
        width = area_per_win / height if area_per_win > 0 else 1.0

        windows_data[name] = {
            'zone': ws_windows.cell(row=row, column=2).value,
            'piso': ws_windows.cell(row=row, column=3).value,
            'ori': ws_windows.cell(row=row, column=4).value,
            'u': ws_windows.cell(row=row, column=5).value or 5.75,
            'shgc': ws_windows.cell(row=row, column=6).value or 0.85,
            'area_total': area_total,
            'qty': qty,
            'height': height,
            'width': round(width, 3),
        }
print(f"    Windows encontradas: {len(windows_data)}")
print(f"    Exemplos:")
for i, (name, data) in enumerate(list(windows_data.items())[:3]):
    print(f"      - {name}: U={data['u']}, SHGC={data['shgc']}, H={data['height']}, W={data['width']}")

# ============================================================
# 5. ANALISAR SHEET ESPACOS
# ============================================================
print("\n[5] Analisando sheet Espacos...")
ws = wb['Espacos']

total_espacos = 0
espacos_com_walls = 0
espacos_com_roofs = 0
total_walls = 0
total_roofs = 0
walls_sem_tipo = 0

for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    total_espacos += 1
    tem_wall = False
    tem_roof = False

    # Contar walls
    for w in range(8):
        col = 52 + w * 9
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value
        wtype = ws.cell(row=row, column=col+2).value

        if exp or area:
            total_walls += 1
            tem_wall = True
            if not wtype:
                walls_sem_tipo += 1

    # Contar roofs
    for r in range(4):
        col = 124 + r * 6
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value

        if exp or area:
            total_roofs += 1
            tem_roof = True

    if tem_wall:
        espacos_com_walls += 1
    if tem_roof:
        espacos_com_roofs += 1

print(f"    Total espaços: {total_espacos}")
print(f"    Espaços com walls: {espacos_com_walls}")
print(f"    Espaços com roofs: {espacos_com_roofs}")
print(f"    Total walls: {total_walls}")
print(f"    Total roofs: {total_roofs}")
print(f"    Walls sem tipo definido: {walls_sem_tipo}")

# ============================================================
# 6. PREENCHER WALL TYPES E ROOF TYPES VAZIOS
# ============================================================
print("\n[6] Preenchendo Wall Types e Roof Types vazios...")

walls_preenchidas = 0
roofs_preenchidas = 0

for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    # Preencher Wall Types
    for w in range(8):
        col = 52 + w * 9
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value
        wtype = ws.cell(row=row, column=col+2).value

        if (exp or area) and not wtype:
            ws.cell(row=row, column=col+2, value=DEFAULT_WALL_TYPE)
            walls_preenchidas += 1

    # Preencher Roof Types
    for r in range(4):
        col = 124 + r * 6
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value
        rtype = ws.cell(row=row, column=col+3).value

        if (exp or area) and not rtype:
            ws.cell(row=row, column=col+3, value=DEFAULT_ROOF_TYPE)
            roofs_preenchidas += 1

print(f"    Wall types preenchidos: {walls_preenchidas}")
print(f"    Roof types preenchidos: {roofs_preenchidas}")

# ============================================================
# 7. RECRIAR SHEET WINDOWS NO FORMATO DO CONVERSOR
# ============================================================
print("\n[7] Recriando sheet Windows no formato do conversor...")

# Apagar sheet existente e criar nova
if 'Windows' in wb.sheetnames:
    del wb['Windows']
ws_win_new = wb.create_sheet('Windows')

# Cabeçalhos (formato esperado pelo excel_to_hap.py)
ws_win_new.cell(row=1, column=1, value='Nome')
ws_win_new.cell(row=1, column=2, value='U-Value')
ws_win_new.cell(row=1, column=3, value='SHGC')
ws_win_new.cell(row=1, column=4, value='Altura')
ws_win_new.cell(row=1, column=5, value='Largura')

# Dados
row = 2
for name, wdata in windows_data.items():
    ws_win_new.cell(row=row, column=1, value=name)
    ws_win_new.cell(row=row, column=2, value=round(wdata['u'], 2))
    ws_win_new.cell(row=row, column=3, value=round(wdata['shgc'], 2))
    ws_win_new.cell(row=row, column=4, value=wdata['height'])
    ws_win_new.cell(row=row, column=5, value=wdata['width'])
    row += 1

print(f"    Windows escritas: {len(windows_data)}")

# ============================================================
# 8. RECRIAR SHEET WALLS NO FORMATO DO CONVERSOR
# ============================================================
print("\n[8] Recriando sheet Walls no formato do conversor...")

if 'Walls' in wb.sheetnames:
    del wb['Walls']
ws_walls_new = wb.create_sheet('Walls')

ws_walls_new.cell(row=1, column=1, value='Nome')
ws_walls_new.cell(row=1, column=2, value='U-Value')
ws_walls_new.cell(row=1, column=3, value='Peso')
ws_walls_new.cell(row=1, column=4, value='Espessura')

row = 2
for name, wdata in wall_types.items():
    ws_walls_new.cell(row=row, column=1, value=name)
    ws_walls_new.cell(row=row, column=2, value=wdata['u'])
    ws_walls_new.cell(row=row, column=3, value=wdata['weight'])
    ws_walls_new.cell(row=row, column=4, value=wdata['thickness'])
    row += 1

print(f"    Wall types escritos: {len(wall_types)}")

# ============================================================
# 9. RECRIAR SHEET ROOFS NO FORMATO DO CONVERSOR
# ============================================================
print("\n[9] Recriando sheet Roofs no formato do conversor...")

if 'Roofs' in wb.sheetnames:
    del wb['Roofs']
ws_roofs_new = wb.create_sheet('Roofs')

ws_roofs_new.cell(row=1, column=1, value='Nome')
ws_roofs_new.cell(row=1, column=2, value='U-Value')
ws_roofs_new.cell(row=1, column=3, value='Peso')
ws_roofs_new.cell(row=1, column=4, value='Espessura')

row = 2
for name, rdata in roof_types.items():
    ws_roofs_new.cell(row=row, column=1, value=name)
    ws_roofs_new.cell(row=row, column=2, value=rdata['u'])
    ws_roofs_new.cell(row=row, column=3, value=rdata['weight'])
    ws_roofs_new.cell(row=row, column=4, value=rdata['thickness'])
    row += 1

print(f"    Roof types escritos: {len(roof_types)}")

# ============================================================
# 10. GUARDAR FICHEIRO PROCESSADO
# ============================================================
output_excel = 'Malhoa22_Processado.xlsx'
print(f"\n[10] Guardando ficheiro: {output_excel}")
wb.save(output_excel)
print(f"     Ficheiro guardado com sucesso!")

# ============================================================
# RESUMO
# ============================================================
print("\n" + "="*70)
print("RESUMO DA ANÁLISE")
print("="*70)
print(f"""
Ficheiro original: HAP_Template_Malhoa22.xlsx
Ficheiro processado: {output_excel}

DADOS ENCONTRADOS:
  - Espaços: {total_espacos}
  - Espaços com paredes exteriores: {espacos_com_walls}
  - Espaços com coberturas: {espacos_com_roofs}
  - Total de paredes: {total_walls}
  - Total de coberturas: {total_roofs}
  - Tipos de janelas: {len(windows_data)}
  - Tipos de paredes: {len(wall_types)}
  - Tipos de coberturas: {len(roof_types)}

CORRECÇÕES APLICADAS:
  - Wall types preenchidos: {walls_preenchidas}
  - Roof types preenchidos: {roofs_preenchidas}

PRÓXIMO PASSO:
  python excel_to_hap.py {output_excel} Modelo_RSECE.E3A Malhoa22.E3A
""")
