"""
Preparar HAP_Template_Malhoa22.xlsx para conversao
- Preenche Walls na sheet Espacos baseado nas Windows
- Ajusta formato das sheets Windows/Walls/Roofs para o conversor
"""
import openpyxl
import math

print("="*60)
print("PREPARAR MALHOA22 PARA CONVERSAO")
print("="*60)

# Carregar ficheiro original
INPUT_FILE = 'C:/Users/pedro/Downloads/HAP_Template_Malhoa22 (2).xlsx'
OUTPUT_FILE = 'Malhoa22_Pronto.xlsx'

print(f"\n1. Carregando {INPUT_FILE}...")
wb = openpyxl.load_workbook(INPUT_FILE)

# ============================================================
# 1. LER WINDOWS COM ZONA E ORIENTACAO
# ============================================================
print("\n2. Lendo janelas (sheet Windows)...")
ws_win = wb['Windows']

# Estrutura: Nome, Zona, Piso, Ori, U-Value, SHGC, Area, Qtd (row 2 = headers, row 3+ = dados)
windows_by_zone = {}  # zona -> {ori: {name, u, shgc, area_total, qty}}

for row in range(3, ws_win.max_row + 1):
    name = ws_win.cell(row=row, column=1).value
    zona = ws_win.cell(row=row, column=2).value
    piso = ws_win.cell(row=row, column=3).value
    ori = ws_win.cell(row=row, column=4).value
    u_value = ws_win.cell(row=row, column=5).value or 5.75
    shgc = ws_win.cell(row=row, column=6).value or 0.85
    area = ws_win.cell(row=row, column=7).value or 0
    qty = ws_win.cell(row=row, column=8).value or 1

    if not name or not zona:
        continue

    if zona not in windows_by_zone:
        windows_by_zone[zona] = {}

    if ori not in windows_by_zone[zona]:
        windows_by_zone[zona][ori] = {
            'name': name,
            'u': u_value,
            'shgc': shgc,
            'area_total': area,
            'qty': qty,
        }
    else:
        # Somar areas se mesma zona e orientacao
        windows_by_zone[zona][ori]['area_total'] += area
        windows_by_zone[zona][ori]['qty'] += qty

print(f"   Zonas com janelas: {len(windows_by_zone)}")

# ============================================================
# 2. LER TIPOS DE PAREDES
# ============================================================
print("\n3. Lendo tipos de paredes (sheet Walls)...")
ws_walls_orig = wb['Walls']

wall_types = {}
for row in range(4, ws_walls_orig.max_row + 1):
    name = ws_walls_orig.cell(row=row, column=1).value
    if name:
        wall_types[name] = {
            'u': ws_walls_orig.cell(row=row, column=2).value or 0.5,
            'weight': ws_walls_orig.cell(row=row, column=3).value or 200,
            'thickness': ws_walls_orig.cell(row=row, column=4).value or 0.3,
        }

DEFAULT_WALL_TYPE = list(wall_types.keys())[0] if wall_types else 'Parede Exterior'
print(f"   Wall types: {list(wall_types.keys())}")

# ============================================================
# 3. LER TIPOS DE COBERTURAS
# ============================================================
print("\n4. Lendo tipos de coberturas (sheet Roofs)...")
ws_roofs_orig = wb['Roofs']

roof_types = {}
for row in range(4, ws_roofs_orig.max_row + 1):
    name = ws_roofs_orig.cell(row=row, column=1).value
    if name:
        roof_types[name] = {
            'u': ws_roofs_orig.cell(row=row, column=2).value or 0.4,
            'weight': ws_roofs_orig.cell(row=row, column=3).value or 300,
            'thickness': ws_roofs_orig.cell(row=row, column=4).value or 0.35,
        }

DEFAULT_ROOF_TYPE = list(roof_types.keys())[0] if roof_types else 'Cobertura Plana'
print(f"   Roof types: {list(roof_types.keys())}")

# ============================================================
# 4. PREENCHER WALLS NA SHEET ESPACOS
# ============================================================
print("\n5. Preenchendo Walls na sheet Espacos...")
ws = wb['Espacos']

espacos_com_walls = 0
total_walls_criadas = 0

# Mapeamento de orientacoes
ORI_MAP = {'N': 'N', 'S': 'S', 'E': 'E', 'W': 'W', 'NE': 'NE', 'NW': 'NW', 'SE': 'SE', 'SW': 'SW'}

for row in range(4, ws.max_row + 1):
    space_name = ws.cell(row=row, column=1).value
    if not space_name:
        continue

    space_area = ws.cell(row=row, column=2).value or 0
    space_height = ws.cell(row=row, column=3).value or 2.8

    # Procurar zona correspondente nas janelas
    # Estrategia: extrair letra inicial + sequencia (001, 007, 008, etc.)
    zona_match = None

    # Parse do nome do espaco: A-06_001_Area_16 -> letra=A, seq=001
    parts = space_name.split('_')
    if len(parts) >= 2:
        space_letra = parts[0][0] if parts[0] else ''  # Primeira letra
        space_seq = parts[1] if len(parts) > 1 else ''  # Sequencia (001, 007, etc.)

        # Procurar zona com mesma letra e sequencia
        for zona in windows_by_zone.keys():
            # Parse da zona: A00_001_Area_24 -> letra=A, seq=001
            zona_parts = zona.split('_')
            if len(zona_parts) >= 2:
                zona_letra = zona_parts[0][0] if zona_parts[0] else ''
                zona_seq = zona_parts[1]

                # Match por letra + sequencia
                if space_letra == zona_letra and space_seq == zona_seq:
                    zona_match = zona
                    break

    if zona_match:
        windows = windows_by_zone[zona_match]
        wall_idx = 0

        for ori, win_data in windows.items():
            if wall_idx >= 8:  # Max 8 walls por espaco
                break

            # Calcular area da parede baseado na area da janela
            win_area = win_data['area_total']
            # Assumir que janela ocupa 30% da parede
            wall_area = win_area / 0.3 if win_area > 0 else space_area * 0.25

            # Coluna base para wall (52 + idx * 9)
            col = 52 + wall_idx * 9

            # Escrever dados da wall
            ws.cell(row=row, column=col, value=ORI_MAP.get(ori, ori))  # Exposure
            ws.cell(row=row, column=col+1, value=round(wall_area, 2))  # Gross Area
            ws.cell(row=row, column=col+2, value=DEFAULT_WALL_TYPE)    # Wall Type
            ws.cell(row=row, column=col+3, value=win_data['name'])     # Window 1
            ws.cell(row=row, column=col+4, value=win_data['qty'])      # Win1 Qty

            wall_idx += 1
            total_walls_criadas += 1

        if wall_idx > 0:
            espacos_com_walls += 1

print(f"   Espacos com walls: {espacos_com_walls}")
print(f"   Total walls criadas: {total_walls_criadas}")

# ============================================================
# 5. RECRIAR SHEET WINDOWS NO FORMATO DO CONVERSOR
# ============================================================
print("\n6. Recriando sheet Windows no formato do conversor...")

# Criar tipos de janelas unicos baseados em nome
unique_windows = {}
for zona, oris in windows_by_zone.items():
    for ori, win_data in oris.items():
        name = win_data['name']
        if name not in unique_windows:
            # Calcular dimensoes baseado na area total e qty
            area_per_win = win_data['area_total'] / win_data['qty'] if win_data['qty'] > 0 else win_data['area_total']
            height = 1.5
            width = area_per_win / height if area_per_win > 0 else 1.0

            unique_windows[name] = {
                'name': name,
                'u': win_data['u'],
                'shgc': win_data['shgc'],
                'height': height,
                'width': round(width, 3),
            }

# Apagar sheet existente e criar nova
if 'Windows' in wb.sheetnames:
    del wb['Windows']
ws_win_new = wb.create_sheet('Windows')

# Cabecalhos (IMPORTANTE: linhas 1-3 como o conversor espera)
ws_win_new.cell(row=1, column=1, value='WINDOWS')
ws_win_new.cell(row=2, column=1, value='')
ws_win_new.cell(row=3, column=1, value='Nome')
ws_win_new.cell(row=3, column=2, value='U-Value')
ws_win_new.cell(row=3, column=3, value='SHGC')
ws_win_new.cell(row=3, column=4, value='Altura')
ws_win_new.cell(row=3, column=5, value='Largura')

# Dados a partir da linha 4
row = 4
for name, wdata in unique_windows.items():
    ws_win_new.cell(row=row, column=1, value=wdata['name'])
    ws_win_new.cell(row=row, column=2, value=round(wdata['u'], 2))
    ws_win_new.cell(row=row, column=3, value=round(wdata['shgc'], 2))
    ws_win_new.cell(row=row, column=4, value=wdata['height'])
    ws_win_new.cell(row=row, column=5, value=wdata['width'])
    row += 1

print(f"   Windows escritas: {len(unique_windows)}")

# ============================================================
# 6. RECRIAR SHEET WALLS NO FORMATO DO CONVERSOR
# ============================================================
print("\n7. Recriando sheet Walls no formato do conversor...")

if 'Walls' in wb.sheetnames:
    del wb['Walls']
ws_walls_new = wb.create_sheet('Walls')

# Cabecalhos (linhas 1-3)
ws_walls_new.cell(row=1, column=1, value='WALLS')
ws_walls_new.cell(row=2, column=1, value='')
ws_walls_new.cell(row=3, column=1, value='Nome')
ws_walls_new.cell(row=3, column=2, value='U-Value')
ws_walls_new.cell(row=3, column=3, value='Peso')
ws_walls_new.cell(row=3, column=4, value='Espessura')

# Dados a partir da linha 4
row = 4
for name, wdata in wall_types.items():
    ws_walls_new.cell(row=row, column=1, value=name)
    ws_walls_new.cell(row=row, column=2, value=wdata['u'])
    ws_walls_new.cell(row=row, column=3, value=wdata['weight'])
    ws_walls_new.cell(row=row, column=4, value=wdata['thickness'])
    row += 1

print(f"   Wall types escritos: {len(wall_types)}")

# ============================================================
# 7. RECRIAR SHEET ROOFS NO FORMATO DO CONVERSOR
# ============================================================
print("\n8. Recriando sheet Roofs no formato do conversor...")

if 'Roofs' in wb.sheetnames:
    del wb['Roofs']
ws_roofs_new = wb.create_sheet('Roofs')

# Cabecalhos (linhas 1-3)
ws_roofs_new.cell(row=1, column=1, value='ROOFS')
ws_roofs_new.cell(row=2, column=1, value='')
ws_roofs_new.cell(row=3, column=1, value='Nome')
ws_roofs_new.cell(row=3, column=2, value='U-Value')
ws_roofs_new.cell(row=3, column=3, value='Peso')
ws_roofs_new.cell(row=3, column=4, value='Espessura')

# Dados a partir da linha 4
row = 4
for name, rdata in roof_types.items():
    ws_roofs_new.cell(row=row, column=1, value=name)
    ws_roofs_new.cell(row=row, column=2, value=rdata['u'])
    ws_roofs_new.cell(row=row, column=3, value=rdata['weight'])
    ws_roofs_new.cell(row=row, column=4, value=rdata['thickness'])
    row += 1

print(f"   Roof types escritos: {len(roof_types)}")

# ============================================================
# 8. CORRIGIR OFFSET DAS COLUNAS (remover Infil Method)
# ============================================================
print("\n9. Corrigindo offset das colunas (removendo Infil Method)...")

# O conversor espera:
# Col 23 = ACH Clg, Col 24 = ACH Htg, Col 25 = ACH Energy
# Mas o ficheiro original tem:
# Col 23 = Infil Method, Col 24 = ACH Clg, Col 25 = ACH Htg, Col 26 = ACH Energy

# Desagrupar celulas merged primeiro
merged_ranges = list(ws.merged_cells.ranges)
for merged_range in merged_ranges:
    ws.unmerge_cells(str(merged_range))

# Shift: mover colunas 24-147 para 23-146 (elimina coluna 23)
for row in range(1, ws.max_row + 1):
    for col in range(23, 147):
        ws.cell(row=row, column=col).value = ws.cell(row=row, column=col+1).value
    # Limpar ultima coluna
    ws.cell(row=row, column=147).value = None

print("   Colunas ajustadas")

# ============================================================
# 9. GUARDAR
# ============================================================
print(f"\n10. Guardando ficheiro: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)

print("\n" + "="*60)
print("PREPARACAO CONCLUIDA!")
print("="*60)
print(f"""
Ficheiro criado: {OUTPUT_FILE}

Resumo:
  - Espacos com walls: {espacos_com_walls}
  - Total walls criadas: {total_walls_criadas}
  - Tipos de janelas: {len(unique_windows)}
  - Tipos de paredes: {len(wall_types)}
  - Tipos de coberturas: {len(roof_types)}

Proximo passo:
  python excel_to_hap.py {OUTPUT_FILE} Modelo_RSECE.E3A Malhoa22_Final.E3A
""")
