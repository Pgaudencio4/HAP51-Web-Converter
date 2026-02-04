"""
Preparar HAP_Template_Malhoa22.xlsx para conversão
VERSÃO 2: Mantém os dados de Walls originais, apenas corrige formato

O ficheiro original já tem:
- 52 espaços com walls definidas (colunas 52+)
- Windows na sheet Windows com zona, orientação, U, SHGC, área, qty
- Walls e Roofs nas respectivas sheets

Este script:
1. Corrige o offset das colunas (remove "Infil Method")
2. Reformata as sheets Windows/Walls/Roofs para o conversor (headers na linha 3)
3. NÃO sobrescreve os dados de walls existentes
"""
import openpyxl

print("="*60)
print("PREPARAR MALHOA22 v2 - MANTER DADOS ORIGINAIS")
print("="*60)

INPUT_FILE = 'C:/Users/pedro/Downloads/HAP_Template_Malhoa22 (2).xlsx'
OUTPUT_FILE = 'Malhoa22_Pronto_v2.xlsx'

print(f"\n1. Carregando {INPUT_FILE}...")
wb = openpyxl.load_workbook(INPUT_FILE)

# ============================================================
# 1. LER DADOS DAS SHEETS AUXILIARES (Windows, Walls, Roofs)
# ============================================================
print("\n2. Lendo dados das sheets auxiliares...")

# Sheet Windows: Nome, Zona, Piso, Ori, U-Value, SHGC, Area, Qtd (row 2 = headers, row 3+ = dados)
ws_win_orig = wb['Windows']
windows_data = {}
for row in range(3, ws_win_orig.max_row + 1):
    name = ws_win_orig.cell(row=row, column=1).value
    if not name:
        continue

    u_value = ws_win_orig.cell(row=row, column=5).value or 5.75
    shgc = ws_win_orig.cell(row=row, column=6).value or 0.85
    area = ws_win_orig.cell(row=row, column=7).value or 0
    qty = ws_win_orig.cell(row=row, column=8).value or 1

    # Calcular dimensões
    area_per_win = area / qty if qty > 0 else area
    height = 1.5
    width = area_per_win / height if area_per_win > 0 else 1.0

    windows_data[name] = {
        'u': u_value,
        'shgc': shgc,
        'height': height,
        'width': round(width, 3),
    }

print(f"   Windows: {len(windows_data)}")

# Sheet Walls (row 4+ = dados)
ws_walls_orig = wb['Walls']
wall_types = {}
for row in range(4, ws_walls_orig.max_row + 1):
    name = ws_walls_orig.cell(row=row, column=1).value
    if name:
        wall_types[name] = {
            'u': ws_walls_orig.cell(row=row, column=2).value or 0.5,
            'weight': ws_walls_orig.cell(row=row, column=3).value or 200,
            'thickness': ws_walls_orig.cell(row=row, column=4).value or 0.3,
        }

print(f"   Wall types: {list(wall_types.keys())}")

# Sheet Roofs (row 4+ = dados)
ws_roofs_orig = wb['Roofs']
roof_types = {}
for row in range(4, ws_roofs_orig.max_row + 1):
    name = ws_roofs_orig.cell(row=row, column=1).value
    if name:
        roof_types[name] = {
            'u': ws_roofs_orig.cell(row=row, column=2).value or 0.4,
            'weight': ws_roofs_orig.cell(row=row, column=3).value or 300,
            'thickness': ws_roofs_orig.cell(row=row, column=4).value or 0.35,
        }

print(f"   Roof types: {list(roof_types.keys())}")

# ============================================================
# 2. CONTAR WALLS EXISTENTES NA SHEET ESPACOS
# ============================================================
print("\n3. Verificando walls existentes na sheet Espacos...")
ws = wb['Espacos']

espacos_com_walls = 0
total_walls = 0

for row in range(4, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if not name:
        continue

    tem_wall = False
    for w in range(8):
        col = 52 + w * 9
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value

        if exp or (area and area != 0):
            total_walls += 1
            tem_wall = True

    if tem_wall:
        espacos_com_walls += 1

print(f"   Espaços com walls: {espacos_com_walls}")
print(f"   Total walls: {total_walls}")

# ============================================================
# 3. PREENCHER WALL TYPE ONDE ESTÁ VAZIO
# ============================================================
print("\n4. Preenchendo Wall Types vazios...")

DEFAULT_WALL_TYPE = list(wall_types.keys())[0] if wall_types else 'Parede Exterior'
walls_preenchidas = 0

for row in range(4, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if not name:
        continue

    for w in range(8):
        col = 52 + w * 9
        exp = ws.cell(row=row, column=col).value
        area = ws.cell(row=row, column=col+1).value
        wtype = ws.cell(row=row, column=col+2).value

        if (exp or (area and area != 0)) and not wtype:
            ws.cell(row=row, column=col+2, value=DEFAULT_WALL_TYPE)
            walls_preenchidas += 1

print(f"   Wall types preenchidos: {walls_preenchidas}")

# ============================================================
# 4. CORRIGIR OFFSET DAS COLUNAS (remover Infil Method col 23)
# ============================================================
print("\n5. Corrigindo offset das colunas...")

# Desagrupar células merged
merged_ranges = list(ws.merged_cells.ranges)
for merged_range in merged_ranges:
    ws.unmerge_cells(str(merged_range))

# Shift: mover colunas 24+ para 23+ (elimina coluna 23 = Infil Method)
for row in range(1, ws.max_row + 1):
    for col in range(23, 150):
        ws.cell(row=row, column=col).value = ws.cell(row=row, column=col+1).value
    ws.cell(row=row, column=150).value = None

print("   Colunas ajustadas (Infil Method removido)")

# ============================================================
# 5. RECRIAR SHEET WINDOWS NO FORMATO DO CONVERSOR
# ============================================================
print("\n6. Recriando sheet Windows...")

if 'Windows' in wb.sheetnames:
    del wb['Windows']
ws_win_new = wb.create_sheet('Windows')

# Headers nas linhas 1-3 (conversor lê a partir da linha 4)
ws_win_new.cell(row=1, column=1, value='WINDOWS')
ws_win_new.cell(row=2, column=1, value='')
ws_win_new.cell(row=3, column=1, value='Nome')
ws_win_new.cell(row=3, column=2, value='U-Value')
ws_win_new.cell(row=3, column=3, value='SHGC')
ws_win_new.cell(row=3, column=4, value='Altura')
ws_win_new.cell(row=3, column=5, value='Largura')

row = 4
for name, wdata in windows_data.items():
    ws_win_new.cell(row=row, column=1, value=name)
    ws_win_new.cell(row=row, column=2, value=round(wdata['u'], 2))
    ws_win_new.cell(row=row, column=3, value=round(wdata['shgc'], 2))
    ws_win_new.cell(row=row, column=4, value=wdata['height'])
    ws_win_new.cell(row=row, column=5, value=wdata['width'])
    row += 1

print(f"   Windows: {len(windows_data)}")

# ============================================================
# 6. RECRIAR SHEET WALLS NO FORMATO DO CONVERSOR
# ============================================================
print("\n7. Recriando sheet Walls...")

if 'Walls' in wb.sheetnames:
    del wb['Walls']
ws_walls_new = wb.create_sheet('Walls')

ws_walls_new.cell(row=1, column=1, value='WALLS')
ws_walls_new.cell(row=2, column=1, value='')
ws_walls_new.cell(row=3, column=1, value='Nome')
ws_walls_new.cell(row=3, column=2, value='U-Value')
ws_walls_new.cell(row=3, column=3, value='Peso')
ws_walls_new.cell(row=3, column=4, value='Espessura')

row = 4
for name, wdata in wall_types.items():
    ws_walls_new.cell(row=row, column=1, value=name)
    ws_walls_new.cell(row=row, column=2, value=wdata['u'])
    ws_walls_new.cell(row=row, column=3, value=wdata['weight'])
    ws_walls_new.cell(row=row, column=4, value=wdata['thickness'])
    row += 1

print(f"   Wall types: {len(wall_types)}")

# ============================================================
# 7. RECRIAR SHEET ROOFS NO FORMATO DO CONVERSOR
# ============================================================
print("\n8. Recriando sheet Roofs...")

if 'Roofs' in wb.sheetnames:
    del wb['Roofs']
ws_roofs_new = wb.create_sheet('Roofs')

ws_roofs_new.cell(row=1, column=1, value='ROOFS')
ws_roofs_new.cell(row=2, column=1, value='')
ws_roofs_new.cell(row=3, column=1, value='Nome')
ws_roofs_new.cell(row=3, column=2, value='U-Value')
ws_roofs_new.cell(row=3, column=3, value='Peso')
ws_roofs_new.cell(row=3, column=4, value='Espessura')

row = 4
for name, rdata in roof_types.items():
    ws_roofs_new.cell(row=row, column=1, value=name)
    ws_roofs_new.cell(row=row, column=2, value=rdata['u'])
    ws_roofs_new.cell(row=row, column=3, value=rdata['weight'])
    ws_roofs_new.cell(row=row, column=4, value=rdata['thickness'])
    row += 1

print(f"   Roof types: {len(roof_types)}")

# ============================================================
# 8. GUARDAR
# ============================================================
print(f"\n9. Guardando ficheiro: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)

print("\n" + "="*60)
print("PREPARAÇÃO CONCLUÍDA!")
print("="*60)
print(f"""
Ficheiro criado: {OUTPUT_FILE}

Dados mantidos do original:
  - Espaços com walls: {espacos_com_walls}
  - Total walls: {total_walls}
  - Wall types preenchidos: {walls_preenchidas}

Tipos definidos:
  - Windows: {len(windows_data)}
  - Wall types: {len(wall_types)}
  - Roof types: {len(roof_types)}

Próximo passo:
  python excel_to_hap.py {OUTPUT_FILE} Modelo_RSECE.E3A Malhoa22_v2.E3A
""")
